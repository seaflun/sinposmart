# -*- coding: utf-8 -*-
from __future__ import annotations

import os
import json
import subprocess
import sys
import tempfile
import threading
import unittest
import zipfile
from datetime import date
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch


PROJECT_ROOT = Path(__file__).resolve().parents[1]
PACKAGE_ROOT = PROJECT_ROOT / "WinPython_公務電腦使用包"
if str(PACKAGE_ROOT) not in sys.path:
    sys.path.insert(0, str(PACKAGE_ROOT))

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")


def run_isolated_python(
    command: str,
    *,
    extra_env: dict[str, str] | None = None,
) -> tuple[int, str]:
    """Run an import-isolation probe without relying on pytest's pipe handles."""
    environment = {**os.environ, "QT_QPA_PLATFORM": "offscreen"}
    environment.update(extra_env or {})
    with tempfile.TemporaryFile(mode="w+b") as output:
        result = subprocess.run(
            [sys.executable, "-c", command],
            cwd=PACKAGE_ROOT,
            env=environment,
            stdin=subprocess.DEVNULL,
            stdout=output,
            stderr=subprocess.STDOUT,
            timeout=30,
            check=False,
        )
        output.seek(0)
        return result.returncode, output.read().decode("utf-8", errors="replace")


class SessionStateTests(unittest.TestCase):
    def test_login_attempt_is_single_flight_and_accepts_current_result(self) -> None:
        from app_core.session import LoginSession, SessionState

        state = SessionState()

        attempt_id = state.begin_login()

        self.assertEqual(attempt_id, 1)
        self.assertIsNone(state.begin_login())
        self.assertTrue(state.login_running)

        session = LoginSession(actor_no="10", user_id="user10", password="secret", verified=True)
        self.assertTrue(state.complete_login(attempt_id, session))
        self.assertFalse(state.login_running)
        self.assertIs(state.session, session)

    def test_timeout_invalidates_late_worker_result(self) -> None:
        from app_core.session import LoginSession, SessionState

        state = SessionState()
        attempt_id = state.begin_login()

        self.assertTrue(state.timeout_login(attempt_id))
        self.assertEqual(state.attempt_id, 2)
        self.assertFalse(
            state.complete_login(
                attempt_id,
                LoginSession(actor_no="10", user_id="user10", password="secret", verified=True),
            )
        )
        self.assertIsNone(state.session)

    def test_failure_and_logout_clear_authenticated_session(self) -> None:
        from app_core.session import LoginSession, SessionState

        state = SessionState()
        attempt_id = state.begin_login()
        self.assertTrue(state.fail_login(attempt_id))
        self.assertIsNone(state.session)

        next_attempt_id = state.begin_login()
        session = LoginSession(actor_no="10", user_id="user10", password="secret", verified=True)
        self.assertTrue(state.complete_login(next_attempt_id, session))
        self.assertIs(state.clear_session(), session)
        self.assertIsNone(state.session)


class LoginVerifierTests(unittest.TestCase):
    class SwitchTo:
        def default_content(self) -> None:
            pass

        def frame(self, _frame) -> None:
            pass

    class Driver:
        def __init__(self) -> None:
            self.switch_to = LoginVerifierTests.SwitchTo()

        def find_elements(self, *_args) -> list[object]:
            return []

        def execute_script(self, _script: str) -> str:
            return "勤務系統\n測試員，您好"

    def test_identity_javascript_keeps_newlines_escaped(self) -> None:
        from app_core.login_verifier import page_identity_hint_text, page_identity_text

        scripts: list[str] = []

        class RecordingDriver:
            def execute_script(self, script: str) -> str:
                scripts.append(script)
                return ""

        driver = RecordingDriver()
        page_identity_text(driver)
        page_identity_hint_text(driver)

        self.assertEqual(len(scripts), 2)
        self.assertTrue(all("join('\\n')" in script for script in scripts))
        self.assertTrue(all("join('\n')" not in script for script in scripts))

    def test_qml_login_driver_uses_the_private_sinposmart_profile(self) -> None:
        from unittest.mock import patch

        from app_core.login_verifier import create_login_webdriver

        driver = object()
        options = SimpleNamespace(
            arguments=["--headless=new", "--window-size=1280,900", "--window-position=-32000,-32000"]
        )
        with patch("duty_rehearsal.build_driver", return_value=driver) as build_driver:
            result = create_login_webdriver(options)

        self.assertIs(result, driver)
        build_driver.assert_called_once_with(
            headless=True,
            option_arguments=("--window-position=-32000,-32000",),
        )

    def test_verifier_owns_one_driver_and_resolves_greeting_name(self) -> None:
        from app_core.login_verifier import LoginVerifier

        driver = self.Driver()
        events: list[str] = []
        verifier = LoginVerifier(
            options_factory=lambda: "options",
            driver_factory=lambda options: events.append(f"create:{options}") or driver,
            configure_driver=lambda value: events.append("configure") if value is driver else None,
            login_function=lambda value, user_id, password: events.append(f"login:{user_id}:{password}") if value is driver else None,
            driver_cleanup=lambda value: events.append("cleanup") if value is driver else None,
        )

        result = verifier.verify(
            typed_actor_no="",
            user_id="user10",
            password="secret",
            actor_no_from_user_id=lambda _user_id: "",
            actor_no_from_name=lambda name: "10" if name == "測試員" else "",
            staff={},
        )

        self.assertEqual(result.actor_no, "10")
        self.assertEqual(result.actor_name, "測試員")
        self.assertEqual(events, ["create:options", "configure", "login:user10:secret", "cleanup"])

    def test_identity_matches_spaced_greeting_against_schedule_staff(self) -> None:
        from app_core.login_verifier import identify_logged_in_actor

        class SpacedGreetingDriver(self.Driver):
            def execute_script(self, _script: str) -> str:
                return "勤務系統\n測 試 員 ， 您好"

        actor_no, actor_name = identify_logged_in_actor(
            SpacedGreetingDriver(),
            lambda _name: "",
            {"10": {"name": "測試員"}, "11": {"name": "其他人"}},
        )

        self.assertEqual(actor_no, "10")
        self.assertEqual(actor_name, "測試員")

    def test_identity_matches_unique_staff_name_outside_duty_tables(self) -> None:
        from app_core.login_verifier import identify_logged_in_actor

        class HeaderIdentityDriver(self.Driver):
            def execute_script(self, script: str) -> str:
                if "closest('table')" in script:
                    return "登入人員\n測 試 員"
                return "勤務表\n測試員\n其他人"

        actor_no, actor_name = identify_logged_in_actor(
            HeaderIdentityDriver(),
            lambda _name: "",
            {"10": {"name": "測試員"}, "11": {"name": "其他人"}},
        )

        self.assertEqual(actor_no, "10")
        self.assertEqual(actor_name, "測試員")

    def test_verifier_rejects_actor_mismatch_and_still_cleans_driver(self) -> None:
        from app_core.login_verifier import LoginVerificationError, LoginVerifier

        driver = self.Driver()
        cleaned: list[object] = []
        verifier = LoginVerifier(
            options_factory=lambda: object(),
            driver_factory=lambda _options: driver,
            configure_driver=lambda _driver: None,
            login_function=lambda _driver, _user_id, _password: None,
            driver_cleanup=cleaned.append,
        )

        with self.assertRaises(LoginVerificationError):
            verifier.verify(
                typed_actor_no="11",
                user_id="user10",
                password="secret",
                actor_no_from_user_id=lambda _user_id: "",
                actor_no_from_name=lambda _name: "10",
                staff={},
            )

        self.assertEqual(cleaned, [driver])

    def test_verifier_queries_actor_no_after_account_login(self) -> None:
        from app_core.login_verifier import LoginVerifier

        driver = self.Driver()
        events: list[str] = []
        verifier = LoginVerifier(
            options_factory=lambda: "options",
            driver_factory=lambda _options: driver,
            configure_driver=lambda _driver: events.append("configure"),
            login_function=lambda _driver, _user_id, _password: events.append("login"),
            actor_no_query=lambda _driver, actor_name: events.append(f"query:{actor_name}") or "10",
            driver_cleanup=lambda _driver: events.append("cleanup"),
        )

        result = verifier.verify(
            typed_actor_no="",
            user_id="user10",
            password="secret",
            actor_no_from_user_id=lambda _user_id: "99",
            actor_no_from_name=lambda _name: "98",
            staff={},
        )

        self.assertEqual(result.actor_no, "10")
        self.assertEqual(result.actor_name, "測試員")
        self.assertEqual(events, ["configure", "login", "query:測試員", "cleanup"])

    def test_verifier_can_defer_actor_resolution_to_existing_schedule_capture(self) -> None:
        from app_core.login_verifier import LoginVerifier

        driver = self.Driver()
        events: list[str] = []
        verifier = LoginVerifier(
            options_factory=lambda: "options",
            driver_factory=lambda _options: driver,
            configure_driver=lambda _driver: events.append("configure"),
            login_function=lambda _driver, _user_id, _password: events.append("login"),
            actor_no_query=lambda _driver, _actor_name: events.append("duplicate-query") or "10",
            defer_actor_resolution=True,
            driver_cleanup=lambda _driver: events.append("cleanup"),
        )

        result = verifier.verify(
            typed_actor_no="",
            user_id="user10",
            password="secret",
            actor_no_from_user_id=lambda _user_id: "99",
            actor_no_from_name=lambda _name: "98",
            staff={},
        )

        self.assertEqual(result.actor_no, "")
        self.assertEqual(result.actor_name, "測試員")
        self.assertEqual(result.warning, "登入成功，正在查詢勤務資料…")
        self.assertEqual(events, ["configure", "login", "cleanup"])

    def test_current_duty_actor_query_matches_authenticated_name(self) -> None:
        from unittest.mock import ANY, patch

        from app_core.login_verifier import query_current_duty_actor_no

        duty_sheet = SimpleNamespace(
            staff={
                "09": {"name": "其他人"},
                "10": {"name": "測 試 員"},
            }
        )
        with (
            patch("app_core.schedule_repository.business_roc_date", return_value="1150729"),
            patch("duty_rehearsal.query_duty_sheet", return_value=duty_sheet) as query,
        ):
            actor_no = query_current_duty_actor_no(object(), "測試員")

        self.assertEqual(actor_no, "10")
        query.assert_called_once_with(ANY, "1150729")

    def test_site_person_query_opens_read_only_insert_and_matches_login_id(self) -> None:
        from unittest.mock import patch

        import duty_rehearsal

        class Driver:
            def __init__(self) -> None:
                self.identity_reads = 0

            def execute_script(self, script: str, *_args):
                if "loginAliases" in script:
                    self.identity_reads += 1
                    return [] if self.identity_reads == 1 else ["測試員"]
                if "_txtDATE" in script and "_areDescription" in script:
                    return False
                raise AssertionError("unexpected website query")

        driver = Driver()
        with (
            patch.object(duty_rehearsal, "ensure_ap", return_value=True) as ensure_ap,
            patch.object(duty_rehearsal, "click_insert_control", return_value={"ok": True}) as click_insert,
            patch.object(duty_rehearsal.time, "sleep"),
        ):
            name = duty_rehearsal.query_authenticated_person_name(driver, "tyfd01510")

        self.assertEqual(name, "測試員")
        ensure_ap.assert_called_once_with(driver, duty_rehearsal.WORK_LOG_AP)
        click_insert.assert_called_once_with(driver)

    def test_actor_query_failure_does_not_turn_successful_login_into_login_failure(self) -> None:
        from app_core.login_verifier import LoginVerifier

        driver = self.Driver()
        verifier = LoginVerifier(
            options_factory=lambda: "options",
            driver_factory=lambda _options: driver,
            configure_driver=lambda _driver: None,
            login_function=lambda _driver, _user_id, _password: None,
            actor_no_query=lambda _driver, _actor_name: (_ for _ in ()).throw(
                RuntimeError("勤務表讀取失敗")
            ),
            allow_post_login_lookup_warning=True,
            driver_cleanup=lambda _driver: None,
        )

        result = verifier.verify(
            typed_actor_no="",
            user_id="user10",
            password="secret",
            actor_no_from_user_id=lambda _user_id: "",
            actor_no_from_name=lambda _name: "",
            staff={},
        )

        self.assertEqual(result.actor_no, "")
        self.assertEqual(result.actor_name, "測試員")
        self.assertEqual(result.warning, "登入成功，但勤務番號查詢失敗；請稍後重新整理勤務資料。")

    def test_read_only_acceptance_can_continue_when_identity_lookup_fails(self) -> None:
        from app_core.login_verifier import LoginVerifier

        class IdentityFailureDriver(self.Driver):
            def execute_script(self, _script: str) -> str:
                raise RuntimeError("登入後頁面結構無法辨識")

        verifier = LoginVerifier(
            options_factory=lambda: "options",
            driver_factory=lambda _options: IdentityFailureDriver(),
            configure_driver=lambda _driver: None,
            login_function=lambda _driver, _user_id, _password: None,
            actor_no_query=lambda _driver, _actor_name: "10",
            allow_post_login_lookup_warning=True,
            driver_cleanup=lambda _driver: None,
        )

        result = verifier.verify(
            typed_actor_no="",
            user_id="user10",
            password="secret",
            actor_no_from_user_id=lambda _user_id: "",
            actor_no_from_name=lambda _name: "",
            staff={},
        )

        self.assertEqual(result.actor_no, "")
        self.assertEqual(result.actor_name, "")
        self.assertEqual(result.warning, "登入成功，但無法辨識登入人員；勤務番號尚未取得。")


class DutyTaskProjectionTests(unittest.TestCase):
    def test_projection_filters_actor_keeps_previous_handoff_and_sorts_rows(self) -> None:
        from datetime import datetime

        from app_core.duty_task_projection import DutyTaskProjectionState, project_duty_tasks

        actions = [
            {
                "kind": "work_log",
                "time": "09:00",
                "actor": "10",
                "target": "10",
                "source": "值班交接",
                "fields": {"工作時間": "09:00", "勤務項目": "巡邏", "服勤人員": ["10"]},
            },
            {
                "kind": "entry_log",
                "time": "07:50",
                "actor": "09",
                "target": "10",
                "source": "值班交接",
                "fields": {"登打時間": "07:50", "出或入": "值班", "領用事由及地點": "交接"},
            },
            {
                "kind": "work_log",
                "time": "08:30",
                "actor": "11",
                "target": "11",
                "fields": {"工作時間": "08:30", "勤務項目": "清點"},
            },
        ]
        state = DutyTaskProjectionState(
            actor_no="10",
            target_roc_date="1150729",
            staff={"09": {"name": "前班"}, "10": {"name": "本班"}},
        )

        rows = project_duty_tasks(actions, state, now=datetime(2026, 7, 29, 8, 0))

        self.assertEqual([row["taskIndex"] for row in rows], [1, 0])
        self.assertEqual(rows[0]["statusText"], "前班手動")
        self.assertEqual(rows[0]["systemText"], "出入")
        self.assertEqual(rows[1]["detailText"], "巡邏")
        self.assertEqual(rows[1]["peopleText"], "10 本班")
        self.assertEqual(rows[1]["statusText"], "等待")

    def test_projection_formats_cross_day_time_and_status_precedence(self) -> None:
        from datetime import datetime

        from app_core.duty_task_projection import DutyTaskProjectionState, project_duty_tasks

        action = {
            "kind": "work_log",
            "time": "25:15",
            "actor": "10",
            "target": "10",
            "fields": {"工作時間": "25:15", "勤務項目": "夜間勤務"},
        }
        state = DutyTaskProjectionState(
            actor_no="10",
            target_roc_date="1150729",
            comparisons={0: {"group": "done", "compare": "已存在(時間不同)"}},
            submitting_indices=frozenset({0}),
        )

        rows = project_duty_tasks([action], state, now=datetime(2026, 7, 30, 2, 0))

        self.assertEqual(rows[0]["timeText"], "30日 01:15")
        self.assertEqual(rows[0]["statusText"], "正在登打")
        self.assertEqual(rows[0]["statusTone"], "running")

    def test_next_task_text_matches_legacy_candidate_and_previous_duty_fallback(self) -> None:
        from datetime import datetime

        from app_core.duty_task_projection import DutyTaskProjectionState, next_duty_task_text

        actions = [
            {
                "kind": "work_log",
                "time": "09:00",
                "actor": "10",
                "target": "10",
                "source": "值班交接",
                "fields": {"工作時間": "09:00", "勤務項目": "巡邏"},
            },
            {
                "kind": "entry_log",
                "time": "07:50",
                "actor": "09",
                "target": "10",
                "source": "值班交接",
                "fields": {"登打時間": "07:50", "出或入": "值班", "領用事由及地點": "交接"},
            },
        ]
        state = DutyTaskProjectionState(actor_no="10", target_roc_date="1150729")

        self.assertEqual(
            next_duty_task_text(actions, state, now=datetime(2026, 7, 29, 8, 0)),
            "09:00  巡邏，約 60 分鐘後",
        )

        completed_state = DutyTaskProjectionState(
            actor_no="10",
            target_roc_date="1150729",
            comparisons={0: {"group": "done", "compare": "已存在"}},
        )
        self.assertEqual(
            next_duty_task_text(actions, completed_state, now=datetime(2026, 7, 29, 8, 0)),
            "前一班尚有 1 筆待手動處理",
        )
        self.assertEqual(
            next_duty_task_text(actions, DutyTaskProjectionState(actor_no="", target_roc_date="1150729")),
            "下一項任務：-",
        )

    def test_comparison_marks_existing_work_row_done(self) -> None:
        from app_core.duty_task_projection import build_schedule_comparisons

        action = {
            "kind": "work_log",
            "time": "08:00",
            "actor": "10",
            "target": "10",
            "fields": {"工作時間": "08:00", "勤務項目": "巡邏"},
        }
        data = {
            "target_date": "1150728",
            "today": {"staff": {"10": {"name": "本班"}}},
            "actions": [action],
        }
        comparison_data = {
            "1150728": {"visible_work_rows": [["115/07/28", "08:00", "巡邏"]]}
        }

        comparisons = build_schedule_comparisons(data, [action], comparison_data)

        self.assertEqual(comparisons[0]["group"], "done")
        self.assertEqual(comparisons[0]["compare"], "已存在")

    def test_comparison_prefers_existing_scheduled_checkout_over_future_time(self) -> None:
        from unittest.mock import patch

        from app_core.duty_task_projection import build_schedule_comparisons

        action = {
            "kind": "entry_log",
            "time": "08:05",
            "actor": "10",
            "target": "10",
            "fields": {"出或入": "值退", "系統寫入時間": "08:05"},
        }
        data = {
            "target_date": "1150807",
            "today": {"staff": {"10": {"name": "測試員"}}},
            "actions": [action],
        }
        comparison_data = {
            "1150807": {"visible_entry_rows": [["115/08/07", "08:05", "-", "測試員", "值退"]]}
        }

        with patch("app_core.duty_task_projection.is_future_action", return_value=True):
            comparisons = build_schedule_comparisons(data, [action], comparison_data)

        self.assertEqual(comparisons[0]["group"], "done")
        self.assertEqual(comparisons[0]["compare"], "已存在")

    def test_audit_projection_includes_existing_rows_and_raw_action_detail(self) -> None:
        from app_core.duty_task_projection import project_audit_tasks

        action = {
            "kind": "work_log",
            "time": "08:00",
            "actor": "10",
            "target": "10",
            "fields": {"工作時間": "08:00", "勤務項目": "巡邏"},
        }
        rows = project_audit_tasks(
            [action],
            target_roc_date="1150729",
            staff={"10": {"name": "本班"}},
            comparisons={
                0: {
                    "group": "done",
                    "compare": "已存在",
                    "matched": ["115/07/29 08:00 巡邏"],
                }
            },
        )

        detail = rows[0]["fullDetailText"]
        self.assertIn("比對：已存在", detail)
        self.assertIn("系統既有紀錄：", detail)
        self.assertIn("115/07/29 08:00 巡邏", detail)
        self.assertIn('"kind": "work_log"', detail)

    def test_due_selection_applies_actor_status_pause_retry_and_time_guards(self) -> None:
        from datetime import datetime, timedelta

        from app_core.duty_task_projection import DueTaskSelectionState, select_due_task_indices

        now = datetime(2026, 7, 29, 9, 0)
        actions = [
            {"kind": "work_log", "time": "08:00", "actor": "10", "source": "在隊訓練"},
            {"kind": "entry_log", "time": "08:10", "actor": "10", "fields": {"出或入": "值班"}},
            {"kind": "work_log", "time": "08:20", "actor": "10", "source": "在隊訓練"},
            {"kind": "work_log", "time": "08:30", "actor": "10", "source": "在隊訓練"},
            {"kind": "work_log", "time": "09:30", "actor": "10", "source": "在隊訓練"},
            {"kind": "work_log", "time": "08:40", "actor": "11", "source": "在隊訓練"},
        ]
        state = DueTaskSelectionState(
            actor_no="10",
            target_roc_date="1150729",
            comparisons={1: {"group": "done"}},
            retry_after={3: now + timedelta(minutes=1)},
        )

        self.assertEqual(select_due_task_indices(actions, state, now=now), [0, 2])

    def test_due_selection_starts_0805_checkout_at_0800(self) -> None:
        from datetime import datetime

        from app_core.duty_task_projection import DueTaskSelectionState, select_due_task_indices

        actions = [
            {
                "kind": "entry_log",
                "time": "08:00",
                "actor": "10",
                "target": "10",
                "fields": {
                    "登打時間": "08:00",
                    "系統寫入時間": "08:05",
                    "出或入": "出",
                    "領用事由及地點": "退勤",
                },
            }
        ]
        state = DueTaskSelectionState(actor_no="10", target_roc_date="1150807")

        self.assertEqual(select_due_task_indices(actions, state, now=datetime(2026, 8, 7, 7, 59)), [])
        self.assertEqual(select_due_task_indices(actions, state, now=datetime(2026, 8, 7, 8, 0)), [0])

    def test_only_known_work_tasks_can_auto_submit_without_generic_pause_controls(self) -> None:
        from datetime import datetime

        from app_core.duty_task_projection import DueTaskSelectionState, select_due_task_indices

        actions = [
            {"kind": "work_log", "time": "08:00", "actor": "10", "source": "在隊訓練"},
            {"kind": "work_log", "time": "08:00", "actor": "10", "source": "無線電試話"},
            {"kind": "work_log", "time": "08:00", "actor": "10", "source": "值班交接"},
            {"kind": "work_log", "time": "08:00", "actor": "10", "source": "未分類工作"},
        ]

        due = select_due_task_indices(
            actions,
            DueTaskSelectionState(actor_no="10", target_roc_date="1150807"),
            now=datetime(2026, 8, 7, 8, 0),
        )

        self.assertEqual(due, [0, 1, 2])

    def test_1800_handoff_keeps_outgoing_incoming_and_work_for_outgoing_actor(self) -> None:
        from datetime import datetime

        from app_core.duty_task_projection import DueTaskSelectionState, select_due_task_indices

        actions = [
            {
                "kind": "entry_log",
                "time": "18:00",
                "actor": "17",
                "target": "17",
                "source": "值班交接",
                "fields": {"出或入": "值退", "領用事由及地點": "值退"},
            },
            {
                "kind": "entry_log",
                "time": "18:00",
                "actor": "17",
                "target": "5",
                "source": "值班交接",
                "fields": {"出或入": "值班", "領用事由及地點": "值班"},
            },
            {
                "kind": "work_log",
                "time": "18:00",
                "actor": "17",
                "target": "17",
                "source": "值班交接",
                "fields": {"工作概述": "交接工作紀錄"},
            },
        ]

        due = select_due_task_indices(
            actions,
            DueTaskSelectionState(actor_no="17", target_roc_date="1150806"),
            now=datetime(2026, 8, 6, 18, 0),
        )

        self.assertEqual(due, [0, 1, 2])

    def test_handoff_adjustment_does_not_block_the_next_person_value_entry(self) -> None:
        from compare_rehearsal_records import is_possible_handoff_adjustment

        incoming_action = {
            "kind": "entry_log",
            "time": "18:00",
            "target": "5",
            "source": "值班交接",
            "fields": {"系統寫入時間": "18:00", "出或入": "值班"},
        }
        rows = ["115/08/06 18:00 | 勤務 | 17番隊員 | 值退 | 值退"]
        staff = {
            "17": {"name": "17番隊員"},
            "5": {"name": "5番隊員"},
        }

        self.assertFalse(
            is_possible_handoff_adjustment(rows, "1150806", staff, incoming_action)
        )
        same_person_action = {**incoming_action, "target": "17"}
        self.assertTrue(
            is_possible_handoff_adjustment(rows, "1150806", staff, same_person_action)
        )

    def test_submit_target_date_overrides_scheduled_action_date(self) -> None:
        from app_core.duty_task_projection import action_target_roc_date

        action = {
            "kind": "work_log",
            "time": "01:02",
            "date_offset": 0,
            "submit_target_date": "1150730",
        }

        self.assertEqual(action_target_roc_date(action, "1150729"), "1150730")


class DutySheetServiceTests(unittest.TestCase):
    def test_loading_execution_core_does_not_import_legacy_tk_ui(self) -> None:
        command = (
            "import sys; "
            "from pathlib import Path; "
            f"package_root = Path({str(PACKAGE_ROOT)!r}); "
            "sys.path.insert(0, str(package_root)); "
            "from app_core.duty_sheet_service import load_legacy_module; "
            "load_legacy_module(package_root / 'duty_sheet_legacy'); "
            "forbidden = {'tkinter', 'customtkinter', 'tkcalendar'}; "
            "raise SystemExit(1 if forbidden.intersection(sys.modules) else 0)"
        )

        return_code, output = run_isolated_python(command)

        self.assertEqual(return_code, 0, output)

    def test_actual_legacy_log_status_forwards_to_runtime_callback_without_tk_root(self) -> None:
        from app_core.duty_sheet_service import load_legacy_module

        module = load_legacy_module(PACKAGE_ROOT / "duty_sheet_legacy")
        messages: list[str] = []
        module._runtime_status_callback = messages.append
        try:
            module.log_status("callback-check")
        finally:
            module._runtime_status_callback = None

        self.assertEqual(messages, ["callback-check"])
        self.assertFalse(hasattr(module, "root"))

    def test_defaults_and_validation_use_non_secret_config_fields(self) -> None:
        from datetime import datetime

        from app_core.duty_sheet_service import DutySheetRequest, DutySheetService

        with tempfile.TemporaryDirectory() as temp_dir:
            package_root = Path(temp_dir)
            project_dir = package_root / "duty_sheet_legacy"
            project_dir.mkdir()
            workbook = project_dir / "duty.xlsm"
            workbook.write_bytes(b"test")
            (project_dir / "config.json").write_text(
                json.dumps(
                    {
                        "login": {"user_id": "must-not-surface", "user_pwd": "secret"},
                        "last_selection": {
                            "workbook_path": "duty.xlsm",
                            "attack": "A",
                            "stop": "S",
                            "amb1": "M1",
                            "amb2": "M2",
                        },
                        "car_options": {"attack": ["A"], "stop": ["S"], "amb": ["M1", "M2"]},
                        "notification": {"enabled": True},
                    }
                ),
                encoding="utf-8",
            )
            service = DutySheetService(
                package_root,
                module_loader=lambda _project_dir: self.fail("saved display settings must not import the legacy GUI"),
            )

            defaults = service.load_defaults(datetime(2026, 7, 29, 12, 0))
            request = service.validate(
                DutySheetRequest(
                    "user10", "password", defaults.workbook_path, defaults.target_date,
                    defaults.attack, defaults.stop, defaults.amb1, defaults.amb2,
                    defaults.notification_enabled,
                )
            )

            self.assertEqual(defaults.target_date, "2026/07/30")
            self.assertEqual(defaults.attack_options, ("A",))
            self.assertNotIn("secret", repr(defaults))
            self.assertNotIn("password", repr(request))
            self.assertEqual(request.workbook_path, str(workbook.resolve()))

    def test_defaults_match_legacy_merge_hidden_options_and_workbook_fallback(self) -> None:
        from datetime import datetime

        from app_core.duty_sheet_service import DutySheetService

        with tempfile.TemporaryDirectory() as temp_dir:
            package_root = Path(temp_dir)
            project_dir = package_root / "duty_sheet_legacy"
            project_dir.mkdir()
            fallback_workbook = project_dir / "legacy-duty.xlsm"
            fallback_workbook.write_bytes(b"test")
            legacy_config = {
                "last_selection": {
                    "workbook_path": "removed-duty.xlsm",
                    "attack": "新坡15/KES-5922",
                    "stop": "新坡11/KEC-2608",
                    "amb1": "新坡91/BGV-2310",
                    "amb2": "新坡95/BPE-5951",
                },
                "car_options": {
                    "attack": ["新坡15/KES-5922", "新坡16/981-S5"],
                    "stop": ["新坡11/KEC-2608"],
                    "amb": ["新坡91/BGV-2310", "新坡95/BPE-5951"],
                },
                "hidden_car_options": {"attack": [], "amb": ["新坡93/BSL-9230"]},
                "notification": {"enabled": True},
            }
            legacy = SimpleNamespace(load_config=lambda: legacy_config)
            service = DutySheetService(package_root, module_loader=lambda _path: legacy)

            defaults = service.load_defaults(datetime(2026, 7, 29, 12, 0))

            self.assertEqual(defaults.workbook_path, str(fallback_workbook.resolve()))
            self.assertEqual(defaults.attack, "新坡15/KES-5922")
            self.assertEqual(defaults.amb_options, ("新坡91/BGV-2310", "新坡95/BPE-5951"))
            self.assertTrue(defaults.notification_enabled)

    def test_vehicle_options_preserve_config_shape_and_hidden_list(self) -> None:
        from app_core.duty_sheet_service import DutySheetService

        with tempfile.TemporaryDirectory() as temp_dir:
            package_root = Path(temp_dir)
            project_dir = package_root / "duty_sheet_legacy"
            project_dir.mkdir()
            config_path = project_dir / "config.json"
            config_path.write_text(
                json.dumps(
                    {
                        "login": {"user_id": "saved", "user_pwd": "preserved"},
                        "last_selection": {
                            "attack": "新坡15/KES-5922",
                            "stop": "新坡11/KEC-2608",
                            "amb1": "新坡93/BSL-9230",
                            "amb2": "新坡91/BGV-2310",
                        },
                        "car_options": {
                            "attack": ["新坡15/KES-5922"],
                            "stop": ["新坡11/KEC-2608"],
                            "amb": ["新坡91/BGV-2310"],
                        },
                        "hidden_car_options": {"attack": [], "stop": [], "amb": ["新坡93/BSL-9230"]},
                        "notification": {"enabled": True},
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            service = DutySheetService(package_root)

            added = service.add_vehicle_option("amb", "新坡93", "BSL-9230")
            self.assertEqual(added, "新坡93/BSL-9230")
            after_add = json.loads(config_path.read_text(encoding="utf-8"))
            self.assertIn(added, after_add["car_options"]["amb"])
            self.assertNotIn(added, after_add["hidden_car_options"]["amb"])
            self.assertEqual(after_add["login"]["user_pwd"], "preserved")

            service.remove_vehicle_option("amb", added)
            after_remove = json.loads(config_path.read_text(encoding="utf-8"))
            self.assertNotIn(added, after_remove["car_options"]["amb"])
            self.assertIn(added, after_remove["hidden_car_options"]["amb"])
            self.assertEqual(after_remove["last_selection"]["amb1"], "新坡91/BGV-2310")

            relay = service.add_vehicle_option("stop", "F-01", "ABC-100")
            self.assertEqual(relay, "F-01/ABC-100")
            after_relay_add = json.loads(config_path.read_text(encoding="utf-8"))
            self.assertIn(relay, after_relay_add["car_options"]["stop"])

            service.remove_vehicle_option("stop", relay)
            after_relay_remove = json.loads(config_path.read_text(encoding="utf-8"))
            self.assertNotIn(relay, after_relay_remove["car_options"]["stop"])
            self.assertIn(relay, after_relay_remove["hidden_car_options"]["stop"])

    def test_execute_calls_legacy_engine_without_persisting_session_password(self) -> None:
        from app_core.duty_sheet_service import DutySheetRequest, DutySheetService

        with tempfile.TemporaryDirectory() as temp_dir:
            package_root = Path(temp_dir)
            project_dir = package_root / "duty_sheet_legacy"
            project_dir.mkdir()
            workbook = project_dir / "duty.xlsm"
            workbook.write_bytes(b"test")
            (project_dir / "sinposmart_1.py").write_text("# placeholder\n", encoding="utf-8")
            saved: list[dict] = []

            def start_automation(_uid, _pwd, _target, _excel, _cars, **kwargs):
                kwargs["status_callback"]("執行中")
                kwargs["success_callback"]("完成")
                return True

            legacy = SimpleNamespace(
                load_config=lambda: {
                    "login": {"user_id": "legacy", "user_pwd": "preserved"},
                    "notification": {"enabled": False},
                    "car_options": {},
                    "hidden_car_options": {},
                },
                save_config=lambda selection, **kwargs: saved.append({"selection": selection, **kwargs}),
                convert_to_minguo=lambda selected: "1150730",
                start_automation=start_automation,
                _runtime_status_callback=None,
            )
            service = DutySheetService(package_root, module_loader=lambda _path: legacy)
            progress: list[str] = []
            request = DutySheetRequest(
                "user10", "session-secret", str(workbook), "2026/07/30",
                "A", "S", "M1", "M2", False,
            )

            result = service.execute(request, status_callback=progress.append)

            self.assertEqual(result, "完成")
            self.assertEqual(progress, ["執行中"])
            self.assertEqual(saved[0]["login_settings"]["user_pwd"], "preserved")
            self.assertNotIn("session-secret", repr(saved))

    def test_browser_start_failure_preserves_safe_shared_driver_message(self) -> None:
        from app_core.duty_sheet_service import (
            DutySheetExecutionError,
            DutySheetRequest,
            DutySheetService,
        )

        class BrowserStartupError(RuntimeError):
            diagnostic_category = "startup_timeout"

            def __str__(self) -> str:
                return (
                    "SinpoSmart 專用瀏覽器啟動失敗，已自動清理暫存資料並重試。"
                    "一般 Chrome 不需關閉；若仍失敗請匯出問題包。"
                )

        def start_automation(*_args, **kwargs):
            kwargs["stage_callback"]("browser_start")
            raise BrowserStartupError()

        with tempfile.TemporaryDirectory() as temp_dir:
            package_root = Path(temp_dir)
            project_dir = package_root / "duty_sheet_legacy"
            project_dir.mkdir()
            workbook = project_dir / "duty.xlsm"
            workbook.write_bytes(b"test")
            (project_dir / "sinposmart_1.py").write_text("# placeholder\n", encoding="utf-8")
            legacy = SimpleNamespace(
                load_config=lambda: {
                    "login": {},
                    "notification": {"enabled": False},
                    "car_options": {},
                    "hidden_car_options": {},
                },
                save_config=lambda *_args, **_kwargs: None,
                convert_to_minguo=lambda _selected: "1150730",
                start_automation=start_automation,
                _runtime_status_callback=None,
            )
            service = DutySheetService(package_root, module_loader=lambda _path: legacy)
            request = DutySheetRequest(
                "user10", "session-secret", str(workbook), "2026/07/30",
                "A", "S", "M1", "M2", False,
            )

            with self.assertRaises(DutySheetExecutionError) as raised:
                service.execute(request)

        self.assertEqual(raised.exception.failure_stage, "browser_start")
        self.assertIn("專用瀏覽器啟動失敗", str(raised.exception))
        self.assertNotIn("session-secret", str(raised.exception))

    def test_existing_tool_cores_load_without_tk_ui(self) -> None:
        command = """
import importlib.util
import sys
from pathlib import Path

package_root = Path.cwd()
sys.path.insert(0, str(package_root))
for name, source_path in (
    ("duty_sheet", package_root / "duty_sheet_legacy" / "sinposmart_1.py"),
    ("rest_monthly", package_root / "rest_time_automation.py"),
):
    spec = importlib.util.spec_from_file_location(f"_qt_ui_boundary_{name}", source_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"cannot load {source_path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)

forbidden = {"tkinter", "customtkinter"}
raise SystemExit(1 if forbidden.intersection(sys.modules) else 0)
"""

        return_code, output = run_isolated_python(command)

        self.assertEqual(return_code, 0, output)


class DutySubmissionServiceTests(unittest.TestCase):
    def test_due_submission_skips_a_stale_fire_day_before_opening_browser(self) -> None:
        from datetime import datetime

        from app_core.duty_submission_service import DutySubmissionRequest, DutySubmissionService

        with tempfile.TemporaryDirectory() as temp_dir:
            driver_started: list[bool] = []
            automation = SimpleNamespace(
                build_driver=lambda *_args, **_kwargs: driver_started.append(True),
            )
            service = DutySubmissionService(
                Path(temp_dir),
                module_loader=lambda: automation,
                now_factory=lambda: datetime(2026, 7, 30, 8, 0),
            )
            data = {
                "target_date": "1150729",
                "actions": [{"kind": "entry_log", "time": "08:00", "actor": "10"}],
            }

            result = service.execute(DutySubmissionRequest("user10", "secret", 0, data))

            self.assertEqual(result.status, "skipped_stale_schedule")
            self.assertEqual(driver_started, [])
            self.assertTrue(result.result_path.is_file())

    def test_submit_checks_duplicates_fills_and_verifies_before_success(self) -> None:
        from datetime import datetime

        from app_core.duty_submission_service import DutySubmissionRequest, DutySubmissionService

        with tempfile.TemporaryDirectory() as temp_dir:
            events: list[str] = []
            query_count = 0

            def query_visible_table(_driver, ap_name, _target_date):
                nonlocal query_count
                query_count += 1
                events.append(f"query:{ap_name}")
                return [["submitted"]] if query_count > 1 else []

            automation = SimpleNamespace(
                WORK_LOG_AP="work",
                ENTRY_LOG_AP="entry",
                build_driver=lambda headless: events.append(f"driver:{headless}") or object(),
                login=lambda _driver, user_id, password: events.append(f"login:{user_id}:{len(password)}"),
                query_visible_table=query_visible_table,
                fill_work_log_form_for_test=lambda _driver, _action, _staff, _date, save: events.append(f"fill:{save}") or {"ok": True},
                fill_entry_log_form_for_test=lambda *_args, **_kwargs: {"ok": True},
                quit_driver=lambda _driver: events.append("quit"),
            )
            comparisons = iter(
                [
                    {0: {"compare": "未找到", "group": "todo", "matched": []}},
                    {0: {"compare": "已存在", "group": "done", "matched": ["submitted"]}},
                ]
            )
            service = DutySubmissionService(
                Path(temp_dir),
                module_loader=lambda: automation,
                now_factory=lambda: datetime(2026, 7, 29, 9, 0),
                comparison_builder=lambda *_args, **_kwargs: next(comparisons),
            )
            data = {
                "target_date": "1150729",
                "today": {"staff": {"10": {"name": "測試員"}}},
                "actions": [{"kind": "work_log", "time": "08:00", "actor": "10"}],
            }
            request = DutySubmissionRequest("user10", "session-secret", 0, data)
            progress: list[str] = []

            result = service.execute(request, status_callback=progress.append)

            self.assertEqual(result.status, "submitted")
            self.assertTrue(result.result_path.is_file())
            self.assertEqual(events, ["driver:True", "login:user10:14", "query:work", "fill:True", "query:work", "quit"])
            self.assertEqual(len(progress), 4)
            self.assertNotIn("session-secret", repr(request))
            self.assertNotIn("session-secret", result.result_path.read_text(encoding="utf-8"))

    def test_submit_retries_post_submit_verification_before_failing(self) -> None:
        from datetime import datetime

        from app_core.duty_submission_service import DutySubmissionRequest, DutySubmissionService

        with tempfile.TemporaryDirectory() as temp_dir:
            waits: list[float] = []
            comparisons = iter(
                [
                    {0: {"compare": "未找到", "group": "todo", "matched": []}},
                    {0: {"compare": "未找到", "group": "todo", "matched": []}},
                    {0: {"compare": "未找到", "group": "todo", "matched": []}},
                    {0: {"compare": "已存在", "group": "done", "matched": ["saved"]}},
                ]
            )
            automation = SimpleNamespace(
                WORK_LOG_AP="work",
                ENTRY_LOG_AP="entry",
                build_driver=lambda *_args, **_kwargs: object(),
                login=lambda *_args: None,
                query_visible_table=lambda *_args: [],
                fill_work_log_form_for_test=lambda *_args, **_kwargs: {"ok": True},
                fill_entry_log_form_for_test=lambda *_args, **_kwargs: {"ok": True},
                quit_driver=lambda *_args: None,
            )
            service = DutySubmissionService(
                Path(temp_dir),
                module_loader=lambda: automation,
                now_factory=lambda: datetime(2026, 8, 7, 8, 0),
                comparison_builder=lambda *_args, **_kwargs: next(comparisons),
                sleeper=waits.append,
            )
            data = {
                "target_date": "1150807",
                "today": {"staff": {"10": {"name": "測試員"}}},
                "actions": [
                    {
                        "kind": "work_log",
                        "time": "08:00",
                        "actor": "10",
                        "target": "10",
                        "fields": {"工作時間": "08:00", "勤務項目": "巡邏"},
                    }
                ],
            }

            result = service.execute(DutySubmissionRequest("user10", "secret", 0, data))

        self.assertEqual(result.status, "submitted")
        self.assertEqual(waits, [1.0, 1.0])

    def test_duplicate_result_skips_form_submission(self) -> None:
        from datetime import datetime

        from app_core.duty_submission_service import DutySubmissionRequest, DutySubmissionService

        with tempfile.TemporaryDirectory() as temp_dir:
            fills: list[bool] = []
            automation = SimpleNamespace(
                WORK_LOG_AP="work",
                ENTRY_LOG_AP="entry",
                build_driver=lambda headless: object(),
                login=lambda *_args: None,
                query_visible_table=lambda *_args: [["existing"]],
                fill_work_log_form_for_test=lambda *_args, **_kwargs: fills.append(True),
                fill_entry_log_form_for_test=lambda *_args, **_kwargs: fills.append(True),
                quit_driver=lambda _driver: None,
            )
            service = DutySubmissionService(
                Path(temp_dir),
                module_loader=lambda: automation,
                now_factory=lambda: datetime(2026, 7, 29, 8, 1),
                comparison_builder=lambda *_args, **_kwargs: {
                    0: {"compare": "已存在", "group": "done", "matched": ["existing"]}
                },
            )
            data = {
                "target_date": "1150729",
                "actions": [{"kind": "entry_log", "time": "08:00", "actor": "10"}],
            }

            result = service.execute(DutySubmissionRequest("user10", "secret", 0, data))

            self.assertEqual(result.status, "skipped_duplicate")
            self.assertEqual(fills, [])

    def test_handoff_action_is_refreshed_before_form_fill(self) -> None:
        from datetime import date, datetime

        from app_core.duty_submission_service import DutySubmissionRequest, DutySubmissionService

        with tempfile.TemporaryDirectory() as temp_dir:
            filled_descriptions: list[str] = []
            automation = SimpleNamespace(
                WORK_LOG_AP="work",
                ENTRY_LOG_AP="entry",
                build_driver=lambda headless: object(),
                login=lambda *_args: None,
                query_visible_table=lambda *_args: [],
                fill_work_log_form_for_test=lambda _driver, action, _staff, _date, save: filled_descriptions.append(action["fields"]["工作概述"]) or {},
                fill_entry_log_form_for_test=lambda *_args, **_kwargs: {},
                quit_driver=lambda _driver: None,
                parse_roc_date=lambda _value: date(2026, 7, 29),
                roc_date=lambda value: f"{value.year - 1911:03d}{value.month:02d}{value.day:02d}",
                query_duty_sheet=lambda *_args: object(),
                query_cases=lambda *_args: [],
                planned_actions=lambda *_args: [
                    {
                        "kind": "work_log",
                        "time": "08:00",
                        "actor": "10",
                        "source": "值班交接",
                        "duplicate_key": "handoff-1",
                        "fields": {"工作概述": "最新內容"},
                    }
                ],
            )
            comparisons = iter(
                [
                    {0: {"compare": "未找到", "group": "todo", "matched": []}},
                    {0: {"compare": "已存在", "group": "done", "matched": ["saved"]}},
                ]
            )
            service = DutySubmissionService(
                Path(temp_dir),
                module_loader=lambda: automation,
                now_factory=lambda: datetime(2026, 7, 29, 8, 1),
                comparison_builder=lambda *_args, **_kwargs: next(comparisons),
            )
            data = {
                "target_date": "1150729",
                "actions": [
                    {
                        "kind": "work_log",
                        "time": "08:00",
                        "actor": "10",
                        "source": "值班交接",
                        "duplicate_key": "handoff-1",
                        "fields": {"工作概述": "舊內容"},
                    }
                ],
            }

            result = service.execute(DutySubmissionRequest("user10", "secret", 0, data))

            self.assertEqual(result.status, "submitted")
            self.assertEqual(filled_descriptions, ["最新內容"])

    def test_due_off_duty_action_pauses_when_external_assignment_is_open(self) -> None:
        from datetime import datetime

        from app_core.duty_submission_service import DutySubmissionRequest, DutySubmissionService

        with tempfile.TemporaryDirectory() as temp_dir:
            fills: list[bool] = []
            checked_minutes: list[int | None] = []

            def open_assignment_checker(_rows, _date, _staff, _action, *, current_minute=None):
                checked_minutes.append(current_minute)
                return True

            automation = SimpleNamespace(
                WORK_LOG_AP="work",
                ENTRY_LOG_AP="entry",
                build_driver=lambda headless: object(),
                login=lambda *_args: None,
                query_visible_table=lambda *_args: [["外勤未返隊"]],
                fill_work_log_form_for_test=lambda *_args, **_kwargs: fills.append(True),
                fill_entry_log_form_for_test=lambda *_args, **_kwargs: fills.append(True),
                quit_driver=lambda _driver: None,
            )
            service = DutySubmissionService(
                Path(temp_dir),
                module_loader=lambda: automation,
                now_factory=lambda: datetime(2026, 7, 29, 9, 15),
                open_assignment_checker=open_assignment_checker,
            )
            data = {
                "target_date": "1150729",
                "today": {"staff": {"10": {"name": "測試員"}}},
                "actions": [
                    {
                        "kind": "entry_log",
                        "time": "09:00",
                        "actor": "10",
                        "target": "10",
                        "fields": {"出或入": "值退", "領用事由及地點": "退勤"},
                    }
                ],
            }

            result = service.execute(DutySubmissionRequest("user10", "secret", 0, data))

            self.assertEqual(result.status, "paused_external")
            self.assertEqual(checked_minutes, [555])
            self.assertEqual(fills, [])

    def test_entry_browser_session_reuses_login_and_keeps_each_action_verification(self) -> None:
        from app_core.duty_submission_service import DutySubmissionRequest, DutySubmissionService

        with tempfile.TemporaryDirectory() as temp_dir:
            driver = object()
            build_calls: list[bool] = []
            login_calls: list[tuple[str, str]] = []
            quit_calls: list[object] = []
            fill_calls: list[int] = []
            query_count = 0

            def query_visible_table(_driver, _ap_name, _action_date):
                nonlocal query_count
                query_count += 1
                return [] if query_count % 2 else [{"verified": True}]

            def comparison_builder(_data, actions, comparison_by_date):
                comparison = next(iter(comparison_by_date.values()))
                group = "done" if comparison["visible_entry_rows"] else "todo"
                return {
                    index: {"group": group, "matched": []}
                    for index, _action in enumerate(actions)
                }

            automation = SimpleNamespace(
                WORK_LOG_AP="work",
                ENTRY_LOG_AP="entry",
                build_driver=lambda headless: build_calls.append(headless) or driver,
                login=lambda _driver, user_id, password: login_calls.append((user_id, password)),
                query_visible_table=query_visible_table,
                fill_entry_log_form_for_test=lambda _driver, action, *_args, **_kwargs: fill_calls.append(
                    int(action["index"])
                ),
                fill_work_log_form_for_test=lambda *_args, **_kwargs: None,
                quit_driver=lambda current_driver: quit_calls.append(current_driver),
            )
            service = DutySubmissionService(
                Path(temp_dir),
                module_loader=lambda: automation,
                comparison_builder=comparison_builder,
            )
            data = {
                "target_date": "1150807",
                "actions": [
                    {"kind": "entry_log", "index": 0, "time": "08:00", "actor": "10", "fields": {}},
                    {"kind": "entry_log", "index": 1, "time": "08:05", "actor": "10", "fields": {}},
                ],
            }
            first_request = DutySubmissionRequest("user10", "secret", 0, data, trigger_type="manual")
            second_request = DutySubmissionRequest("user10", "secret", 1, data, trigger_type="manual")

            session = service.open_browser_session(first_request)
            first_result = service.execute_with_browser_session(first_request, session)
            second_result = service.execute_with_browser_session(second_request, session)
            service.close_browser_session(session)

        self.assertEqual(first_result.status, "submitted")
        self.assertEqual(second_result.status, "submitted")
        self.assertEqual(build_calls, [True])
        self.assertEqual(login_calls, [("user10", "secret")])
        self.assertEqual(fill_calls, [0, 1])
        self.assertEqual(quit_calls, [driver])

    def test_recovery_off_duty_action_still_checks_open_external_assignment(self) -> None:
        from datetime import datetime

        from app_core.duty_submission_service import DutySubmissionRequest, DutySubmissionService

        with tempfile.TemporaryDirectory() as temp_dir:
            fills: list[bool] = []
            checked_minutes: list[int | None] = []

            def open_assignment_checker(_rows, _date, _staff, _action, *, current_minute=None):
                checked_minutes.append(current_minute)
                return True

            automation = SimpleNamespace(
                WORK_LOG_AP="work",
                ENTRY_LOG_AP="entry",
                build_driver=lambda headless: object(),
                login=lambda *_args: None,
                query_visible_table=lambda *_args: [["外勤未返隊"]],
                fill_work_log_form_for_test=lambda *_args, **_kwargs: fills.append(True),
                fill_entry_log_form_for_test=lambda *_args, **_kwargs: fills.append(True),
                quit_driver=lambda _driver: None,
            )
            service = DutySubmissionService(
                Path(temp_dir),
                module_loader=lambda: automation,
                now_factory=lambda: datetime(2026, 7, 29, 9, 15),
                open_assignment_checker=open_assignment_checker,
            )
            data = {
                "target_date": "1150729",
                "today": {"staff": {"10": {"name": "測試員"}}},
                "actions": [
                    {
                        "kind": "entry_log",
                        "time": "09:15",
                        "actor": "10",
                        "target": "10",
                        "fields": {"出或入": "值退", "領用事由及地點": "退勤"},
                    }
                ],
            }

            result = service.execute(
                DutySubmissionRequest("user10", "secret", 0, data, trigger_type="recovery")
            )

            self.assertEqual(result.status, "paused_external")
            self.assertEqual(checked_minutes, [555])
            self.assertEqual(fills, [])

    def test_handoff_preflight_checks_incoming_staff_without_writing_a_form(self) -> None:
        from datetime import datetime

        from app_core.duty_submission_service import DutySubmissionRequest, DutySubmissionService

        with tempfile.TemporaryDirectory() as temp_dir:
            fills: list[bool] = []
            checked_actions: list[dict] = []

            def open_assignment_checker(_rows, _date, _staff, action, *, current_minute=None):
                checked_actions.append(dict(action))
                return True

            automation = SimpleNamespace(
                ENTRY_LOG_AP="entry",
                WORK_LOG_AP="work",
                build_driver=lambda *_args, **_kwargs: object(),
                login=lambda *_args: None,
                query_visible_table=lambda *_args: [],
                fill_entry_log_form_for_test=lambda *_args, **_kwargs: fills.append(True),
                quit_driver=lambda _driver: None,
            )
            service = DutySubmissionService(
                Path(temp_dir),
                module_loader=lambda: automation,
                now_factory=lambda: datetime(2026, 8, 7, 18, 0),
                open_assignment_checker=open_assignment_checker,
            )
            data = {
                "target_date": "1150807",
                "today": {"staff": {"11": {"name": "接班"}}},
                "actions": [
                    {
                        "kind": "handoff_preflight",
                        "time": "18:00",
                        "actor": "10",
                        "target": "11",
                        "source": "值班交接",
                        "fields": {"出或入": "值班", "領用事由及地點": "值班"},
                    }
                ],
            }

            result = service.execute(DutySubmissionRequest("user10", "secret", 0, data))

            self.assertEqual(result.status, "paused_external")
            self.assertEqual(checked_actions[0]["target"], "11")
            self.assertEqual(fills, [])

    def test_unreturned_return_queue_keeps_fixed_expiry_and_changes_handoff_interval(self) -> None:
        from datetime import datetime

        from app_core.unreturned_return_queue import UnreturnedReturnQueue

        with tempfile.TemporaryDirectory() as temp_dir:
            action = {
                "kind": "entry_log",
                "time": "08:05",
                "actor": "10",
                "target": "10",
                "fields": {"出或入": "值退", "領用事由及地點": "退勤"},
            }
            schedule = {
                "target_date": "1150729",
                "today": {"staff": {"10": {"name": "測試員"}}},
            }
            started_at = datetime(2026, 7, 29, 8, 0)
            queue = UnreturnedReturnQueue(Path(temp_dir), now_factory=lambda: started_at)

            record, created = queue.pause(action, schedule, owner_actor_no="10", now=started_at)

            self.assertTrue(created)
            self.assertEqual(record["next_retry_at"], "2026-07-29T08:05:00")
            self.assertEqual(record["expires_at"], "2026-07-30T02:00:00")
            self.assertTrue((Path(temp_dir) / "unreturned_return_queue.json").is_file())
            restarted = UnreturnedReturnQueue(Path(temp_dir), now_factory=lambda: started_at)
            self.assertEqual(len(restarted.active_records()), 1)

            current_shift = restarted.claim_due("10", now=datetime(2026, 7, 29, 8, 5))
            self.assertEqual(current_shift["retry_interval_minutes"], 5)
            self.assertEqual(current_shift["next_retry_at"], "2026-07-29T08:10:00")
            restarted.defer(record["queue_id"], "10", now=datetime(2026, 7, 29, 8, 5))
            handoff = restarted.claim_due("11", now=datetime(2026, 7, 29, 8, 10))
            self.assertEqual(handoff["retry_interval_minutes"], 10)
            self.assertEqual(handoff["next_retry_at"], "2026-07-29T08:20:00")
            self.assertEqual(
                [item["queue_id"] for item in restarted.expire_due(now=datetime(2026, 7, 30, 2, 0))],
                [record["queue_id"]],
            )

    def test_handoff_group_queue_resolves_only_after_every_component_succeeds(self) -> None:
        from app_core.unreturned_return_queue import UnreturnedReturnQueue

        with tempfile.TemporaryDirectory() as temp_dir:
            queue = UnreturnedReturnQueue(Path(temp_dir))
            actions = [
                {
                    "kind": "entry_log",
                    "time": "18:00",
                    "actor": "10",
                    "target": "10",
                    "duplicate_key": "entry:out",
                    "fields": {"出或入": "值退"},
                },
                {
                    "kind": "entry_log",
                    "time": "18:00",
                    "actor": "10",
                    "target": "11",
                    "duplicate_key": "entry:in",
                    "fields": {"出或入": "值班"},
                },
                {
                    "kind": "work_log",
                    "time": "18:00",
                    "actor": "10",
                    "target": "10",
                    "duplicate_key": "work:handoff",
                    "fields": {"工作時間": "18:00"},
                },
            ]
            record, created = queue.pause_group(actions, {"target_date": "1150807"}, owner_actor_no="10")

            self.assertTrue(created)
            queue.claim_manual(record["queue_id"], "10")
            partial, resolved = queue.complete_action(record["queue_id"], actions[0], "submitted")
            self.assertFalse(resolved)
            self.assertEqual(partial["completed_keys"], ["entry:out"])
            queue.complete_action(record["queue_id"], actions[1], "skipped_duplicate")
            completed, resolved = queue.complete_action(record["queue_id"], actions[2], "submitted")

            self.assertTrue(resolved)
            self.assertEqual(completed["queue_id"], record["queue_id"])
            self.assertEqual(queue.active_records(), [])

    def test_due_checkout_starts_at_0800_and_writes_0805(self) -> None:
        from datetime import datetime

        from app_core.duty_submission_service import DutySubmissionRequest, DutySubmissionService

        with tempfile.TemporaryDirectory() as temp_dir:
            events: list[str] = []
            automation = SimpleNamespace(
                ENTRY_LOG_AP="entry",
                WORK_LOG_AP="work",
                build_driver=lambda headless: events.append(f"driver:{headless}") or object(),
                login=lambda _driver, _user_id, _password: events.append("login"),
                query_visible_table=lambda _driver, _ap_name, _date: [],
                fill_entry_log_form_for_test=lambda _driver, action, _staff, _date, save: events.append(
                    f"fill:{action['fields']['系統寫入時間']}:{save}"
                )
                or {"ok": True},
                quit_driver=lambda _driver: events.append("quit"),
            )
            comparisons = iter(
                [
                    {0: {"compare": "尚未到點", "group": "future", "matched": []}},
                    {0: {"compare": "已存在", "group": "done", "matched": ["saved"]}},
                ]
            )
            service = DutySubmissionService(
                Path(temp_dir),
                module_loader=lambda: automation,
                now_factory=lambda: datetime(2026, 8, 7, 8, 0),
                comparison_builder=lambda *_args, **_kwargs: next(comparisons),
            )
            data = {
                "target_date": "1150807",
                "today": {"staff": {"10": {"name": "測試員"}}},
                "actions": [
                    {
                        "kind": "entry_log",
                        "time": "08:00",
                        "actor": "10",
                        "target": "10",
                        "source": "昨日在勤且今日未在勤",
                        "fields": {
                            "登打時間": "08:00",
                            "系統寫入時間": "08:05",
                            "出或入": "出",
                            "領用事由及地點": "退勤",
                        },
                    }
                ],
            }

            result = service.execute(DutySubmissionRequest("user10", "secret", 0, data))

            self.assertEqual(result.status, "submitted")
            self.assertEqual(events, ["driver:True", "login", "fill:08:05:True", "quit"])

    def test_manual_external_review_action_can_submit_after_confirmation(self) -> None:
        from datetime import datetime

        from app_core.duty_submission_service import DutySubmissionRequest, DutySubmissionService

        with tempfile.TemporaryDirectory() as temp_dir:
            fills: list[str] = []
            comparisons = iter(
                [
                    {0: {"compare": "外勤確認", "group": "review", "matched": []}},
                    {0: {"compare": "已存在", "group": "done", "matched": ["submitted"]}},
                ]
            )
            automation = SimpleNamespace(
                WORK_LOG_AP="work",
                ENTRY_LOG_AP="entry",
                build_driver=lambda headless: object(),
                login=lambda *_args: None,
                query_visible_table=lambda *_args: [],
                fill_work_log_form_for_test=lambda *_args, **_kwargs: {},
                fill_entry_log_form_for_test=lambda _driver, action, *_args, **_kwargs: fills.append(
                    action["source"]
                ) or {},
                quit_driver=lambda _driver: None,
            )
            service = DutySubmissionService(
                Path(temp_dir),
                module_loader=lambda: automation,
                now_factory=lambda: datetime(2026, 7, 29, 8, 1),
                comparison_builder=lambda *_args, **_kwargs: next(comparisons),
            )
            data = {
                "target_date": "1150729",
                "actions": [
                    {
                        "kind": "entry_log",
                        "time": "08:00",
                        "actor": "10",
                        "target": "10",
                        "source": "外勤支援",
                        "fields": {"出或入": "出", "領用事由及地點": "外勤"},
                    }
                ],
            }

            result = service.execute(
                DutySubmissionRequest("user10", "test-secret", 0, data, trigger_type="manual")
            )

            self.assertEqual(result.status, "submitted")
            self.assertEqual(fills, ["外勤支援"])

    def test_manual_group_action_can_submit_after_confirmation(self) -> None:
        from datetime import datetime

        from app_core.duty_submission_service import DutySubmissionRequest, DutySubmissionService

        with tempfile.TemporaryDirectory() as temp_dir:
            fills: list[str] = []
            comparisons = iter(
                [
                    {0: {"compare": "手動登打", "group": "manual", "matched": []}},
                    {0: {"compare": "已存在", "group": "done", "matched": ["submitted"]}},
                ]
            )
            automation = SimpleNamespace(
                WORK_LOG_AP="work",
                ENTRY_LOG_AP="entry",
                build_driver=lambda headless: object(),
                login=lambda *_args: None,
                query_visible_table=lambda *_args: [],
                fill_work_log_form_for_test=lambda _driver, action, *_args, **_kwargs: fills.append(
                    action["source"]
                ) or {},
                fill_entry_log_form_for_test=lambda *_args, **_kwargs: {},
                quit_driver=lambda _driver: None,
            )
            service = DutySubmissionService(
                Path(temp_dir),
                module_loader=lambda: automation,
                now_factory=lambda: datetime(2026, 7, 29, 8, 1),
                comparison_builder=lambda *_args, **_kwargs: next(comparisons),
            )
            data = {
                "target_date": "1150729",
                "actions": [
                    {
                        "kind": "work_log",
                        "time": "08:00",
                        "actor": "10",
                        "target": "10",
                        "source": "人工補登",
                        "fields": {"勤務項目": "工作紀錄"},
                    }
                ],
            }

            result = service.execute(
                DutySubmissionRequest("user10", "test-secret", 0, data, trigger_type="manual")
            )

            self.assertEqual(result.status, "submitted")
            self.assertEqual(fills, ["人工補登"])

    def test_manual_adjust_group_action_can_submit_after_confirmation(self) -> None:
        from datetime import datetime

        from app_core.duty_submission_service import DutySubmissionRequest, DutySubmissionService

        with tempfile.TemporaryDirectory() as temp_dir:
            comparisons = iter(
                [
                    {0: {"compare": "可能臨時調整", "group": "adjust", "matched": []}},
                    {0: {"compare": "已存在", "group": "done", "matched": ["submitted"]}},
                ]
            )
            automation = SimpleNamespace(
                WORK_LOG_AP="work",
                ENTRY_LOG_AP="entry",
                build_driver=lambda headless: object(),
                login=lambda *_args: None,
                query_visible_table=lambda *_args: [],
                fill_work_log_form_for_test=lambda *_args, **_kwargs: {},
                fill_entry_log_form_for_test=lambda *_args, **_kwargs: {},
                quit_driver=lambda _driver: None,
            )
            service = DutySubmissionService(
                Path(temp_dir),
                module_loader=lambda: automation,
                now_factory=lambda: datetime(2026, 7, 29, 8, 1),
                comparison_builder=lambda *_args, **_kwargs: next(comparisons),
            )
            data = {
                "target_date": "1150729",
                "actions": [
                    {
                        "kind": "work_log",
                        "time": "08:00",
                        "actor": "10",
                        "target": "10",
                        "source": "值班交接",
                        "fields": {"勤務項目": "工作紀錄"},
                    }
                ],
            }

            result = service.execute(
                DutySubmissionRequest("user10", "test-secret", 0, data, trigger_type="manual")
            )

            self.assertEqual(result.status, "submitted")

    def test_manual_near_group_action_remains_blocked(self) -> None:
        from datetime import datetime

        from app_core.duty_submission_service import DutySubmissionRequest, DutySubmissionService

        with tempfile.TemporaryDirectory() as temp_dir:
            fills: list[str] = []
            automation = SimpleNamespace(
                WORK_LOG_AP="work",
                ENTRY_LOG_AP="entry",
                build_driver=lambda headless: object(),
                login=lambda *_args: None,
                query_visible_table=lambda *_args: [],
                fill_work_log_form_for_test=lambda *_args, **_kwargs: fills.append("filled") or {},
                fill_entry_log_form_for_test=lambda *_args, **_kwargs: {},
                quit_driver=lambda _driver: None,
            )
            service = DutySubmissionService(
                Path(temp_dir),
                module_loader=lambda: automation,
                now_factory=lambda: datetime(2026, 7, 29, 8, 1),
                comparison_builder=lambda *_args, **_kwargs: {
                    0: {"compare": "時間近似", "group": "near", "matched": ["existing"]}
                },
            )
            data = {
                "target_date": "1150729",
                "actions": [
                    {
                        "kind": "work_log",
                        "time": "08:00",
                        "actor": "10",
                        "target": "10",
                        "source": "值班交接",
                        "fields": {"勤務項目": "工作紀錄"},
                    }
                ],
            }

            result = service.execute(
                DutySubmissionRequest("user10", "test-secret", 0, data, trigger_type="manual")
            )

            self.assertEqual(result.status, "review_required")
            self.assertEqual(fills, [])

    def test_login_failure_is_classified_without_exposing_password(self) -> None:
        from datetime import datetime

        from app_core.duty_submission_service import (
            DutySubmissionExecutionError,
            DutySubmissionRequest,
            DutySubmissionService,
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            automation = SimpleNamespace(
                build_driver=lambda headless: object(),
                login=lambda *_args: (_ for _ in ()).throw(RuntimeError("登入失敗：帳號或密碼")),
                quit_driver=lambda _driver: None,
            )
            service = DutySubmissionService(
                Path(temp_dir),
                module_loader=lambda: automation,
                now_factory=lambda: datetime(2026, 7, 29, 8, 1),
            )
            data = {
                "target_date": "1150729",
                "actions": [{"kind": "work_log", "time": "08:00", "actor": "10"}],
            }

            with self.assertRaises(DutySubmissionExecutionError) as caught:
                service.execute(DutySubmissionRequest("user10", "session-secret", 0, data))

            self.assertEqual(caught.exception.error_code, "login_failed")
            self.assertNotIn("session-secret", str(caught.exception))
            self.assertTrue(caught.exception.result_path.is_file())
            result_files = list((Path(temp_dir) / "runtime_outputs" / "form_tests").glob("*.json"))
            self.assertEqual(len(result_files), 1)
            self.assertNotIn("session-secret", result_files[0].read_text(encoding="utf-8"))


class RestMonthlyServiceTests(unittest.TestCase):
    def test_loading_execution_core_does_not_import_legacy_tk_ui(self) -> None:
        command = (
            "import sys; "
            "from pathlib import Path; "
            f"package_root = Path({str(PACKAGE_ROOT)!r}); "
            "sys.path.insert(0, str(package_root)); "
            "from app_core.rest_monthly_service import load_legacy_module; "
            "load_legacy_module(package_root); "
            "forbidden = {'tkinter', 'customtkinter'}; "
            "raise SystemExit(1 if forbidden.intersection(sys.modules) else 0)"
        )

        return_code, output = run_isolated_python(command)

        self.assertEqual(return_code, 0, output)

    def test_defaults_keep_current_month_without_loading_legacy_ui_module(self) -> None:
        from datetime import date

        import openpyxl

        from app_core.rest_monthly_service import RestMonthlyService

        with tempfile.TemporaryDirectory() as temp_dir:
            package_root = Path(temp_dir)
            workbook = package_root / "duty.xlsx"
            book = openpyxl.Workbook()
            book.active.cell(row=2, column=5).value = 8
            book.save(workbook)
            (package_root / "rest_time_automation_config.json").write_text(
                json.dumps({"workbook_path": str(workbook)}),
                encoding="utf-8",
            )
            legacy_loads: list[Path] = []
            service = RestMonthlyService(
                package_root,
                module_loader=lambda path: legacy_loads.append(path),
            )

            defaults = service.load_rest_defaults(date(2026, 7, 29))

            self.assertEqual(defaults.roc_year, 115)
            self.assertEqual(defaults.month_options, ("06", "07", "08"))
            self.assertEqual(defaults.selected_month, "07")
            self.assertEqual(defaults.workbook_path, str(workbook.resolve()))
            self.assertEqual(legacy_loads, [])

    def test_monthly_defaults_offer_previous_current_next_with_current_selected(self) -> None:
        from datetime import date

        from app_core.rest_monthly_service import RestMonthlyService

        with tempfile.TemporaryDirectory() as temp_dir:
            package_root = Path(temp_dir)
            (package_root / "rest_time_automation_config.json").write_text(
                json.dumps({"rest_month": "07", "monthly_base_month": "09"}),
                encoding="utf-8",
            )
            service = RestMonthlyService(package_root)

            rest_defaults = service.load_rest_defaults(date(2026, 8, 5))
            monthly_defaults = service.load_monthly_defaults(date(2026, 8, 5))

            self.assertEqual(rest_defaults.month_options, ("07", "08", "09"))
            self.assertEqual(monthly_defaults.month_options, ("07", "08", "09"))
            self.assertEqual(rest_defaults.selected_month, "08")
            self.assertEqual(monthly_defaults.selected_month, "08")

    def test_default_rest_workbook_matches_legacy_xlsm_only_fallback(self) -> None:
        from datetime import date

        from app_core.rest_monthly_service import RestMonthlyService

        with tempfile.TemporaryDirectory() as temp_dir:
            package_root = Path(temp_dir)
            legacy_workbook = package_root / "legacy-duty.xlsm"
            newer_xlsx = package_root / "newer-duty.xlsx"
            legacy_workbook.write_bytes(b"legacy")
            newer_xlsx.write_bytes(b"newer")
            service = RestMonthlyService(package_root)

            defaults = service.load_rest_defaults(date(2026, 7, 29))

            self.assertEqual(defaults.workbook_path, str(legacy_workbook.resolve()))

    def test_selecting_rest_workbook_updates_month_and_saved_path(self) -> None:
        from datetime import date

        import openpyxl

        from app_core.rest_monthly_service import RestMonthlyService

        with tempfile.TemporaryDirectory() as temp_dir:
            package_root = Path(temp_dir)
            workbook = package_root / "selected.xlsx"
            book = openpyxl.Workbook()
            book.active.cell(row=2, column=5).value = 8
            book.save(workbook)
            service = RestMonthlyService(package_root)

            defaults = service.select_rest_workbook(workbook, date(2026, 7, 29))

            self.assertEqual(defaults.selected_month, "08")
            self.assertEqual(defaults.workbook_path, str(workbook.resolve()))
            saved = json.loads(
                (package_root / "rest_time_automation_config.json").read_text(encoding="utf-8")
            )
            self.assertEqual(saved["workbook_path"], str(workbook.resolve()))

    def test_execute_uses_existing_engines_and_closes_browser(self) -> None:
        from app_core.rest_monthly_service import (
            MonthlyBaseRequest,
            RestMonthlyService,
            RestTimeRequest,
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            package_root = Path(temp_dir)
            workbook = package_root / "duty.xlsm"
            workbook.write_bytes(b"test")
            (package_root / "rest_time_automation.py").write_text("# placeholder\n", encoding="utf-8")
            calls: list[tuple[str, tuple, dict]] = []

            def submit_rest(*args, **kwargs):
                calls.append(("rest", args, kwargs))
                kwargs["status"]("休息時間執行中") if "status" in kwargs else args[4]("休息時間執行中")
                return "休息完成"

            def submit_monthly(*args, **kwargs):
                calls.append(("monthly", args, kwargs))
                args[4]("勤務基準執行中")
                return "勤務基準完成"

            legacy = SimpleNamespace(
                submit_rest_entries=submit_rest,
                submit_monthly_base_entries=submit_monthly,
                format_automation_error=lambda exc: str(exc),
            )
            service = RestMonthlyService(package_root, module_loader=lambda _path: legacy)
            progress: list[str] = []

            rest_request = RestTimeRequest("user10", "secret", "10", str(workbook), 115, 7)
            monthly_request = MonthlyBaseRequest("user10", "secret", "10", 115, 7)
            rest_result = service.execute_rest(
                rest_request,
                status_callback=progress.append,
            )
            monthly_result = service.execute_monthly(
                monthly_request,
                status_callback=progress.append,
            )

            self.assertEqual(rest_result, "休息完成")
            self.assertEqual(monthly_result, "勤務基準完成")
            self.assertEqual(progress, ["休息時間執行中", "勤務基準執行中"])
            self.assertFalse(calls[0][2]["keep_browser_open"])
            self.assertFalse(calls[1][2]["keep_browser_open"])
            self.assertNotIn("secret", repr(rest_request))
            self.assertNotIn("secret", repr(monthly_request))

    def test_qml_rest_monthly_requests_use_name_before_actor_number(self) -> None:
        from app_core.rest_monthly_service import (
            MonthlyBaseRequest,
            RestMonthlyService,
            RestTimeRequest,
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            package_root = Path(temp_dir)
            workbook = package_root / "duty.xlsm"
            workbook.write_bytes(b"test")
            (package_root / "rest_time_automation.py").write_text("# placeholder\n", encoding="utf-8")
            calls: list[tuple[tuple, dict]] = []
            legacy = SimpleNamespace(
                submit_rest_entries=lambda *args, **kwargs: calls.append((args, kwargs)) or "完成",
                submit_monthly_base_entries=lambda *args, **kwargs: calls.append((args, kwargs)) or "完成",
                format_automation_error=lambda exc: str(exc),
            )
            service = RestMonthlyService(package_root, module_loader=lambda _path: legacy)

            service.execute_rest(
                RestTimeRequest("user10", "secret", "10", str(workbook), 115, 8, "王小明")
            )
            service.execute_monthly(
                MonthlyBaseRequest("user10", "secret", "10", 115, 8, "王小明")
            )

            self.assertEqual(calls[0][1]["actor_no"], "")
            self.assertEqual(calls[0][1]["actor_name"], "王小明")
            self.assertEqual(calls[1][0][2], "")
            self.assertEqual(calls[1][1]["actor_name"], "王小明")

    def test_monthly_base_source_requires_a_unique_name_match(self) -> None:
        from unittest.mock import patch

        import rest_time_automation as module
        source = "來源標題\n,,10,11\n,,王小明,王小華\n1,,,\n"
        month_match = SimpleNamespace(group=lambda index: "115" if index == 1 else "8")
        with (
            patch.object(module, "download_monthly_base_csv", return_value=source),
            patch.object(module.re, "search", return_value=month_match),
        ):
            plan = module.fetch_monthly_base_plan("", actor_name="王小明")

        self.assertEqual(plan.actor_no, "10")
        self.assertEqual(plan.name, "王小明")

        duplicate_source = "來源標題\n,,10,11\n,,王小明,王小明\n1,,,\n"
        with (
            patch.object(module, "download_monthly_base_csv", return_value=duplicate_source),
            patch.object(module.re, "search", return_value=month_match),
        ):
            with self.assertRaisesRegex(RuntimeError, "唯一"):
                module.fetch_monthly_base_plan("", actor_name="王小明")

    def test_rest_and_monthly_save_last_confirmed_month_without_reusing_as_default(self) -> None:
        from datetime import date

        from app_core.rest_monthly_service import (
            MonthlyBaseRequest,
            RestMonthlyService,
            RestTimeRequest,
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            package_root = Path(temp_dir)
            workbook = package_root / "duty.xlsm"
            workbook.write_bytes(b"test")
            (package_root / "rest_time_automation.py").write_text("# placeholder\n", encoding="utf-8")
            legacy = SimpleNamespace(
                submit_rest_entries=lambda *_args, **_kwargs: "休息時間完成",
                submit_monthly_base_entries=lambda *_args, **_kwargs: "勤務基準表完成",
                format_automation_error=lambda exc: str(exc),
            )
            service = RestMonthlyService(package_root, module_loader=lambda _path: legacy)

            service.execute_rest(
                RestTimeRequest("user10", "secret", "10", str(workbook), 115, 8)
            )
            service.execute_monthly(MonthlyBaseRequest("user10", "secret", "10", 115, 6))

            restarted = RestMonthlyService(package_root, module_loader=lambda _path: legacy)
            rest_defaults = restarted.load_rest_defaults(date(2026, 7, 29))
            monthly_defaults = restarted.load_monthly_defaults(date(2026, 7, 29))
            saved = json.loads(
                (package_root / "rest_time_automation_config.json").read_text(encoding="utf-8")
            )

            self.assertEqual(rest_defaults.selected_month, "07")
            self.assertEqual(monthly_defaults.selected_month, "07")
            self.assertEqual(saved["workbook_path"], str(workbook.resolve()))
            self.assertEqual(saved["rest_month"], "08")
            self.assertEqual(saved["monthly_base_month"], "06")
            self.assertNotIn("user10", json.dumps(saved))
            self.assertNotIn("secret", json.dumps(saved))

    def test_monthly_browser_start_failure_is_safe_and_classified(self) -> None:
        from app_core.rest_monthly_service import (
            MonthlyBaseRequest,
            RestMonthlyExecutionError,
            RestMonthlyService,
        )

        class BrowserStartupError(RuntimeError):
            diagnostic_category = "startup_timeout"

        def submit_monthly(*_args, **kwargs):
            kwargs["stage_callback"]("browser_start")
            raise BrowserStartupError()

        with tempfile.TemporaryDirectory() as temp_dir:
            package_root = Path(temp_dir)
            (package_root / "rest_time_automation.py").write_text("# placeholder\n", encoding="utf-8")
            legacy = SimpleNamespace(
                submit_monthly_base_entries=submit_monthly,
                format_automation_error=lambda _exc: "不應顯示原始瀏覽器錯誤",
            )
            service = RestMonthlyService(package_root, module_loader=lambda _path: legacy)

            with self.assertRaises(RestMonthlyExecutionError) as raised:
                service.execute_monthly(MonthlyBaseRequest("user10", "secret", "10", 115, 8))

        self.assertEqual(raised.exception.failure_stage, "browser_start")
        self.assertEqual(raised.exception.failure_detail, "browser_startup")
        self.assertIn("專用瀏覽器啟動失敗", str(raised.exception))
        self.assertNotIn("原始瀏覽器錯誤", str(raised.exception))


class DailyVehicleServiceTests(unittest.TestCase):
    def test_execute_uses_existing_script_with_ephemeral_credentials_and_cleanup(self) -> None:
        from datetime import date

        from app_core.daily_vehicle_service import DailyVehicleRequest, DailyVehicleService

        with tempfile.TemporaryDirectory() as temp_dir:
            package_root = Path(temp_dir)
            project_dir = package_root / "daily_vehicle_legacy"
            script = project_dir / "automation" / "ppe_selenium_daily.py"
            script.parent.mkdir(parents=True)
            script.write_text("# placeholder\n", encoding="utf-8")
            launches: list[dict] = []

            class FakeProcess:
                pid = 12345
                returncode = 0

                def communicate(self, timeout=None):
                    return "[done] automation finished", None

                def kill(self):
                    raise AssertionError("successful process must not be killed")

            def process_factory(command, **kwargs):
                launches.append({"command": command, **kwargs})
                return FakeProcess()

            service = DailyVehicleService(
                package_root,
                process_factory=process_factory,
                process_checker=lambda _pid: False,
            )
            defaults = service.load_defaults(date(2026, 7, 29))
            request = DailyVehicleRequest("user10", "session-secret")
            confirmation = service.confirmation_summary(request)
            progress: list[str] = []

            result = service.execute(request, status_callback=progress.append)

            self.assertEqual(defaults.target_date, "2026/07/29")
            self.assertEqual(confirmation, "將開啟瀏覽器執行車輛保養清點，是否繼續？")
            self.assertEqual(result, "車輛保養清點已完成。")
            self.assertEqual(len(launches), 1)
            self.assertEqual(launches[0]["env"]["PPE_ACCOUNT"], "user10")
            self.assertEqual(launches[0]["env"]["PPE_PASSWORD"], "session-secret")
            self.assertEqual(launches[0]["env"]["KEEP_BROWSER_OPEN"], "true")
            self.assertEqual(
                launches[0]["creationflags"],
                getattr(subprocess, "CREATE_NO_WINDOW", 0),
            )
            self.assertFalse((project_dir / ".daily_vehicle_runner.pid").exists())
            self.assertNotIn("session-secret", repr(request))
            self.assertTrue(progress)

    def test_browser_start_failure_uses_safe_shared_driver_message(self) -> None:
        from app_core.daily_vehicle_service import (
            DailyVehicleExecutionError,
            DailyVehicleRequest,
            DailyVehicleService,
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            package_root = Path(temp_dir)
            project_dir = package_root / "daily_vehicle_legacy"
            script = project_dir / "automation" / "ppe_selenium_daily.py"
            script.parent.mkdir(parents=True)
            script.write_text("# placeholder\n", encoding="utf-8")

            class FailedProcess:
                pid = 12345
                returncode = 1

                def communicate(self, timeout=None):
                    return (
                        "[sinposmart-stage] browser_start\n"
                        "SinpoSmart 專用瀏覽器啟動失敗，已自動清理暫存資料並重試。\n"
                        "WebDriverException: raw startup detail",
                        None,
                    )

                def kill(self):
                    raise AssertionError("failed process must not be killed after it exits")

            service = DailyVehicleService(
                package_root,
                process_factory=lambda *_args, **_kwargs: FailedProcess(),
                process_checker=lambda _pid: False,
            )

            with self.assertRaises(DailyVehicleExecutionError) as raised:
                service.execute(DailyVehicleRequest("user10", "session-secret"))

        self.assertEqual(raised.exception.failure_stage, "browser_start")
        self.assertIn("專用瀏覽器啟動失敗", str(raised.exception))
        self.assertNotIn("raw startup detail", str(raised.exception))

    def test_validation_rejects_existing_runner(self) -> None:
        from app_core.daily_vehicle_service import (
            DailyVehicleRequest,
            DailyVehicleService,
            DailyVehicleValidationError,
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            package_root = Path(temp_dir)
            project_dir = package_root / "daily_vehicle_legacy"
            script = project_dir / "automation" / "ppe_selenium_daily.py"
            script.parent.mkdir(parents=True)
            script.write_text("# placeholder\n", encoding="utf-8")
            (project_dir / ".daily_vehicle_runner.pid").write_text("12345\n", encoding="utf-8")
            service = DailyVehicleService(package_root, process_checker=lambda _pid: True)

            with self.assertRaises(DailyVehicleValidationError):
                service.validate(DailyVehicleRequest("user10", "secret"))


class RescueVideoServiceTests(unittest.TestCase):
    def test_loading_existing_core_does_not_import_tk_ui(self) -> None:
        command = (
            "import sys; "
            "from pathlib import Path; "
            f"package_root = Path({str(PACKAGE_ROOT)!r}); "
            "sys.path.insert(0, str(package_root)); "
            "from app_core.rescue_video_service import load_rescue_video_core; "
            "load_rescue_video_core(package_root); "
            "forbidden = {'tkinter', 'customtkinter'}; "
            "raise SystemExit(1 if forbidden.intersection(sys.modules) else 0)"
        )

        return_code, output = run_isolated_python(command)

        self.assertEqual(return_code, 0, output)

    def test_preview_reuses_existing_classifier_and_projects_results(self) -> None:
        import argparse
        from datetime import datetime

        from app_core.rescue_video_service import RescueVideoRequest, RescueVideoService

        captured_args = []
        result_row = SimpleNamespace(
            source=Path("video001.TS"),
            adjusted_time=datetime(2026, 7, 29, 9, 30),
            case=SimpleNamespace(name="07290900-92"),
            destination=Path("case") / "車" / "video001.TS",
            status="預計複製",
            note="dry-run",
        )

        def build_args(values, mode):
            payload = dict(values)
            payload.update(
                apply=mode != "preview",
                delete_source=mode == "delete",
                offset_minutes=float(values.get("offset_minutes") or 0),
                report=Path(values["report"]),
            )
            return argparse.Namespace(**payload)

        core = SimpleNamespace(
            DEFAULT_WORK_LOG_ROOT=Path("work_logs"),
            validate_form=lambda _values, _mode: ([], ["測試警告"]),
            build_args=build_args,
            choose_runtime_offset=lambda _args: 6,
            run_classification=lambda args: captured_args.append(args) or [result_row],
            summarize_results=lambda _results: {"預計複製": 1},
            format_summary=lambda _summary: "預計複製: 1",
            status_tag=lambda _status: "",
            build_public_duty_report_path=lambda _date, _vehicle: Path("report.csv"),
        )
        service = RescueVideoService(PACKAGE_ROOT, module_loader=lambda _root: core)
        request = RescueVideoRequest(
            source_path="source",
            destination_path="destination",
            target_date="2026-07-29",
            vehicle="92",
            offset_text="",
            repair_mismatch=False,
            mode="preview",
        )
        progress = []

        result = service.execute(request, status_callback=progress.append)

        self.assertFalse(captured_args[0].apply)
        self.assertFalse(captured_args[0].delete_source)
        self.assertEqual(captured_args[0].offset_minutes, 6)
        self.assertEqual(result.summary_text, "預計複製: 1")
        self.assertEqual(result.rows[0]["sourceText"], "video001.TS")
        self.assertEqual(result.rows[0]["caseText"], "07290900-92")
        self.assertIn("測試警告", result.warning_text)
        self.assertTrue(progress)

    def test_public_defaults_run_the_legacy_automatic_preflight_contract(self) -> None:
        import argparse
        from datetime import date

        from app_core.rescue_video_service import RescueVideoService

        state = SimpleNamespace(
            vehicles=["92", "93"],
            ready=True,
            checks={
                "source": SimpleNamespace(detail="來源：X:/DCIM/100CAREC"),
                "destination": SimpleNamespace(detail="案件目的地可存取"),
                "work_log": SimpleNamespace(detail="工作／返隊紀錄可存取"),
            },
        )
        core = SimpleNamespace(
            DEFAULT_DESTINATION=Path("Z:/救護行車影片"),
            DEFAULT_WORK_LOG_ROOT=Path("runtime_outputs/comparison"),
            DEFAULT_REPORT=Path("分類結果.csv"),
            classifier=SimpleNamespace(resolve_source=lambda _source: Path("X:/DCIM/100CAREC")),
            evaluate_preflight=lambda _values: state,
            build_public_duty_report_path=lambda _date, _vehicle: Path("report.csv"),
            build_args=lambda values, _mode: argparse.Namespace(**values),
            choose_runtime_offset=lambda _args: 6,
        )
        service = RescueVideoService(PACKAGE_ROOT, module_loader=lambda _root: core)

        defaults = service.load_defaults(date(2026, 7, 29), vehicle="93")

        self.assertEqual(defaults.target_date, "2026-07-29")
        self.assertEqual(defaults.vehicle_options, ("92", "93"))
        self.assertEqual(defaults.selected_vehicle, "93")
        self.assertEqual(defaults.offset_text, "6")
        self.assertTrue(defaults.is_ready)
        self.assertEqual(defaults.status_text, "自動檢查通過")
        self.assertIn("工作／返隊紀錄可存取", defaults.check_text)
        self.assertIn("自動採用記憶卡偏移：6 分鐘", defaults.check_text)
        self.assertEqual(
            [card.key for card in defaults.check_cards],
            ["source", "destination", "work_log", "vehicle_date", "report", "videos"],
        )
        self.assertEqual(defaults.check_cards[0].title, "記憶卡來源")

    def test_copy_reuses_existing_classifier_without_deleting_source(self) -> None:
        import argparse

        from app_core.rescue_video_service import RescueVideoRequest, RescueVideoService

        captured_args = []

        def build_args(values, mode):
            payload = dict(values)
            payload.update(
                apply=mode != "preview",
                delete_source=mode == "delete",
                offset_minutes=float(values.get("offset_minutes") or 0),
                report=Path(values["report"]),
            )
            return argparse.Namespace(**payload)

        core = SimpleNamespace(
            DEFAULT_WORK_LOG_ROOT=Path("work_logs"),
            validate_form=lambda _values, _mode: ([], []),
            build_args=build_args,
            choose_runtime_offset=lambda _args: 0,
            run_classification=lambda args: captured_args.append(args) or [],
            summarize_results=lambda _results: {},
            format_summary=lambda _summary: "複製完成",
            status_tag=lambda _status: "",
            build_public_duty_report_path=lambda _date, _vehicle: Path("report.csv"),
        )
        service = RescueVideoService(PACKAGE_ROOT, module_loader=lambda _root: core)
        request = RescueVideoRequest(
            source_path="source",
            destination_path="destination",
            target_date="2026-07-29",
            vehicle="92",
            offset_text="0",
            repair_mismatch=False,
            mode="copy",
        )

        result = service.execute(request)

        self.assertTrue(captured_args[0].apply)
        self.assertFalse(captured_args[0].delete_source)
        self.assertEqual(result.summary_text, "複製完成")

    def test_result_model_has_stable_qml_roles(self) -> None:
        from qt_app.models.rescue_video_result_model import RescueVideoResultModel

        model = RescueVideoResultModel()
        roles = {bytes(value).decode("utf-8") for value in model.roleNames().values()}
        self.assertEqual(
            roles,
            {"sourceText", "timeText", "caseText", "statusText", "destinationText", "noteText", "tone"},
        )
        model.replace_rows([{"sourceText": "video001.TS", "statusText": "預計複製"}])
        self.assertEqual(model.rowCount(), 1)
        self.assertEqual(model.data(model.index(0, 0), model.SourceTextRole), "video001.TS")


class ScheduleRepositoryTests(unittest.TestCase):
    def test_available_dates_preserve_legacy_schedule_and_rehearsal_choices(self) -> None:
        from app_core.schedule_repository import ScheduleRepository

        with tempfile.TemporaryDirectory() as temp_dir:
            runtime_dir = Path(temp_dir)
            schedule_dir = runtime_dir / "schedule"
            rehearsal_dir = runtime_dir / "rehearsal"
            schedule_dir.mkdir()
            rehearsal_dir.mkdir()
            (schedule_dir / "schedule_output_1150728.json").write_text("{}", encoding="utf-8")
            (rehearsal_dir / "rehearsal_output_1150729.json").write_text("{}", encoding="utf-8")
            (schedule_dir / "schedule_output_invalid.json").write_text("{}", encoding="utf-8")

            self.assertEqual(
                ScheduleRepository(runtime_dir).available_dates(max_roc_date="1150730"),
                ["1150728", "1150729"],
            )

    def test_schedule_file_has_priority_over_legacy_rehearsal(self) -> None:
        from app_core.schedule_repository import ScheduleRepository

        with tempfile.TemporaryDirectory() as temp_dir:
            runtime_dir = Path(temp_dir)
            schedule_dir = runtime_dir / "schedule"
            rehearsal_dir = runtime_dir / "rehearsal"
            schedule_dir.mkdir()
            rehearsal_dir.mkdir()
            target_date = "1150729"
            (rehearsal_dir / f"rehearsal_output_{target_date}.json").write_text(
                json.dumps({"target_date": target_date, "actions": [{"kind": "work_log", "time": "08:00", "actor": "10", "target": "10", "source": "legacy", "fields": {}}]}),
                encoding="utf-8",
            )
            schedule_path = schedule_dir / f"schedule_output_{target_date}.json"
            schedule_path.write_text(
                json.dumps({"target_date": target_date, "actions": [{"kind": "work_log", "time": "08:00", "actor": "10", "target": "10", "source": "current", "fields": {}}]}),
                encoding="utf-8",
            )

            snapshot = ScheduleRepository(runtime_dir).load_for_date(target_date)

            self.assertEqual(snapshot.path, schedule_path)
            self.assertEqual(snapshot.data["actions"][0]["source"], "current")

    def test_invalid_schedule_json_raises_safe_load_error(self) -> None:
        from app_core.schedule_repository import ScheduleLoadError, ScheduleRepository

        with tempfile.TemporaryDirectory() as temp_dir:
            runtime_dir = Path(temp_dir)
            schedule_dir = runtime_dir / "schedule"
            schedule_dir.mkdir()
            (schedule_dir / "schedule_output_1150729.json").write_text("{broken", encoding="utf-8")

            with self.assertRaisesRegex(ScheduleLoadError, "排程資料無法讀取"):
                ScheduleRepository(runtime_dir).load_for_date("1150729")

    def test_explicit_preview_file_uses_payload_date_and_read_only_validation(self) -> None:
        from app_core.schedule_repository import ScheduleRepository

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            runtime_dir = root / "runtime_outputs"
            preview_path = root / "selected_preview.json"
            preview_path.write_text(
                json.dumps(
                    {
                        "target_date": "1150730",
                        "today": {"staff": {"10": {"name": "本班"}}},
                        "actions": [
                            {
                                "kind": "work_log",
                                "time": "09:00",
                                "actor": "10",
                                "target": "10",
                                "fields": {"勤務項目": "巡邏"},
                            }
                        ],
                    }
                ),
                encoding="utf-8",
            )

            snapshot = ScheduleRepository(runtime_dir).load_path(preview_path)

            self.assertEqual(snapshot.path, preview_path)
            self.assertEqual(snapshot.target_roc_date, "1150730")
            self.assertEqual(snapshot.data["actions"][0]["fields"]["勤務項目"], "巡邏")


class ScheduleCaptureServiceTests(unittest.TestCase):
    def test_validation_allows_initial_capture_before_actor_no_is_known(self) -> None:
        from app_core.schedule_capture_service import ScheduleCaptureRequest, ScheduleCaptureService

        service = ScheduleCaptureService(Path("package"))

        request = service.validate(ScheduleCaptureRequest("user10", "secret", "", "1150729"))

        self.assertEqual(request.actor_no, "")

    def test_capture_writes_live_schedule_and_comparison_without_credentials(self) -> None:
        from dataclasses import dataclass
        from datetime import date, datetime
        from threading import Event

        from app_core.schedule_capture_service import ScheduleCaptureRequest, ScheduleCaptureService

        @dataclass
        class Sheet:
            date: str
            staff: dict
            rows: list

        @dataclass
        class Action:
            kind: str
            time: str
            actor: str
            target: str
            fields: dict

        events: list[str] = []
        comparison_started = Event()

        def query_duty_sheet(_driver, value):
            events.append(f"sheet:{value}")
            if value == "1150729":
                self.assertTrue(
                    comparison_started.wait(1),
                    "勤務與比對查詢應比照舊 GUI 使用兩個背景工作並行",
                )
            return Sheet(value, {"10": {"name": "本班"}}, [])

        def query_visible_table(_driver, ap_name, _date):
            comparison_started.set()
            events.append(f"query:{ap_name}")
            return []

        automation = SimpleNamespace(
            WORK_LOG_AP="work",
            ENTRY_LOG_AP="entry",
            build_driver=lambda headless: events.append(f"driver:{headless}") or object(),
            login=lambda _driver, user_id, password: events.append(f"login:{user_id}:{len(password)}"),
            parse_roc_date=lambda _value: date(2026, 7, 29),
            roc_date=lambda value: f"{value.year - 1911:03d}{value.month:02d}{value.day:02d}",
            query_duty_sheet=query_duty_sheet,
            query_cases=lambda _driver, value: events.append(f"cases:{value}") or [],
            planned_actions=lambda *_args: [
                Action("work_log", "00:00", "10", "10", {"勤務項目": "巡邏"})
            ],
            query_visible_table=query_visible_table,
            query_authenticated_person_name=lambda _driver, user_id: events.append(f"identity:{user_id}") or "本班",
            quit_driver=lambda _driver: events.append("quit"),
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            service = ScheduleCaptureService(
                Path(temp_dir),
                module_loader=lambda: automation,
                now_factory=lambda: datetime(2026, 7, 29, 9, 0),
                identity_resolver=lambda _driver, _staff: ("", ""),
            )
            request = ScheduleCaptureRequest(
                "user10",
                "session-secret",
                "10",
                "1150729",
                "本班",
            )
            progress: list[str] = []

            snapshot = service.capture(request, status_callback=progress.append)

            self.assertTrue(snapshot.found)
            self.assertTrue(snapshot.path.is_file())
            self.assertEqual(snapshot.target_roc_date, "1150729")
            expected_comparison_dates = {"1150728", "1150729", "1150730"}
            self.assertEqual(set(snapshot.schedule_data_by_date), {"1150729"})
            schedule_path = Path(temp_dir) / "runtime_outputs" / "schedule" / "schedule_output_1150729.json"
            self.assertTrue(schedule_path.is_file())
            for target_date in expected_comparison_dates:
                comparison_path = Path(temp_dir) / "runtime_outputs" / "comparison" / f"comparison_output_{target_date}.json"
                self.assertTrue(comparison_path.is_file())
            self.assertEqual(sum(event.startswith("sheet:") for event in events), 3)
            self.assertEqual(sum(event.startswith("cases:") for event in events), 2)
            self.assertEqual(events.count("query:work"), 3)
            self.assertEqual(events.count("query:entry"), 3)
            self.assertNotIn("identity:user10", events)
            self.assertEqual(events[-1], "quit")
            self.assertNotIn("session-secret", repr(request))
            self.assertNotIn("session-secret", snapshot.path.read_text(encoding="utf-8"))
            self.assertNotIn("authenticated_actor", snapshot.path.read_text(encoding="utf-8"))
            self.assertEqual(snapshot.authenticated_actor_no, "10")
            self.assertEqual(snapshot.authenticated_actor_name, "本班")
            self.assertTrue(progress)

    def test_capture_classifies_login_failure_and_closes_driver(self) -> None:
        from app_core.schedule_capture_service import (
            ScheduleCaptureError,
            ScheduleCaptureRequest,
            ScheduleCaptureService,
        )

        events = []
        automation = SimpleNamespace(
            build_driver=lambda headless: events.append(f"driver:{headless}") or object(),
            login=lambda *_args: (_ for _ in ()).throw(RuntimeError("登入失敗：帳號或密碼")),
            quit_driver=lambda _driver: events.append("quit"),
        )
        service = ScheduleCaptureService(Path("package"), module_loader=lambda: automation)

        with self.assertRaises(ScheduleCaptureError) as caught:
            service.capture(ScheduleCaptureRequest("user10", "session-secret", "10", "1150729"))

        self.assertEqual(caught.exception.error_code, "login_failed")
        self.assertNotIn("session-secret", str(caught.exception))
        self.assertEqual(events.count("driver:True"), 2)
        self.assertEqual(events.count("quit"), 2)

    def test_duty_sheet_query_waits_for_asynchronous_result_rows(self) -> None:
        import duty_rehearsal

        class Driver:
            def __init__(self) -> None:
                self.wait_checks = 0

            def get(self, _url: str) -> None:
                return None

            def execute_script(self, script: str, *_args):
                if "const dateField" in script:
                    return True
                if "return Array.from(document.querySelectorAll('table')).some" in script:
                    self.wait_checks += 1
                    return self.wait_checks >= 2
                if "function cellText(cell)" in script:
                    return {
                        "unit": "新坡分隊",
                        "rows": [{"slot": "8-9", "columns": {"值班": "10"}}],
                        "summary": {},
                        "staff": {"10": {"role": "隊員", "name": "測試員"}},
                        "checkNums": [],
                    }
                return True

        test_case = self

        class ImmediateWait:
            def __init__(self, _driver, _timeout, *, poll_frequency) -> None:
                self.driver = _driver
                self.poll_frequency = poll_frequency

            def until(self, condition):
                first_result = condition(self.driver)
                if self.driver.wait_checks:
                    test_case.assertFalse(first_result)
                    return condition(self.driver)
                return first_result

        driver = Driver()
        with (
            patch("duty_rehearsal.time.sleep"),
            patch("duty_rehearsal.WebDriverWait", ImmediateWait),
        ):
            sheet = duty_rehearsal.query_duty_sheet(driver, "1150808")

        self.assertEqual(driver.wait_checks, 2)
        self.assertEqual(sheet.roc_date, "1150808")
        self.assertEqual(sheet.rows[0].columns["值班"], ["10"])
    def test_capture_persists_redacted_failure_stage(self) -> None:
        from app_core.schedule_capture_service import (
            ScheduleCaptureError,
            ScheduleCaptureRequest,
            ScheduleCaptureService,
        )

        automation = SimpleNamespace(
            build_driver=lambda headless: object(),
            login=lambda *_args: (_ for _ in ()).throw(RuntimeError("login failed: session-secret")),
            quit_driver=lambda _driver: None,
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            service = ScheduleCaptureService(Path(temp_dir), module_loader=lambda: automation)
            request = ScheduleCaptureRequest("user10", "session-secret", "10", "1150808")

            with self.assertRaises(ScheduleCaptureError):
                service.capture_schedule(request)

            diagnostic_path = Path(temp_dir) / "runtime_outputs" / "browser" / "schedule_capture_failure.json"
            diagnostic = json.loads(diagnostic_path.read_text(encoding="utf-8"))

        self.assertEqual(diagnostic["stage"], "login")
        self.assertEqual(diagnostic["exception_type"], "RuntimeError")
        self.assertIn("[REDACTED]", diagnostic["message"])
        self.assertNotIn("session-secret", diagnostic["message"])

    def test_capture_comparisons_persists_precise_redacted_failure_stage(self) -> None:
        from datetime import date

        from app_core.schedule_capture_service import (
            ScheduleCaptureError,
            ScheduleCaptureRequest,
            ScheduleCaptureService,
        )

        def query_visible_table(_driver, ap_name, _target_roc_date):
            if ap_name == "entry":
                raise RuntimeError("entry query failed: session-secret")
            return []

        automation = SimpleNamespace(
            WORK_LOG_AP="work",
            ENTRY_LOG_AP="entry",
            build_driver=lambda headless: object(),
            login=lambda *_args: None,
            parse_roc_date=lambda _value: date(2026, 8, 8),
            roc_date=lambda value: f"{value.year - 1911:03d}{value.month:02d}{value.day:02d}",
            query_visible_table=query_visible_table,
            quit_driver=lambda _driver: None,
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            service = ScheduleCaptureService(Path(temp_dir), module_loader=lambda: automation)
            request = ScheduleCaptureRequest("user10", "session-secret", "10", "1150808")

            with self.assertRaises(ScheduleCaptureError):
                service.capture_comparisons(request)

            diagnostic_path = Path(temp_dir) / "runtime_outputs" / "browser" / "schedule_capture_failure.json"
            diagnostic = json.loads(diagnostic_path.read_text(encoding="utf-8"))

        self.assertEqual(diagnostic["stage"], "comparison_entry_rows_1150807")
        self.assertEqual(diagnostic["exception_type"], "RuntimeError")
        self.assertIn("[REDACTED]", diagnostic["message"])
        self.assertNotIn("session-secret", diagnostic["message"])


class WorkLogSettingsServiceTests(unittest.TestCase):
    def test_save_preserves_case_overrides_and_validates_numbers(self) -> None:
        from app_core.work_log_settings_service import (
            NUMERIC_FIELDS,
            WorkLogSettingsService,
            WorkLogSettingsValidationError,
        )

        current = {key: 1 for key in NUMERIC_FIELDS}
        current.update(
            {
                "important_note": "原記事",
                "case_vehicle_overrides": {"1150729": {"case-1": 3}},
                "future_key": "preserved",
            }
        )
        saved: list[dict] = []
        automation = SimpleNamespace(
            DEFAULT_WORK_LOG_DEFAULTS={**current, "important_note": "預設記事"},
            load_work_log_defaults=lambda: dict(current),
            save_work_log_defaults=lambda payload: saved.append(dict(payload)),
            work_handoff_description=lambda payload, count: f"無線電 {payload['radio_count']}；{payload['important_note']}；{count}",
        )
        service = WorkLogSettingsService(module_loader=lambda: automation)
        values = {key: 2 for key in NUMERIC_FIELDS}

        result = service.save(values, "新記事")

        self.assertEqual(result.values["radio_count"], 2)
        self.assertEqual(saved[0]["case_vehicle_overrides"], {"1150729": {"case-1": 3}})
        self.assertEqual(saved[0]["future_key"], "preserved")
        self.assertIn("無線電 2", service.preview(values, "新記事"))
        with self.assertRaises(WorkLogSettingsValidationError):
            service.save({**values, "radio_count": "-1"}, "記事")

    def test_case_vehicle_items_and_save_override_follow_schedule_data(self) -> None:
        from app_core.work_log_settings_service import NUMERIC_FIELDS, WorkLogSettingsService

        class Case:
            def __init__(self, report_time="", return_time="", category="", raw=None):
                self.report_time = report_time
                self.return_time = return_time
                self.category = category
                self.raw = list(raw or [])

        current = {key: 1 for key in NUMERIC_FIELDS}
        current.update({"important_note": "原記事", "case_vehicle_overrides": {"1150728": {"old": 4}}})
        saved: list[dict] = []

        def case_items(cases, settings, target_date):
            return [
                {
                    "key": f"{target_date}|{case.report_time}|{case.category}",
                    "date": target_date,
                    "report_time": case.report_time,
                    "category": case.category,
                    "default_count": 2,
                    "count": settings.get("case_vehicle_overrides", {}).get(target_date, {}).get(
                        f"{target_date}|{case.report_time}|{case.category}", 2
                    ),
                }
                for case in cases
                if not case.return_time
            ]

        automation = SimpleNamespace(
            CaseRecord=Case,
            DEFAULT_WORK_LOG_DEFAULTS=dict(current),
            load_work_log_defaults=lambda: dict(current),
            save_work_log_defaults=lambda payload: saved.append(dict(payload)),
            work_handoff_description=lambda payload, count: f"出勤 {count} 台；{payload['important_note']}",
            unreturned_case_vehicle_items=case_items,
            roc_date_after=lambda value, offset: "1150728" if offset == -1 else value,
        )
        service = WorkLogSettingsService(module_loader=lambda: automation)
        schedule_data = {
            "target_date": "1150729",
            "yesterday_cases": [
                {"report_time": "22:10", "return_time": "", "category": "火警", "raw": ["A"]}
            ],
            "cases": [
                {"report_time": "09:30", "return_time": "", "category": "緊急救護", "raw": ["B"]}
            ],
        }

        items = service.case_items(schedule_data, current)

        self.assertEqual([item["date"] for item in items], ["1150728", "1150729"])
        values = {key: 2 for key in NUMERIC_FIELDS}
        service.save(values, "新記事", {items[0]["key"]: 3, items[1]["key"]: 1})
        self.assertEqual(saved[0]["case_vehicle_overrides"]["1150728"]["old"], 4)
        self.assertEqual(saved[0]["case_vehicle_overrides"]["1150728"][items[0]["key"]], 3)
        self.assertEqual(saved[0]["case_vehicle_overrides"]["1150729"][items[1]["key"]], 1)
        self.assertEqual(service.preview(values, "新記事", vehicle_out_count=4), "出勤 4 台；新記事")


class WorkLogSettingsControllerTests(unittest.TestCase):
    def test_controller_replaces_empty_primary_values_with_source_defaults(self) -> None:
        from app_core.work_log_settings_service import NUMERIC_FIELDS, WorkLogSettings
        from qt_app.controllers.work_log_settings_controller import WorkLogSettingsController

        empty_values = {key: 0 for key in NUMERIC_FIELDS}
        default_values = {
            **empty_values,
            "radio_count": 34,
            "emergency_vehicles_in_station": 6,
            "ems_case_vehicles": 1,
            "fire_case_vehicles": 2,
            "support_vehicles_in_station": 5,
            "rescue_equipment_in_station": 2,
            "tic_count": 5,
        }

        class Service:
            def __init__(self) -> None:
                self.default_calls = 0

            def load(self):
                return WorkLogSettings(empty_values, "", {})

            def defaults(self):
                self.default_calls += 1
                return WorkLogSettings(default_values, "預設記事", {})

            def case_items(self, _schedule_data, _settings):
                return []

            def preview(self, values, note, *, vehicle_out_count=0):
                return f"{values['radio_count']} {note}"

        service = Service()
        controller = WorkLogSettingsController(service)

        controller.load()
        self.assertEqual(service.default_calls, 1)
        self.assertEqual(controller.values["radio_count"], 34)
        self.assertEqual(controller.values["emergency_vehicles_in_station"], 6)
        self.assertEqual(controller.values["support_vehicles_in_station"], 5)
        self.assertEqual(controller.values["rescue_equipment_in_station"], 2)
        self.assertEqual(controller.values["tic_count"], 5)

    def test_controller_keeps_the_panel_available_when_source_data_fails(self) -> None:
        from PySide6.QtTest import QSignalSpy

        from app_core.work_log_settings_service import WorkLogSettingsError
        from qt_app.controllers.work_log_settings_controller import WorkLogSettingsController

        class Service:
            def load(self):
                raise WorkLogSettingsError("設定檔無法讀取")

            def defaults(self):
                raise WorkLogSettingsError("預設內容無法讀取")

            def case_items(self, _schedule_data, _settings):
                raise WorkLogSettingsError("案件無法讀取")

            def preview(self, values, note, *, vehicle_out_count=0):
                return f"{values['radio_count']} {note}"

        controller = WorkLogSettingsController(Service())
        error_spy = QSignalSpy(controller.errorOccurred)

        controller.load()

        self.assertEqual(controller.values["radio_count"], 34)
        self.assertEqual(controller.values["emergency_vehicles_in_station"], 6)
        self.assertEqual(controller.values["support_vehicles_in_station"], 5)
        self.assertEqual(controller.values["rescue_equipment_in_station"], 2)
        self.assertEqual(controller.values["tic_count"], 5)
        self.assertEqual(controller.caseItems, [])
        self.assertEqual(error_spy.count(), 0)

    def test_controller_exposes_and_saves_case_vehicle_counts(self) -> None:
        from app_core.work_log_settings_service import NUMERIC_FIELDS, WorkLogSettings
        from qt_app.controllers.work_log_settings_controller import WorkLogSettingsController

        values = {key: 1 for key in NUMERIC_FIELDS}
        persisted = {"case_vehicle_overrides": {}}
        saved: list[dict[str, int]] = []

        class Service:
            def load(self):
                return WorkLogSettings(values, "記事", persisted)

            def defaults(self):
                return WorkLogSettings(values, "預設記事", {"case_vehicle_overrides": {}})

            def case_items(self, schedule_data, settings):
                self.schedule_data = dict(schedule_data)
                self.settings = dict(settings)
                return [
                    {
                        "key": "1150729|09:30|緊急救護",
                        "date": "1150729",
                        "report_time": "09:30",
                        "category": "緊急救護",
                        "default_count": 1,
                        "count": 1,
                    }
                ]

            def save(self, new_values, note, case_vehicle_counts=None):
                saved.append(dict(case_vehicle_counts or {}))
                return WorkLogSettings(dict(new_values), note, persisted)

            def preview(self, new_values, note, *, vehicle_out_count=0):
                return f"案件出勤 {vehicle_out_count} 台；{note}"

        service = Service()
        controller = WorkLogSettingsController(service)
        controller.set_schedule_data({"target_date": "1150729", "cases": [{}]})

        controller.load()

        self.assertEqual(len(controller.caseItems), 1)
        self.assertEqual(controller.caseItems[0]["label"], "115/07/29 09:30 緊急救護")
        self.assertEqual(controller.previewText, "案件出勤 1 台；記事")
        controller.setCaseVehicleCount("1150729|09:30|緊急救護", "3")
        self.assertEqual(controller.caseItems[0]["count"], 3)
        self.assertEqual(controller.previewText, "案件出勤 3 台；記事")
        controller.save()
        self.assertEqual(saved, [{"1150729|09:30|緊急救護": 3}])

    def test_save_failure_keeps_panel_state_available_for_correction(self) -> None:
        from PySide6.QtTest import QSignalSpy

        from app_core.work_log_settings_service import (
            NUMERIC_FIELDS,
            WorkLogSettings,
            WorkLogSettingsError,
        )
        from qt_app.controllers.work_log_settings_controller import WorkLogSettingsController

        values = {key: 1 for key in NUMERIC_FIELDS}

        class Service:
            def load(self):
                return WorkLogSettings(values, "記事", {})

            def case_items(self, _schedule_data, _settings):
                return []

            def preview(self, _values, note, *, vehicle_out_count=0):
                return note

            def save(self, _values, _note, _case_vehicle_counts=None):
                raise WorkLogSettingsError("工作紀錄預設內容無法儲存。")

        controller = WorkLogSettingsController(Service())
        saved_spy = QSignalSpy(controller.settingsSaved)
        error_spy = QSignalSpy(controller.errorOccurred)
        controller.load()

        self.assertFalse(controller.save())
        self.assertEqual(saved_spy.count(), 0)
        self.assertEqual(error_spy.at(0)[0], "工作紀錄預設內容無法儲存。")


class ScheduleCaptureWorkerTests(unittest.TestCase):
    def test_worker_finishes_schedule_before_starting_same_account_comparisons(self) -> None:
        from app_core.schedule_capture_service import ScheduleCaptureRequest
        from app_core.schedule_repository import ScheduleSnapshot
        from qt_app.workers.schedule_capture_worker import ScheduleCaptureWorker

        class FakeCaptureService:
            def __init__(self) -> None:
                self.calls: list[str] = []
                self.schedule_active = threading.Event()
                self.comparison_started = threading.Event()

            def capture_schedule(self, request, *, status_callback=None):
                self.calls.append("schedule")
                self.schedule_active.set()
                self.comparison_started.wait(0.1)
                self.schedule_active.clear()
                return ScheduleSnapshot(
                    Path(f"schedule_output_{request.target_roc_date}.json"),
                    {"target_date": request.target_roc_date, "actions": []},
                    request.target_roc_date,
                )

            def capture_comparisons(self, request, *, status_callback=None):
                self.calls.append("comparisons")
                self.comparison_started.set()
                if self.schedule_active.is_set():
                    raise AssertionError("勤務表與比對不得同時登入同一個勤務系統帳號")
                return {request.target_roc_date: {}}

            @staticmethod
            def combine_capture(snapshot, comparison_data):
                return ScheduleSnapshot(
                    snapshot.path,
                    snapshot.data,
                    snapshot.target_roc_date,
                    {},
                    comparison_data=comparison_data,
                )

        service = FakeCaptureService()
        worker = ScheduleCaptureWorker(
            1,
            service,
            ScheduleCaptureRequest("user10", "secret", "10", "1150808"),
        )
        succeeded: list[object] = []
        failed: list[object] = []
        worker.succeeded.connect(lambda *_args: succeeded.append(True))
        worker.failed.connect(lambda *_args: failed.append(True))

        worker.run()

        self.assertEqual(service.calls, ["schedule", "comparisons"])
        self.assertEqual(succeeded, [True])
        self.assertEqual(failed, [])


class ScheduledFolderServiceTests(unittest.TestCase):
    def test_opens_each_scheduled_folder_once_per_minute(self) -> None:
        from datetime import datetime

        from app_core.scheduled_folder_service import ScheduledFolderService

        events: list[object] = []
        with tempfile.TemporaryDirectory() as temp_dir:
            service = ScheduledFolderService(
                Path(temp_dir),
                show_desktop=lambda: events.append("desktop"),
                open_folder=lambda path: events.append(path),
            )

            self.assertIsNone(service.check_and_open(datetime(2026, 7, 29, 16, 29, 59)))
            daily = service.check_and_open(datetime(2026, 7, 29, 16, 30, 0))
            self.assertEqual(daily, Path(temp_dir) / "每日勤務表")
            self.assertTrue(daily.is_dir())
            self.assertIsNone(service.check_and_open(datetime(2026, 7, 29, 16, 30, 45)))
            night = service.check_and_open(datetime(2026, 7, 29, 21, 55, 0))
            self.assertEqual(night, Path(temp_dir) / "夜間勤務")
            self.assertEqual(events, ["desktop", daily, "desktop", night])

            next_day = service.check_and_open(datetime(2026, 7, 30, 16, 30, 0))
            self.assertEqual(next_day, daily)
            self.assertEqual(events[-2:], ["desktop", daily])


class ToolControllerTests(unittest.TestCase):
    def test_tool_model_exposes_catalog_roles(self) -> None:
        from qt_app.controllers.tool_controller import TOOL_CATALOG
        from qt_app.models.tool_model import ToolListModel

        model = ToolListModel(TOOL_CATALOG)
        roles = {bytes(value).decode("utf-8") for value in model.roleNames().values()}

        self.assertEqual(
            roles,
            {"toolId", "label", "description", "statusText", "tone", "available"},
        )
        self.assertEqual(model.rowCount(), 5)

    def test_tool_controller_never_launches_external_tk_tool(self) -> None:
        from PySide6.QtTest import QSignalSpy

        from qt_app.controllers.tool_controller import ToolController

        with tempfile.TemporaryDirectory() as temp_dir:
            package_root = Path(temp_dir)
            tool_dir = package_root / "rescue_video"
            tool_dir.mkdir()
            tool_path = tool_dir / "救護影片分類GUI.py"
            tool_path.write_text("# test tool\n", encoding="utf-8")
            controller = ToolController(package_root)
            error_spy = QSignalSpy(controller.errorOccurred)
            source = (PACKAGE_ROOT / "qt_app" / "controllers" / "tool_controller.py").read_text(
                encoding="utf-8"
            )

            self.assertNotIn("import subprocess", source)
            self.assertNotIn("Popen(", source)
            controller.launch("duty_sheet")
            self.assertEqual(error_spy.count(), 1)

            controller.launch("rescue_video")
            self.assertEqual(error_spy.count(), 2)
            self.assertIn("原生 QML", controller.statusText)

    def test_tool_usage_history_preserves_legacy_json_and_latest_finished_card(self) -> None:
        from datetime import datetime

        from PySide6.QtTest import QSignalSpy

        from qt_app.controllers.tool_controller import ToolController

        with tempfile.TemporaryDirectory() as temp_dir:
            package_root = Path(temp_dir)
            history_path = package_root / "runtime_outputs" / "tool_usage_history.json"
            history_path.parent.mkdir()
            history_path.write_text(
                json.dumps(
                    [
                        {
                            "id": "legacy-entry",
                            "time": "2026-07-29 08:30",
                            "business_roc_date": "1150729",
                            "usage_period": "1150729",
                            "tool_name": "duty_sheet",
                            "tool_label": "勤務表登打",
                            "people": "隊員 測試員",
                            "operator": "隊員 測試員",
                            "report": "已完成",
                        }
                    ],
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            controller = ToolController(
                package_root,
                now_factory=lambda: datetime(2026, 7, 30, 9, 15),
            )
            usage_spy = QSignalSpy(controller.usageChanged)

            self.assertEqual(
                controller.usage("duty_sheet"),
                {
                    "time": "2026-07-29 08:30",
                    "people": "隊員 測試員",
                    "report": "1150729 勤務表已登打完成",
                    "tone": "success",
                },
            )

            controller.record_started(
                "rest_time",
                "休息時間登打",
                "8番 測試員",
                "11508",
            )
            controller.record_finished("rest_time", "completed", "115年08月休息時間完成")

            saved = json.loads(history_path.read_text(encoding="utf-8"))
            self.assertEqual(saved[-1]["tool_name"], "rest_time")
            self.assertEqual(saved[-1]["operator"], "8番 測試員")
            self.assertEqual(saved[-1]["usage_period"], "11508")
            self.assertEqual(saved[-1]["report"], "已完成")
            self.assertEqual(controller.usage("rest_time")["report"], "11508 休息時間已登打完成")
            self.assertEqual(usage_spy.count(), 2)

            controller.record_started(
                "daily_vehicle",
                "車輛保養清點",
                "8番 測試員",
                "1150805",
            )
            controller.record_finished("daily_vehicle", "completed")
            self.assertEqual(
                controller.usage("daily_vehicle")["report"],
                "1150805 車輛保養清點已登打完成",
            )

            controller.record_started(
                "monthly_base",
                "勤務基準表登打",
                "8番 測試員",
                "11508",
            )
            controller.record_finished("monthly_base", "completed")
            self.assertEqual(
                controller.usage("monthly_base")["report"],
                "11508 勤務基準表已登打完成",
            )

    def test_tool_usage_history_hides_rank_but_keeps_the_operator_number_and_name(self) -> None:
        from qt_app.controllers.tool_controller import ToolController
        from qt_app.models.tool_model import ToolUsageListModel

        with tempfile.TemporaryDirectory() as temp_dir:
            controller = ToolController(Path(temp_dir))
            controller.record_started(
                "daily_vehicle",
                "車輛保養清點",
                "8番 隊員 曾彥綸",
                actor_no="8",
            )
            controller.record_finished("daily_vehicle", "completed")

            model = controller.usageModel("daily_vehicle", "", "", "", False)

        self.assertEqual(
            model.data(model.index(0, 0), ToolUsageListModel.PeopleRole),
            "8番 曾彥綸",
        )

    def test_usage_model_shows_latest_daily_result_and_current_monthly_operator(self) -> None:
        from datetime import datetime

        from qt_app.controllers.tool_controller import ToolController
        from qt_app.models.tool_model import ToolUsageListModel

        with tempfile.TemporaryDirectory() as temp_dir:
            controller = ToolController(
                Path(temp_dir),
                now_factory=lambda: datetime(2026, 7, 30, 9, 15),
            )

            controller.record_started(
                "duty_sheet", "勤務表登打", "10番 王小明", actor_no="10", user_id="user10"
            )
            controller.record_finished("duty_sheet", "completed")
            controller.record_started(
                "duty_sheet", "勤務表登打", "11番 李小華", actor_no="11", user_id="user11"
            )
            controller.record_finished("duty_sheet", "failed")
            controller.record_started(
                "duty_sheet", "勤務表登打", "10番 王小明", actor_no="10", user_id="user10"
            )
            controller.record_finished("duty_sheet", "failed")
            controller.record_started(
                "rest_time", "休息時間登打", "10番 王小明", actor_no="10", user_id="user10"
            )
            controller.record_finished("rest_time", "completed")
            controller.record_started(
                "rest_time", "休息時間登打", "11番 李小華", actor_no="11", user_id="user11"
            )
            controller.record_finished("rest_time", "failed")

            daily = controller.usageModel("duty_sheet", "", "", "", False)
            monthly = controller.usageModel("rest_time", "10", "user10", "10番 王小明", True)

            self.assertEqual(daily.rowCount(), 1)
            self.assertEqual(
                daily.data(daily.index(0, 0), ToolUsageListModel.PeopleRole), "10番 王小明"
            )
            self.assertEqual(
                daily.data(daily.index(0, 0), ToolUsageListModel.ResultRole), "失敗"
            )
            self.assertEqual(monthly.rowCount(), 1)
            self.assertEqual(
                monthly.data(monthly.index(0, 0), ToolUsageListModel.PeopleRole), "10番 王小明"
            )
            self.assertEqual(
                monthly.data(monthly.index(0, 0), ToolUsageListModel.ResultRole), "11507 休息時間已登打完成"
            )

    def test_usage_model_monthly_matches_legacy_actor_prefix_after_title_changes(self) -> None:
        from qt_app.controllers.tool_controller import ToolController

        with tempfile.TemporaryDirectory() as temp_dir:
            package_root = Path(temp_dir)
            history_path = package_root / "runtime_outputs" / "tool_usage_history.json"
            history_path.parent.mkdir()
            history_path.write_text(
                json.dumps(
                    [
                        {
                            "time": "2026-07-29 08:30",
                            "tool_name": "rest_time",
                            "operator": "10番 王小明",
                            "report": "已完成",
                        }
                    ],
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            controller = ToolController(package_root)

            model = controller.usageModel("rest_time", "10", "user10", "10番 王小明 隊員", True)

            self.assertEqual(model.rowCount(), 1)


class TrayControllerTests(unittest.TestCase):
    class Window:
        def __init__(self) -> None:
            self.hidden = False
            self.shown = False

        def hide(self) -> None:
            self.hidden = True

        def show(self) -> None:
            self.shown = True

        def raise_(self) -> None:
            pass

        def requestActivate(self) -> None:
            pass

    def test_close_is_intercepted_only_when_system_tray_is_available(self) -> None:
        from PySide6.QtWidgets import QApplication

        from qt_app.controllers.tray_controller import TrayController

        app = QApplication.instance() or QApplication(["test_tray_controller"])
        window = self.Window()
        controller = TrayController(app, tray_available=True)
        controller.attach_window(window)

        self.assertTrue(controller.interceptClose())
        self.assertTrue(window.hidden)

        unavailable_window = self.Window()
        unavailable = TrayController(app, tray_available=False)
        unavailable.attach_window(unavailable_window)
        self.assertFalse(unavailable.interceptClose())
        self.assertFalse(unavailable_window.hidden)

        window.hidden = False
        controller.initialize(window)
        action_labels = [action.text() for action in controller._menu.actions() if not action.isSeparator()]
        self.assertEqual(action_labels, ["顯示 SinpoSmart", "縮小到背景", "結束程式"])
        next(action for action in controller._menu.actions() if action.text() == "縮小到背景").trigger()
        self.assertTrue(window.hidden)
        controller.shutdown()

    def test_notification_prefers_native_windows_toast_then_uses_system_tray(self) -> None:
        from PySide6.QtWidgets import QApplication

        from qt_app.controllers.tray_controller import TrayController

        app = QApplication.instance() or QApplication(["test_tray_notification"])
        sent: list[tuple[str, str]] = []
        controller = TrayController(
            app,
            tray_available=False,
            native_notifier=lambda title, message: sent.append((title, message)) or True,
        )
        controller.notify("SinpoSmart", "原生通知")
        self.assertEqual(sent, [("SinpoSmart", "原生通知")])

        class FakeTray:
            def __init__(self) -> None:
                self.messages: list[tuple[str, str]] = []

            def showMessage(self, title, message, *_args) -> None:
                self.messages.append((title, message))

        fallback = TrayController(
            app,
            tray_available=True,
            native_notifier=lambda _title, _message: False,
        )
        fake_tray = FakeTray()
        fallback._tray = fake_tray
        fallback._available = True
        fallback.notify("SinpoSmart", "系統匣備援")
        self.assertEqual(fake_tray.messages, [("SinpoSmart", "系統匣備援")])


class UpdateControllerTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        from PySide6.QtWidgets import QApplication

        cls.app = QApplication.instance() or QApplication(["test_update_controller"])

    def test_update_repository_compares_valid_release_versions(self) -> None:
        from app_core.update_repository import UpdateRepository

        with tempfile.TemporaryDirectory() as temp_dir:
            version_path = Path(temp_dir) / "VERSION.txt"
            version_path.write_text("2026.07.29.1000\n", encoding="utf-8")
            requests: list[tuple[str, int]] = []
            repository = UpdateRepository(
                version_path,
                remote_version_url="https://example.invalid/version.txt",
                text_fetcher=lambda url, timeout: requests.append((url, timeout)) or "2026.07.29.1100",
            )

            info = repository.check()

            self.assertTrue(info.update_available)
            self.assertEqual(info.current_version, "2026.07.29.1000")
            self.assertEqual(info.latest_version, "2026.07.29.1100")
            self.assertEqual(requests, [("https://example.invalid/version.txt", 10)])

    def test_update_controller_checks_in_worker_and_closes_thread(self) -> None:
        from PySide6.QtTest import QSignalSpy, QTest

        from app_core.update_repository import UpdateRepository
        from qt_app.controllers.update_controller import UpdateController

        with tempfile.TemporaryDirectory() as temp_dir:
            version_path = Path(temp_dir) / "VERSION.txt"
            version_path.write_text("2026.07.29.1000\n", encoding="utf-8")
            repository = UpdateRepository(
                version_path,
                text_fetcher=lambda _url, _timeout: "2026.07.29.1100",
            )
            controller = UpdateController(repository)
            spy = QSignalSpy(controller.stateChanged)

            controller.check()

            for _ in range(20):
                if controller.updateAvailable and not controller._workers:
                    break
                spy.wait(250)
                QTest.qWait(10)
            self.assertTrue(controller.updateAvailable)
            self.assertEqual(controller.latestVersion, "2026.07.29.1100")
            self.assertFalse(controller._workers)

    def test_update_controller_launches_existing_script_only_after_update_is_available(self) -> None:
        from app_core.update_repository import UpdateRepository, VersionInfo
        from qt_app.controllers.update_controller import UpdateController

        with tempfile.TemporaryDirectory() as temp_dir:
            package_root = Path(temp_dir)
            version_path = package_root / "VERSION.txt"
            version_path.write_text("2026.07.29.1000\n", encoding="utf-8")
            script_path = package_root / "update_package.ps1"
            script_path.write_text("# test updater\n", encoding="utf-8")
            launched: list[Path] = []
            controller = UpdateController(
                UpdateRepository(version_path),
                process_launcher=lambda path: launched.append(path),
            )

            controller.launchUpdate()
            self.assertEqual(launched, [])
            controller._check_succeeded(
                0,
                VersionInfo("2026.07.29.1000", "2026.07.29.1100", True),
            )
            controller.launchUpdate()

            self.assertEqual(launched, [script_path])
            self.assertIn("已開啟更新程式", controller.statusText)

    def test_update_controller_emits_update_prompt_or_completed_status(self) -> None:
        from PySide6.QtTest import QSignalSpy

        from app_core.update_repository import UpdateRepository, VersionInfo
        from qt_app.controllers.update_controller import UpdateController

        with tempfile.TemporaryDirectory() as temp_dir:
            version_path = Path(temp_dir) / "VERSION.txt"
            version_path.write_text("2026.07.29.1000\n", encoding="utf-8")
            controller = UpdateController(UpdateRepository(version_path))
            ready_spy = QSignalSpy(controller.updateReady)
            completed_spy = QSignalSpy(controller.checkCompleted)

            controller._check_succeeded(
                0,
                VersionInfo("2026.07.29.1000", "2026.07.29.1000", False),
            )
            self.assertEqual(ready_spy.count(), 0)
            self.assertEqual(completed_spy.count(), 1)
            self.assertEqual(completed_spy.at(0)[0], "目前已是最新版")
            controller._check_succeeded(
                0,
                VersionInfo("2026.07.29.1000", "2026.07.29.1100", True),
            )
            self.assertEqual(ready_spy.count(), 1)
            self.assertEqual(ready_spy.at(0)[0], "2026.07.29.1100")
            self.assertEqual(completed_spy.count(), 1)


class DiagnosticsServiceTests(unittest.TestCase):
    def test_issue_package_uses_allowlist_and_excludes_credentials(self) -> None:
        from app_core.diagnostics_service import DiagnosticsService, DiagnosticSnapshot

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            (root / "runtime_outputs" / "schedule").mkdir(parents=True)
            (root / "runtime_outputs" / "schedule" / "schedule_output_1150729.json").write_text(
                '{"target_date":"1150729"}', encoding="utf-8"
            )
            (root / "runtime_outputs" / "browser").mkdir(parents=True)
            (root / "runtime_outputs" / "browser" / "browser_startup.jsonl").write_text(
                '{"category":"startup_timeout"}\n', encoding="utf-8"
            )
            (root / "runtime_outputs" / "sinposmart_operational_sync_status.json").write_text(
                '{"event":{"state":"failed","detail":"safe"}}', encoding="utf-8"
            )
            (root / "VERSION.txt").write_text("2026.07.29.1000", encoding="utf-8")
            (root / ".env").write_text("SECRET=do-not-package", encoding="utf-8")
            (root / "saved_login.json").write_text('{"password":"secret"}', encoding="utf-8")

            package_path = DiagnosticsService(root).export(
                DiagnosticSnapshot(target_date="1150729", session_actor="12", session_verified=True)
            )

            with zipfile.ZipFile(package_path) as archive:
                names = set(archive.namelist())
                manifest = json.loads(archive.read("manifest.json").decode("utf-8"))
            self.assertIn("runtime_outputs/schedule/schedule_output_1150729.json", names)
            self.assertIn("runtime_outputs/browser/browser_startup.jsonl", names)
            self.assertIn("runtime_outputs/sinposmart_operational_sync_status.json", names)
            self.assertIn("VERSION.txt", names)
            self.assertNotIn(".env", names)
            self.assertNotIn("saved_login.json", names)
            self.assertEqual(manifest["session_actor"], "12")
            self.assertNotIn("password", manifest)


class OperationalSyncServiceTests(unittest.TestCase):
    def test_board_payload_is_stable_and_sync_deduplicates_successful_content(self) -> None:
        from app_core.operational_sync_service import OperationalSyncService, build_duty_board_payload

        schedule = {
            "today": {
                "roc_date": "1150729",
                "rows": [{"slot": "8-9", "columns": {"值班": ["12"]}}],
                "staff": {"12": {"name": "測試員"}},
            }
        }
        payload = build_duty_board_payload(schedule)
        posted = []
        with tempfile.TemporaryDirectory() as temp_dir:
            service = OperationalSyncService(Path(temp_dir), board_poster=lambda value: posted.append(value) or {"ok": True})
            service.sync_board_payload(payload)
            self.assertFalse(service.sync_board_async(schedule))

        self.assertEqual(payload["days"][0]["slots"][0]["names"], ["測試員"])
        self.assertEqual(len(posted), 1)
        self.assertEqual(posted[0]["content_hash"], payload["content_hash"])

    def test_board_payload_accepts_legacy_time_separators(self) -> None:
        from app_core.operational_sync_service import build_duty_board_payload

        for slot in ("08-10", "08~10", "08～10"):
            payload = build_duty_board_payload(
                {
                    "today": {
                        "roc_date": "1150806",
                        "rows": [{"slot": slot, "columns": {"值班": ["12"]}}],
                        "staff": {"12": {"name": "測試員"}},
                    }
                }
            )

            self.assertEqual(payload["days"][0]["slots"][0]["slot"], slot)
            self.assertEqual(payload["days"][0]["slots"][0]["start_hour"], 8)
            self.assertEqual(payload["days"][0]["slots"][0]["end_hour"], 10)

    def test_managed_board_sync_is_synchronous_and_deduplicates_content(self) -> None:
        from app_core.operational_sync_service import OperationalSyncService

        schedule = {
            "today": {
                "roc_date": "1150729",
                "rows": [{"slot": "8-9", "columns": {"勤務": ["12"]}}],
                "staff": {"12": {"name": "測試員"}},
            }
        }
        posted = []
        with tempfile.TemporaryDirectory() as temp_dir:
            service = OperationalSyncService(
                Path(temp_dir),
                board_poster=lambda value: posted.append(value) or {"ok": True},
            )

            self.assertTrue(service.sync_board(schedule))
            self.assertFalse(service.sync_board(schedule))

        self.assertEqual(len(posted), 1)

    def test_event_queue_acks_payload_without_sensitive_fields(self) -> None:
        from app_core.operational_sync_service import OperationalSyncService

        posted = []
        with tempfile.TemporaryDirectory() as temp_dir:
            service = OperationalSyncService(
                Path(temp_dir),
                event_poster=lambda payload: posted.append(payload) or {"ok": True, "ack_id": payload["event_id"]},
            )
            payload = service.build_event_payload(
                "login",
                actor_no="12",
                user_id="safe-user",
                content="工具執行完成",
                password="must-not-appear",
                snapshot={"token": "hidden", "result": "ok"},
            )
            service.send_event_payload(payload)
            pending = service.pending_path.read_text(encoding="utf-8")

        serialized = json.dumps(posted[0], ensure_ascii=False)
        self.assertNotIn("must-not-appear", serialized)
        self.assertNotIn("hidden", serialized)
        self.assertEqual(posted[0]["snapshot"]["result"], "ok")
        self.assertEqual(posted[0]["content"], "工具執行完成")
        self.assertEqual(pending, "")

    def test_action_event_payload_retains_legacy_identity_and_action_fields(self) -> None:
        from app_core.operational_sync_service import OperationalSyncService

        with tempfile.TemporaryDirectory() as temp_dir:
            payload = OperationalSyncService(Path(temp_dir)).build_event_payload(
                "action_result",
                status="submitted",
                trigger_type="manual",
                actor_no="10",
                user_id="user10",
                display_name="10番 測試員",
                target="10番 測試員（隊員）",
                action={
                    "kind": "work_log",
                    "source": "值班交接",
                    "target": "10",
                    "time": "08:00",
                    "fields": {"勤務項目": "值班交接", "工作內容": "交接完成"},
                },
            )

        self.assertEqual(payload["display_name"], "10番 測試員")
        self.assertEqual(payload["item_kind"], "工作")
        self.assertIn("值班交接", payload["item_title"])
        self.assertEqual(payload["content"], "交接完成")
        self.assertEqual(payload["target"], "10番 測試員（隊員）")
        self.assertEqual(payload["target_time"], "08:00")

    def test_arrival_event_payload_appends_target_person_to_item_title(self) -> None:
        from app_core.operational_sync_service import OperationalSyncService

        with tempfile.TemporaryDirectory() as temp_dir:
            payload = OperationalSyncService(Path(temp_dir)).build_event_payload(
                "action_result",
                status="submitted",
                trigger_type="due",
                target="10番 測試員",
                action={
                    "kind": "entry_log",
                    "target": "10",
                    "time": "07:55",
                    "fields": {
                        "登打時間": "07:55",
                        "出或入": "入",
                        "領用事由及地點": "到勤",
                    },
                },
            )

        self.assertEqual(payload["item_title"], "入 / 到勤 ｜ 10番 測試員")
        self.assertEqual(payload["target"], "10番 測試員")

    def test_immediate_event_is_persisted_and_acknowledged_before_return(self) -> None:
        from app_core.operational_sync_service import OperationalSyncService

        posted = []
        with tempfile.TemporaryDirectory() as temp_dir:
            service = OperationalSyncService(
                Path(temp_dir),
                event_poster=lambda payload: posted.append(payload) or {"ok": True, "ack_id": payload["event_id"]},
            )
            payload = service.enqueue_event(
                "logout",
                status="ok",
                trigger_type="update",
                actor_no="10",
                user_id="user10",
                content="更新前登出",
                immediate=True,
            )
            pending = service.pending_path.read_text(encoding="utf-8")

        self.assertEqual(posted, [payload])
        self.assertEqual(pending, "")

    def test_event_uses_legacy_default_url_when_only_token_is_configured(self) -> None:
        from app_core.operational_sync_service import OperationalSyncService

        with tempfile.TemporaryDirectory() as temp_dir:
            with patch.dict(
                os.environ,
                {"SINPOSMART_CREDENTIAL_SYNC_TOKEN": "test-token"},
                clear=True,
            ):
                service = OperationalSyncService(Path(temp_dir))

                self.assertTrue(service.event_enabled)
                self.assertEqual(
                    service.event_url,
                    "http://10.30.65.30:8080/api/sinposmart/events",
                )

    def test_event_failure_is_safely_recorded_and_kept_for_retry(self) -> None:
        from app_core.operational_sync_service import OperationalSyncService

        def fail_post(_payload):
            raise RuntimeError("connection failure secret-password")

        with tempfile.TemporaryDirectory() as temp_dir:
            service = OperationalSyncService(Path(temp_dir), event_poster=fail_post)
            payload = service.build_event_payload("login", actor_no="12")

            service.send_event_payload(payload)

            status = json.loads(service.status_path.read_text(encoding="utf-8"))
            pending = service.pending_path.read_text(encoding="utf-8")

        self.assertEqual(status["event"]["state"], "failed")
        self.assertIn("同步失敗", status["event"]["detail"])
        self.assertNotIn("secret-password", json.dumps(status, ensure_ascii=False))
        self.assertIn(payload["event_id"], pending)

    def test_board_failure_is_safely_recorded_for_later_retry(self) -> None:
        from app_core.operational_sync_service import OperationalSyncService

        schedule = {
            "today": {
                "roc_date": "1150729",
                "rows": [{"slot": "8-9", "columns": {"值班": ["12"]}}],
                "staff": {"12": {"name": "測試員"}},
            }
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            service = OperationalSyncService(
                Path(temp_dir),
                board_poster=lambda _payload: (_ for _ in ()).throw(RuntimeError("board error")),
            )

            self.assertFalse(service.sync_board(schedule))
            status = json.loads(service.status_path.read_text(encoding="utf-8"))

        self.assertEqual(status["board"]["state"], "failed")
        self.assertIn("同步失敗", status["board"]["detail"])


class CredentialSyncServiceTests(unittest.TestCase):
    def test_sync_preserves_legacy_multi_account_payload_without_logging_secrets(self) -> None:
        from app_core.credential_sync_service import CredentialSyncService

        posted = []
        service = CredentialSyncService(poster=lambda payload: posted.append(payload) or {"ok": True})
        count = service.sync(
            [
                {"actor_no": "12", "user_id": "user12", "password": "secret12", "name": "甲"},
                {"actor_no": "13", "user_id": "user13", "password": "secret13", "name": "乙"},
            ],
            sync_code="test-code",
        )

        self.assertEqual(count, 2)
        self.assertEqual(posted[0]["sync_code"], "test-code")
        self.assertEqual(posted[0]["user_id"], "user12")
        self.assertEqual(len(posted[0]["accounts"]), 2)


class QtShellTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        from PySide6.QtWidgets import QApplication

        cls.app = QApplication.instance() or QApplication(["test_qt_shell"])

    def test_qml_visual_literals_are_centralized_in_design_tokens(self) -> None:
        qml_path = PACKAGE_ROOT / "qt_app" / "qml" / "Main.qml"
        design_path = PACKAGE_ROOT / "qt_app" / "qml" / "styles" / "Design.qml"
        components_path = PACKAGE_ROOT / "qt_app" / "qml" / "components"
        qml = qml_path.read_text(encoding="utf-8")
        design = design_path.read_text(encoding="utf-8")
        component_sources = "\n".join(
            path.read_text(encoding="utf-8") for path in sorted(components_path.glob("*.qml"))
        )
        dialog_sources = "\n".join(
            path.read_text(encoding="utf-8")
            for path in sorted((qml_path.parent / "dialogs").glob("*.qml"))
        )
        page_sources = "\n".join(
            path.read_text(encoding="utf-8")
            for path in sorted((qml_path.parent / "pages").glob("*.qml"))
        )
        runtime_start = qml.index("\n    Connections {")
        runtime_qml = qml[runtime_start:]

        self.assertIn('import "styles"', qml)
        self.assertNotIn("readonly property var design: Design", qml)
        self.assertIn("pragma Singleton", design)
        self.assertEqual(
            (design_path.parent / "qmldir").read_text(encoding="utf-8").strip(),
            "singleton Design 1.0 Design.qml",
        )
        self.assertRegex(design, r"#[0-9A-Fa-f]{6,8}")
        qml_without_design = qml + component_sources + dialog_sources + page_sources
        self.assertNotRegex(qml_without_design, r"#[0-9A-Fa-f]{6,8}")
        self.assertNotIn('"transparent"', qml_without_design)
        self.assertNotRegex(qml_without_design, r"font\.pixelSize:\s*\d+")
        self.assertNotRegex(qml_without_design, r"radius:\s*[1-9]\d*")
        self.assertNotRegex(qml_without_design, r"border\.width:\s*\d+")
        self.assertNotRegex(
            runtime_qml,
            r"(?m)^\s+(?:fillColor|hoverColor|strokeColor|textColor):",
        )

    def test_qml_tool_buttons_and_audit_mode_preserve_released_gui_contract(self) -> None:
        qml_root = PACKAGE_ROOT / "qt_app" / "qml"
        design = (qml_root / "styles" / "Design.qml").read_text(encoding="utf-8")
        apple_button = (qml_root / "components" / "AppleButton.qml").read_text(
            encoding="utf-8"
        )
        apple_calendar = (qml_root / "components" / "AppleCalendarButton.qml").read_text(
            encoding="utf-8"
        )
        apple_tab = (qml_root / "components" / "AppleTabButton.qml").read_text(
            encoding="utf-8"
        )
        apple_combo = (qml_root / "components" / "AppleComboBox.qml").read_text(
            encoding="utf-8"
        )
        apple_dialog = (qml_root / "components" / "AppleDialog.qml").read_text(
            encoding="utf-8"
        )
        side_panel = (qml_root / "components" / "ToolSidePanel.qml").read_text(
            encoding="utf-8"
        )
        task_card = (qml_root / "components" / "DutyTaskCard.qml").read_text(
            encoding="utf-8"
        )
        quick_tools = (qml_root / "pages" / "DutyQuickToolsPanel.qml").read_text(
            encoding="utf-8"
        )
        tool_header = (qml_root / "components" / "ToolPanelHeader.qml").read_text(
            encoding="utf-8"
        )
        tool_usage = (qml_root / "components" / "ToolUsageHistory.qml").read_text(
            encoding="utf-8"
        )
        tool_status = (qml_root / "components" / "ToolStatusBar.qml").read_text(
            encoding="utf-8"
        )
        tool_run = (qml_root / "components" / "ToolRunButton.qml").read_text(
            encoding="utf-8"
        )
        audit_panel = (qml_root / "pages" / "AuditFilterPanel.qml").read_text(
            encoding="utf-8"
        )
        task_area = (qml_root / "pages" / "DutyTaskArea.qml").read_text(encoding="utf-8")
        operation_bar = (qml_root / "pages" / "DutyOperationBar.qml").read_text(
            encoding="utf-8"
        )
        main_qml = (qml_root / "Main.qml").read_text(encoding="utf-8")

        for token in (
            "property bool selectedState: false",
            'property string iconKind: ""',
            "Design.appWindowControlSymbolSize",
            "Design.appWindowControlSymbolStrokeWidth",
            'tone === "menu" ? Design.transparent',
            "hoverEnabled: true",
            "appleButton.hovered",
            "appleButton.activeFocus",
            "property string disabledHint: \"\"",
            "scale: down ? 0.98 : 1",
            "ToolTip.visible: appleButton.hovered && !appleButton.enabled",
            "selectedState: dutyQuickToolsPanel.isSelected",
            'objectName: "toolPanelBackButton"',
            "property var closeAction: null",
        ):
            self.assertIn(token, apple_button + quick_tools + tool_header)
        for token in (
            "hoverEnabled: true",
            "appleCombo.hovered",
            "comboDelegate.hovered",
            "highlighted: appleCombo.highlightedIndex === comboDelegate.index",
            "Design.buttonFeedbackDuration",
            "Design.buttonColorTransitionDuration",
        ):
            self.assertIn(token, apple_tab + apple_combo)
        self.assertNotIn(
            "comboDelegate.index === appleCombo.currentIndex ? Design.comboSelected",
            apple_combo,
        )
        self.assertNotIn(
            "comboDelegate.hovered || comboDelegate.highlighted",
            apple_combo,
        )
        self.assertNotIn("Behavior on color", apple_combo)
        self.assertNotIn("Behavior on border.color", apple_combo)
        self.assertIn("property string acceptText", apple_dialog)
        self.assertIn("property string acceptTone", apple_dialog)
        self.assertIn("footer: Item", apple_dialog)
        self.assertIn("height: Design.borderWidth", apple_dialog)
        self.assertNotIn("footer: Rectangle", apple_dialog)
        self.assertIn("tone: appleDialog.acceptTone", apple_dialog)
        self.assertIn("property bool hasBeenOpened: false", side_panel)
        self.assertIn("visible: opened || (hasBeenOpened && opacity > 0)", side_panel)
        self.assertIn("Design.sidePanelTransitionDuration", side_panel)
        self.assertIn("readonly property int sidePanelTransitionDuration", design)
        self.assertNotIn('text: "✓"', task_card)
        self.assertIn('objectName: "selectedTaskActions"', task_area)
        self.assertIn("selectedTaskCount > 0", task_area)
        self.assertIn("toolSidePanelCloseTimer", main_qml)
        for token in (
            "DayOfWeekRow {",
            "MonthGrid {",
            "signal dateSelected(string value)",
            'dateFormat === "roc"',
            'dateFormat === "iso"',
            "calendarPopup.open()",
            "calendarPopup.close()",
        ):
            self.assertIn(token, apple_calendar)
        self.assertIn("visible: auditFilterPanel.auditModeActive", audit_panel)
        self.assertNotIn("auditPreview", audit_panel)
        self.assertIn("dutyTaskArea.modeIndex === 1", task_area)
        self.assertIn('objectName: "auditEmptyState"', task_area)
        self.assertIn("visible: taskList.count === 0", task_area)
        self.assertIn(
            'text: dutyTaskArea.modeIndex === 1 ? "此日期尚無審核任務" : "目前沒有勤務任務"',
            task_area,
        )
        self.assertIn("font.pixelSize: Design.sectionTitleSize", task_area)
        self.assertNotIn("可切換日期或重新查詢。", task_area)
        self.assertNotIn("勤務資料載入後會顯示於此。", task_area)
        self.assertIn(
            "visible: dutyOperationBar.backend.sessionController.isLoggedIn",
            operation_bar,
        )
        for token in (
            'objectName: "modeMenuButton"',
            'objectName: "systemMenuButton"',
            'objectName: "windowMenuButton"',
            'objectName: "modeCommandMenu"',
            'objectName: "systemCommandMenu"',
            'objectName: "windowCommandMenu"',
            'objectName: "checkForUpdatesMenuItem"',
            'objectName: "exportIssuePackageMenuItem"',
            'objectName: "hideToBackgroundMenuItem"',
            'objectName: "logoutMenuItem"',
            'objectName: "quitApplicationMenuItem"',
            "function openModeMenu()",
            "function openSystemMenu()",
            "function openWindowMenu()",
            "function closeOpenMenus()",
            "function toggleModeMenu()",
            "function toggleSystemMenu()",
            "function toggleWindowMenu()",
            "modeMenu.popup(modeMenuButton, 0, modeMenuButton.height + 2)",
            "systemMenu.popup(systemMenuButton, 0, systemMenuButton.height + 2)",
            "modal: false",
            "closePolicy: Popup.CloseOnReleaseOutside | Popup.CloseOnEscape",
        ):
            self.assertIn(token, operation_bar)
        self.assertIn('text: "模式"', operation_bar)
        self.assertIn('text: "系統"', operation_bar)
        self.assertNotIn("▾", operation_bar)
        self.assertNotIn("✓", operation_bar)
        self.assertIn("enabled: !currentMode", operation_bar)
        self.assertIn("hoverEnabled: !currentMode", operation_bar)
        self.assertIn("Design.titleMenuCurrentSurface", operation_bar)
        self.assertIn("Design.titleMenuHover", apple_button)
        self.assertIn("Design.appTitleMenuButtonWidth", operation_bar)
        self.assertIn("Design.appTitleMenuButtonHeight", operation_bar)
        self.assertNotIn("visible: dutyOperationBar.modeIndex === 1", operation_bar)
        self.assertIn("cornerRadius: implicitHeight / 2", tool_header)
        self.assertNotIn("padding: 10", tool_usage)
        self.assertIn("spacing: 8", tool_usage)
        self.assertEqual(tool_usage.count("ToolFieldLabel {"), 6)
        self.assertIn("readonly property int toolFieldLabelWidth: 50", design)
        self.assertIn("visible: modeTabs.currentIndex === 0", main_qml)
        self.assertNotIn("auditPreviewDialog", main_qml)
        self.assertIn("window.backend.openAuditMode()", main_qml)
        for token in (
            'text: "上次使用"',
            'text: "時間"',
            'text: "人員"',
            'text: "結果"',
            "ToolFormCard {",
            "color: Design.divider",
            'id: emptyUsageContent',
        ):
            self.assertIn(token, tool_usage)
        self.assertIn("markerColor: Design.successAction", tool_usage)
        work_log_panel = (qml_root / "pages" / "WorkLogSettingsPanel.qml").read_text(encoding="utf-8")
        self.assertEqual(work_log_panel.count("ToolFormCard {"), 3)
        for object_name in (
            "workLogDefaultsCard",
            "workLogCaseCard",
            "workLogPreviewCard",
        ):
            self.assertIn(f'objectName: "{object_name}"', work_log_panel)
        self.assertIn('text: "未返隊案件出勤估算"', work_log_panel)
        self.assertIn('text: "工作紀錄預覽"', work_log_panel)
        self.assertIn("wrapMode: Text.WrapAnywhere", work_log_panel)
        self.assertIn("id: workLogSettingsScroll", work_log_panel)
        self.assertIn("contentWidth: availableWidth", work_log_panel)
        self.assertIn("width: workLogSettingsScroll.availableWidth", work_log_panel)
        self.assertIn("ScrollBar.horizontal.policy: ScrollBar.AlwaysOff", work_log_panel)
        for panel in (
            (qml_root / "pages" / "DutySheetToolPanel.qml").read_text(encoding="utf-8"),
            (qml_root / "pages" / "RestTimeToolPanel.qml").read_text(encoding="utf-8"),
            (qml_root / "pages" / "MonthlyBaseToolPanel.qml").read_text(encoding="utf-8"),
            (qml_root / "pages" / "DailyVehicleToolPanel.qml").read_text(encoding="utf-8"),
        ):
            self.assertLess(panel.index("ToolUsageHistory {"), panel.index("ToolRunButton"))
            self.assertLess(panel.index("ToolStatusBar {"), panel.index("ToolRunButton"))
            self.assertNotIn("ToolCloseButton {", panel)
            if "Design." in panel:
                self.assertIn('import "../styles"', panel)
        for token in (
            "sideStatusReadySurface",
            "sideStatusProgressSurface",
            "sideStatusSuccessSurface",
            "sideStatusWarningSurface",
            "sideStatusErrorSurface",
            "statusCategory",
        ):
            self.assertIn(token, tool_status)
        self.assertIn("implicitHeight: Design.toolRunButtonHeight", tool_run)
        self.assertIn('tone: "primary"', tool_run)

    def test_qml_runtime_uses_global_control_and_tool_styles(self) -> None:
        qml = (PACKAGE_ROOT / "qt_app" / "qml" / "Main.qml").read_text(encoding="utf-8")
        components_path = PACKAGE_ROOT / "qt_app" / "qml" / "components"
        runtime_start = qml.index("\n    Connections {")
        runtime_qml = qml[runtime_start:]

        self.assertNotRegex(
            runtime_qml,
            r"(?m)^\s*(?:Button|TextField|TextArea|ComboBox|CheckBox)\s*\{",
        )
        self.assertNotRegex(runtime_qml, r"(?m)^\s*Dialog\s*\{")
        self.assertIn('import "components"', qml)
        text_field_style = (components_path / "AppleTextField.qml").read_text(encoding="utf-8")
        self.assertIn("color: enabled ? Design.text : Design.muted", text_field_style)
        self.assertIn("placeholderTextColor: Design.muted", text_field_style)
        self.assertEqual(
            (components_path / "qmldir").read_text(encoding="utf-8").splitlines(),
            [
                "AppleButton 1.0 AppleButton.qml",
                "AppleCalendarButton 1.0 AppleCalendarButton.qml",
                "AppleCheckBox 1.0 AppleCheckBox.qml",
                "AppleComboBox 1.0 AppleComboBox.qml",
                "AppleDialog 1.0 AppleDialog.qml",
                "AppleTabButton 1.0 AppleTabButton.qml",
                "AppleTextArea 1.0 AppleTextArea.qml",
                "AppleTextField 1.0 AppleTextField.qml",
                "AuditSummaryCard 1.0 AuditSummaryCard.qml",
                "DataSectionTitle 1.0 DataSectionTitle.qml",
                "DataTableCell 1.0 DataTableCell.qml",
                "DangerButton 1.0 DangerButton.qml",
                "DutyActionButton 1.0 DutyActionButton.qml",
                "DutyTaskCard 1.0 DutyTaskCard.qml",
                "DutyTaskStatusPill 1.0 DutyTaskStatusPill.qml",
                "FormFieldTitle 1.0 FormFieldTitle.qml",
                "PrimaryButton 1.0 PrimaryButton.qml",
                "SettingsButton 1.0 SettingsButton.qml",
                "StrongHeaderTitle 1.0 StrongHeaderTitle.qml",
                "ToolAddButton 1.0 ToolAddButton.qml",
                "ToolBrowseButton 1.0 ToolBrowseButton.qml",
                "ToolCloseButton 1.0 ToolCloseButton.qml",
                "ToolDateStepButton 1.0 ToolDateStepButton.qml",
                "ToolFieldLabel 1.0 ToolFieldLabel.qml",
                "ToolFormCard 1.0 ToolFormCard.qml",
                "ToolMonthCombo 1.0 ToolMonthCombo.qml",
                "ToolPanelContent 1.0 ToolPanelContent.qml",
                "ToolPanelHeader 1.0 ToolPanelHeader.qml",
                "ToolPanelTitle 1.0 ToolPanelTitle.qml",
                "ToolRemoveButton 1.0 ToolRemoveButton.qml",
                "ToolRunButton 1.0 ToolRunButton.qml",
                "ToolSectionTitle 1.0 ToolSectionTitle.qml",
                "ToolSidePanel 1.0 ToolSidePanel.qml",
                "ToolStatusBar 1.0 ToolStatusBar.qml",
                "ToolUsageHistory 1.0 ToolUsageHistory.qml",
                "WorkLogValueControl 1.0 WorkLogValueControl.qml",
            ],
        )
        for component_name in (
            "AppleButton",
            "AppleCalendarButton",
            "AppleCheckBox",
            "AppleComboBox",
            "AppleDialog",
            "AppleTabButton",
            "AppleTextArea",
            "AppleTextField",
            "AuditSummaryCard",
            "DataSectionTitle",
            "DataTableCell",
            "DangerButton",
            "DutyActionButton",
            "DutyTaskCard",
            "DutyTaskStatusPill",
            "FormFieldTitle",
            "PrimaryButton",
            "SettingsButton",
            "StrongHeaderTitle",
            "ToolAddButton",
            "ToolBrowseButton",
            "ToolCloseButton",
            "ToolDateStepButton",
            "ToolFieldLabel",
            "ToolFormCard",
            "ToolMonthCombo",
            "ToolPanelContent",
            "ToolPanelHeader",
            "ToolPanelTitle",
            "ToolRemoveButton",
            "ToolRunButton",
            "ToolSectionTitle",
            "ToolSidePanel",
            "ToolStatusBar",
            "ToolUsageHistory",
            "WorkLogValueControl",
        ):
            self.assertTrue((components_path / f"{component_name}.qml").is_file())
            self.assertNotIn(f"component {component_name}:", qml)
        dialogs_path = PACKAGE_ROOT / "qt_app" / "qml" / "dialogs"
        self.assertEqual(
            (dialogs_path / "qmldir").read_text(encoding="utf-8").splitlines(),
            [
                "AccountManagerWindow 1.0 AccountManagerWindow.qml",
                "RescueVideoWindow 1.0 RescueVideoWindow.qml",
                "ActionConfirmations 1.0 ActionConfirmations.qml",
            ],
        )
        self.assertIn('import "dialogs"', qml)
        for dialog_name in (
            "AccountManagerWindow",
            "RescueVideoWindow",
            "ActionConfirmations",
        ):
            self.assertTrue((dialogs_path / f"{dialog_name}.qml").is_file())
            self.assertIn(f"{dialog_name} {{", qml)
        pages_path = PACKAGE_ROOT / "qt_app" / "qml" / "pages"
        self.assertEqual(
            (pages_path / "qmldir").read_text(encoding="utf-8").splitlines(),
            [
                "SessionHeader 1.0 SessionHeader.qml",
                "DutySheetToolPanel 1.0 DutySheetToolPanel.qml",
                "RestTimeToolPanel 1.0 RestTimeToolPanel.qml",
                "MonthlyBaseToolPanel 1.0 MonthlyBaseToolPanel.qml",
                "DailyVehicleToolPanel 1.0 DailyVehicleToolPanel.qml",
                "AuditFilterPanel 1.0 AuditFilterPanel.qml",
                "WorkLogSettingsPanel 1.0 WorkLogSettingsPanel.qml",
                "DutyQuickToolsPanel 1.0 DutyQuickToolsPanel.qml",
                "DutyOperationBar 1.0 DutyOperationBar.qml",
                "DutyTaskArea 1.0 DutyTaskArea.qml",
            ],
        )
        self.assertIn('import "pages"', qml)
        self.assertTrue((pages_path / "SessionHeader.qml").is_file())
        duty_sheet_panel = (pages_path / "DutySheetToolPanel.qml").read_text(
            encoding="utf-8"
        )
        rest_time_panel = (pages_path / "RestTimeToolPanel.qml").read_text(
            encoding="utf-8"
        )
        monthly_base_panel = (pages_path / "MonthlyBaseToolPanel.qml").read_text(
            encoding="utf-8"
        )
        daily_vehicle_panel = (pages_path / "DailyVehicleToolPanel.qml").read_text(
            encoding="utf-8"
        )
        audit_filter_panel = (pages_path / "AuditFilterPanel.qml").read_text(
            encoding="utf-8"
        )
        work_log_settings_panel = (
            pages_path / "WorkLogSettingsPanel.qml"
        ).read_text(encoding="utf-8")
        duty_quick_tools_panel = (
            pages_path / "DutyQuickToolsPanel.qml"
        ).read_text(encoding="utf-8")
        duty_operation_bar = (pages_path / "DutyOperationBar.qml").read_text(
            encoding="utf-8"
        )
        duty_task_area = (pages_path / "DutyTaskArea.qml").read_text(encoding="utf-8")
        self.assertTrue((pages_path / "DutySheetToolPanel.qml").is_file())
        self.assertTrue((pages_path / "RestTimeToolPanel.qml").is_file())
        self.assertTrue((pages_path / "MonthlyBaseToolPanel.qml").is_file())
        self.assertTrue((pages_path / "DailyVehicleToolPanel.qml").is_file())
        self.assertTrue((pages_path / "AuditFilterPanel.qml").is_file())
        self.assertTrue((pages_path / "WorkLogSettingsPanel.qml").is_file())
        self.assertTrue((pages_path / "DutyQuickToolsPanel.qml").is_file())
        self.assertTrue((pages_path / "DutyOperationBar.qml").is_file())
        self.assertTrue((pages_path / "DutyTaskArea.qml").is_file())
        self.assertIn("SessionHeader {", qml)

        panel_sources = {
            "dutySheetDialog": duty_sheet_panel,
            "restTimeDialog": rest_time_panel,
            "monthlyBaseDialog": monthly_base_panel,
            "dailyVehicleDialog": daily_vehicle_panel,
        }
        for panel_id, panel_source in panel_sources.items():
            panel = panel_source.split(f"id: {panel_id}", 1)[1]
            self.assertNotIn("background: Rectangle", panel)
            self.assertNotRegex(panel, r"font\.pixelSize:\s*design\.")
            self.assertNotRegex(panel, r"(?m)^\s*tone:")

    def test_default_login_timeout_exceeds_selenium_page_timeout(self) -> None:
        from app_core.credential_repository import CredentialRepository
        from qt_app.controllers.session_controller import SessionController

        with tempfile.TemporaryDirectory() as temp_dir:
            controller = SessionController(
                repository=CredentialRepository(Path(temp_dir) / "saved_login.json", "SinpoSmart", None),
            )

            self.assertGreaterEqual(controller._login_timeout_ms, 120_000)

    def test_qml_close_hides_to_tray_and_show_restores_window(self) -> None:
        from PySide6.QtTest import QTest

        from app_core.credential_repository import CredentialRepository
        from qt_app.controllers.app_controller import AppController
        from qt_app.controllers.tray_controller import TrayController
        from qt_app.main import create_engine

        with tempfile.TemporaryDirectory() as temp_dir:
            tray = TrayController(self.app, tray_available=True)
            controller = AppController(
                repository=CredentialRepository(Path(temp_dir) / "saved_login.json", "SinpoSmart", None),
                credential_sync_service=SimpleNamespace(enabled=False),
                tray_controller=tray,
            )
            engine = create_engine(controller)
            root = engine.rootObjects()[0]
            tray.attach_window(root)

            try:
                self.assertTrue(root.isVisible())
                root.close()
                QTest.qWait(50)
                self.assertFalse(root.isVisible())
                self.assertFalse(tray.quitRequested)

                tray.showWindow()
                QTest.qWait(50)
                self.assertTrue(root.isVisible())
            finally:
                tray._quit_requested = True
                root.close()
                controller.shutdown()

    def test_qml_preserves_legacy_account_manager_and_confirmed_update_launch(self) -> None:
        source = (PACKAGE_ROOT / "qt_app" / "qml" / "Main.qml").read_text(encoding="utf-8")
        account_manager = (
            PACKAGE_ROOT / "qt_app" / "qml" / "dialogs" / "AccountManagerWindow.qml"
        ).read_text(encoding="utf-8")
        rescue_video = (
            PACKAGE_ROOT / "qt_app" / "qml" / "dialogs" / "RescueVideoWindow.qml"
        ).read_text(encoding="utf-8")
        action_confirmations = (
            PACKAGE_ROOT / "qt_app" / "qml" / "dialogs" / "ActionConfirmations.qml"
        ).read_text(encoding="utf-8")
        session_header = (
            PACKAGE_ROOT / "qt_app" / "qml" / "pages" / "SessionHeader.qml"
        ).read_text(encoding="utf-8")
        duty_sheet_panel = (
            PACKAGE_ROOT / "qt_app" / "qml" / "pages" / "DutySheetToolPanel.qml"
        ).read_text(encoding="utf-8")
        rest_time_panel = (
            PACKAGE_ROOT / "qt_app" / "qml" / "pages" / "RestTimeToolPanel.qml"
        ).read_text(encoding="utf-8")
        monthly_base_panel = (
            PACKAGE_ROOT / "qt_app" / "qml" / "pages" / "MonthlyBaseToolPanel.qml"
        ).read_text(encoding="utf-8")
        daily_vehicle_panel = (
            PACKAGE_ROOT / "qt_app" / "qml" / "pages" / "DailyVehicleToolPanel.qml"
        ).read_text(encoding="utf-8")
        audit_filter_panel = (
            PACKAGE_ROOT / "qt_app" / "qml" / "pages" / "AuditFilterPanel.qml"
        ).read_text(encoding="utf-8")
        work_log_settings_panel = (
            PACKAGE_ROOT / "qt_app" / "qml" / "pages" / "WorkLogSettingsPanel.qml"
        ).read_text(encoding="utf-8")
        duty_quick_tools_panel = (
            PACKAGE_ROOT / "qt_app" / "qml" / "pages" / "DutyQuickToolsPanel.qml"
        ).read_text(encoding="utf-8")
        duty_operation_bar = (
            PACKAGE_ROOT / "qt_app" / "qml" / "pages" / "DutyOperationBar.qml"
        ).read_text(encoding="utf-8")
        duty_task_area = (
            PACKAGE_ROOT / "qt_app" / "qml" / "pages" / "DutyTaskArea.qml"
        ).read_text(encoding="utf-8")
        source = "\n".join(
            (
                source,
                account_manager,
                rescue_video,
                action_confirmations,
                session_header,
                duty_sheet_panel,
                rest_time_panel,
                monthly_base_panel,
                daily_vehicle_panel,
                audit_filter_panel,
                work_log_settings_panel,
                duty_quick_tools_panel,
                duty_operation_bar,
                duty_task_area,
            )
        )
        design = (PACKAGE_ROOT / "qt_app" / "qml" / "styles" / "Design.qml").read_text(
            encoding="utf-8"
        )
        entrypoint = (PACKAGE_ROOT / "qt_app" / "main.py").read_text(encoding="utf-8")
        tray_controller = (
            PACKAGE_ROOT / "qt_app" / "controllers" / "tray_controller.py"
        ).read_text(encoding="utf-8")

        self.assertIn(
            "accountManagerWindow.sessionController.deleteSavedAccount(",
            account_manager,
        )
        self.assertNotIn("id: actorNoField", source)
        self.assertNotIn("登入成功後自動查詢勤務番號", source)
        self.assertNotIn('passwordField.text,\n                        "",', source)
        self.assertIn('objectName: "savedAccountManagerButton"', source)
        self.assertIn('objectName: "accountManagerWindow"', account_manager)
        self.assertIn('id: accountGrid', account_manager)
        self.assertIn('objectName: "savedAccountGrid"', account_manager)
        self.assertIn('maximumRows: Math.min(15, accountCount)', account_manager)
        self.assertIn('columnCount: Math.max(2, Math.ceil(accountCount / 15))', account_manager)
        self.assertIn('Layout.row: index % 15', account_manager)
        self.assertIn('Layout.preferredWidth: accountManagerWindow.accountCardWidth', account_manager)
        self.assertIn('function positionInAvailableWorkArea()', account_manager)
        self.assertIn('Qt.callLater(accountManagerWindow.positionInAvailableWorkArea)', account_manager)
        self.assertIn('maximumListHeight', account_manager)
        self.assertIn('objectName: "savedAccountViewport"', account_manager)
        self.assertIn('objectName: "loginStatusLabel"', source)
        self.assertIn('objectName: "loggedInStatusLabel"', source)
        logged_in_status = session_header.split('objectName: "loggedInStatusLabel"', 1)[1].split(
            "DangerButton {", 1
        )[0]
        self.assertIn("font.bold: true", logged_in_status)
        self.assertNotIn('text: window.errorMessage.length > 0', source)
        self.assertIn('text: sessionHeader.backend.sessionController.loginStatus', source)
        self.assertIn('loginStatusTone === "error"', source)
        self.assertIn('loginStatusTone === "warning"', source)
        self.assertNotIn('objectName: "appErrorBanner"', source)
        self.assertIn("function onSavedAccountSelected(_actorNo, userId, password)", source)
        self.assertIn("sessionHeader.passwordText = password", source)
        self.assertNotIn('objectName: "savedPasswordMask"', source)
        self.assertNotIn('placeholderText: "勤務系統帳號"', source)
        self.assertNotIn('placeholderText: "密碼"', source)
        self.assertNotIn("已儲存密碼；留空即可使用", source)
        self.assertIn("if (!window.backend.sessionController.isLoggedIn)", source)
        self.assertIn('title: "SinpoSmart - 帳號管理"', account_manager)
        self.assertIn("flags: Qt.Dialog | Qt.FramelessWindowHint", account_manager)
        self.assertIn('objectName: "accountTitleBar"', account_manager)
        self.assertIn('objectName: "accountTitleCloseButton"', account_manager)
        self.assertNotIn("nativeTitleBarConfigurator", account_manager)
        self.assertNotIn("configureNativeTitleBar(accountManagerWindow)", source)
        self.assertIn("readonly property bool usesCustomTitleBar: true", rescue_video)
        self.assertIn('flags: Qt.Window | Qt.FramelessWindowHint', rescue_video)
        self.assertIn('objectName: "rescueVideoTitleBar"', rescue_video)
        self.assertIn('objectName: "rescueVideoTitleCloseButton"', rescue_video)
        self.assertIn('"登入頁面"', source)
        self.assertIn("maximumWidth: 550", source)
        self.assertIn("maximumHeight: 320 + Design.appTitleBarHeight", source)
        self.assertIn("flags: Qt.Window | Qt.FramelessWindowHint", source)
        self.assertIn("dutyOperationBar.closeOpenMenus()", source)
        self.assertIn('objectName: "titleDragRegion"', source)
        self.assertIn("DragHandler {", source)
        self.assertIn("window.startSystemMove()", source)
        self.assertNotIn("onPositionChanged: function(mouse)", source)
        self.assertNotIn("movingWindow", source)
        self.assertIn('iconKind: "minimize"', source)
        self.assertIn('iconKind: "maximize"', source)
        self.assertIn('iconKind: "close"', source)
        for object_name in (
            "appTitleBar",
            "appTitleLabel",
            "titleMinimizeButton",
            "titleMaximizeButton",
            "titleCloseButton",
        ):
            self.assertIn(f'objectName: "{object_name}"', source)
        self.assertIn('text: "帳號選擇"', account_manager)
        self.assertIn('text: "選擇已儲存帳號，或刪除不再使用的項目。"', account_manager)
        self.assertIn('text: "目前沒有已儲存帳號。"', account_manager)
        self.assertIn('title: "確認刪除"', account_manager)
        self.assertNotIn("id: savedAccountCombo", source)
        for tool_text in (
            "勤務表登打",
            "車輛保養清點",
            "行車紀錄器",
            "休息時間登打",
            "勤務基準表登打",
        ):
            self.assertIn(f'text: "{tool_text}"', source)
        quick_tools = (
            PACKAGE_ROOT / "qt_app" / "qml" / "pages" / "DutyQuickToolsPanel.qml"
        ).read_text(encoding="utf-8")
        self.assertEqual(quick_tools.count("Layout.preferredWidth: 1"), 5)
        self.assertNotIn("Item { Layout.fillWidth: true }", quick_tools)
        self.assertNotIn('objectName: "rescueVideoCopyButton"', rescue_video)
        self.assertIn('text: "預覽分類"', rescue_video)
        self.assertNotIn('text: "執行複製"', rescue_video)
        self.assertIn('text: "複製後刪除已驗證來源"', rescue_video)
        self.assertIn("id: updateConfirmation", source)
        self.assertIn("function onUpdateReady(_latestVersion)", source)
        self.assertIn("actionConfirmations.openUpdateConfirmation()", source)
        self.assertIn("function onCheckCompleted(message)", source)
        self.assertIn("actionConfirmations.openUpdateStatus(message)", source)
        self.assertIn("id: updateStatusDialog", source)
        self.assertIn("updateController.launchUpdate()", source)
        self.assertIn("dutyOperationBar.backend.exportIssuePackage()", source)
        self.assertIn("未返隊案件出勤估算", source)
        self.assertIn("workLogSettingsDialog.controller.caseItems", source)
        self.assertIn("setCaseVehicleCount(", source)
        self.assertIn("onTextEdited:", source)
        self.assertIn("if (text.length > 0 && acceptableInput)", source)
        for theme_value in (
            'readonly property color background: "#F5F7FB"',
            'readonly property color panel: "#FFFFFF"',
            'readonly property color border: "#D7E2F0"',
            'readonly property color text: "#172033"',
            'readonly property color muted: "#64748B"',
            'readonly property color blue: "#2563EB"',
        ):
            self.assertIn(theme_value, design)
        self.assertNotIn("fontFamily", source)
        self.assertNotIn("font.family:", source)
        self.assertIn("font.pixelSize: Design.bodySize", source)
        settings_button = (
            PACKAGE_ROOT / "qt_app" / "qml" / "components" / "SettingsButton.qml"
        ).read_text(encoding="utf-8")
        self.assertIn("implicitWidth: Design.settingsButtonSize", settings_button)
        self.assertIn("implicitHeight: Design.settingsButtonSize", settings_button)
        self.assertIn("border.color: settingsButton.enabled ? Design.controlText : Design.muted", settings_button)
        self.assertNotIn("font.family", settings_button)

        components_path = PACKAGE_ROOT / "qt_app" / "qml" / "components"
        work_log_value_control = (components_path / "WorkLogValueControl.qml").read_text(
            encoding="utf-8"
        )
        apple_text_field = (components_path / "AppleTextField.qml").read_text(encoding="utf-8")
        self.assertIn("TextField {", apple_text_field)
        self.assertIn("HoverHandler {", apple_text_field)
        self.assertIn("cursorShape: Qt.IBeamCursor", apple_text_field)
        self.assertIn("onTextEdited:", work_log_value_control)
        self.assertIn("acceptableInput", work_log_value_control)
        self.assertIn("TextArea {", (components_path / "AppleTextArea.qml").read_text(encoding="utf-8"))
        self.assertIn("ComboBox {", (components_path / "AppleComboBox.qml").read_text(encoding="utf-8"))
        tool_usage = (components_path / "ToolUsageHistory.qml").read_text(encoding="utf-8")
        self.assertIn("ToolFormCard {", tool_usage)
        self.assertIn("markerColor: Design.successAction", tool_usage)
        self.assertIn("readonly property int toolCompactControlHeight: toolBrowseButtonHeight", design)
        self.assertIn("readonly property int toolUsageMaxHeight: 104", design)
        self.assertNotIn("delegate: Rectangle", tool_usage)
        self.assertNotIn("emptyUsageCard", tool_usage)
        self.assertIn('app.setApplicationDisplayName("SinpoSmart")', entrypoint)
        self.assertIn("configure_windows_notification_identity()", entrypoint)
        self.assertIn('APP_USER_MODEL_ID = "TYFD.DutyAutomation"', tray_controller)
        self.assertIn("ensure_windows_notification_shortcut()", tray_controller)
        self.assertIn("self._tray.showMessage(APP_DISPLAY_NAME", tray_controller)
        self.assertNotIn("SetWindowTextW", entrypoint)
        self.assertIn("def schedule_windows_title_bar", entrypoint)
        self.assertIn("QTimer.singleShot(delay", entrypoint)
        self.assertIn("root_window.windowTitleChanged.connect", entrypoint)
        self.assertNotRegex(source, r"(?m)^\s+(?:TextField|TextArea|ComboBox) \{")
        self.assertIn('objectName: "workLogSettingsHeader"', source)
        self.assertIn('text: "工作紀錄預設"', source)
        self.assertNotIn("消防救護車出勤由未返隊案件帶入，例外可調整單筆案件台數。", source)
        self.assertIn("Layout.preferredWidth: Design.workLogValueFieldWidth", work_log_value_control)
        self.assertIn("Layout.preferredHeight: Design.workLogValueFieldHeight", work_log_value_control)
        self.assertIn("required property var settingsController", work_log_value_control)
        self.assertIn("if (workLogSettingsDialog.controller.save())", source)
        self.assertIn(
            "window.backend.workLogSettingsController.load()\n"
            "                workLogSettingsDialog.open()",
            source,
        )
        self.assertNotIn("if (window.backend.workLogSettingsController.load())", source)
        for work_log_label in (
            'text: "無線電"',
            'text: "消防及救護車"',
            'text: "後勤車"',
            'text: "救災器材"',
            'text: "TIC"',
            'text: "重要記事"',
            'text: "未返隊案件出勤估算"',
            'text: "取消"',
        ):
            self.assertIn(work_log_label, source)
        self.assertNotIn('text: "放棄變更"', source)
        self.assertNotIn('settingKey: "ems_case_vehicles"', source)
        self.assertNotIn('settingKey: "fire_case_vehicles"', source)
        self.assertNotIn('label: "救災案件預設車輛"', source)
        self.assertIn("id: auditDetailDialog", source)
        self.assertIn("taskRow.fullDetailText", source)
        self.assertNotIn("actorResolutionCard", source)
        self.assertNotIn("resolveSessionActor", source)
        self.assertNotIn('objectName: "manualPauseButton"', source)
        self.assertNotIn('objectName: "resumeScheduleButton"', source)
        self.assertNotIn('objectName: "unreturnedReturnQueuePanel"', source)
        self.assertIn('objectName: "manualSubmitButton"', source)
        self.assertIn('objectName: "manualSubmissionConfirmation"', source)
        self.assertIn('objectName: "dutyTaskHeader"', source)
        self.assertIn('objectName: "auditTaskHeader"', source)
        for object_name in (
            "auditTodoSummaryCard",
            "auditReviewSummaryCard",
            "auditReadySummaryCard",
            "auditDoneSummaryCard",
            "auditDateCard",
            "auditFilterCard",
            "auditDateField",
            "auditDateCalendarButton",
        ):
            self.assertIn(f'objectName: "{object_name}"', source)
        self.assertIn('dateFormat: "roc"', audit_filter_panel)
        self.assertIn(
            "auditFilterPanel.backend.refreshAuditDate(value)",
            audit_filter_panel,
        )
        self.assertIn("triggerOnly: true", audit_filter_panel)
        self.assertIn("auditDateCalendar.openForCurrentDate()", audit_filter_panel)
        self.assertIn("clickAction: function()", audit_filter_panel)
        for column_text in (
            "時間",
            "類型",
            "任務內容",
            "人員",
            "狀態",
            "比對",
            "登打時間",
            "登打人",
            "對象/服勤",
            "內容",
        ):
            self.assertIn(f'text: "{column_text}"', source)
        for legacy_audit_text in (
            "日期切換",
            "勤務日期",
            "重新查詢",
            "篩選條件",
            "未找到 ",
            "人工確認 ",
            "尚未到點 ",
            "已登打 ",
        ):
            self.assertIn(legacy_audit_text, source)
        self.assertIn(
            '["需處理", "全部", "已登打", "手動", "尚未到點", "疑似異動", "時間近似", "人工確認"]',
            source,
        )
        self.assertIn('["全部", "工作", "出入", "案件工作"]', source)
        self.assertNotIn('text: "只看本人"', source)
        for unauthorized_text in (
            'text: "同步帳密"',
            '"今日任務"',
            '"審核任務"',
            '"點選任務可手動暫停或繼續排程"',
            '"登入後將以即時勤務資料載入今日任務。"',
        ):
            self.assertNotIn(unauthorized_text, source)
        self.assertIn('objectName: "dutyVehicleAddButton"', source)
        self.assertIn('objectName: "dutyVehicleRemoveButton"', source)
        self.assertIn("dutySheetDialog.controller.addVehicleOption(", source)
        self.assertIn("dutySheetDialog.controller.removeVehicleOption(", source)
        self.assertIn('objectName: "rescueVideoDialog"', rescue_video)
        self.assertIn("controller.preparePreview(", rescue_video)
        self.assertIn("controller.prepareDelete(", rescue_video)
        self.assertNotIn('text: "工具中心"', source)
        self.assertNotIn("modeTabs.currentIndex === 2", source)
        for object_name in (
            "modeTabs",
            "dutyModeTab",
            "auditModeTab",
            "auditFilterPanel",
            "auditPreviousDayButton",
            "auditNextDayButton",
            "auditRefreshButton",
            "dutySheetRunButton",
            "dutySheetConfirmation",
            "restTimeRunButton",
            "monthlyBaseRunButton",
            "restMonthlyConfirmation",
            "dailyVehicleRunButton",
            "dailyVehicleConfirmation",
            "dutyQuickToolsPanel",
            "quickDutySheetToolButton",
            "quickDailyVehicleToolButton",
            "quickRescueVideoToolButton",
            "quickRestTimeToolButton",
            "quickMonthlyBaseToolButton",
            "settingsTab",
            "auditDetailTextArea",
            "caseSettingsRepeater",
            "caseSettingsRow",
            "caseVehicleCountField",
            "workLogSettingsDiscardButton",
            "workLogSettingsSaveButton",
        ):
            self.assertIn(f'objectName: "{object_name}"', source)
        self.assertEqual(
            source.count("enabled: !auditFilterPanel.backend.dutyController.isRefreshing"),
            5,
        )
        self.assertIn('text: auditFilterPanel.backend.dutyController.isRefreshing ? "更新中…" : "重新查詢"', source)
        self.assertIn("function showAppError(message)", source)
        self.assertIn("function showDutyStatusError(message)", source)
        self.assertIn("window.errorMessage = normalized", source)
        self.assertNotIn('setOperationalStatus(normalized, "warning")', source)
        self.assertIn("Design.auditRefreshButtonWidth", source)
        self.assertIn("Design.monthlySourceOpenButtonWidth", source)
        self.assertNotIn('objectName: "auditPreviewCard"', source)
        self.assertNotIn("dutyController.loadPreviewPath(auditPreviewField.text)", source)
        self.assertIn("window.backend.openAuditMode()", source)
        duty_sheet_panel = (
            PACKAGE_ROOT / "qt_app" / "qml" / "pages" / "DutySheetToolPanel.qml"
        ).read_text(encoding="utf-8")
        self.assertIn('id: dutyVehicleButtons', duty_sheet_panel)
        duty_vehicle_buttons = duty_sheet_panel[
            duty_sheet_panel.index('id: dutyVehicleButtons'):
        ]
        self.assertEqual(duty_vehicle_buttons.count("Layout.preferredWidth: 0"), 2)
        self.assertIn("Layout.fillWidth: true", duty_sheet_panel)
        rest_time_panel = (
            PACKAGE_ROOT / "qt_app" / "qml" / "pages" / "RestTimeToolPanel.qml"
        ).read_text(encoding="utf-8")
        monthly_base_panel = (
            PACKAGE_ROOT / "qt_app" / "qml" / "pages" / "MonthlyBaseToolPanel.qml"
        ).read_text(encoding="utf-8")
        daily_vehicle_panel = (
            PACKAGE_ROOT / "qt_app" / "qml" / "pages" / "DailyVehicleToolPanel.qml"
        ).read_text(encoding="utf-8")
        self.assertIn("DutySheetToolPanel {\n        id: dutySheetDialog", source)
        self.assertIn("ToolSidePanel {\n    id: dutySheetDialog", duty_sheet_panel)
        tool_side_panel = (
            PACKAGE_ROOT / "qt_app" / "qml" / "components" / "ToolSidePanel.qml"
        ).read_text(encoding="utf-8")
        self.assertIn("RestTimeToolPanel {\n        id: restTimeDialog", source)
        self.assertIn("ToolSidePanel {\n    id: restTimeDialog", rest_time_panel)
        self.assertIn("MonthlyBaseToolPanel {\n        id: monthlyBaseDialog", source)
        self.assertIn("ToolSidePanel {\n    id: monthlyBaseDialog", monthly_base_panel)
        self.assertIn("DailyVehicleToolPanel {\n        id: dailyVehicleDialog", source)
        self.assertIn("ToolSidePanel {\n    id: dailyVehicleDialog", daily_vehicle_panel)
        self.assertIn("RescueVideoWindow {\n        id: rescueVideoDialog", source)
        self.assertIn("Window {", rescue_video)
        self.assertIn('title: "SinpoSmart - 行車紀錄器"', rescue_video)
        self.assertIn('objectName: "rescueVideoWindowTitleLabel"', rescue_video)
        self.assertIn('text: "SinpoSmart - 行車紀錄器"', rescue_video)
        self.assertNotIn("Drawer {", source)
        self.assertIn("readonly property int dutyMainWidth: 550", source)
        self.assertIn("readonly property int auditMainWidth: 780", source)
        self.assertIn("readonly property int toolSideWidth: 400", source)
        self.assertIn("readonly property int dutyExpandedWidth: 964", source)
        self.assertIn('objectName: "mainContentHost"', source)
        self.assertIn("function showToolSidePanel(panel)", source)
        self.assertIn("function hideToolSidePanel(panel)", source)
        self.assertIn("x: hostWindow.dutyMainWidth", tool_side_panel)
        self.assertNotIn("x: hostWindow.dutyMainWidth + 14", tool_side_panel)
        self.assertIn(
            "y: Design.appTitleBarHeight + Design.appContentTopSpacing",
            tool_side_panel,
        )
        self.assertIn(
            "height: hostWindow.height - Design.appTitleBarHeight\n"
            "            - Design.appContentTopSpacing - Design.appContentBottomSpacing",
            tool_side_panel,
        )
        self.assertNotIn("component ToolUsageCard:", source)
        self.assertNotIn("ToolUsageCard {", source)
        usage_component = (
            PACKAGE_ROOT / "qt_app" / "qml" / "components" / "ToolUsageHistory.qml"
        ).read_text(encoding="utf-8")
        self.assertIn('text: "上次使用"', usage_component)
        self.assertIn('text: "時間"', usage_component)
        self.assertIn('text: "人員"', usage_component)
        self.assertIn('text: "結果"', usage_component)
        self.assertIn("function onErrorOccurred(message)", source)
        self.assertIn("function onActionFailed(_index, message, _errorCode)", source)
        self.assertIn(
            'objectName: dutyTaskArea.modeIndex === 1 ? "auditTaskRow" : "dutyTaskRow"',
            source,
        )
        self.assertIn("Accessible.name: caseSettingsRow.modelData.label +", source)
        self.assertIn("!dutyTaskArea.backend.readOnlyAcceptance", source)
        self.assertNotIn("toolRow.toolId", source)
        self.assertIn('QQuickStyle.setStyle("Basic")', (PACKAGE_ROOT / "qt_app" / "main.py").read_text(encoding="utf-8"))

    def test_qml_login_control_sizes_match_finalized_legacy_gui(self) -> None:
        source = (
            PACKAGE_ROOT / "qt_app" / "qml" / "pages" / "SessionHeader.qml"
        ).read_text(encoding="utf-8")
        user_block = source.split('objectName: "loginUserIdField"', 1)[1].split(
            'objectName: "savedAccountManagerButton"', 1
        )[0]
        password_block = source.split('objectName: "loginPasswordField"', 1)[1].split(
            'id: rememberLoginCheck', 1
        )[0]
        remember_block = source.split('id: rememberLoginCheck', 1)[1].split(
            'id: loginButton', 1
        )[0]
        status_block = source.split('objectName: "loginStatusLabel"', 1)[1].split("}", 1)[0]
        self.assertIn("Layout.preferredHeight: 38", user_block)
        self.assertIn("Layout.preferredHeight: 36", password_block)
        self.assertIn("font.pixelSize: Design.captionSize", remember_block)
        self.assertIn("font.pixelSize: Design.controlSize", status_block)
        self.assertNotIn("signal auditModeRequested()", source)
        self.assertNotIn('objectName: "loggedOutAuditModeButton"', source)
        self.assertIn(
            "currentIndex === 1 && !window.backend.sessionController.isLoggedIn",
            (PACKAGE_ROOT / "qt_app" / "qml" / "Main.qml").read_text(encoding="utf-8"),
        )
        self.assertIn(
            "window.backend.setDutyModeActive(currentIndex === 0)",
            (PACKAGE_ROOT / "qt_app" / "qml" / "Main.qml").read_text(encoding="utf-8"),
        )
        self.assertIn(
            "auditFilterPanel.backend.refreshAuditLiveData()",
            (PACKAGE_ROOT / "qt_app" / "qml" / "pages" / "AuditFilterPanel.qml").read_text(
                encoding="utf-8"
            ),
        )

    def test_qml_logged_out_audit_entry_is_not_available(self) -> None:
        from PySide6.QtCore import QObject
        from PySide6.QtTest import QTest

        from qt_app.main import (
            STARTUP_SMOKE_ARG,
            cleanup_acceptance_directory,
            create_app_controller,
            create_engine,
        )

        controller = create_app_controller([STARTUP_SMOKE_ARG])
        engine = create_engine(controller)
        root = engine.rootObjects()[0]
        try:
            QTest.qWait(50)
            audit_button = root.findChild(QObject, "loggedOutAuditModeButton")
            mode_tabs = root.findChild(QObject, "modeTabs")
            self.assertIsNone(audit_button)
            self.assertIsNotNone(mode_tabs)
            mode_tabs.setProperty("currentIndex", 1)
            QTest.qWait(50)
            self.assertEqual(mode_tabs.property("currentIndex"), 0)
        finally:
            root.close()
            controller.shutdown()
            cleanup_acceptance_directory(controller)

    def test_qml_controls_apply_hover_feedback_from_design_tokens(self) -> None:
        from PySide6.QtCore import QObject, QPointF, Qt
        from PySide6.QtTest import QTest

        from qt_app.main import (
            STARTUP_SMOKE_ARG,
            cleanup_acceptance_directory,
            create_app_controller,
            create_engine,
        )

        def background_color(control: QObject) -> str:
            background = control.property("background")
            return background.property("color").name().upper()

        controller = create_app_controller([STARTUP_SMOKE_ARG])
        engine = create_engine(controller)
        root = engine.rootObjects()[0]
        try:
            QTest.qWait(50)
            login_button = root.findChild(QObject, "loginSubmitButton")
            self.assertIsNotNone(login_button)

            login_point = login_button.mapToScene(
                QPointF(login_button.width() / 2, login_button.height() / 2)
            )
            QTest.mouseMove(root, login_point.toPoint())
            QTest.qWait(150)
            self.assertTrue(login_button.property("hovered"))
            self.assertEqual(background_color(login_button), "#1D4ED8")
        finally:
            root.close()
            controller.shutdown()
            cleanup_acceptance_directory(controller)

    def test_qml_duty_task_visuals_use_central_semantic_components(self) -> None:
        source = (PACKAGE_ROOT / "qt_app" / "qml" / "Main.qml").read_text(encoding="utf-8")
        operation_bar = (
            PACKAGE_ROOT / "qt_app" / "qml" / "pages" / "DutyOperationBar.qml"
        ).read_text(encoding="utf-8")
        task_area = (PACKAGE_ROOT / "qt_app" / "qml" / "pages" / "DutyTaskArea.qml").read_text(
            encoding="utf-8"
        )
        design = (PACKAGE_ROOT / "qt_app" / "qml" / "styles" / "Design.qml").read_text(
            encoding="utf-8"
        )
        task_delegate = task_area.split("delegate: DutyTaskCard {", 1)[1].split(
            "RowLayout {\n                    visible: dutyTaskArea.modeIndex === 1", 1
        )[0]
        action_row = task_area.split('id: selectedTaskActions', 1)[1]

        for token in (
            'readonly property color taskReadySurface: "#D1D5DB"',
            'readonly property color taskReadyText: "#111827"',
            'readonly property color taskTriggeredSurface: "#D1FAE5"',
            "readonly property int dutyActionButtonHeight: 38",
            "readonly property int dutyActionButtonWidth: 104",
            "readonly property int dutyModeButtonWidth: 112",
        ):
            self.assertIn(token, design)
        components_path = PACKAGE_ROOT / "qt_app" / "qml" / "components"
        self.assertIn("Rectangle {", (components_path / "DutyTaskCard.qml").read_text(encoding="utf-8"))
        self.assertIn("Rectangle {", (components_path / "DutyTaskStatusPill.qml").read_text(encoding="utf-8"))
        self.assertIn("AppleButton {", (components_path / "DutyActionButton.qml").read_text(encoding="utf-8"))
        self.assertLess(source.index("DutyOperationBar {"), source.index("DutyTaskArea {"))
        self.assertIn("DutyTaskStatusPill {", task_delegate)
        self.assertIn("errorText: taskRow.errorText", task_delegate)
        self.assertIn('objectName: "dutyTaskErrorText"', task_area)
        self.assertIn("implicitWidth: Design.externalReturnManualButtonWidth", action_row)
        self.assertIn(
            'color: taskRow.comparisonText === "尚未到點" ? Design.blueHover',
            task_area,
        )
        self.assertEqual(
            task_area.count('color: taskRow.systemText === "出入" ? Design.blueHover : Design.success'),
            2,
        )
        self.assertIn("selectedTaskCount > 0", action_row)
        self.assertNotIn("canAdjustSelectedSchedule", action_row)
        self.assertIn(
            "enabled: dutyTaskArea.backend.dutyController.canManualSubmitSelected",
            action_row,
        )
        self.assertIn('text: "確認返隊手動登打"', action_row)
        self.assertIn(
            "enabled: dutyTaskArea.backend.dutyController.canConfirmExternalReturnManualSubmissionSelected",
            action_row,
        )
        self.assertEqual(
            action_row.count("visible: !dutyTaskArea.backend.dutyController.hasExternalReturnPauseSelected"),
            1,
        )
        self.assertIn("emphasizedBorder: true", action_row)
        self.assertEqual(action_row.count("DutyActionButton {"), 2)

    def test_qml_duty_sheet_panel_preserves_finalized_legacy_text_and_order(self) -> None:
        source = (
            PACKAGE_ROOT / "qt_app" / "qml" / "pages" / "DutySheetToolPanel.qml"
        ).read_text(encoding="utf-8")
        panel = source.split('id: dutySheetDialog', 1)[1].split(
            'id: dutySheetConfirmation', 1
        )[0]

        for expected in (
            'text: "日期"',
            'text: "中繼車"',
            'text: "救護 1 車"',
            'text: "救護 2 車"',
            'text: "完成後發送勤務表截圖"',
            'text: dutySheetDialog.controller.isRunning ? "啟動中..." : "啟動登打"',
            'objectName: "dutyPreviousDateButton"',
            'objectName: "dutyNextDateButton"',
            'objectName: "dutyDateCalendarButton"',
            'objectName: "dutyVehicleAddFunction"',
            'objectName: "dutyVehicleRemoveFunction"',
            'model: ["攻擊車", "中繼車"]',
            'dateFormat: "slash"',
            "onToggled: dutySheetDialog.controller.setNotificationEnabled(checked)",
            "dutySheetDialog.controller.notificationEnabled)",
            'objectName: "dutySheetStatusBar"',
            'toolId: "duty_sheet"',
            "ToolUsageHistory {",
        ):
            self.assertIn(expected, panel)
        self.assertIn('"attack" : "stop"', panel)
        self.assertIn("dutySheetDialog.controller.stopOptions", panel)
        for unauthorized in (
            'text: "勤務日期"',
            'text: "指揮車"',
            'text: "救護車 1"',
            'text: "救護車 2"',
            'text: "完成後發送群組通知"',
            'text: "取消"',
            '"確認並執行"',
            "ToolUsageCard {",
        ):
            self.assertNotIn(unauthorized, panel)
        self.assertIn("ToolPanelHeader {", panel)
        self.assertEqual(panel.count("ToolFormCard {"), 2)
        self.assertLess(panel.index('text: "來源檔案及日期設定"'), panel.index('text: "主力車設定"'))
        self.assertNotIn('objectName: "dutySheetCloseButton"', panel)
        self.assertLess(panel.index('objectName: "dutySheetStatusBar"'), panel.index('objectName: "dutySheetRunButton"'))

    def test_qml_rescue_video_window_preserves_public_duty_gui_contract(self) -> None:
        source = (
            PACKAGE_ROOT / "qt_app" / "qml" / "dialogs" / "RescueVideoWindow.qml"
        ).read_text(encoding="utf-8")
        panel = source.split('id: rescueVideoWindow', 1)[1].split(
            'id: rescueVideoDeleteConfirmation', 1
        )[0]

        for expected in (
            'title: "SinpoSmart - 行車紀錄器"',
            "width: Design.rescueWindowWidth",
            "height: Design.rescueWindowHeight + Design.appTitleBarHeight",
            'text: "救護行車影片分類"',
            'text: "車號"',
            'text: "日期"',
            'text: "車號由當日案件資料夾自動取得；工作紀錄、報告位置與時間偏移均自動處理。"',
            'text: "自動檢查"',
            'model: rescueVideoWindow.controller.checkCards',
            'objectName: "rescueVideoCheckCard_" + rescueVideoCheckCard.modelData.key',
            'text: "選擇資料夾"',
            'objectName: "rescueVideoVehicleCombo"',
            'objectName: "rescueVideoDateField"',
            'objectName: "rescueVideoDateCalendarButton"',
            'dateFormat: "iso"',
            'objectName: "rescueVideoPreviewButton"',
            'text: "預覽分類"',
            'objectName: "rescueVideoDeleteButton"',
            'text: "複製後刪除已驗證來源"',
            'text: "分類結果"',
            'text: "來源檔案"',
            'text: "校正後時間"',
            'text: "案件資料夾"',
            'text: "狀態"',
            'text: "目的地"',
            'text: "備註"',
            "rescueVideoWindow.controller.isReady",
        ):
            self.assertIn(expected, panel)
        for unauthorized in (
            'objectName: "rescueVideoSourceField"',
            'objectName: "rescueVideoDestinationField"',
            'objectName: "rescueVideoOffsetField"',
            'objectName: "rescueVideoRepairCheck"',
            'objectName: "rescueVideoCopyButton"',
            'text: "執行複製"',
            'text: "複製並刪除已驗證來源"',
            'text: "修正案件時間不一致"',
            'text: "關閉"',
            '"報表："',
            '"先執行預覽',
            "ToolUsageCard {",
        ):
            self.assertNotIn(unauthorized, panel)
        self.assertNotIn("rescueTitleSize", source)
        self.assertNotIn("ToolUsageHistory {", panel)
        self.assertNotIn('toolId: "rescue_video"', panel)
        self.assertNotIn("Timer {", panel)
        self.assertNotIn("transientParent: hostWindow", source)
        self.assertIn("function positionBesideHost()", source)
        self.assertNotIn("availableGeometry", source)
        for expected in (
            "screenInfo.virtualX",
            "screenInfo.virtualY",
            "screenInfo.desktopAvailableWidth",
            "screenInfo.desktopAvailableHeight",
        ):
            self.assertIn(expected, source)
        main_source = (PACKAGE_ROOT / "qt_app" / "qml" / "Main.qml").read_text(encoding="utf-8")
        self.assertNotIn("availableGeometry", main_source)
        for expected in (
            "function positionInAvailableWorkArea()",
            "function positionDutyWindowAtTopLeft()",
            "Qt.callLater(window.positionDutyWindowAtTopLeft)",
            "screenInfo.virtualX",
            "screenInfo.virtualY",
            "screenInfo.desktopAvailableWidth",
            "screenInfo.desktopAvailableHeight",
        ):
            self.assertIn(expected, main_source)
        components_path = PACKAGE_ROOT / "qt_app" / "qml" / "components"
        self.assertIn(
            "Label {",
            (components_path / "StrongHeaderTitle.qml").read_text(encoding="utf-8"),
        )
        self.assertIn(
            "Label {",
            (components_path / "DataSectionTitle.qml").read_text(encoding="utf-8"),
        )
        self.assertIn('objectName: "rescueVideoTitleLabel"', panel)
        self.assertIn("StrongHeaderTitle {", panel)
        self.assertIn('objectName: "rescueVideoResultTitle"', panel)
        self.assertIn("DataSectionTitle {", panel)
        self.assertNotIn("section: true", panel)
        self.assertIn('title: "選擇記憶卡 DCIM\\\\100CAREC 資料夾"', source)
        self.assertIn('title: "確認刪除記憶卡來源"', source)

    def test_qml_rest_time_panel_preserves_finalized_legacy_contract(self) -> None:
        source = (PACKAGE_ROOT / "qt_app" / "qml" / "Main.qml").read_text(encoding="utf-8")
        design = (PACKAGE_ROOT / "qt_app" / "qml" / "styles" / "Design.qml").read_text(
            encoding="utf-8"
        )
        panel = (
            PACKAGE_ROOT / "qt_app" / "qml" / "pages" / "RestTimeToolPanel.qml"
        ).read_text(encoding="utf-8")

        for expected in (
            'text: "休息時間登打"',
            'text: "來源檔案及月份設定"',
            'text: "登打以勤務表為主；若個人有補欠時數歸還，請自行修正。"',
            'objectName: "restWorkbookTitle"',
            'text: "Excel"',
            'text: "年月"',
            'text: "年"',
            'text: "月"',
            'objectName: "restWorkbookField"',
            'objectName: "restWorkbookBrowseButton"',
            'objectName: "restMonthCombo"',
            'objectName: "restTimeRunButton"',
            'text: restTimeDialog.controller.isRunning ? "啟動中..." : "啟動登打"',
            'objectName: "restTimeStatusBar"',
            'toolId: "rest_time"',
            "currentOperatorOnly: true",
        ):
            self.assertIn(expected, panel)
        for unauthorized in (
            'text: "取消"',
            '"確認並執行"',
            "ToolUsageCard {",
        ):
            self.assertNotIn(unauthorized, panel)
        self.assertNotIn('objectName: "restTimeCloseButton"', panel)
        self.assertLess(panel.index('objectName: "restTimeStatusBar"'), panel.index('objectName: "restTimeRunButton"'))
        self.assertIn("readonly property int toolMonthComboWidth: 78", design)
        self.assertIn("readonly property int toolMonthComboHeight: toolCompactControlHeight", design)
        self.assertIn(
            "onAccepted: window.backend.restMonthlyController.selectRestWorkbook(selectedFile)",
            source,
        )

    def test_qml_monthly_base_panel_preserves_finalized_legacy_contract(self) -> None:
        source = (PACKAGE_ROOT / "qt_app" / "qml" / "Main.qml").read_text(encoding="utf-8")
        panel = (
            PACKAGE_ROOT / "qt_app" / "qml" / "pages" / "MonthlyBaseToolPanel.qml"
        ).read_text(encoding="utf-8")

        for expected in (
            'text: "勤務基準表登打"',
            'text: "來源及月份設定"',
            'text: "登打以 Google 試算表為主；若後續有更改假別，請自行修正。"',
            'objectName: "monthlySourceTitle"',
            'objectName: "monthlySourceLabel"',
            'text: "Google 試算表 / 輪休基準表"',
            'objectName: "monthlySourceOpenButton"',
            'text: "開啟試算表"',
            'Qt.openUrlExternally("https://docs.google.com/spreadsheets/d/1m-zy4KNR8_GMO94dYtFotyWPIvuT_tt32J9l7hhGZt0/edit#gid=1587057625")',
            'text: "年月"',
            'text: "年"',
            'text: "月"',
            'objectName: "monthlyMonthCombo"',
            'objectName: "monthlyBaseRunButton"',
            'text: monthlyBaseDialog.controller.isRunning ? "啟動中..." : "啟動登打"',
            'objectName: "monthlyBaseStatusBar"',
            'toolId: "monthly_base"',
            "currentOperatorOnly: true",
        ):
            self.assertIn(expected, panel)
        for unauthorized in (
            'text: "取消"',
            '"確認並執行"',
            "ToolUsageCard {",
            "資料來源：Google 勤務基準表",
            "人員：",
        ):
            self.assertNotIn(unauthorized, panel)
        self.assertNotIn('objectName: "monthlyBaseCloseButton"', panel)
        self.assertLess(panel.index('objectName: "monthlyBaseStatusBar"'), panel.index('objectName: "monthlyBaseRunButton"'))

    def test_qml_daily_vehicle_panel_preserves_legacy_prompt_and_browser_lifecycle(self) -> None:
        panel = (
            PACKAGE_ROOT / "qt_app" / "qml" / "pages" / "DailyVehicleToolPanel.qml"
        ).read_text(encoding="utf-8")
        confirmation = (
            PACKAGE_ROOT / "qt_app" / "qml" / "dialogs" / "ActionConfirmations.qml"
        ).read_text(encoding="utf-8").split("id: dailyVehicleConfirmation", 1)[1]

        for expected in (
            'text: "車輛保養清點"',
            'objectName: "dailyVehiclePromptText"',
            'text: "會使用目前登入帳密開啟瀏覽器。依序至車輛平日保養檢查清點、定期保養檢查頁，勾選保養（日、週、月、半年）；再至隨車器材清點頁，勾選清點。"',
            'objectName: "dailyVehicleRunButton"',
            'text: dailyVehicleDialog.controller.isRunning ? "啟動中..." : "啟動登打"',
            'objectName: "dailyVehicleStatusBar"',
            'toolId: "daily_vehicle"',
            "ToolUsageHistory {",
        ):
            self.assertIn(expected, panel)
        for unauthorized in (
            "ToolUsageCard {",
            "執行日期：",
            "完成後自動關閉",
            'text: "取消"',
            '"確認並執行"',
        ):
            self.assertNotIn(unauthorized, panel)
        self.assertNotIn('objectName: "dailyVehicleCloseButton"', panel)
        self.assertLess(panel.index('objectName: "dailyVehicleStatusBar"'), panel.index('objectName: "dailyVehicleRunButton"'))
        self.assertIn('title: "車輛保養清點"', confirmation)
        self.assertIn(
            "text: actionConfirmations.backend.dailyVehicleController.confirmationSummary",
            confirmation,
        )
        self.assertIn('"KEEP_BROWSER_OPEN": "true"', (PACKAGE_ROOT / "app_core" / "daily_vehicle_service.py").read_text(encoding="utf-8"))

    def test_qt_application_font_can_render_traditional_chinese(self) -> None:
        from unittest.mock import patch

        from qt_app.main import configure_application_font

        applied = []
        application = SimpleNamespace(setFont=applied.append)
        with patch(
            "qt_app.main.QFontDatabase.families",
            return_value=["Segoe UI", "Microsoft JhengHei UI"],
        ):
            family = configure_application_font(application)

            self.assertEqual(family, "Microsoft JhengHei UI")
        self.assertEqual(applied[0].family(), "Microsoft JhengHei UI")

    def test_qt_application_font_registers_windows_traditional_chinese_fallback(self) -> None:
        from unittest.mock import patch

        from qt_app.main import configure_application_font

        applied = []
        application = SimpleNamespace(setFont=applied.append)
        with (
            patch("qt_app.main.QFontDatabase.families", return_value=[]),
            patch("qt_app.main.Path.is_file", return_value=True),
            patch("qt_app.main.QFontDatabase.addApplicationFont", return_value=7) as add_font,
            patch(
                "qt_app.main.QFontDatabase.applicationFontFamilies",
                return_value=["Microsoft JhengHei UI"],
            ),
        ):
            family = configure_application_font(application)

        self.assertEqual(family, "Microsoft JhengHei UI")
        self.assertEqual(applied[0].family(), "Microsoft JhengHei UI")
        add_font.assert_called_once()

    def test_qt_combo_indicator_is_font_independent(self) -> None:
        combo = (
            PACKAGE_ROOT / "qt_app" / "qml" / "components" / "AppleComboBox.qml"
        ).read_text(encoding="utf-8")
        design = (
            PACKAGE_ROOT / "qt_app" / "qml" / "styles" / "Design.qml"
        ).read_text(encoding="utf-8")

        self.assertIn("Canvas {", combo)
        self.assertIn("onPaint:", combo)
        self.assertNotIn('text: "▾"', combo)
        for token in (
            "comboArrowWidth",
            "comboArrowHeight",
            "comboArrowInset",
            "comboArrowStrokeWidth",
        ):
            self.assertIn(f"Design.{token}", combo)
            self.assertIn(f"readonly property int {token}", design)

    def test_local_instance_server_routes_update_logout_without_showing_window(self) -> None:
        from qt_app.main import show_existing_window_requests

        events: list[str] = []

        class Connection:
            def waitForReadyRead(self, _timeout):
                return True

            def readAll(self):
                return b"update_logout\n"

            def write(self, payload):
                events.append(bytes(payload).decode("utf-8").strip())

            def waitForBytesWritten(self, _timeout):
                return True

            def disconnectFromServer(self):
                events.append("disconnect")

        class Server:
            def __init__(self):
                self.pending = True

            def hasPendingConnections(self):
                return self.pending

            def nextPendingConnection(self):
                self.pending = False
                return Connection()

        controller = SimpleNamespace(
            recordUpdateLogout=lambda: events.append("update_logout") or True,
            trayController=SimpleNamespace(showWindow=lambda: events.append("show")),
        )

        show_existing_window_requests(Server(), controller)

        self.assertEqual(events, ["update_logout", "ok", "disconnect"])

    def test_package_env_does_not_override_existing_values(self) -> None:
        from qt_app.main import load_package_env

        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            (root / ".env").write_text(
                "SINPOSMART_TEST_KEEP=file-value\nSINPOSMART_TEST_ADD=added-value\n",
                encoding="utf-8",
            )
            previous_keep = os.environ.get("SINPOSMART_TEST_KEEP")
            previous_add = os.environ.get("SINPOSMART_TEST_ADD")
            os.environ["SINPOSMART_TEST_KEEP"] = "process-value"
            os.environ.pop("SINPOSMART_TEST_ADD", None)
            try:
                load_package_env(root)
                self.assertEqual(os.environ["SINPOSMART_TEST_KEEP"], "process-value")
                self.assertEqual(os.environ["SINPOSMART_TEST_ADD"], "added-value")
            finally:
                if previous_keep is None:
                    os.environ.pop("SINPOSMART_TEST_KEEP", None)
                else:
                    os.environ["SINPOSMART_TEST_KEEP"] = previous_keep
                if previous_add is None:
                    os.environ.pop("SINPOSMART_TEST_ADD", None)
                else:
                    os.environ["SINPOSMART_TEST_ADD"] = previous_add

    def test_app_controller_opens_claimed_scheduled_folder_off_ui_thread(self) -> None:
        from PySide6.QtTest import QTest

        from qt_app.controllers.app_controller import AppController

        class FakeScheduledFolderService:
            def __init__(self):
                self.claim_count = 0
                self.opened = threading.Event()
                self.folder = None

            def claim_due_folder(self, _now):
                self.claim_count += 1
                return Path("scheduled-folder") if self.claim_count == 1 else None

            def open(self, folder):
                self.folder = folder
                self.opened.set()

        service = FakeScheduledFolderService()
        controller = AppController(scheduled_folder_service=service)
        try:
            controller._check_scheduled_folders()
            self.assertTrue(service.opened.wait(2))
            for _ in range(20):
                if not controller._scheduled_folder_workers:
                    break
                QTest.qWait(25)
            self.assertEqual(service.folder, Path("scheduled-folder"))
            self.assertFalse(controller._scheduled_folder_workers)
            self.assertTrue(controller._scheduled_folder_timer.isActive())
        finally:
            controller.shutdown()

    def test_qt_controllers_do_not_start_unmanaged_python_threads(self) -> None:
        controller_root = PACKAGE_ROOT / "qt_app" / "controllers"

        for path in controller_root.glob("*.py"):
            source = path.read_text(encoding="utf-8")
            self.assertNotIn("threading.Thread", source, path.name)
            self.assertNotIn("import threading", source, path.name)

    def test_hourly_refresh_retries_same_slot_when_capture_is_busy(self) -> None:
        from datetime import datetime as RealDateTime
        from unittest.mock import patch

        from app_core.session import LoginSession
        from qt_app.controllers.app_controller import AppController

        controller = AppController()
        results = iter((False, True))
        calls = []
        controller.dutyController.refresh_live_schedule = (
            lambda *args, **kwargs: calls.append((args, kwargs)) or next(results)
        )
        attempt_id = controller._session_state.begin_login()
        controller._session_state.complete_login(
            attempt_id,
            LoginSession("10", "user10", "secret", verified=True),
        )
        try:
            with patch("qt_app.controllers.app_controller.datetime") as clock:
                clock.now.return_value = RealDateTime(2026, 7, 29, 10, 2)
                controller._refresh_hourly_live_schedule()
                controller._refresh_hourly_live_schedule()
        finally:
            controller.shutdown()

        self.assertEqual(len(calls), 2)
        self.assertEqual(controller._last_hourly_refresh_key, "2026072910")

    def test_hourly_audit_refresh_updates_comparisons_without_refetching_schedule(self) -> None:
        from datetime import datetime as RealDateTime
        from unittest.mock import patch

        from app_core.session import LoginSession
        from qt_app.controllers.app_controller import AppController

        controller = AppController()
        schedule_calls = []
        comparison_calls = []
        controller.dutyController.refresh_live_schedule = (
            lambda *args, **kwargs: schedule_calls.append((args, kwargs)) or True
        )
        controller.dutyController.refresh_live_comparisons = (
            lambda *args, **kwargs: comparison_calls.append((args, kwargs)) or True
        )
        attempt_id = controller._session_state.begin_login()
        controller._session_state.complete_login(
            attempt_id,
            LoginSession("10", "user10", "session-secret", verified=True),
        )
        try:
            controller.setDutyModeActive(False)
            with patch("qt_app.controllers.app_controller.datetime") as clock:
                clock.now.return_value = RealDateTime(2026, 7, 29, 10, 2)
                controller._refresh_hourly_live_schedule()

            self.assertEqual(schedule_calls, [])
            self.assertEqual(len(comparison_calls), 1)
            self.assertEqual(controller._last_hourly_refresh_key, "2026072910")
        finally:
            controller.shutdown()

    def test_evening_capture_writes_missing_tomorrow_schedule_without_replacing_duty_view(self) -> None:
        from datetime import datetime as RealDateTime
        from unittest.mock import patch

        from PySide6.QtTest import QTest
        from app_core.schedule_repository import ScheduleSnapshot, business_roc_date
        from app_core.session import LoginSession
        from qt_app.controllers.app_controller import AppController

        class FakeCaptureService:
            def __init__(self, runtime_dir: Path) -> None:
                self.runtime_dir = runtime_dir
                self.requests = []
                self.comparison_calls = 0

            def capture_schedule(self, request, *, status_callback=None):
                self.requests.append(request)
                if status_callback is not None:
                    status_callback("正在建立明日勤務快照")
                path = self.runtime_dir / "schedule" / f"schedule_output_{request.target_roc_date}.json"
                path.parent.mkdir(parents=True, exist_ok=True)
                path.write_text("{}", encoding="utf-8")
                return ScheduleSnapshot(
                    path,
                    {"target_date": request.target_roc_date, "actions": []},
                    request.target_roc_date,
                    schedule_data_by_date={request.target_roc_date: {"target_date": request.target_roc_date, "actions": []}},
                )

            def capture_comparisons(self, *_args, **_kwargs):
                self.comparison_calls += 1
                raise AssertionError("明日快照不得啟動比較查詢")

        class FakeOperationalSyncService:
            def __init__(self) -> None:
                self.events = []

            def enqueue_event(self, *args, **kwargs):
                self.events.append((args, kwargs))
                return {}

            def sync_board_async(self, _schedule_data):
                return False

        with tempfile.TemporaryDirectory() as temp_dir:
            service = FakeCaptureService(Path(temp_dir))
            operational_sync = FakeOperationalSyncService()
            controller = AppController(
                schedule_capture_service=service,
                operational_sync_service=operational_sync,
            )
            attempt_id = controller._session_state.begin_login()
            controller._session_state.complete_login(
                attempt_id,
                LoginSession("10", "user10", "session-secret", verified=True),
            )
            controller.dutyController.replace_schedule_data(
                {"target_date": "1150729", "today": {"staff": {}}, "actions": []}
            )
            try:
                with patch("qt_app.controllers.app_controller.datetime") as clock:
                    clock.now.return_value = RealDateTime(2026, 7, 29, 18, 1)
                    controller._capture_evening_tomorrow_schedule()
                    for _ in range(40):
                        if service.requests and not controller._tomorrow_schedule_workers:
                            break
                        QTest.qWait(25)
                    controller._capture_evening_tomorrow_schedule()

                self.assertEqual([request.target_roc_date for request in service.requests], ["1150730"])
                self.assertEqual(service.comparison_calls, 0)
                self.assertEqual(operational_sync.events, [])
                self.assertEqual(controller.dutyController.targetDateText, "1150729")
                self.assertFalse(controller.dutyController.isPreviewLoaded)
                self.assertFalse(controller._tomorrow_schedule_workers)
            finally:
                controller.shutdown()

    def test_read_only_acceptance_blocks_submission_and_external_sync(self) -> None:
        from app_core.session import LoginSession
        from qt_app.controllers.app_controller import AppController

        class OperationalSync:
            def __init__(self):
                self.events = []

            def enqueue_event(self, *args, **kwargs):
                self.events.append((args, kwargs))
                return {}

            def sync_board_async(self, _schedule_data):
                self.events.append(("board", {}))
                return True

        operational_sync = OperationalSync()
        controller = AppController(
            operational_sync_service=operational_sync,
            read_only_acceptance=True,
        )
        try:
            attempt_id = controller._session_state.begin_login()
            controller._session_state.complete_login(
                attempt_id,
                LoginSession("10", "test-user", "test-password", verified=True),
            )
            controller._send_operational_event("login", status="ok")
            controller._enqueue_due_tasks([0])
            controller._enqueue_manual_tasks([0])

            self.assertTrue(controller.readOnlyAcceptance)
            self.assertEqual(operational_sync.events, [])
            self.assertFalse(controller.dutyExecutionController.isBusy)
        finally:
            controller.shutdown()

    def test_verified_login_captures_schedule_then_resolves_actor_from_existing_staff(self) -> None:
        from unittest.mock import patch

        from app_core.session import LoginSession
        from qt_app.controllers.app_controller import AppController

        class OperationalSync:
            def enqueue_event(self, *_args, **_kwargs):
                raise AssertionError("只讀驗收不得送出事件")

            def sync_board_async(self, _schedule_data):
                raise AssertionError("只讀驗收不得同步看板")

        controller = AppController(
            operational_sync_service=OperationalSync(),
            read_only_acceptance=True,
        )
        captures: list[tuple[str, str, str, str]] = []
        controller.dutyController.refresh_live_schedule = (
            lambda user_id, password, actor_no, _target="", actor_name="": captures.append(
                (user_id, password, actor_no, actor_name)
            )
        )
        try:
            attempt_id = controller._session_state.begin_login()
            controller._session_state.complete_login(
                attempt_id,
                LoginSession(
                    "",
                    "test-user",
                    "test-password",
                    verified=True,
                    actor_name="測試員",
                ),
            )

            controller._sync_session_actor()
            self.assertEqual(captures, [("test-user", "test-password", "", "測試員")])

            with patch(
                "qt_app.controllers.app_controller.business_roc_date",
                return_value="1150729",
            ):
                controller._live_schedule_captured(
                    {
                        "target_date": "1150729",
                        "today": {
                            "staff": {
                                "09": {"name": "其他人"},
                                "10": {"name": "測 試 員", "role": "隊員"},
                            }
                        },
                        "actions": [],
                    }
                )

            self.assertEqual(controller.sessionController.actorNo, "10")
            self.assertEqual(controller.sessionController.displayName, "10番 測試員")
            self.assertEqual(
                controller.sessionController.loginStatus,
                "已登入：隊員 測 試 員，今日無值班時段。",
            )
            self.assertEqual(controller._synced_actor_no, "10")
            self.assertEqual(controller.dutyController._actor_no, "10")
            self.assertEqual(captures, [("test-user", "test-password", "", "測試員")])
        finally:
            controller.shutdown()

    def test_live_schedule_replaces_actor_when_fire_day_number_changes(self) -> None:
        from unittest.mock import patch

        from app_core.session import LoginSession
        from qt_app.controllers.app_controller import AppController

        controller = AppController(read_only_acceptance=True)
        try:
            attempt_id = controller._session_state.begin_login()
            controller._session_state.complete_login(
                attempt_id,
                LoginSession(
                    "10",
                    "test-user",
                    "test-password",
                    verified=True,
                    actor_name="測試員",
                ),
            )
            controller._synced_actor_no = "10"
            controller._synced_user_id = "test-user"
            controller.dutyController.set_actor_no("10")
            controller.dutyController.enable_auto_execution()

            with patch(
                "qt_app.controllers.app_controller.business_roc_date",
                return_value="1150729",
            ):
                controller._live_schedule_captured(
                    {
                        "target_date": "1150729",
                        "today": {"staff": {"11": {"name": "測試員", "role": "隊員"}}},
                        "actions": [],
                        "_authenticated_actor": {"actor_no": "11", "actor_name": "測試員"},
                    }
                )

            self.assertEqual(controller.sessionController.actorNo, "11")
            self.assertEqual(controller.dutyController._actor_no, "11")
            self.assertEqual(controller._synced_actor_no, "11")
            self.assertFalse(controller.dutyController._auto_execution_enabled)
        finally:
            controller.shutdown()

    def test_live_schedule_disables_auto_execution_when_actor_cannot_be_resolved(self) -> None:
        from unittest.mock import patch

        from app_core.session import LoginSession
        from qt_app.controllers.app_controller import AppController

        controller = AppController(read_only_acceptance=True)
        try:
            attempt_id = controller._session_state.begin_login()
            controller._session_state.complete_login(
                attempt_id,
                LoginSession(
                    "10",
                    "test-user",
                    "test-password",
                    verified=True,
                    actor_name="測試員",
                ),
            )
            controller.dutyController.set_actor_no("10")
            controller.dutyController.enable_auto_execution()
            controller.dutyController.refresh_live_schedule = lambda *_args, **_kwargs: False

            with patch(
                "qt_app.controllers.app_controller.business_roc_date",
                return_value="1150729",
            ):
                controller._live_schedule_captured(
                    {
                        "target_date": "1150729",
                        "today": {"staff": {"11": {"name": "其他人", "role": "隊員"}}},
                        "actions": [],
                    }
                )

            self.assertEqual(controller.sessionController.actorNo, "10")
            self.assertEqual(controller.dutyController._actor_no, "")
            self.assertFalse(controller.dutyController._auto_execution_enabled)
            self.assertEqual(
                controller.sessionController.loginStatus,
                "已登入：登入身分待確認，已暫停自動登打。",
            )
        finally:
            controller.shutdown()

    def test_live_schedule_ignores_a_stale_fire_day_identity_result(self) -> None:
        from unittest.mock import patch

        from app_core.session import LoginSession
        from qt_app.controllers.app_controller import AppController

        controller = AppController(read_only_acceptance=True)
        try:
            attempt_id = controller._session_state.begin_login()
            controller._session_state.complete_login(
                attempt_id,
                LoginSession(
                    "10",
                    "test-user",
                    "test-password",
                    verified=True,
                    actor_name="測試員",
                ),
            )
            controller.dutyController.set_actor_no("10")
            controller.dutyController.enable_auto_execution()

            with patch(
                "qt_app.controllers.app_controller.business_roc_date",
                return_value="1150730",
            ):
                controller._live_schedule_captured(
                    {
                        "target_date": "1150729",
                        "today": {"staff": {"11": {"name": "測試員"}}},
                        "actions": [],
                        "_authenticated_actor": {"actor_no": "11", "actor_name": "測試員"},
                    }
                )

            self.assertEqual(controller.sessionController.actorNo, "10")
            self.assertEqual(controller.dutyController._actor_no, "10")
            self.assertFalse(controller.dutyController._auto_execution_enabled)
        finally:
            controller.shutdown()

    def test_fire_day_change_disables_due_execution_and_requests_current_schedule(self) -> None:
        from app_core.session import LoginSession
        from qt_app.controllers.app_controller import AppController

        controller = AppController()
        calls: list[tuple[tuple[object, ...], dict[str, object]]] = []
        controller.dutyController.refresh_live_schedule = (
            lambda *args, **kwargs: calls.append((args, kwargs)) or True
        )
        try:
            attempt_id = controller._session_state.begin_login()
            controller._session_state.complete_login(
                attempt_id,
                LoginSession("10", "test-user", "test-password", verified=True),
            )
            calls.clear()
            controller.dutyController.enable_auto_execution()
            controller.dutyController.fireDayChanged.emit("1150730")

            self.assertFalse(controller.dutyController._auto_execution_enabled)
            self.assertEqual(len(calls), 1)
            args, kwargs = calls[0]
            self.assertEqual(args[:3], ("test-user", "test-password", "10"))
            self.assertEqual(kwargs["target_roc_date"], "1150730")
        finally:
            controller.shutdown()

    def test_saved_actor_loads_cached_schedule_before_full_live_capture(self) -> None:
        from app_core.session import LoginSession
        from qt_app.controllers.app_controller import AppController

        controller = AppController(read_only_acceptance=True)
        cached_loads: list[bool] = []
        captures: list[tuple[str, str, str, str]] = []
        controller.sessionController._accounts = [
            {"actor_no": "10", "user_id": "test-user", "display_name": "10番 測試員"}
        ]
        controller.dutyController.load_current_schedule = lambda: cached_loads.append(True)
        controller.dutyController.refresh_live_schedule = (
            lambda user_id, password, actor_no, _target="", actor_name="": captures.append(
                (user_id, password, actor_no, actor_name)
            )
        )
        try:
            attempt_id = controller._session_state.begin_login()
            controller._session_state.complete_login(
                attempt_id,
                LoginSession("", "test-user", "test-password", verified=True, actor_name="測試員"),
            )

            controller._sync_session_actor()

            self.assertEqual(controller.sessionController.actorNo, "")
            self.assertEqual(controller.dutyController._actor_no, "10")
            self.assertEqual(cached_loads, [True])
            self.assertEqual(captures, [("test-user", "test-password", "10", "測試員")])

            controller._cached_schedule_loaded(
                {
                    "target_date": "1150729",
                    "today": {
                        "staff": {"10": {"name": "測試員", "role": "隊員"}},
                        "rows": [],
                    },
                    "actions": [],
                }
            )
            self.assertEqual(
                controller.sessionController.loginStatus,
                "已登入：隊員 測試員，今日無值班時段。",
            )
        finally:
            controller.shutdown()

    def test_unknown_actor_still_loads_cached_schedule_before_full_live_capture(self) -> None:
        from app_core.session import LoginSession
        from qt_app.controllers.app_controller import AppController

        controller = AppController(read_only_acceptance=True)
        cached_loads: list[bool] = []
        captures: list[tuple[str, str, str, str]] = []
        controller.dutyController.load_current_schedule = lambda: cached_loads.append(True)
        controller.dutyController.refresh_live_schedule = (
            lambda user_id, password, actor_no, _target="", actor_name="": captures.append(
                (user_id, password, actor_no, actor_name)
            )
        )
        try:
            attempt_id = controller._session_state.begin_login()
            controller._session_state.complete_login(
                attempt_id,
                LoginSession("", "test-user", "test-password", verified=True, actor_name="測試員"),
            )

            controller._sync_session_actor()

            self.assertEqual(cached_loads, [True])
            self.assertEqual(captures, [("test-user", "test-password", "", "測試員")])
        finally:
            controller.shutdown()

    def test_cached_schedule_resolves_unknown_actor_with_role_before_live_capture_finishes(self) -> None:
        from app_core.session import LoginSession
        from qt_app.controllers.app_controller import AppController

        controller = AppController(read_only_acceptance=True)
        controller.sessionController.sessionChanged.disconnect(controller._sync_session_actor)
        try:
            attempt_id = controller._session_state.begin_login()
            controller._session_state.complete_login(
                attempt_id,
                LoginSession("", "test-user", "test-password", verified=True, actor_name="測試員"),
            )
            controller._cached_schedule_loaded(
                {
                    "target_date": "1150730",
                    "today": {
                        "staff": {
                            "09": {"name": "其他人", "role": "隊員"},
                            "10": {"name": "測試員", "role": "分隊長"},
                        },
                        "rows": [],
                    },
                    "actions": [],
                }
            )

            self.assertEqual(controller.sessionController.actorNo, "10")
            self.assertEqual(controller.dutyController._actor_no, "10")
            self.assertEqual(
                controller.sessionController.loginStatus,
                "已登入：分隊長 測試員，今日無值班時段。",
            )
        finally:
            controller.shutdown()

    def test_schedule_capture_metadata_resolves_actor_when_login_page_has_no_name(self) -> None:
        from unittest.mock import patch

        from app_core.session import LoginSession
        from qt_app.controllers.app_controller import AppController

        controller = AppController(read_only_acceptance=True)
        try:
            attempt_id = controller._session_state.begin_login()
            controller._session_state.complete_login(
                attempt_id,
                LoginSession("", "test-user", "test-password", verified=True),
            )
            controller._synced_user_id = "test-user"

            with patch(
                "qt_app.controllers.app_controller.business_roc_date",
                return_value="1150729",
            ):
                controller._live_schedule_captured(
                    {
                        "target_date": "1150729",
                        "today": {"staff": {"10": {"name": "測試員"}}},
                        "actions": [],
                        "_authenticated_actor": {"actor_no": "10", "actor_name": "測試員"},
                    }
                )

            self.assertEqual(controller.sessionController.actorNo, "10")
            self.assertEqual(controller.sessionController.displayName, "10番 測試員")
            self.assertNotIn(
                "_authenticated_actor",
                controller.workLogSettingsController._schedule_data,
            )
        finally:
            controller.shutdown()

    def test_read_only_acceptance_temp_directory_is_explicitly_cleaned(self) -> None:
        from qt_app.main import (
            READ_ONLY_ACCEPTANCE_ARG,
            cleanup_acceptance_directory,
            create_app_controller,
        )

        controller = create_app_controller([READ_ONLY_ACCEPTANCE_ARG])
        temporary_root = Path(controller.acceptance_temporary_directory.name)
        try:
            self.assertTrue(temporary_root.is_dir())
            cleanup_acceptance_directory(controller)

            self.assertFalse(temporary_root.exists())
            self.assertIsNone(controller.acceptance_temporary_directory)
        finally:
            controller.shutdown()
            cleanup_acceptance_directory(controller)

    def test_read_only_acceptance_uses_foreground_chrome_login(self) -> None:
        from qt_app.main import (
            READ_ONLY_ACCEPTANCE_ARG,
            cleanup_acceptance_directory,
            create_app_controller,
        )

        controller = create_app_controller([READ_ONLY_ACCEPTANCE_ARG])
        try:
            options = controller.sessionController._verifier.options_factory()
            self.assertNotIn("--headless=new", options.arguments)
            self.assertIn("--window-size=1280,900", options.arguments)
            self.assertIn("--window-position=80,80", options.arguments)
            self.assertTrue(controller.sessionController._verifier.allow_post_login_lookup_warning)
        finally:
            controller.shutdown()
            cleanup_acceptance_directory(controller)

    def test_duty_identity_label_keeps_role_before_name(self) -> None:
        from qt_app.controllers.app_controller import duty_identity_label

        schedule_data = {
            "today": {
                "staff": {
                    "10": {
                        "name": "驗收人員",
                        "role": "分隊長",
                    }
                }
            }
        }

        self.assertEqual(
            duty_identity_label(schedule_data, "10", "備援名稱"),
            "分隊長 驗收人員",
        )

    def test_audit_date_shift_handles_minguo_year_and_month_boundaries(self) -> None:
        from qt_app.controllers.app_controller import clamp_audit_roc_date, shift_roc_date

        self.assertEqual(shift_roc_date("1150729", 1), "1150730")
        self.assertEqual(shift_roc_date("1150731", 1), "1150801")
        self.assertEqual(shift_roc_date("1150101", -1), "1141231")
        self.assertEqual(clamp_audit_roc_date("1150731", date(2026, 7, 29)), "1150730")
        self.assertEqual(clamp_audit_roc_date("1150728", date(2026, 7, 29)), "1150728")

    def test_historical_capture_never_enables_auto_execution(self) -> None:
        from app_core.schedule_repository import ScheduleSnapshot, business_roc_date
        from qt_app.controllers.duty_controller import DutyController

        controller = DutyController()
        controller.set_actor_no("10")
        controller._active_capture_request = 1
        controller._capture_succeeded(
            1,
            "10",
            ScheduleSnapshot(
                Path("schedule_output_1000101.json"),
                {"target_date": "1000101", "actions": []},
                "1000101",
                {},
            ),
        )

        self.assertFalse(controller._auto_execution_enabled)

    def test_current_capture_without_resolved_actor_keeps_auto_execution_disabled(self) -> None:
        from app_core.schedule_repository import ScheduleSnapshot, business_roc_date
        from qt_app.controllers.duty_controller import DutyController

        target_date = business_roc_date()
        controller = DutyController()
        controller._active_capture_request = 1
        controller._capture_succeeded(
            1,
            "",
            ScheduleSnapshot(
                Path(f"schedule_output_{target_date}.json"),
                {
                    "target_date": target_date,
                    "today": {"staff": {"10": {"name": "測試員"}}},
                    "actions": [],
                },
                target_date,
            ),
        )

        self.assertFalse(controller._auto_execution_enabled)

    def test_current_capture_enables_auto_execution_after_metadata_resolves_actor(self) -> None:
        from app_core.schedule_repository import ScheduleSnapshot, business_roc_date
        from qt_app.controllers.duty_controller import DutyController

        target_date = business_roc_date()
        controller = DutyController()
        controller.liveScheduleCaptured.connect(
            lambda data: controller.set_actor_no(data["_authenticated_actor"]["actor_no"])
        )
        controller._active_schedule_request = 7
        controller._active_capture_request = 1
        controller._capture_succeeded(
            1,
            "",
            ScheduleSnapshot(
                Path(f"schedule_output_{target_date}.json"),
                {"target_date": target_date, "actions": []},
                target_date,
                authenticated_actor_no="10",
                authenticated_actor_name="測試員",
            ),
        )

        self.assertTrue(controller._auto_execution_enabled)
        controller._schedule_loaded(
            7,
            ScheduleSnapshot(
                Path("schedule_output_1000101.json"),
                {"target_date": "1000101", "actions": []},
                "1000101",
            ),
        )
        self.assertEqual(controller.targetDateText, target_date)

    def test_schedule_becomes_usable_before_background_comparison_finishes(self) -> None:
        from PySide6.QtTest import QSignalSpy

        from app_core.schedule_repository import ScheduleSnapshot, business_roc_date
        from qt_app.controllers.duty_controller import DutyController

        target_date = business_roc_date()
        controller = DutyController()
        controller.liveScheduleCaptured.connect(
            lambda data: controller.set_actor_no(data["_authenticated_actor"]["actor_no"])
        )
        controller._active_capture_request = 1
        controller._capture_targets[1] = target_date
        error_spy = QSignalSpy(controller.errorOccurred)

        controller._capture_schedule_ready(
            1,
            "",
            ScheduleSnapshot(
                Path(f"schedule_output_{target_date}.json"),
                {
                    "target_date": target_date,
                    "today": {"staff": {"10": {"name": "測試員"}}},
                    "actions": [],
                },
                target_date,
                schedule_data_by_date={target_date: {"target_date": target_date, "actions": []}},
                authenticated_actor_no="10",
                authenticated_actor_name="測試員",
            ),
        )

        self.assertTrue(controller._auto_execution_enabled)
        self.assertIn("正在背景比對", controller.scheduleStatus)
        controller._capture_failed(
            1,
            "",
            "已登打資料比對逾時，勤務資料仍可使用。",
            "comparison_timeout",
        )

        self.assertTrue(controller._auto_execution_enabled)
        self.assertEqual(error_spy.count(), 1)
        self.assertEqual(controller.targetDateText, target_date)

    def test_duty_refresh_state_stays_busy_until_worker_cleanup(self) -> None:
        from PySide6.QtTest import QSignalSpy

        from qt_app.controllers.duty_controller import DutyController

        class FinishedThread:
            def quit(self):
                pass

            def wait(self, _timeout):
                return True

            def deleteLater(self):
                pass

        controller = DutyController()
        spy = QSignalSpy(controller.scheduleChanged)
        controller._capture_workers[7] = (FinishedThread(), object())
        controller._capture_targets[7] = "1150729"

        self.assertTrue(controller.isRefreshing)
        self.assertFalse(controller.refresh_live_schedule("user10", "secret", "10"))
        controller._capture_worker_finished(7)

        self.assertFalse(controller.isRefreshing)
        self.assertEqual(spy.count(), 1)

    def test_duty_failures_emit_frontend_error_signal(self) -> None:
        from PySide6.QtTest import QSignalSpy

        from qt_app.controllers.duty_controller import DutyController

        controller = DutyController()
        error_spy = QSignalSpy(controller.errorOccurred)

        controller._active_capture_request = 1
        controller._capture_failed(1, "", "勤務查詢失敗", "unknown_error")
        controller.handle_submission_failure(0, "勤務登打失敗", "timeout")

        self.assertEqual(error_spy.count(), 2)
        self.assertEqual(error_spy.at(0)[0], "勤務查詢失敗")
        self.assertEqual(error_spy.at(1)[0], "勤務登打失敗")

    def test_historical_capture_failure_does_not_fallback_to_current_schedule(self) -> None:
        from qt_app.controllers.duty_controller import DutyController

        controller = DutyController()
        controller.set_actor_no("10")
        controller._active_capture_request = 1
        controller._capture_targets[1] = "1000101"
        fallback_calls = []
        controller.load_current_schedule = lambda: fallback_calls.append(True)

        controller._capture_failed(1, "10", "歷史查詢失敗", "unknown_error")

        self.assertEqual(fallback_calls, [])
        self.assertEqual(controller.scheduleStatus, "歷史查詢失敗")

    def test_second_qt_instance_signals_existing_server(self) -> None:
        from uuid import uuid4

        from PySide6.QtNetwork import QLocalServer

        from qt_app.main import create_instance_server

        server_name = f"SinpoSmart-test-{uuid4().hex}"
        first = create_instance_server(server_name)
        self.assertIsNotNone(first)
        try:
            second = create_instance_server(server_name)
            self.assertIsNone(second)
            self.assertTrue(first.waitForNewConnection(1_000))
        finally:
            first.close()
            QLocalServer.removeServer(server_name)

    def test_windowed_entry_runs_full_qt_startup_and_clean_shutdown(self) -> None:
        from qt_app.main import STARTUP_SMOKE_ARG

        entry_path = PACKAGE_ROOT / "duty_gui.pyw"
        command = f"""
import runpy
import sys

sys.argv = [{str(entry_path)!r}, {STARTUP_SMOKE_ARG!r}]
try:
    runpy.run_path({str(entry_path)!r}, run_name="__main__")
except SystemExit as exc:
    return_code = int(exc.code or 0)
else:
    return_code = 0

forbidden = {{
    "duty_gui",
    "tkinter",
    "_tkinter",
    "customtkinter",
    "pystray",
    "PIL.ImageTk",
    "selenium",
}}
loaded = sorted(forbidden.intersection(sys.modules))
if return_code != 0 or loaded:
    print({{"returncode": return_code, "forbidden_loaded": loaded}})
    raise SystemExit(1)
"""
        return_code, output = run_isolated_python(
            command,
            extra_env={"PYTHONDONTWRITEBYTECODE": "1"},
        )

        self.assertEqual(return_code, 0, output)
        self.assertNotIn("QQmlApplicationEngine failed", output)
        self.assertNotIn("QThread: Destroyed", output)

    def test_qt_modules_do_not_import_legacy_duty_gui(self) -> None:
        command = (
            "import sys; "
            f"sys.path.insert(0, {str(PACKAGE_ROOT)!r}); "
            "from PySide6.QtCore import QTimer; "
            "from PySide6.QtWidgets import QApplication; "
            "from pathlib import Path; "
            "import tempfile; "
            "from app_core.credential_repository import CredentialRepository; "
            "from qt_app.controllers.app_controller import AppController; "
            "from qt_app.main import create_engine; "
            "app = QApplication(['qt-shell-check']); "
            "repository = CredentialRepository(Path(tempfile.gettempdir()) / 'sinposmart-qt-shell-nonexistent.json', 'SinpoSmart', None); "
            "engine = create_engine(AppController(repository=repository)); "
            "assert engine.rootObjects(); "
            "QTimer.singleShot(50, app.quit); "
            "app.exec(); "
            "forbidden = {'duty_gui', 'tkinter', 'customtkinter', 'selenium'}; "
            "raise SystemExit(1 if forbidden.intersection(sys.modules) else 0)"
        )

        return_code, output = run_isolated_python(command)

        self.assertEqual(return_code, 0, output)

    def test_task_model_exposes_stable_roles_and_replaces_rows(self) -> None:
        from PySide6.QtTest import QSignalSpy

        from qt_app.models.duty_task_model import DutyTaskListModel

        model = DutyTaskListModel()
        roles = {bytes(value).decode("utf-8") for value in model.roleNames().values()}

        self.assertEqual(
            roles,
            {
                "taskIndex",
                "timeText",
                "systemText",
                "kindText",
                "detailText",
                "peopleText",
                "statusText",
                "statusTone",
                "selected",
                "actorText",
                "targetText",
                "comparisonText",
                "group",
                "fullDetailText",
                "errorText",
            },
        )
        self.assertEqual(model.rowCount(), 0)

        model.replace_tasks([{"taskIndex": 7, "timeText": "08:00", "selected": True}])

        self.assertEqual(model.rowCount(), 1)
        self.assertEqual(model.data(model.index(0, 0), model.TaskIndexRole), 7)
        self.assertEqual(model.data(model.index(0, 0), model.TimeTextRole), "08:00")
        self.assertTrue(model.data(model.index(0, 0), model.SelectedRole))

        reset_spy = QSignalSpy(model.modelReset)
        changed_spy = QSignalSpy(model.dataChanged)
        model.replace_tasks([{"taskIndex": 7, "timeText": "08:00", "selected": False}])

        self.assertEqual(reset_spy.count(), 0)
        self.assertEqual(changed_spy.count(), 1)
        self.assertEqual(changed_spy.at(0)[2], [model.SelectedRole])

    def test_duty_controller_projects_schedule_after_actor_is_known(self) -> None:
        from qt_app.controllers.duty_controller import DutyController

        controller = DutyController()
        controller.replace_schedule_data(
            {
                "target_date": "1150729",
                "today": {"staff": {"10": {"name": "本班"}}},
                "actions": [
                    {
                        "kind": "work_log",
                        "time": "09:00",
                        "actor": "10",
                        "target": "10",
                        "source": "值班交接",
                        "fields": {"工作時間": "09:00", "勤務項目": "巡邏"},
                    }
                ],
            }
        )

        self.assertEqual(controller.taskModel.rowCount(), 0)
        controller.set_actor_no("10")
        self.assertEqual(controller.targetDateText, "1150729")
        self.assertEqual(controller.taskModel.rowCount(), 1)
        self.assertEqual(
            controller.taskModel.data(controller.taskModel.index(0, 0), controller.taskModel.DetailTextRole),
            "巡邏",
        )
        self.assertIn("巡邏", controller.nextTaskText)

    def test_audit_model_filters_shared_comparison_rows(self) -> None:
        from qt_app.controllers.duty_controller import DutyController

        controller = DutyController()
        controller.replace_schedule_data(
            {
                "target_date": "1150729",
                "today": {"staff": {"10": {"name": "本班"}, "11": {"name": "他班"}}},
                "actions": [
                    {"kind": "work_log", "time": "08:00", "actor": "10", "target": "10", "fields": {"工作時間": "08:00", "勤務項目": "巡邏"}},
                    {"kind": "entry_log", "time": "09:00", "actor": "11", "target": "11", "fields": {"登打時間": "09:00", "出或入": "出", "領用事由及地點": "公出"}},
                ],
            },
            comparisons={
                0: {"compare": "已存在", "group": "done", "matched": []},
                1: {"compare": "人工確認", "group": "review", "matched": []},
            },
        )

        self.assertEqual(controller.auditModel.rowCount(), 1)
        self.assertEqual(
            controller.auditModel.data(controller.auditModel.index(0, 0), controller.auditModel.ActorTextRole),
            "11 他班",
        )
        controller.setAuditStatusFilter("全部")
        controller.setAuditKindFilter("工作")
        self.assertEqual(controller.auditModel.rowCount(), 1)
        self.assertEqual(
            controller.auditModel.data(controller.auditModel.index(0, 0), controller.auditModel.GroupRole),
            "done",
        )

    def test_audit_model_includes_case_work_from_schedule_cases(self) -> None:
        from qt_app.controllers.duty_controller import DutyController

        controller = DutyController()
        controller.replace_schedule_data(
            {
                "target_date": "1150729",
                "today": {"staff": {"10": {"name": "本班"}}},
                "actions": [],
                "cases": [
                    {
                        "category": "緊急救護",
                        "report_time": "08:30",
                        "location": "中正路",
                        "return_time": "09:05",
                    }
                ],
            },
            comparison_data={
                "1150729": {"visible_work_rows": [], "visible_entry_rows": []}
            },
        )

        self.assertEqual(controller.auditModel.rowCount(), 1)
        row = controller.auditModel.index(0, 0)
        self.assertEqual(
            controller.auditModel.data(row, controller.auditModel.SystemTextRole),
            "案件工作",
        )
        self.assertIn(
            "緊急救護",
            controller.auditModel.data(row, controller.auditModel.DetailTextRole),
        )
        self.assertEqual(
            controller.auditModel.data(row, controller.auditModel.GroupRole),
            "todo",
        )
        controller.set_actor_no("10")
        self.assertEqual(controller.taskModel.rowCount(), 0)

    def test_audit_model_compares_case_work_from_schedule_cases(self) -> None:
        from qt_app.controllers.duty_controller import DutyController

        controller = DutyController()
        controller.replace_schedule_data(
            {
                "target_date": "1150729",
                "actions": [],
                "cases": [
                    {
                        "category": "緊急救護",
                        "report_time": "08:30",
                        "location": "中正路",
                    }
                ],
            },
            comparison_data={
                "1150729": {
                    "visible_work_rows": [["115/07/29", "08:30", "緊急救護", "中正路"]],
                    "visible_entry_rows": [],
                }
            },
        )
        controller.setAuditStatusFilter("全部")
        controller.setAuditKindFilter("案件工作")

        self.assertEqual(controller.auditModel.rowCount(), 1)
        self.assertEqual(
            controller.auditModel.data(
                controller.auditModel.index(0, 0),
                controller.auditModel.GroupRole,
            ),
            "done",
        )

    def test_audit_summary_and_filters_preserve_legacy_groups(self) -> None:
        from qt_app.controllers.duty_controller import DutyController

        controller = DutyController()
        actions = [
            {
                "kind": "work_log",
                "source": "一般勤務",
                "time": f"0{index}:00",
                "actor": "10",
                "target": "10",
                "fields": {"工作時間": f"0{index}:00", "勤務項目": "巡邏"},
            }
            for index in range(1, 8)
        ]
        actions.append(
            {
                "kind": "work_log",
                "source": "案件工作審核",
                "time": "08:00",
                "actor": "10",
                "target": "10",
                "fields": {"工作時間": "08:00", "事由": "救護案件"},
            }
        )
        comparisons = {
            0: {"compare": "未找到", "group": "todo", "matched": []},
            1: {"compare": "人工確認", "group": "review", "matched": []},
            2: {"compare": "可能臨時調整", "group": "adjust", "matched": []},
            3: {"compare": "手動", "group": "manual", "matched": []},
            4: {"compare": "尚未到點", "group": "future", "matched": []},
            5: {"compare": "已存在", "group": "done", "matched": []},
            6: {"compare": "時間近似", "group": "near", "matched": []},
            7: {"compare": "未找到", "group": "todo", "matched": []},
        }
        controller.replace_schedule_data(
            {
                "target_date": "1150729",
                "today": {"staff": {"10": {"name": "本班"}}},
                "actions": actions,
            },
            comparisons=comparisons,
        )

        self.assertEqual(controller.auditTodoCount, 2)
        self.assertEqual(controller.auditReviewCount, 3)
        self.assertEqual(controller.auditReadyCount, 1)
        self.assertEqual(controller.auditDoneCount, 1)

        controller.setAuditStatusFilter("時間近似")
        self.assertEqual(controller.auditModel.rowCount(), 1)
        controller.setAuditStatusFilter("疑似異動")
        self.assertEqual(controller.auditModel.rowCount(), 1)
        controller.setAuditStatusFilter("手動")
        self.assertEqual(controller.auditModel.rowCount(), 1)
        controller.setAuditStatusFilter("全部")
        controller.setAuditKindFilter("案件工作")
        self.assertEqual(controller.auditModel.rowCount(), 1)
        self.assertEqual(
            controller.auditModel.data(
                controller.auditModel.index(0, 0),
                controller.auditModel.SystemTextRole,
            ),
            "案件工作",
        )

    def test_duty_controller_loads_schedule_in_worker_and_closes_thread(self) -> None:
        from PySide6.QtTest import QSignalSpy, QTest

        from app_core.schedule_repository import ScheduleSnapshot, business_roc_date
        from qt_app.controllers.duty_controller import DutyController

        class FakeScheduleRepository:
            def load_current(self) -> ScheduleSnapshot:
                return ScheduleSnapshot(
                    Path("schedule_output_1150729.json"),
                    {
                        "target_date": "1150729",
                        "today": {"staff": {"10": {"name": "本班"}}},
                        "actions": [
                            {
                                "kind": "work_log",
                                "time": "09:00",
                                "actor": "10",
                                "target": "10",
                                "fields": {"工作時間": "09:00", "勤務項目": "巡邏"},
                            }
                        ],
                    },
                    "1150729",
                )

        controller = DutyController(repository=FakeScheduleRepository())
        spy = QSignalSpy(controller.scheduleChanged)
        cached_spy = QSignalSpy(controller.cachedScheduleLoaded)

        controller.load_current_schedule()

        for _ in range(10):
            if not controller._schedule_workers:
                break
            spy.wait(250)
            QTest.qWait(10)
        self.assertEqual(cached_spy.count(), 1)
        self.assertEqual(controller.taskModel.rowCount(), 0)
        controller.set_actor_no("10")
        self.assertEqual(controller.taskModel.rowCount(), 1)
        self.assertEqual(controller.scheduleStatus, "已載入 schedule_output_1150729.json")
        self.assertFalse(controller._schedule_workers)

    def test_duty_controller_loads_selected_preview_without_enabling_automation(self) -> None:
        from PySide6.QtCore import QUrl
        from PySide6.QtTest import QSignalSpy, QTest

        from app_core.schedule_repository import ScheduleSnapshot
        from qt_app.controllers.duty_controller import DutyController

        class FakeScheduleRepository:
            def __init__(self):
                self.paths = []

            def load_path(self, path):
                self.paths.append(Path(path))
                return ScheduleSnapshot(
                    Path(path),
                    {
                        "target_date": "1150730",
                        "today": {"staff": {"10": {"name": "本班"}}},
                        "actions": [
                            {
                                "kind": "work_log",
                                "time": "09:00",
                                "actor": "10",
                                "target": "10",
                                "fields": {"勤務項目": "離線預演"},
                            }
                        ],
                    },
                    "1150730",
                )

        repository = FakeScheduleRepository()
        controller = DutyController(repository=repository)
        controller.set_actor_no("10")
        controller.enable_auto_execution()
        spy = QSignalSpy(controller.scheduleChanged)
        preview_path = Path("selected_preview.json").resolve()

        controller.loadPreviewPath(str(preview_path))

        for _ in range(20):
            if controller.taskModel.rowCount() == 1 and not controller._schedule_workers:
                break
            spy.wait(250)
            QTest.qWait(10)
        self.assertEqual(repository.paths, [preview_path])
        self.assertEqual(controller.taskModel.rowCount(), 1)
        self.assertTrue(controller.isPreviewLoaded)
        self.assertFalse(controller._auto_execution_enabled)
        self.assertEqual(controller.scheduleStatus, "已載入預演檔 selected_preview.json")
        self.assertEqual(controller.automationStatus, "預演模式；自動登打已停用")
        self.assertFalse(controller._schedule_workers)

    def test_duty_controller_loads_saved_audit_date_without_login(self) -> None:
        from PySide6.QtTest import QSignalSpy, QTest

        from app_core.schedule_repository import ScheduleSnapshot
        from qt_app.controllers.duty_controller import DutyController

        class FakeScheduleRepository:
            def __init__(self) -> None:
                self.target_dates: list[str] = []

            def load_for_date(self, target_roc_date: str) -> ScheduleSnapshot:
                self.target_dates.append(target_roc_date)
                return ScheduleSnapshot(
                    Path(f"schedule_output_{target_roc_date}.json"),
                    {
                        "target_date": target_roc_date,
                        "actions": [
                            {
                                "kind": "work_log",
                                "time": "09:00",
                                "actor": "10",
                                "target": "10",
                                "fields": {"勤務項目": "勤務"},
                            }
                        ],
                    },
                    target_roc_date,
                )

        repository = FakeScheduleRepository()
        controller = DutyController(repository=repository)
        spy = QSignalSpy(controller.scheduleChanged)

        controller.load_audit_schedule("1150729")

        for _ in range(20):
            if controller.auditModel.rowCount() == 1 and not controller._schedule_workers:
                break
            spy.wait(250)
            QTest.qWait(10)

        self.assertEqual(repository.target_dates, ["1150729"])
        self.assertEqual(controller.auditModel.rowCount(), 1)
        self.assertFalse(controller.isPreviewLoaded)
        self.assertFalse(controller._auto_execution_enabled)
        self.assertFalse(controller._schedule_workers)

    def test_duty_controller_preserves_audit_date_when_schedule_is_missing(self) -> None:
        from app_core.schedule_repository import ScheduleSnapshot
        from qt_app.controllers.duty_controller import DutyController

        controller = DutyController()
        controller._active_schedule_request = 1

        try:
            controller._schedule_loaded(1, ScheduleSnapshot(None, {}, "1150730"))

            self.assertEqual(controller.targetDateText, "1150730")
            self.assertEqual(controller.taskModel.rowCount(), 0)
            self.assertEqual(controller.auditModel.rowCount(), 0)
            self.assertEqual(controller.scheduleStatus, "1150730 尚無排程資料")
        finally:
            controller.shutdown()

    def test_duty_controller_refreshes_comparisons_without_replacing_schedule(self) -> None:
        from PySide6.QtTest import QTest

        from qt_app.controllers.duty_controller import DutyController

        class FakeCaptureService:
            def __init__(self) -> None:
                self.schedule_calls = 0
                self.comparison_requests = []

            def capture_schedule(self, *_args, **_kwargs):
                self.schedule_calls += 1
                raise AssertionError("審核模式的比對更新不得重新查詢勤務表")

            def capture_comparisons(self, request, *, status_callback=None):
                self.comparison_requests.append(request)
                if status_callback is not None:
                    status_callback("正在更新已登打比對資料…")
                return {
                    "1150729": {
                        "visible_work_rows": [],
                        "visible_entry_rows": [],
                    }
                }

        service = FakeCaptureService()
        controller = DutyController(capture_service=service)
        controller.replace_schedule_data(
            {
                "target_date": "1150729",
                "today": {"staff": {"10": {"name": "測試員"}}},
                "actions": [
                    {
                        "kind": "work_log",
                        "time": "09:00",
                        "actor": "10",
                        "target": "10",
                        "fields": {"工作時間": "09:00", "勤務項目": "勤務"},
                    }
                ],
            }
        )
        try:
            self.assertTrue(
                controller.refresh_live_comparisons(
                    "user10",
                    "test-password",
                    "10",
                    target_roc_date="1150729",
                )
            )
            for _ in range(20):
                if not controller._comparison_workers:
                    break
                QTest.qWait(50)

            self.assertEqual(service.schedule_calls, 0)
            self.assertEqual(len(service.comparison_requests), 1)
            self.assertEqual(service.comparison_requests[0].target_roc_date, "1150729")
            self.assertEqual(controller.targetDateText, "1150729")
            self.assertIn(0, controller._comparisons)
            self.assertFalse(controller._comparison_workers)
        finally:
            controller.shutdown()

    def test_duty_controller_audit_capture_does_not_emit_operational_snapshots(self) -> None:
        from PySide6.QtTest import QSignalSpy, QTest

        from app_core.schedule_repository import ScheduleSnapshot, business_roc_date
        from qt_app.controllers.duty_controller import DutyController

        class FakeCaptureService:
            def capture_schedule(self, request, *, status_callback=None):
                if status_callback is not None:
                    status_callback("正在讀取勤務表…")
                return ScheduleSnapshot(
                    Path(f"schedule_output_{request.target_roc_date}.json"),
                    {
                        "target_date": request.target_roc_date,
                        "today": {"staff": {"10": {"name": "測試員"}}},
                        "actions": [],
                    },
                    request.target_roc_date,
                )

            def capture_comparisons(self, request, *, status_callback=None):
                if status_callback is not None:
                    status_callback("正在讀取已登打資料…")
                return {
                    request.target_roc_date: {
                        "visible_work_rows": [],
                        "visible_entry_rows": [],
                    }
                }

            @staticmethod
            def combine_capture(snapshot, comparison_data):
                return ScheduleSnapshot(
                    snapshot.path,
                    snapshot.data,
                    snapshot.target_roc_date,
                    {},
                    comparison_data=comparison_data,
                    schedule_data_by_date={snapshot.target_roc_date: snapshot.data},
                )

        target_roc_date = business_roc_date()
        controller = DutyController(capture_service=FakeCaptureService())
        controller.set_actor_no("10")
        schedule_spy = QSignalSpy(controller.liveScheduleCaptured)
        snapshot_spy = QSignalSpy(controller.liveSnapshotCaptured)
        try:
            self.assertTrue(
                controller.refresh_live_schedule(
                    "user10",
                    "test-password",
                    "10",
                    target_roc_date=target_roc_date,
                    publish_events=False,
                    allow_auto_execution=False,
                )
            )
            for _ in range(20):
                if not controller._capture_workers:
                    break
                QTest.qWait(50)

            self.assertEqual(controller.targetDateText, target_roc_date)
            self.assertEqual(schedule_spy.count(), 0)
            self.assertEqual(snapshot_spy.count(), 0)
            self.assertFalse(controller._auto_execution_enabled)
            self.assertFalse(controller._capture_workers)
        finally:
            controller.shutdown()

    def test_app_controller_uses_saved_audit_schedule_without_login(self) -> None:
        from qt_app.controllers.app_controller import AppController

        controller = AppController(read_only_acceptance=True)
        saved_schedule_calls: list[str] = []
        live_schedule_calls: list[tuple[object, ...]] = []
        controller.dutyController.load_audit_schedule = saved_schedule_calls.append
        controller.dutyController.refresh_live_schedule = (
            lambda *args, **_kwargs: live_schedule_calls.append(args)
        )

        try:
            self.assertFalse(controller.sessionController.isLoggedIn)
            controller.refreshAuditDate("1150729")

            self.assertEqual(saved_schedule_calls, ["1150729"])
            self.assertEqual(live_schedule_calls, [])
        finally:
            controller.shutdown()

    def test_app_controller_uses_saved_audit_schedule_after_login_too(self) -> None:
        from app_core.session import LoginSession
        from qt_app.controllers.app_controller import AppController

        controller = AppController(read_only_acceptance=True)
        saved_schedule_calls: list[str] = []
        live_schedule_calls: list[tuple[object, ...]] = []
        controller.dutyController.load_audit_schedule = saved_schedule_calls.append
        controller.dutyController.refresh_live_schedule = (
            lambda *args, **_kwargs: live_schedule_calls.append(args)
        )
        attempt_id = controller._session_state.begin_login()
        controller._session_state.complete_login(
            attempt_id,
            LoginSession("10", "user10", "session-secret", verified=True),
        )

        try:
            controller.refreshAuditDate("1150729")

            self.assertEqual(saved_schedule_calls, ["1150729"])
            self.assertEqual(live_schedule_calls, [])
        finally:
            controller.shutdown()

    def test_app_controller_opens_audit_mode_with_pending_status_and_all_kinds(self) -> None:
        from qt_app.controllers.app_controller import AppController

        controller = AppController(read_only_acceptance=True)
        loaded_dates: list[str] = []
        controller.dutyController._target_date_text = "1150729"
        controller.dutyController.load_audit_schedule = loaded_dates.append

        try:
            self.assertEqual(controller.dutyController.auditStatusFilter, "需處理")
            controller.dutyController.setAuditKindFilter("工作")
            controller.openAuditMode()

            self.assertEqual(controller.dutyController.auditStatusFilter, "需處理")
            self.assertEqual(controller.dutyController.auditKindFilter, "全部")
            self.assertEqual(loaded_dates, ["1150729"])
        finally:
            controller.shutdown()

    def test_app_controller_refreshes_audit_data_without_publishing_nas_events(self) -> None:
        from app_core.session import LoginSession
        from qt_app.controllers.app_controller import AppController

        controller = AppController()
        calls: list[tuple[tuple[object, ...], dict[str, object]]] = []
        controller.dutyController.refresh_live_schedule = (
            lambda *args, **kwargs: calls.append((args, kwargs)) or True
        )
        attempt_id = controller._session_state.begin_login()
        controller._session_state.complete_login(
            attempt_id,
            LoginSession("10", "user10", "session-secret", verified=True),
        )
        try:
            calls.clear()
            controller.dutyController._target_date_text = "1150729"

            self.assertTrue(controller.refreshAuditLiveData())

            self.assertEqual(len(calls), 1)
            args, kwargs = calls[0]
            self.assertEqual(args[:3], ("user10", "session-secret", "10"))
            self.assertEqual(kwargs["target_roc_date"], "1150729")
            self.assertFalse(kwargs["publish_events"])
            self.assertFalse(kwargs["allow_auto_execution"])
        finally:
            controller.shutdown()

    def test_app_controller_reports_why_audit_refresh_cannot_start(self) -> None:
        from qt_app.controllers.app_controller import AppController

        controller = AppController()
        try:
            self.assertFalse(controller.refreshAuditLiveData())
            self.assertEqual(
                controller.dutyController.scheduleStatus,
                "請先完成勤務系統登入驗證後再重新查詢。",
            )
        finally:
            controller.shutdown()

    def test_duty_controller_has_no_generic_pause_or_resume_controls(self) -> None:
        from qt_app.controllers.duty_controller import DutyController

        controller = DutyController()
        controller.set_actor_no("10")
        controller.replace_schedule_data(
            {
                "target_date": "1150729",
                "today": {"staff": {"10": {"name": "本班"}}},
                "actions": [
                    {
                        "kind": "work_log",
                        "time": "00:00",
                        "actor": "10",
                        "target": "10",
                        "source": "在隊訓練",
                        "fields": {"勤務項目": "巡邏"},
                    }
                ],
            }
        )
        self.assertEqual(controller.dueTaskCount, 1)

        controller.toggleTaskSelection(0)
        self.assertEqual(controller.selectedTaskCount, 1)
        self.assertEqual(controller.dueTaskCount, 1)
        self.assertFalse(hasattr(controller, "pauseSelectedTasks"))
        self.assertFalse(hasattr(controller, "resumeSelectedTasks"))

    def test_duty_controller_locks_auto_execution_when_fire_day_changes(self) -> None:
        from unittest.mock import patch

        from qt_app.controllers.duty_controller import DutyController

        controller = DutyController()
        changes: list[str] = []
        controller.fireDayChanged.connect(changes.append)
        controller._observed_fire_day = "1150729"
        controller._current_date_text = ""
        controller.enable_auto_execution()
        try:
            with patch(
                "qt_app.controllers.duty_controller.business_roc_date",
                return_value="1150730",
            ):
                controller._update_clock()

            self.assertEqual(changes, ["1150730"])
            self.assertFalse(controller._auto_execution_enabled)
        finally:
            controller.shutdown()

    def test_duty_sheet_controller_adds_and_removes_vehicle_options(self) -> None:
        from app_core.duty_sheet_service import DutySheetDefaults
        from app_core.session import SessionState
        from qt_app.controllers.duty_sheet_controller import DutySheetController

        class FakeService:
            def __init__(self):
                self.attack_options = ["A"]
                self.stop_options = ["S"]
                self.amb_options = ["M1"]

            def load_defaults(self):
                return DutySheetDefaults(
                    "duty.xlsm", "2026/07/30", "A", "S", "M1", "M1",
                    tuple(self.attack_options), tuple(self.stop_options), tuple(self.amb_options), False,
                )

            def add_vehicle_option(self, group, code, plate):
                value = f"{code}/{plate}"
                {
                    "attack": self.attack_options,
                    "stop": self.stop_options,
                    "amb": self.amb_options,
                }[group].append(value)
                return value

            def remove_vehicle_option(self, group, value):
                {
                    "attack": self.attack_options,
                    "stop": self.stop_options,
                    "amb": self.amb_options,
                }[group].remove(value)
                return value

        service = FakeService()
        controller = DutySheetController(SessionState(), service)
        controller.loadDefaults()
        self.assertEqual(controller.statusText, "準備就緒。")

        controller.addVehicleOption("amb", "新坡93", "BSL-9230")
        self.assertIn("新坡93/BSL-9230", controller.ambOptions)
        self.assertIn("已新增救護車", controller.statusText)

        controller.removeVehicleOption("amb", "新坡93/BSL-9230")
        self.assertNotIn("新坡93/BSL-9230", controller.ambOptions)
        self.assertIn("已移除救護車", controller.statusText)

        controller.addVehicleOption("stop", "新坡16", "KET-0001")
        self.assertIn("新坡16/KET-0001", controller.stopOptions)
        self.assertIn("消防車（中繼車）", controller.statusText)

        controller.removeVehicleOption("stop", "新坡16/KET-0001")
        self.assertNotIn("新坡16/KET-0001", controller.stopOptions)

    def test_duty_sheet_controller_requires_confirmation_and_runs_in_worker(self) -> None:
        from PySide6.QtTest import QSignalSpy, QTest

        from app_core.duty_sheet_service import DutySheetDefaults, DutySheetRequest
        from app_core.session import LoginSession, SessionState
        from qt_app.controllers.duty_sheet_controller import DutySheetController

        class FakeService:
            def load_defaults(self):
                return DutySheetDefaults(
                    "duty.xlsm", "2026/07/30", "A", "S", "M1", "M2",
                    ("A",), ("S",), ("M1", "M2"), False,
                )

            def validate(self, request):
                return request

            def confirmation_summary(self, request):
                return f"確認 {request.target_date}"

            def execute(self, request, *, status_callback=None):
                status_callback("執行中")
                return f"完成 {request.target_date}"

        state = SessionState()
        attempt_id = state.begin_login()
        state.complete_login(
            attempt_id,
            LoginSession("10", "user10", "secret", verified=True),
        )
        controller = DutySheetController(state, FakeService())
        confirmation_spy = QSignalSpy(controller.confirmationRequested)
        success_spy = QSignalSpy(controller.runSucceeded)

        controller.loadDefaults()
        controller.setNotificationEnabled(True)
        self.assertTrue(controller.notificationEnabled)
        controller.setNotificationEnabled(False)
        controller.prepareRun("duty.xlsm", "2026/07/30", "A", "S", "M1", "M2", False)
        self.assertEqual(confirmation_spy.count(), 1)
        self.assertIn("確認 2026/07/30", controller.confirmationSummary)
        self.assertFalse(controller._pending_request.notification_enabled)

        controller.confirmRun()
        for _ in range(20):
            if success_spy.count() and not controller._workers:
                break
            success_spy.wait(250)
            QTest.qWait(10)
        self.assertEqual(success_spy.count(), 1)
        self.assertFalse(controller._workers)

    def test_rest_monthly_controller_requires_confirmation_and_runs_in_worker(self) -> None:
        from PySide6.QtTest import QSignalSpy, QTest

        from app_core.rest_monthly_service import (
            MonthlyBaseRequest,
            RestMonthlyDefaults,
            RestTimeRequest,
        )
        from app_core.session import LoginSession, SessionState
        from qt_app.controllers.rest_monthly_controller import RestMonthlyController

        class FakeService:
            def load_rest_defaults(self):
                return RestMonthlyDefaults(115, ("06", "07", "08"), "07", "duty.xlsm")

            def load_monthly_defaults(self):
                return RestMonthlyDefaults(115, ("06", "07", "08"), "07")

            def select_rest_workbook(self, path):
                return RestMonthlyDefaults(115, ("06", "07", "08"), "08", path)

            def validate_rest(self, request):
                return request

            def validate_monthly(self, request):
                return request

            def confirmation_summary(self, request):
                return f"確認 {request.roc_year}/{request.month:02d}"

            def execute_rest(self, request, *, status_callback=None):
                status_callback("休息時間執行中")
                return "休息時間完成"

            def execute_monthly(self, request, *, status_callback=None):
                status_callback("勤務基準執行中")
                return "勤務基準完成"

        state = SessionState()
        attempt_id = state.begin_login()
        state.complete_login(
            attempt_id,
            LoginSession("10", "user10", "secret", verified=True, actor_name="王小明隊員"),
        )
        controller = RestMonthlyController(state, FakeService())
        confirmation_spy = QSignalSpy(controller.confirmationRequested)
        success_spy = QSignalSpy(controller.runSucceeded)

        controller.loadRestDefaults()
        self.assertEqual(controller.statusText, "準備就緒。10番 王小明隊員")
        from PySide6.QtCore import QUrl

        controller.selectRestWorkbook(QUrl.fromLocalFile("selected.xlsm"))
        self.assertEqual(controller.restWorkbookPath, "selected.xlsm")
        self.assertEqual(controller.restMonth, "08")
        self.assertEqual(controller.statusText, "已選擇勤務表 Excel。")
        controller.prepareRestRun("selected.xlsm", "08")
        self.assertEqual(confirmation_spy.count(), 1)
        self.assertEqual(confirmation_spy.at(0)[0], "rest_time")
        self.assertIn("確認 115/08", controller.confirmationSummary)

        controller.confirmRun()
        for _ in range(20):
            if success_spy.count() and not controller._workers:
                break
            success_spy.wait(250)
            QTest.qWait(10)
        self.assertEqual(success_spy.count(), 1)
        self.assertEqual(success_spy.at(0)[0], "rest_time")
        self.assertFalse(controller._workers)

        controller.prepareMonthlyRun("06")
        self.assertEqual(confirmation_spy.count(), 2)
        self.assertEqual(confirmation_spy.at(1)[0], "monthly_base")
        self.assertEqual(controller.monthlyMonth, "06")
        controller.cancelPendingRun()

    def test_rest_monthly_qml_rejects_a_session_without_actor_name(self) -> None:
        from PySide6.QtTest import QSignalSpy

        from app_core.session import LoginSession, SessionState
        from qt_app.controllers.rest_monthly_controller import RestMonthlyController

        state = SessionState()
        attempt_id = state.begin_login()
        state.complete_login(
            attempt_id,
            LoginSession("10", "user10", "secret", verified=True),
        )
        controller = RestMonthlyController(state, object())
        errors = QSignalSpy(controller.errorOccurred)

        controller.prepareMonthlyRun("08")

        self.assertEqual(errors.count(), 1)
        self.assertIn("姓名", errors.at(0)[0])

    def test_daily_vehicle_controller_requires_confirmation_and_runs_in_worker(self) -> None:
        from PySide6.QtTest import QSignalSpy, QTest

        from app_core.daily_vehicle_service import DailyVehicleDefaults
        from app_core.session import LoginSession, SessionState
        from qt_app.controllers.daily_vehicle_controller import DailyVehicleController

        class FakeService:
            def load_defaults(self):
                return DailyVehicleDefaults("2026/07/29", ("車輛保養檢查", "車輛器材清點"))

            def validate(self, request):
                return request

            def confirmation_summary(self, request):
                return "將開啟瀏覽器執行車輛保養清點，是否繼續？"

            def execute(self, request, *, status_callback=None):
                status_callback("車輛保養執行中")
                return "車輛保養清點已完成。"

        state = SessionState()
        attempt_id = state.begin_login()
        state.complete_login(attempt_id, LoginSession("10", "user10", "secret", verified=True))
        controller = DailyVehicleController(state, FakeService())
        confirmation_spy = QSignalSpy(controller.confirmationRequested)
        success_spy = QSignalSpy(controller.runSucceeded)

        controller.loadDefaults()
        self.assertEqual(controller.statusText, "準備就緒。")
        controller.prepareRun()
        self.assertEqual(confirmation_spy.count(), 1)
        self.assertEqual(controller.confirmationSummary, "將開啟瀏覽器執行車輛保養清點，是否繼續？")

        controller.confirmRun()
        for _ in range(20):
            if success_spy.count() and not controller._workers:
                break
            success_spy.wait(250)
            QTest.qWait(10)
        self.assertEqual(success_spy.count(), 1)
        self.assertFalse(controller._workers)

    def test_daily_vehicle_controller_requires_resolved_actor_no(self) -> None:
        from PySide6.QtTest import QSignalSpy

        from app_core.daily_vehicle_service import DailyVehicleDefaults
        from app_core.session import LoginSession, SessionState
        from qt_app.controllers.daily_vehicle_controller import DailyVehicleController

        class FakeService:
            def load_defaults(self):
                return DailyVehicleDefaults("2026/07/29", ())

            def validate(self, request):
                return request

            def confirmation_summary(self, request):
                return "should not be called"

        state = SessionState()
        attempt_id = state.begin_login()
        state.complete_login(
            attempt_id,
            LoginSession("", "user10", "secret", verified=True, actor_name="測試員"),
        )
        controller = DailyVehicleController(state, FakeService())
        errors = QSignalSpy(controller.errorOccurred)

        controller.prepareRun()

        self.assertEqual(errors.count(), 1)
        self.assertIn("番號確認", errors.at(0)[0])

    def test_rescue_video_controller_requires_resolved_actor_no_before_running(self) -> None:
        from PySide6.QtTest import QSignalSpy

        from app_core.rescue_video_service import RescueVideoRequest
        from app_core.session import LoginSession, SessionState
        from qt_app.controllers.rescue_video_controller import RescueVideoController

        state = SessionState()
        attempt_id = state.begin_login()
        state.complete_login(
            attempt_id,
            LoginSession("", "user10", "secret", verified=True, actor_name="測試員"),
        )
        controller = RescueVideoController(object(), session_state=state)
        errors = QSignalSpy(controller.errorOccurred)

        controller._start_worker(
            "execute",
            RescueVideoRequest("source", "destination", "1150808", "92", "", False, "preview"),
        )

        self.assertEqual(errors.count(), 1)
        self.assertFalse(controller._workers)

    def test_rescue_video_controller_runs_preview_and_confirms_delete(self) -> None:
        from dataclasses import replace

        from PySide6.QtTest import QSignalSpy, QTest

        from app_core.rescue_video_service import (
            RescueVideoCheckCard,
            RescueVideoDefaults,
            RescueVideoRunResult,
        )
        from qt_app.controllers.rescue_video_controller import RescueVideoController

        class FakeService:
            def load_defaults(self):
                defaults = RescueVideoDefaults(
                    "source",
                    "destination",
                    "2026-07-29",
                    ("92", "93"),
                    "92",
                    "6",
                    False,
                    "來源可用\n自動採用記憶卡偏移：6 分鐘",
                    True,
                    "自動檢查通過",
                )
                return replace(
                    defaults,
                    check_cards=(
                        RescueVideoCheckCard("source", "source", "ready", "ok"),
                    ),
                )

            def validate(self, request):
                return request, [], {}

            def confirmation_summary(self, request):
                return f"確認刪除 {request.vehicle}"

            def execute(self, request, *, status_callback=None):
                status_callback("分類中")
                return RescueVideoRunResult(
                    summary_text=f"{request.mode} 完成",
                    warning_text="",
                    report_path="report.csv",
                    rows=({"sourceText": "video001.TS", "statusText": "預計複製"},),
                )

        controller = RescueVideoController(FakeService())
        success_spy = QSignalSpy(controller.runSucceeded)
        delete_spy = QSignalSpy(controller.deleteConfirmationRequested)

        controller.loadDefaults()
        for _ in range(20):
            if not controller._workers:
                break
            QTest.qWait(50)
        self.assertEqual(controller.vehicleOptions, ["92", "93"])
        self.assertTrue(controller.isReady)
        self.assertIn("自動採用記憶卡偏移", controller.checkText)
        self.assertEqual(controller.checkCards[0]["key"], "source")

        controller.preparePreview("source", "destination", "2026-07-29", "92", "", False)
        for _ in range(20):
            if success_spy.count() and not controller._workers:
                break
            success_spy.wait(250)
            QTest.qWait(10)
        self.assertEqual(success_spy.count(), 1)
        self.assertEqual(controller.resultModel.rowCount(), 1)
        self.assertEqual(controller.reportPath, "report.csv")

        controller.prepareDelete("source", "destination", "2026-07-29", "92", "", False)
        self.assertEqual(delete_spy.count(), 1)
        self.assertIn("確認刪除 92", controller.confirmationSummary)
        controller.cancelDelete()
        self.assertEqual(controller.confirmationSummary, "")
        self.assertEqual(controller.statusText, "自動檢查通過")

    def test_duty_execution_controller_runs_single_entry_lane_and_work_lane_and_deduplicates_queue(self) -> None:
        from PySide6.QtTest import QSignalSpy, QTest

        from app_core.duty_submission_service import (
            DutySubmissionRequest,
            DutySubmissionResult,
        )
        from qt_app.controllers.duty_execution_controller import DutyExecutionController

        barrier = threading.Barrier(1)

        class FakeService:
            def validate(self, request):
                return request

            def execute(self, request, *, status_callback=None):
                status_callback(f"執行 {request.action_index}")
                barrier.wait(timeout=2)
                return DutySubmissionResult(
                    request.action_index,
                    "submitted",
                    f"完成 {request.action_index}",
                    Path(f"result-{request.action_index}.json"),
                    {"group": "done"},
                )

        data = {
            "target_date": "1150729",
            "actions": [
                {"kind": "entry_log", "time": "08:00", "actor": "10"},
                {"kind": "entry_log", "time": "08:00", "actor": "11"},
                {"kind": "work_log", "time": "08:00", "actor": "10"},
            ],
        }
        controller = DutyExecutionController(FakeService())
        finished_spy = QSignalSpy(controller.actionFinished)
        first_entry_request = DutySubmissionRequest("user10", "secret", 0, data)
        second_entry_request = DutySubmissionRequest("user10", "secret", 1, data)
        work_request = DutySubmissionRequest("user10", "secret", 2, data)

        self.assertTrue(controller.enqueue(first_entry_request))
        self.assertTrue(controller.enqueue(second_entry_request))
        self.assertTrue(controller.enqueue(work_request))
        self.assertFalse(controller.enqueue(work_request))
        for _ in range(20):
            if finished_spy.count() == 3 and not controller.isBusy:
                break
            finished_spy.wait(250)
            QTest.qWait(10)

        self.assertEqual(finished_spy.count(), 3)
        self.assertFalse(controller.isBusy)
        self.assertEqual({finished_spy.at(index)[0] for index in range(finished_spy.count())}, {0, 1, 2})
        controller.shutdown()

    def test_duty_execution_controller_reuses_one_browser_for_serial_entry_queue(self) -> None:
        from PySide6.QtTest import QSignalSpy, QTest

        from app_core.duty_submission_service import DutySubmissionRequest, DutySubmissionResult
        from qt_app.controllers.duty_execution_controller import DutyExecutionController

        first_entry_started = threading.Event()
        release_first_entry = threading.Event()
        work_started = threading.Event()

        class FakeService:
            def __init__(self) -> None:
                self.entry_sessions: list[object] = []
                self.entry_calls: list[int] = []
                self.entry_trigger_types: list[str] = []
                self.work_calls: list[int] = []

            def validate(self, request):
                return request

            def open_browser_session(self, request, *, status_callback=None):
                session = SimpleNamespace(user_id=request.user_id, visible=request.visible)
                self.entry_sessions.append(session)
                if status_callback:
                    status_callback("browser ready")
                return session

            def execute_with_browser_session(self, request, _session, *, status_callback=None):
                self.entry_calls.append(request.action_index)
                self.entry_trigger_types.append(request.trigger_type)
                if request.action_index == 0:
                    first_entry_started.set()
                    release_first_entry.wait(timeout=2)
                return DutySubmissionResult(
                    request.action_index,
                    "submitted",
                    "entry submitted",
                    Path(f"entry-{request.action_index}.json"),
                    {"group": "done"},
                )

            def close_browser_session(self, _session):
                return None

            def execute(self, request, *, status_callback=None):
                self.work_calls.append(request.action_index)
                work_started.set()
                return DutySubmissionResult(
                    request.action_index,
                    "submitted",
                    "work submitted",
                    Path(f"work-{request.action_index}.json"),
                    {"group": "done"},
                )

        data = {
            "target_date": "1150807",
            "actions": [
                {"kind": "entry_log", "time": "08:00", "actor": "10"},
                {"kind": "entry_log", "time": "08:05", "actor": "10"},
                {"kind": "work_log", "time": "08:00", "actor": "10"},
            ],
        }
        service = FakeService()
        controller = DutyExecutionController(service)
        finished_spy = QSignalSpy(controller.actionFinished)

        self.assertTrue(controller.enqueue(DutySubmissionRequest("user10", "secret", 0, data)))
        self.assertTrue(first_entry_started.wait(timeout=2))
        self.assertTrue(
            controller.enqueue(DutySubmissionRequest("user10", "secret", 1, data, trigger_type="recovery"))
        )
        self.assertTrue(controller.enqueue(DutySubmissionRequest("user10", "secret", 2, data)))
        self.assertTrue(work_started.wait(timeout=2))
        release_first_entry.set()
        for _ in range(30):
            if finished_spy.count() == 3 and not controller.isBusy:
                break
            finished_spy.wait(250)
            QTest.qWait(10)

        self.assertEqual(service.entry_calls, [0, 1])
        self.assertEqual(service.entry_trigger_types, ["due", "recovery"])
        self.assertEqual(len(service.entry_sessions), 1)
        self.assertEqual(service.work_calls, [2])
        self.assertFalse(controller.isBusy)
        controller.shutdown()

    def test_duty_execution_controller_runs_manual_entry_next_without_interrupting_active_one(self) -> None:
        from PySide6.QtTest import QSignalSpy, QTest

        from app_core.duty_submission_service import DutySubmissionRequest, DutySubmissionResult
        from qt_app.controllers.duty_execution_controller import DutyExecutionController

        first_entry_started = threading.Event()
        release_first_entry = threading.Event()

        class FakeService:
            def __init__(self) -> None:
                self.calls: list[int] = []

            def validate(self, request):
                return request

            def open_browser_session(self, request, *, status_callback=None):
                return SimpleNamespace(user_id=request.user_id, visible=request.visible)

            def execute_with_browser_session(self, request, _session, *, status_callback=None):
                self.calls.append(request.action_index)
                if request.action_index == 0:
                    first_entry_started.set()
                    release_first_entry.wait(timeout=2)
                return DutySubmissionResult(
                    request.action_index,
                    "submitted",
                    "submitted",
                    Path(f"manual-priority-{request.action_index}.json"),
                    {"group": "done"},
                )

            def close_browser_session(self, _session):
                return None

        data = {
            "target_date": "1150807",
            "actions": [
                {"kind": "entry_log", "time": "08:00", "actor": "10"},
                {"kind": "entry_log", "time": "08:05", "actor": "10"},
                {"kind": "entry_log", "time": "08:10", "actor": "10"},
            ],
        }
        service = FakeService()
        controller = DutyExecutionController(service)
        finished_spy = QSignalSpy(controller.actionFinished)

        self.assertTrue(controller.enqueue(DutySubmissionRequest("user10", "secret", 0, data)))
        self.assertTrue(first_entry_started.wait(timeout=2))
        self.assertTrue(controller.enqueue(DutySubmissionRequest("user10", "secret", 1, data)))
        self.assertTrue(
            controller.enqueue(DutySubmissionRequest("user10", "secret", 2, data, trigger_type="manual"))
        )
        release_first_entry.set()
        for _ in range(30):
            if finished_spy.count() == 3 and not controller.isBusy:
                break
            finished_spy.wait(250)
            QTest.qWait(10)

        self.assertEqual(service.calls, [0, 2, 1])
        self.assertFalse(controller.isBusy)
        controller.shutdown()

    def test_duty_execution_controller_keeps_stale_entry_precheck_before_browser_start(self) -> None:
        from PySide6.QtTest import QSignalSpy, QTest

        from app_core.duty_submission_service import DutySubmissionRequest, DutySubmissionResult
        from qt_app.controllers.duty_execution_controller import DutyExecutionController

        class FakeService:
            def __init__(self) -> None:
                self.browser_started = False
                self.executed: list[int] = []

            def validate(self, request):
                return request

            def is_stale_due_request(self, _request):
                return True

            def open_browser_session(self, *_args, **_kwargs):
                self.browser_started = True
                raise AssertionError("stale entry must not start a browser")

            def execute_with_browser_session(self, *_args, **_kwargs):
                raise AssertionError("stale entry must not reuse a browser")

            def close_browser_session(self, _session):
                return None

            def execute(self, request, *, status_callback=None):
                self.executed.append(request.action_index)
                return DutySubmissionResult(
                    request.action_index,
                    "skipped_stale_schedule",
                    "stale",
                    Path("stale.json"),
                    {"group": "stale"},
                )

        data = {
            "target_date": "1150806",
            "actions": [{"kind": "entry_log", "time": "08:00", "actor": "10"}],
        }
        service = FakeService()
        controller = DutyExecutionController(service)
        finished_spy = QSignalSpy(controller.actionFinished)

        self.assertTrue(controller.enqueue(DutySubmissionRequest("user10", "secret", 0, data)))
        for _ in range(20):
            if finished_spy.count() == 1 and not controller.isBusy:
                break
            finished_spy.wait(250)
            QTest.qWait(10)

        self.assertEqual(service.executed, [0])
        self.assertFalse(service.browser_started)
        controller.shutdown()

    def test_duty_execution_controller_reopens_expired_entry_session_once(self) -> None:
        from PySide6.QtTest import QSignalSpy, QTest

        from app_core.duty_submission_service import (
            DutySubmissionExecutionError,
            DutySubmissionRequest,
            DutySubmissionResult,
        )
        from qt_app.controllers.duty_execution_controller import DutyExecutionController

        class FakeService:
            def __init__(self) -> None:
                self.opens = 0
                self.calls: list[tuple[int, int]] = []

            def validate(self, request):
                return request

            def open_browser_session(self, request, *, status_callback=None):
                session = SimpleNamespace(
                    user_id=request.user_id,
                    visible=request.visible,
                    generation=self.opens,
                )
                self.opens += 1
                return session

            def execute_with_browser_session(self, request, session, *, status_callback=None):
                self.calls.append((request.action_index, session.generation))
                if request.action_index == 1 and session.generation == 0:
                    raise DutySubmissionExecutionError("login expired", "login_failed")
                return DutySubmissionResult(
                    request.action_index,
                    "submitted",
                    "submitted",
                    Path(f"relogin-{request.action_index}.json"),
                    {"group": "done"},
                )

            def close_browser_session(self, _session):
                return None

        data = {
            "target_date": "1150807",
            "actions": [
                {"kind": "entry_log", "time": "08:00", "actor": "10"},
                {"kind": "entry_log", "time": "08:05", "actor": "10"},
            ],
        }
        service = FakeService()
        controller = DutyExecutionController(service)
        finished_spy = QSignalSpy(controller.actionFinished)
        failed_spy = QSignalSpy(controller.actionFailed)

        self.assertTrue(controller.enqueue(DutySubmissionRequest("user10", "secret", 0, data)))
        self.assertTrue(controller.enqueue(DutySubmissionRequest("user10", "secret", 1, data)))
        for _ in range(30):
            if finished_spy.count() == 2 and not controller.isBusy:
                break
            finished_spy.wait(250)
            QTest.qWait(10)

        self.assertEqual(service.opens, 2)
        self.assertEqual(service.calls, [(0, 0), (1, 0), (1, 1)])
        self.assertEqual(failed_spy.count(), 0)
        controller.shutdown()

    def test_duty_notification_identifies_the_completed_action_and_target(self) -> None:
        from qt_app.controllers.app_controller import AppController

        message = AppController._format_duty_notification(
            {
                "kind": "entry_log",
                "target": "8",
                "fields": {"出或入": "值班", "領用事由及地點": "值班"},
            },
            {"8": {"name": "曾彥綸"}},
            "登打完成",
        )

        self.assertEqual(message, "出入｜值班 / 值班 08 曾彥綸｜登打完成")

    def test_duty_execution_controller_retries_work_browser_without_using_entry_channel(self) -> None:
        from PySide6.QtTest import QSignalSpy, QTest

        from app_core.duty_submission_service import (
            DutySubmissionExecutionError,
            DutySubmissionRequest,
            DutySubmissionResult,
        )
        from qt_app.controllers.duty_execution_controller import DutyExecutionController

        first_work_started = threading.Event()
        release_first_work = threading.Event()

        class FakeService:
            def __init__(self) -> None:
                self.calls: list[int] = []
                self.first_work_failed = False

            def validate(self, request):
                return request

            def execute(self, request, *, status_callback=None):
                self.calls.append(request.action_index)
                if request.action_index == 0 and not self.first_work_failed:
                    first_work_started.set()
                    release_first_work.wait(timeout=2)
                    self.first_work_failed = True
                    raise DutySubmissionExecutionError("背景瀏覽器啟動失敗", "browser_startup")
                return DutySubmissionResult(
                    request.action_index,
                    "submitted",
                    f"完成 {request.action_index}",
                    Path(f"fallback-{request.action_index}.json"),
                    {"group": "done"},
                )

        data = {
            "target_date": "1150806",
            "actions": [
                {"kind": "work_log", "time": "18:00", "actor": "17"},
                {"kind": "work_log", "time": "18:00", "actor": "17"},
                {"kind": "entry_log", "time": "18:00", "actor": "17"},
            ],
        }
        service = FakeService()
        controller = DutyExecutionController(service)
        finished_spy = QSignalSpy(controller.actionFinished)
        failed_spy = QSignalSpy(controller.actionFailed)

        self.assertTrue(controller.enqueue(DutySubmissionRequest("user17", "secret", 0, data)))
        self.assertTrue(first_work_started.wait(timeout=2))
        self.assertTrue(controller.enqueue(DutySubmissionRequest("user17", "secret", 1, data)))
        self.assertTrue(controller.enqueue(DutySubmissionRequest("user17", "secret", 2, data)))
        release_first_work.set()
        for _ in range(40):
            if finished_spy.count() == 3 and not controller.isBusy:
                break
            finished_spy.wait(250)
            QTest.qWait(10)

        self.assertEqual(service.calls.count(0), 2)
        self.assertEqual(set(service.calls), {0, 1, 2})
        self.assertEqual({finished_spy.at(index)[0] for index in range(finished_spy.count())}, {0, 1, 2})
        self.assertEqual(failed_spy.count(), 0)
        self.assertFalse(controller.isBusy)
        controller.reset_parallel_lanes()
        self.assertEqual(controller._disabled_lanes, set())
        controller.shutdown()

    def test_duty_execution_controller_blocks_manual_duplicate_while_due_task_is_running(self) -> None:
        from PySide6.QtTest import QTest

        from app_core.duty_submission_service import DutySubmissionRequest, DutySubmissionResult
        from qt_app.controllers.duty_execution_controller import DutyExecutionController

        started = threading.Event()
        release = threading.Event()

        class FakeService:
            def validate(self, request):
                return request

            def execute(self, request, *, status_callback=None):
                started.set()
                release.wait(timeout=2)
                return DutySubmissionResult(
                    request.action_index,
                    "submitted",
                    "完成",
                    Path("duplicate.json"),
                    {"group": "done"},
                )

        data = {
            "target_date": "1150806",
            "actions": [{"kind": "entry_log", "time": "18:00", "actor": "17"}],
        }
        controller = DutyExecutionController(FakeService())
        due_request = DutySubmissionRequest("user17", "secret", 0, data, trigger_type="due")
        manual_request = DutySubmissionRequest("user17", "secret", 0, data, trigger_type="manual")

        self.assertTrue(controller.enqueue(due_request))
        self.assertTrue(started.wait(timeout=2))
        self.assertFalse(controller.enqueue(manual_request))
        release.set()
        for _ in range(20):
            if not controller.isBusy:
                break
            QTest.qWait(50)
        self.assertFalse(controller.isBusy)
        controller.shutdown()

    def test_duty_execution_controller_keeps_all_three_1800_handoff_actions(self) -> None:
        from PySide6.QtTest import QSignalSpy, QTest

        from app_core.duty_submission_service import DutySubmissionRequest, DutySubmissionResult
        from qt_app.controllers.duty_execution_controller import DutyExecutionController

        class FakeService:
            def __init__(self) -> None:
                self.executed: list[int] = []

            def validate(self, request):
                return request

            def execute(self, request, *, status_callback=None):
                self.executed.append(request.action_index)
                if status_callback:
                    status_callback(f"執行 {request.action_index}")
                return DutySubmissionResult(
                    request.action_index,
                    "submitted",
                    f"完成 {request.action_index}",
                    Path(f"handoff-{request.action_index}.json"),
                    {"group": "done"},
                )

        data = {
            "target_date": "1150806",
            "actions": [
                {"kind": "entry_log", "time": "18:00", "actor": "17", "target": "17", "source": "值班交接"},
                {"kind": "entry_log", "time": "18:00", "actor": "17", "target": "5", "source": "值班交接"},
                {"kind": "work_log", "time": "18:00", "actor": "17", "target": "17"},
            ],
        }
        service = FakeService()
        controller = DutyExecutionController(service)
        finished_spy = QSignalSpy(controller.actionFinished)

        for index in (0, 1, 2):
            self.assertTrue(controller.enqueue(DutySubmissionRequest("user17", "secret", index, data)))
        for _ in range(30):
            if finished_spy.count() == 3 and not controller.isBusy:
                break
            finished_spy.wait(250)
            QTest.qWait(10)

        self.assertEqual(set(service.executed), {0, 1, 2})
        self.assertEqual([index for index in service.executed if index in (0, 1)], [0, 1])
        self.assertEqual(finished_spy.count(), 3)
        self.assertFalse(controller.isBusy)
        controller.shutdown()

    def test_app_controller_enqueues_due_task_and_applies_verified_result(self) -> None:
        from PySide6.QtTest import QSignalSpy, QTest

        from app_core.credential_repository import CredentialRepository
        from app_core.duty_submission_service import DutySubmissionResult
        from app_core.schedule_capture_service import ScheduleCaptureRequest
        from app_core.schedule_repository import ScheduleSnapshot, business_roc_date
        from app_core.session import LoginSession
        from qt_app.controllers.app_controller import AppController

        target_roc_date = business_roc_date()

        class FakeScheduleRepository:
            def load_current(self):
                payload = {
                    "target_date": target_roc_date,
                    "today": {"staff": {"10": {"name": "本班"}}},
                    "actions": [
                        {
                            "kind": "work_log",
                            "time": "00:00",
                            "actor": "10",
                            "target": "10",
                            "source": "在隊訓練",
                            "fields": {"勤務項目": "巡邏"},
                        }
                    ],
                }
                return ScheduleSnapshot(
                    Path(f"schedule_output_{target_roc_date}.json"),
                    payload,
                    target_roc_date,
                    comparison_data={
                        target_roc_date: {
                            "visible_work_rows": [{"勤務項目": "巡邏"}],
                            "visible_entry_rows": [{"出或入": "入"}],
                        }
                    },
                    schedule_data_by_date={
                        target_roc_date: payload,
                    },
                )

        class FakeSubmissionService:
            def __init__(self):
                self.requests = []

            def validate(self, request):
                return request

            def execute(self, request, *, status_callback=None):
                self.requests.append(request)
                status_callback("假執行中")
                return DutySubmissionResult(
                    request.action_index,
                    "submitted",
                    "假登打完成",
                    Path("fake-result.json"),
                    {"group": "done"},
                )

        class FakeCaptureService:
            def current_request(self, user_id, password, actor_no, actor_name=""):
                return ScheduleCaptureRequest(
                    user_id,
                    password,
                    actor_no,
                    target_roc_date,
                    actor_name,
                )

            def capture(self, request, *, status_callback=None):
                status_callback("假即時查詢中")
                return FakeScheduleRepository().load_current()

        class FakeOperationalSyncService:
            def __init__(self):
                self.events = []
                self.boards = []

            def enqueue_event(self, record_type, **fields):
                self.events.append((record_type, fields))
                return {"record_type": record_type}

            def sync_board_async(self, schedule_data):
                self.boards.append(schedule_data)
                return True

        with tempfile.TemporaryDirectory() as temp_dir:
            credentials = CredentialRepository(Path(temp_dir) / "saved.json", "SinpoSmart", None)
            submission = FakeSubmissionService()
            operational_sync = FakeOperationalSyncService()
            controller = AppController(
                repository=credentials,
                schedule_repository=FakeScheduleRepository(),
                duty_submission_service=submission,
                schedule_capture_service=FakeCaptureService(),
                operational_sync_service=operational_sync,
            )
            finished_spy = QSignalSpy(controller.dutyExecutionController.actionFinished)
            attempt_id = controller._session_state.begin_login()
            controller._session_state.complete_login(
                attempt_id,
                LoginSession("10", "user10", "secret", verified=True, actor_name="本班"),
            )

            controller.sessionController.sessionChanged.emit()
            for _ in range(30):
                if finished_spy.count() and not controller.dutyExecutionController.isBusy:
                    break
                finished_spy.wait(250)
                QTest.qWait(10)

            self.assertEqual(len(submission.requests), 1)
            self.assertEqual(finished_spy.count(), 1)
            self.assertEqual(controller.dutyController.dueTaskCount, 0)
            self.assertIn("假登打完成", controller.dutyController.scheduleStatus)
            self.assertEqual(len(operational_sync.boards), 1)
            for _ in range(100):
                if (
                    not controller._operational_sync_workers
                    and not controller._operational_sync_queue
                ):
                    break
                QTest.qWait(10)
            record_types = [record_type for record_type, _fields in operational_sync.events]
            self.assertIn("login", record_types)
            self.assertIn("schedule_snapshot", record_types)
            self.assertIn("comparison_snapshot", record_types)
            self.assertNotIn("schedule_refresh", record_types)
            self.assertIn("action_queued", record_types)
            self.assertIn("action_result", record_types)
            self.assertNotIn("action_finished", record_types)
            schedule_fields = next(
                fields
                for record_type, fields in operational_sync.events
                if record_type == "schedule_snapshot"
            )
            comparison_fields = next(
                fields
                for record_type, fields in operational_sync.events
                if record_type == "comparison_snapshot"
            )
            self.assertEqual(
                [day["target_date"] for day in schedule_fields["snapshot"]["days"]],
                [target_roc_date],
            )
            self.assertEqual([day["action_count"] for day in schedule_fields["snapshot"]["days"]], [1])
            self.assertEqual(comparison_fields["snapshot"]["days"][0]["work_count"], 1)
            self.assertEqual(comparison_fields["snapshot"]["days"][0]["entry_count"], 1)
            result_fields = next(
                fields
                for record_type, fields in operational_sync.events
                if record_type == "action_result"
            )
            self.assertEqual(result_fields["status"], "submitted")
            self.assertEqual(result_fields["trigger_type"], "due")
            self.assertEqual(result_fields["action"]["fields"]["勤務項目"], "巡邏")
            self.assertEqual(result_fields["target"], "10番 本班")
            self.assertTrue(result_fields["snapshot"]["completion_key"])
            self.assertEqual(
                controller.workLogSettingsController._schedule_data.get("target_date"),
                target_roc_date,
            )

    def test_app_controller_preserves_update_logout_identity_and_legacy_event_contract(self) -> None:
        from app_core.credential_repository import CredentialRepository
        from app_core.session import LoginSession
        from qt_app.controllers.app_controller import AppController

        class FakeOperationalSyncService:
            def __init__(self) -> None:
                self.events = []

            def enqueue_event(self, record_type, **fields):
                self.events.append((record_type, fields))
                return {"record_type": record_type}

            def sync_board_async(self, _schedule_data):
                return True

        with tempfile.TemporaryDirectory() as temp_dir:
            operational_sync = FakeOperationalSyncService()
            controller = AppController(
                repository=CredentialRepository(Path(temp_dir) / "saved.json", "SinpoSmart", None),
                operational_sync_service=operational_sync,
            )
            controller.dutyController.refresh_live_schedule = lambda *_args, **_kwargs: None
            try:
                attempt_id = controller._session_state.begin_login()
                controller._session_state.complete_login(
                    attempt_id,
                    LoginSession("10", "user10", "secret", verified=True),
                )
                controller.sessionController._display_name = "10番 隊員 測試員"
                controller._sync_session_actor()
                from PySide6.QtTest import QTest

                for _ in range(50):
                    if not controller._operational_sync_workers:
                        break
                    QTest.qWait(10)
                operational_sync.events.clear()
                controller.sessionController.logout()

                for _ in range(50):
                    if not controller._operational_sync_workers:
                        break
                    QTest.qWait(10)
                logout_events = list(operational_sync.events)
                operational_sync.events.clear()

                recorded = controller.recordUpdateLogout()
            finally:
                controller.shutdown()

        self.assertTrue(recorded)
        self.assertEqual(len(logout_events), 1)
        self.assertEqual(logout_events[0][0], "logout")
        self.assertEqual(logout_events[0][1]["actor_no"], "10")
        self.assertEqual(logout_events[0][1]["display_name"], "10番 測試員")
        self.assertEqual(len(operational_sync.events), 1)
        record_type, fields = operational_sync.events[0]
        self.assertEqual(record_type, "logout")
        self.assertEqual(fields["trigger_type"], "update")
        self.assertEqual(fields["content"], "更新前登出")
        self.assertEqual(fields["actor_no"], "10")
        self.assertEqual(fields["user_id"], "user10")
        self.assertEqual(fields["display_name"], "10番 測試員")
        self.assertTrue(fields["immediate"])

    def test_app_controller_persists_queued_logout_before_shutdown(self) -> None:
        from qt_app.controllers.app_controller import AppController

        class FakeOperationalSyncService:
            def __init__(self) -> None:
                self.events = []

            def enqueue_event(self, record_type, **fields):
                self.events.append((record_type, fields))
                return {"record_type": record_type}

        operational_sync = FakeOperationalSyncService()
        controller = AppController(operational_sync_service=operational_sync)
        controller._operational_sync_queue.append(
            (
                1,
                "event",
                "logout",
                {
                    "status": "ok",
                    "trigger_type": "logout",
                    "actor_no": "10",
                    "user_id": "user10",
                    "display_name": "10番 測試員",
                },
                {},
            )
        )
        controller.shutdown()

        self.assertEqual(len(operational_sync.events), 1)
        record_type, fields = operational_sync.events[0]
        self.assertEqual(record_type, "logout")
        self.assertTrue(fields["immediate"])
        self.assertEqual(fields["display_name"], "10番 測試員")

    def test_app_controller_uses_managed_qthreads_for_operational_sync(self) -> None:
        from PySide6.QtTest import QTest

        from qt_app.controllers.app_controller import AppController

        class FakeOperationalSyncService:
            def __init__(self) -> None:
                self.events = []
                self.boards = []

            def enqueue_event(self, record_type, **fields):
                self.events.append((record_type, fields, QThread.currentThread()))
                return {"record_type": record_type}

            def sync_board_async(self, schedule_data):
                self.boards.append((schedule_data, QThread.currentThread()))
                return True

        from PySide6.QtCore import QThread

        operational_sync = FakeOperationalSyncService()
        controller = AppController(operational_sync_service=operational_sync)
        try:
            controller._send_operational_event("login", status="ok")
            controller._start_operational_sync(
                "board",
                schedule_data={"target_date": "1150729"},
            )
            for _ in range(100):
                if not controller._operational_sync_workers:
                    break
                QTest.qWait(10)

            self.assertFalse(controller._operational_sync_workers)
            self.assertEqual(operational_sync.events[0][0], "login")
            self.assertEqual(operational_sync.boards[0][0]["target_date"], "1150729")
            self.assertIsNot(operational_sync.events[0][2], self.app.thread())
            self.assertIsNot(operational_sync.boards[0][1], self.app.thread())
        finally:
            controller.shutdown()

    def test_app_controller_retries_current_google_duty_board_after_a_failed_post(self) -> None:
        from PySide6.QtTest import QTest

        from app_core.schedule_repository import business_roc_date
        from app_core.session import LoginSession
        from qt_app.controllers.app_controller import AppController

        class FakeOperationalSyncService:
            board_enabled = True

            def __init__(self) -> None:
                self.boards = []

            def enqueue_event(self, _record_type, **_fields):
                return {}

            def sync_board_async(self, schedule_data):
                self.boards.append(schedule_data)
                return True

        service = FakeOperationalSyncService()
        controller = AppController(operational_sync_service=service)
        try:
            attempt_id = controller._session_state.begin_login()
            controller._session_state.complete_login(
                attempt_id,
                LoginSession("10", "user10", "secret", verified=True),
            )
            controller._duty_controller._schedule_data = {
                "target_date": business_roc_date(),
            }

            controller._retry_current_duty_board()
            for _ in range(100):
                if service.boards and not controller._operational_sync_workers:
                    break
                QTest.qWait(10)
        finally:
            controller.shutdown()

        self.assertEqual(len(service.boards), 1)
        self.assertEqual(service.boards[0]["target_date"], business_roc_date())

    def test_app_controller_has_no_direct_async_operational_sync_call(self) -> None:
        source = (
            PACKAGE_ROOT / "qt_app" / "controllers" / "app_controller.py"
        ).read_text(encoding="utf-8")

        self.assertNotIn(".sync_board_async(", source)

    def test_app_controller_shutdown_waits_for_operational_sync_worker(self) -> None:
        from PySide6.QtTest import QTest

        from qt_app.controllers.app_controller import AppController

        started = threading.Event()
        release = threading.Event()
        finished = threading.Event()

        class FakeOperationalSyncService:
            def enqueue_event(self, _record_type, **_fields):
                started.set()
                release.wait(2)
                finished.set()
                return {}

            def sync_board_async(self, _schedule_data):
                return True

        controller = AppController(operational_sync_service=FakeOperationalSyncService())
        controller._send_operational_event("login", status="ok")
        for _ in range(100):
            if started.is_set():
                break
            QTest.qWait(10)
        self.assertTrue(started.is_set())

        release_timer = threading.Timer(0.05, release.set)
        release_timer.start()
        controller.shutdown()
        release_timer.join()

        self.assertTrue(finished.is_set())
        self.assertFalse(controller._operational_sync_workers)

    def test_app_controller_reports_submission_validation_failure_as_legacy_action_result(self) -> None:
        from app_core.credential_repository import CredentialRepository
        from app_core.duty_submission_service import DutySubmissionRequest
        from qt_app.controllers.app_controller import AppController

        class FakeOperationalSyncService:
            def __init__(self) -> None:
                self.events = []

            def enqueue_event(self, record_type, **fields):
                self.events.append((record_type, fields))
                return {"record_type": record_type}

            def sync_board_async(self, _schedule_data):
                return True

        request = DutySubmissionRequest(
            "",
            "",
            0,
            {
                "target_date": "1150729",
                "actions": [
                    {
                        "kind": "work_log",
                        "source": "值班交接",
                        "actor": "10",
                        "target": "10",
                        "time": "08:00",
                        "duplicate_key": "work:1150729:0800:值班交接:10",
                        "fields": {"工作項目": "值班交接"},
                    }
                ],
            },
            trigger_type="manual",
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            operational_sync = FakeOperationalSyncService()
            controller = AppController(
                repository=CredentialRepository(Path(temp_dir) / "saved.json", "SinpoSmart", None),
                operational_sync_service=operational_sync,
            )
            try:
                accepted = controller.dutyExecutionController.enqueue(request)
            finally:
                controller.shutdown()

        self.assertFalse(accepted)
        self.assertEqual(len(operational_sync.events), 1)
        record_type, fields = operational_sync.events[0]
        self.assertEqual(record_type, "action_result")
        self.assertEqual(fields["status"], "failed")
        self.assertEqual(fields["trigger_type"], "manual")
        self.assertEqual(fields["snapshot"]["completion_key"], "work:1150729:0800:值班交接:10")
        self.assertEqual(fields["snapshot"]["error_code"], "validation_error")

    def test_app_controller_exports_issue_package_for_worker_submission_failure(self) -> None:
        from app_core.credential_repository import CredentialRepository
        from app_core.duty_submission_service import DutySubmissionRequest
        from qt_app.controllers.app_controller import AppController

        class FakeOperationalSyncService:
            def __init__(self) -> None:
                self.events = []

            def enqueue_event(self, record_type, **fields):
                self.events.append((record_type, fields))
                return {"record_type": record_type}

            def sync_board_async(self, _schedule_data):
                return True

        class FakeDiagnosticsService:
            def __init__(self) -> None:
                self.snapshots = []

            def export(self, snapshot):
                self.snapshots.append(snapshot)
                return Path("issue_reports/issue_report_test.zip")

        request = DutySubmissionRequest(
            "user10",
            "secret",
            0,
            {
                "target_date": "1150729",
                "actions": [
                    {
                        "kind": "work_log",
                        "source": "值班交接",
                        "actor": "10",
                        "target": "10",
                        "time": "08:00",
                        "duplicate_key": "work:1150729:0800:值班交接:10",
                        "fields": {"勤務項目": "值班交接"},
                    }
                ],
            },
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            operational_sync = FakeOperationalSyncService()
            diagnostics = FakeDiagnosticsService()
            controller = AppController(
                repository=CredentialRepository(Path(temp_dir) / "saved.json", "SinpoSmart", None),
                operational_sync_service=operational_sync,
                diagnostics_service=diagnostics,
            )
            try:
                controller.dutyExecutionController.submissionFailed.emit(
                    request,
                    "勤務系統登打失敗。",
                    "timeout",
                    str(Path(temp_dir) / "runtime_outputs" / "form_tests" / "result.json"),
                )
            finally:
                controller.shutdown()

        self.assertEqual(len(diagnostics.snapshots), 1)
        self.assertEqual(controller.diagnosticsStatus, "問題包已匯出：issue_report_test.zip")
        record_type, fields = operational_sync.events[0]
        self.assertEqual(record_type, "action_result")
        self.assertEqual(fields["result_ref"], "result.json")
        self.assertEqual(fields["snapshot"]["error_code"], "timeout")

    def test_app_controller_reports_safe_login_failure_without_password(self) -> None:
        from app_core.credential_repository import CredentialRepository
        from qt_app.controllers.app_controller import AppController

        class FakeOperationalSyncService:
            def __init__(self) -> None:
                self.events = []

            def enqueue_event(self, record_type, **fields):
                self.events.append((record_type, fields))
                return {"record_type": record_type}

            def sync_board_async(self, _schedule_data):
                return True

        with tempfile.TemporaryDirectory() as temp_dir:
            operational_sync = FakeOperationalSyncService()
            controller = AppController(
                repository=CredentialRepository(Path(temp_dir) / "saved.json", "SinpoSmart", None),
                operational_sync_service=operational_sync,
            )
            try:
                controller.sessionController.loginAttemptFailed.emit(
                    "user10",
                    "登入失敗：請確認帳號密碼。",
                    "login_failed",
                )
            finally:
                controller.shutdown()

        self.assertEqual(len(operational_sync.events), 1)
        record_type, fields = operational_sync.events[0]
        self.assertEqual(record_type, "login_failed")
        self.assertEqual(fields["trigger_type"], "login")
        self.assertEqual(fields["user_id"], "user10")
        self.assertEqual(fields["snapshot"]["error_code"], "login_failed")
        self.assertNotIn("password", json.dumps(fields, ensure_ascii=False).lower())

    def test_app_controller_reports_and_clears_expired_live_capture_session(self) -> None:
        from app_core.credential_repository import CredentialRepository
        from app_core.schedule_repository import business_roc_date
        from app_core.session import LoginSession
        from qt_app.controllers.app_controller import AppController

        class FakeOperationalSyncService:
            def __init__(self) -> None:
                self.events = []

            def enqueue_event(self, record_type, **fields):
                self.events.append((record_type, fields))
                return {"record_type": record_type}

            def sync_board_async(self, _schedule_data):
                return True

        with tempfile.TemporaryDirectory() as temp_dir:
            operational_sync = FakeOperationalSyncService()
            controller = AppController(
                repository=CredentialRepository(Path(temp_dir) / "saved.json", "SinpoSmart", None),
                operational_sync_service=operational_sync,
            )
            controller.dutyController.refresh_live_schedule = lambda *_args, **_kwargs: None
            try:
                attempt_id = controller._session_state.begin_login()
                controller._session_state.complete_login(
                    attempt_id,
                    LoginSession("10", "user10", "secret", verified=True),
                )
                controller._sync_session_actor()
                from PySide6.QtTest import QTest

                for _ in range(50):
                    if not controller._operational_sync_workers:
                        break
                    QTest.qWait(10)
                operational_sync.events.clear()
                controller.dutyController._active_capture_request = 1
                controller.dutyController._capture_targets[1] = business_roc_date()

                controller.dutyController._capture_failed(
                    1,
                    "10",
                    "登入失敗：請重新登入。",
                    "login_failed",
                )
            finally:
                controller.shutdown()

        self.assertFalse(controller.sessionController.isLoggedIn)
        self.assertEqual(
            [record_type for record_type, _fields in operational_sync.events],
            ["error", "login_expired", "logout"],
        )
        self.assertEqual(operational_sync.events[0][1]["trigger_type"], "schedule")
        self.assertEqual(operational_sync.events[0][1]["snapshot"]["error_code"], "login_failed")
        self.assertEqual(operational_sync.events[1][1]["trigger_type"], "login")

    def test_app_controller_preserves_legacy_tool_failure_event_contract(self) -> None:
        from app_core.credential_repository import CredentialRepository
        from qt_app.controllers.app_controller import AppController
        from qt_app.controllers.tool_controller import ToolController

        class FakeOperationalSyncService:
            def __init__(self) -> None:
                self.events = []

            def enqueue_event(self, record_type, **fields):
                self.events.append((record_type, fields))
                return {"record_type": record_type}

            def sync_board_async(self, _schedule_data):
                return True

        with tempfile.TemporaryDirectory() as temp_dir:
            operational_sync = FakeOperationalSyncService()
            controller = AppController(
                repository=CredentialRepository(Path(temp_dir) / "saved.json", "SinpoSmart", None),
                operational_sync_service=operational_sync,
                tool_controller=ToolController(Path(temp_dir)),
            )
            try:
                controller.dutySheetController.errorOccurred.emit("表單驗證失敗")
                self.assertEqual(operational_sync.events, [])

                controller.dutySheetController.runStarted.emit()
                controller.dutySheetController.runFailed.emit("工具執行失敗")
            finally:
                controller.shutdown()

        self.assertEqual(
            [record_type for record_type, _fields in operational_sync.events],
            ["tool_action_started", "tool_action_finished"],
        )
        self.assertEqual(
            [fields["trigger_type"] for _record_type, fields in operational_sync.events],
            ["tool_start", "tool_finish"],
        )
        self.assertEqual(
            [fields["status"] for _record_type, fields in operational_sync.events],
            ["started", "failed"],
        )
        self.assertEqual(operational_sync.events[0][1]["snapshot"]["tool_name"], "duty_sheet")
        self.assertEqual(operational_sync.events[1][1]["error"], "工具執行失敗")
        self.assertTrue(controller.toolController.usage("duty_sheet")["report"].endswith("勤務表失敗"))

    def test_tool_failure_event_includes_browser_detail_for_nas_backend(self) -> None:
        from app_core.credential_repository import CredentialRepository
        from qt_app.controllers.app_controller import AppController
        from qt_app.controllers.tool_controller import ToolController

        class FakeOperationalSyncService:
            def __init__(self) -> None:
                self.events = []

            def enqueue_event(self, record_type, **fields):
                self.events.append((record_type, fields))
                return {"record_type": record_type}

            def sync_board_async(self, _schedule_data):
                return True

        with tempfile.TemporaryDirectory() as temp_dir:
            operational_sync = FakeOperationalSyncService()
            controller = AppController(
                repository=CredentialRepository(Path(temp_dir) / "saved.json", "SinpoSmart", None),
                operational_sync_service=operational_sync,
                tool_controller=ToolController(Path(temp_dir)),
            )
            try:
                controller.restMonthlyController._failure_stage = "browser_start"
                controller.restMonthlyController._failure_detail = "browser_startup"
                controller.restMonthlyController.runStarted.emit("monthly_base")
                controller.restMonthlyController.runFailed.emit("monthly_base", "browser failed")
            finally:
                controller.shutdown()

        self.assertEqual(
            [record_type for record_type, _fields in operational_sync.events],
            ["tool_action_started", "tool_action_finished"],
        )
        failure = operational_sync.events[1][1]
        self.assertEqual(failure["status"], "failed")
        self.assertEqual(failure["error"], "browser failed")
        self.assertEqual(failure["snapshot"]["tool_name"], "monthly_base")
        self.assertEqual(failure["snapshot"]["failure_stage"], "browser_start")
        self.assertEqual(failure["snapshot"]["failure_detail"], "browser_startup")

    def test_all_tool_failures_include_error_and_stage_for_nas_backend(self) -> None:
        """Every QML tool failure keeps a displayable NAS event contract."""
        from app_core.credential_repository import CredentialRepository
        from qt_app.controllers.app_controller import AppController
        from qt_app.controllers.tool_controller import ToolController

        class FakeOperationalSyncService:
            def __init__(self) -> None:
                self.events = []

            def enqueue_event(self, record_type, **fields):
                self.events.append((record_type, fields))
                return {"record_type": record_type}

            def sync_board_async(self, _schedule_data):
                return True

        with tempfile.TemporaryDirectory() as temp_dir:
            operational_sync = FakeOperationalSyncService()
            controller = AppController(
                repository=CredentialRepository(Path(temp_dir) / "saved.json", "SinpoSmart", None),
                operational_sync_service=operational_sync,
                tool_controller=ToolController(Path(temp_dir)),
            )
            try:
                controller.dutySheetController._failure_stage = "submit"
                controller.dutySheetController.runFailed.emit("勤務表失敗")
                controller.restMonthlyController._failure_stage = "source_read"
                controller.restMonthlyController.runFailed.emit("rest_time", "休息時間失敗")
                controller.restMonthlyController._failure_stage = "browser_start"
                controller.restMonthlyController.runFailed.emit("monthly_base", "勤務基準表失敗")
                controller.dailyVehicleController._failure_stage = "login"
                controller.dailyVehicleController.runFailed.emit("每日車輛失敗")
                controller.rescueVideoController._failure_stage = "file_scan"
                controller.rescueVideoController.runFailed.emit("preview", "行車紀錄器失敗")
            finally:
                controller.shutdown()

        failures = [fields for record_type, fields in operational_sync.events if record_type == "tool_action_finished"]
        self.assertEqual(
            {failure["snapshot"]["tool_name"] for failure in failures},
            {"duty_sheet", "rest_time", "monthly_base", "daily_vehicle", "rescue_video"},
        )
        for failure in failures:
            self.assertEqual(failure["status"], "failed")
            self.assertTrue(failure["error"])
            self.assertTrue(failure["snapshot"]["failure_stage"])

    def test_auto_logout_requires_post_login_handoff_and_completed_group(self) -> None:
        from datetime import datetime

        from PySide6.QtTest import QSignalSpy

        from qt_app.controllers.duty_controller import DutyController

        action = {
            "kind": "entry_log",
            "time": "08:00",
            "actor": "10",
            "target": "10",
            "source": "值班交接",
            "fields": {"出或入": "值退", "領用事由及地點": "退勤"},
        }
        controller = DutyController()
        controller.set_actor_no("10")
        external_action = {
            "kind": "entry_log",
            "time": "08:00",
            "actor": "10",
            "target": "10",
            "source": "外勤支援",
            "fields": {"出或入": "出", "領用事由及地點": "外勤"},
        }
        controller.replace_schedule_data(
            {"target_date": "1150729", "actions": [action, external_action]}
        )
        logout_spy = QSignalSpy(controller.autoLogoutRequested)

        controller._login_started_at = datetime(2026, 7, 29, 7, 59)
        controller.handle_submission_result(0, "submitted", "完成", "result.json")
        self.assertTrue(controller._auto_logout_timer.isActive())
        controller._check_auto_logout()
        self.assertEqual(logout_spy.count(), 1)

        controller.replace_schedule_data({"target_date": "1150730", "actions": [action]})
        controller._login_started_at = datetime(2026, 7, 30, 9, 0)
        controller.handle_submission_result(0, "skipped_duplicate", "已存在", "result.json")
        self.assertFalse(controller._auto_logout_timer.isActive())

    def test_fire_day_reload_preserves_executed_actions_by_completion_key(self) -> None:
        from qt_app.controllers.duty_controller import DutyController

        previous_action = {
            "kind": "entry_log",
            "time": "07:55",
            "date_offset": 1,
            "actor": "10",
            "target": "10",
            "duplicate_key": "entry:2026-08-07:755:in:10:到勤",
            "fields": {"登打時間": "07:55", "系統寫入時間": "07:55", "出或入": "入", "領用事由及地點": "到勤"},
        }
        carried_action = {**previous_action, "date_offset": 0}
        new_action = {
            "kind": "entry_log",
            "time": "08:00",
            "actor": "10",
            "target": "10",
            "duplicate_key": "entry:2026-08-07:8:值班:10",
            "fields": {"登打時間": "08:00", "系統寫入時間": "08:00", "出或入": "值班", "領用事由及地點": "值班"},
        }
        controller = DutyController()
        controller.set_actor_no("10")
        try:
            controller.replace_schedule_data({"target_date": "1150806", "actions": [previous_action]})
            controller.handle_submission_result(0, "submitted", "完成", "result.json")

            controller.replace_schedule_data({"target_date": "1150807", "actions": [carried_action, new_action]})

            self.assertEqual(controller._executed_indices, {0})
            self.assertEqual(controller._comparisons[0]["group"], "done")
            self.assertNotIn(1, controller._executed_indices)
        finally:
            controller.shutdown()

    def test_manual_submission_includes_external_review_task(self) -> None:
        from PySide6.QtTest import QSignalSpy

        from qt_app.controllers.duty_controller import DutyController

        controller = DutyController()
        controller.set_actor_no("10")
        controller.replace_schedule_data(
            {
                "target_date": "1150729",
                "actions": [
                    {
                        "kind": "entry_log",
                        "time": "08:00",
                        "actor": "10",
                        "target": "10",
                        "source": "外勤支援",
                        "fields": {"出或入": "出", "領用事由及地點": "外勤"},
                    }
                ],
            },
            comparisons={0: {"compare": "外勤確認", "group": "review", "matched": []}},
        )
        confirmation_spy = QSignalSpy(controller.manualSubmissionConfirmationRequested)
        controller.toggleTaskSelection(0)

        controller.prepareManualSubmission()

        self.assertEqual(confirmation_spy.count(), 1)
        self.assertEqual(controller._pending_manual_indices, [0])

    def test_manual_submission_includes_manual_group_task(self) -> None:
        from PySide6.QtTest import QSignalSpy

        from qt_app.controllers.duty_controller import DutyController

        controller = DutyController()
        controller.set_actor_no("10")
        controller.replace_schedule_data(
            {
                "target_date": "1150729",
                "actions": [
                    {
                        "kind": "work_log",
                        "time": "08:00",
                        "actor": "9",
                        "target": "10",
                        "source": "人工補登",
                        "fields": {"勤務項目": "工作紀錄"},
                    }
                ],
            },
            comparisons={0: {"compare": "手動登打", "group": "manual", "matched": []}},
        )
        confirmation_spy = QSignalSpy(controller.manualSubmissionConfirmationRequested)
        controller.toggleTaskSelection(0)

        controller.prepareManualSubmission()

        self.assertEqual(confirmation_spy.count(), 1)
        self.assertEqual(controller._pending_manual_indices, [0])

    def test_external_return_pause_uses_dedicated_manual_confirmation(self) -> None:
        from PySide6.QtTest import QSignalSpy

        from app_core.unreturned_return_queue import UnreturnedReturnQueue
        from qt_app.controllers.duty_controller import DutyController

        temporary_queue_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temporary_queue_dir.cleanup)
        controller = DutyController(
            unreturned_return_queue=UnreturnedReturnQueue(Path(temporary_queue_dir.name))
        )
        controller.set_actor_no("10")
        controller.replace_schedule_data(
            {
                "target_date": "1150807",
                "today": {"staff": {"8": {"name": "測試員"}}},
                "actions": [
                    {
                        "kind": "entry_log",
                        "time": "08:00",
                        "actor": "9",
                        "target": "8",
                        "source": "昨日在勤且今日未在勤",
                        "fields": {"出或入": "出", "領用事由及地點": "退勤"},
                    }
                ],
            },
            comparisons={0: {"compare": "未返隊，暫停登打", "group": "paused", "matched": []}},
        )
        confirmation_spy = QSignalSpy(controller.externalReturnManualSubmissionConfirmationRequested)
        requested_spy = QSignalSpy(controller.externalReturnQueueManualSubmissionRequested)
        controller.handle_submission_result(0, "paused_external", "人員尚未返隊", "")
        controller.toggleTaskSelection(0)

        self.assertTrue(controller.hasExternalReturnPauseSelected)
        self.assertTrue(controller.canConfirmExternalReturnManualSubmissionSelected)
        controller.prepareExternalReturnManualSubmission()

        self.assertEqual(confirmation_spy.count(), 1)
        self.assertEqual(controller._pending_external_return_indices, [0])
        self.assertIn("確認人員已返隊", controller.externalReturnConfirmationSummary)
        self.assertIn("08 測試員", controller.externalReturnConfirmationSummary)
        controller.confirmExternalReturnManualSubmission()

        self.assertEqual(requested_spy.count(), 1)
        self.assertEqual(
            requested_spy.at(0),
            [controller._external_return_queue_ids_by_action_index[0]],
        )
        self.assertEqual(controller.selectedTaskCount, 0)

    def test_external_return_pause_mixed_selection_disables_confirmation(self) -> None:
        from qt_app.controllers.duty_controller import DutyController

        controller = DutyController()
        controller.set_actor_no("10")
        controller.replace_schedule_data(
            {
                "target_date": "1150807",
                "actions": [
                    {
                        "kind": "entry_log",
                        "time": "08:00",
                        "actor": "9",
                        "target": "8",
                        "fields": {"出或入": "出", "領用事由及地點": "退勤"},
                    },
                    {
                        "kind": "work_log",
                        "time": "09:00",
                        "actor": "10",
                        "target": "10",
                        "fields": {"勤務項目": "一般勤務"},
                    },
                ],
            },
            comparisons={0: {"compare": "未返隊，暫停登打", "group": "paused", "matched": []}},
        )
        controller.toggleTaskSelection(0)
        controller.toggleTaskSelection(1)

        self.assertTrue(controller.hasExternalReturnPauseSelected)
        self.assertFalse(controller.canConfirmExternalReturnManualSubmissionSelected)

    def test_persisted_external_return_queue_uses_special_confirmation_and_current_time(self) -> None:
        from datetime import datetime

        from PySide6.QtTest import QSignalSpy

        from app_core.unreturned_return_queue import UnreturnedReturnQueue
        from qt_app.controllers.duty_controller import DutyController

        with tempfile.TemporaryDirectory() as temp_dir:
            queue = UnreturnedReturnQueue(Path(temp_dir))
            action = {
                "kind": "entry_log",
                "time": "08:05",
                "actor": "10",
                "target": "8",
                "fields": {"出或入": "值退", "領用事由及地點": "退勤"},
            }
            queue.pause(
                action,
                {
                    "target_date": "1150807",
                    "today": {"staff": {"8": {"name": "測試員"}}},
                },
                owner_actor_no="10",
            )
            controller = DutyController(unreturned_return_queue=queue)
            controller.set_actor_no("11")
            controller.replace_schedule_data(
                {
                    "target_date": "1150807",
                    "today": {"staff": {"8": {"name": "測試員"}}},
                    "actions": [action],
                }
            )
            queue_id = controller._external_return_queue_ids_by_action_index[0]
            confirmation_spy = QSignalSpy(controller.externalReturnManualSubmissionConfirmationRequested)
            requested_spy = QSignalSpy(controller.externalReturnQueueManualSubmissionRequested)

            controller.toggleTaskSelection(0)
            self.assertTrue(controller.canConfirmExternalReturnManualSubmissionSelected)
            controller.prepareExternalReturnManualSubmission()

            self.assertEqual(confirmation_spy.count(), 1)
            self.assertIn("確認人員已返隊", controller.externalReturnConfirmationSummary)
            controller.confirmExternalReturnManualSubmission()
            self.assertEqual(requested_spy.at(0), [queue_id])
            request = controller.queued_external_return_manual_submission_request(
                "user11",
                "secret",
                queue_id,
                submit_at=datetime(2026, 8, 7, 1, 2),
            )
            self.assertIsNotNone(request)
            queued_action = request.schedule_data["actions"][0]
            self.assertEqual(request.trigger_type, "manual")
            self.assertEqual(queued_action["time"], "01:02")
            self.assertEqual(queued_action["fields"]["登打時間"], "01:02")
            self.assertEqual(queued_action["fields"]["系統寫入時間"], "01:02")
            self.assertEqual(queued_action["submit_target_date"], "1150807")

    def test_handoff_external_pause_groups_three_items_and_manual_uses_actual_time(self) -> None:
        from datetime import datetime

        from PySide6.QtTest import QSignalSpy

        from app_core.unreturned_return_queue import UnreturnedReturnQueue
        from qt_app.controllers.duty_controller import DutyController

        temporary_queue_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temporary_queue_dir.cleanup)
        controller = DutyController(
            unreturned_return_queue=UnreturnedReturnQueue(Path(temporary_queue_dir.name))
        )
        controller.set_actor_no("10")
        controller.replace_schedule_data(
            {
                "target_date": "1150807",
                "today": {
                    "staff": {
                        "10": {"name": "原值班"},
                        "11": {"name": "接班"},
                    }
                },
                "actions": [
                    {
                        "kind": "entry_log",
                        "time": "00:00",
                        "actor": "10",
                        "target": "10",
                        "source": "值班交接",
                        "duplicate_key": "entry:1150807:00:值退:10",
                        "fields": {"出或入": "值退", "領用事由及地點": "值退"},
                    },
                    {
                        "kind": "entry_log",
                        "time": "00:00",
                        "actor": "10",
                        "target": "11",
                        "source": "值班交接",
                        "duplicate_key": "entry:1150807:00:值班:11",
                        "fields": {"出或入": "值班", "領用事由及地點": "值班"},
                    },
                    {
                        "kind": "work_log",
                        "time": "00:00",
                        "actor": "10",
                        "target": "10",
                        "source": "值班交接",
                        "duplicate_key": "work:1150807:00:值班交接:10",
                        "fields": {
                            "工作時間": "00:00",
                            "處理情形": "一、時間:16-18\n二、交接完成",
                        },
                    },
                ],
            }
        )
        controller._due_task_indices = [0, 1, 2]
        controller.enable_auto_execution()

        preflight_requests = controller.due_submission_requests("user10", "secret", [0, 1, 2])

        self.assertEqual(len(preflight_requests), 1)
        self.assertEqual(
            preflight_requests[0].schedule_data["actions"][preflight_requests[0].action_index]["kind"],
            "handoff_preflight",
        )
        controller.mark_submission_enqueued(preflight_requests[0].action_index)
        controller.handle_handoff_preflight_paused(preflight_requests[0])

        records = controller._unreturned_return_queue.active_records()
        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["record_type"], "handoff_group")
        self.assertEqual(len(records[0]["actions"]), 3)
        self.assertEqual(set(controller._external_return_queue_ids_by_action_index), {0, 1, 2})

        controller.toggleTaskSelection(1)
        self.assertEqual(controller.selectedTaskCount, 3)
        self.assertTrue(controller.canConfirmExternalReturnManualSubmissionSelected)
        confirmation_spy = QSignalSpy(controller.externalReturnManualSubmissionConfirmationRequested)
        controller.prepareExternalReturnManualSubmission()
        self.assertEqual(confirmation_spy.count(), 1)

        queue_id = controller._external_return_queue_ids_by_action_index[1]
        requests = controller.queued_external_return_manual_submission_requests(
            "user10",
            "secret",
            queue_id,
            submit_at=datetime(2026, 8, 7, 18, 25),
        )

        self.assertEqual(len(requests), 3)
        stamped_actions = [request.schedule_data["actions"][0] for request in requests]
        self.assertEqual(
            [action["kind"] for action in stamped_actions],
            ["entry_log", "entry_log", "work_log"],
        )
        self.assertTrue(
            all(action["time"] == "18:25" for action in stamped_actions)
        )
        work_action = stamped_actions[2]
        self.assertEqual(work_action["fields"]["工作時間"], "18:25")
        self.assertEqual(work_action["fields"]["處理情形"].splitlines()[0], "一、時間:16:00-18:25")

    def test_handoff_selection_groups_checkout_on_duty_and_work(self) -> None:
        from qt_app.controllers.duty_controller import DutyController

        controller = DutyController()
        controller.replace_schedule_data(
            {
                "target_date": "1150807",
                "actions": [
                    {
                        "kind": "entry_log",
                        "time": "08:00",
                        "actor": "10",
                        "target": "10",
                        "source": "\u503c\u73ed\u4ea4\u63a5",
                    },
                    {
                        "kind": "entry_log",
                        "time": "08:00",
                        "actor": "10",
                        "target": "11",
                        "source": "\u503c\u73ed\u4ea4\u63a5",
                    },
                    {
                        "kind": "work_log",
                        "time": "08:00",
                        "actor": "10",
                        "target": "10",
                        "source": "\u503c\u73ed\u4ea4\u63a5",
                    },
                ],
            }
        )

        controller.toggleTaskSelection(1)
        self.assertEqual(controller.selectedTaskCount, 3)
        controller.toggleTaskSelection(0)
        self.assertEqual(controller.selectedTaskCount, 0)

    def test_manual_submission_includes_adjust_group_task(self) -> None:
        from PySide6.QtTest import QSignalSpy

        from qt_app.controllers.duty_controller import DutyController

        controller = DutyController()
        controller.set_actor_no("10")
        controller.replace_schedule_data(
            {
                "target_date": "1150729",
                "actions": [
                    {
                        "kind": "entry_log",
                        "time": "08:00",
                        "actor": "10",
                        "target": "10",
                        "source": "值班交接",
                        "fields": {"出或入": "值退", "領用事由及地點": "退勤"},
                    }
                ],
            },
            comparisons={0: {"compare": "可能臨時調整", "group": "adjust", "matched": []}},
        )
        confirmation_spy = QSignalSpy(controller.manualSubmissionConfirmationRequested)
        controller.toggleTaskSelection(0)

        controller.prepareManualSubmission()

        self.assertEqual(confirmation_spy.count(), 1)
        self.assertEqual(controller._pending_manual_indices, [0])

    def test_completed_selected_task_disables_manual_submission(self) -> None:
        from PySide6.QtTest import QSignalSpy

        from qt_app.controllers.duty_controller import DutyController

        controller = DutyController()
        controller.set_actor_no("10")
        controller.replace_schedule_data(
            {
                "target_date": "1150806",
                "actions": [
                    {
                        "kind": "entry_log",
                        "time": "08:00",
                        "actor": "10",
                        "target": "10",
                        "source": "值班交接",
                        "fields": {},
                    }
                ],
            },
            comparisons={0: {"compare": "已登打", "group": "done", "matched": []}},
        )
        confirmation_spy = QSignalSpy(controller.manualSubmissionConfirmationRequested)
        controller.toggleTaskSelection(0)

        self.assertFalse(controller.canManualSubmitSelected)
        controller.prepareManualSubmission()

        self.assertEqual(confirmation_spy.count(), 0)
        self.assertEqual(controller._pending_manual_indices, [])

    def test_manual_submission_excludes_near_group_task(self) -> None:
        from PySide6.QtTest import QSignalSpy

        from qt_app.controllers.duty_controller import DutyController

        controller = DutyController()
        controller.set_actor_no("10")
        controller.replace_schedule_data(
            {
                "target_date": "1150729",
                "actions": [
                    {
                        "kind": "entry_log",
                        "time": "08:00",
                        "actor": "10",
                        "target": "10",
                        "source": "值班交接",
                        "fields": {"出或入": "值退", "領用事由及地點": "退勤"},
                    }
                ],
            },
            comparisons={0: {"compare": "時間近似", "group": "near", "matched": ["existing"]}},
        )
        confirmation_spy = QSignalSpy(controller.manualSubmissionConfirmationRequested)
        controller.toggleTaskSelection(0)

        controller.prepareManualSubmission()

        self.assertEqual(confirmation_spy.count(), 0)
        self.assertEqual(controller._pending_manual_indices, [])

    def test_login_failure_requests_relogin_instead_of_retry(self) -> None:
        from PySide6.QtTest import QSignalSpy

        from qt_app.controllers.duty_controller import DutyController

        controller = DutyController()
        controller.set_actor_no("10")
        controller.replace_schedule_data(
            {
                "target_date": "1150729",
                "actions": [
                    {"kind": "work_log", "time": "00:00", "actor": "10", "source": "在隊訓練"}
                ],
            }
        )
        controller.mark_submission_enqueued(0)
        relogin_spy = QSignalSpy(controller.reloginRequired)

        controller.handle_submission_failure(0, "登入失敗", "login_failed")

        self.assertEqual(relogin_spy.count(), 1)
        self.assertEqual(controller.dueTaskCount, 1)
        self.assertEqual(controller._retry_after, {})

    def test_live_capture_login_failure_requests_relogin_without_loading_cached_schedule(self) -> None:
        from PySide6.QtTest import QSignalSpy

        from app_core.schedule_repository import business_roc_date
        from qt_app.controllers.duty_controller import DutyController

        class Repository:
            def __init__(self) -> None:
                self.loads = 0

            def load_current(self):
                self.loads += 1
                raise AssertionError("登入失效時不應改載離線排程")

        repository = Repository()
        controller = DutyController(repository=repository)
        controller.set_actor_no("10")
        controller._active_capture_request = 1
        controller._capture_targets[1] = business_roc_date()
        relogin_spy = QSignalSpy(controller.reloginRequired)
        failed_spy = QSignalSpy(controller.liveCaptureFailed)

        controller._capture_failed(1, "10", "登入失敗：請重新登入。", "login_failed")

        self.assertEqual(relogin_spy.count(), 1)
        self.assertEqual(failed_spy.at(0), ["登入失敗：請重新登入。", "login_failed"])
        self.assertEqual(repository.loads, 0)

    def test_work_log_settings_controller_loads_edits_and_saves(self) -> None:
        from PySide6.QtTest import QSignalSpy

        from app_core.work_log_settings_service import NUMERIC_FIELDS, WorkLogSettings
        from qt_app.controllers.work_log_settings_controller import WorkLogSettingsController

        class FakeService:
            def __init__(self):
                self.saved = []

            def load(self):
                return WorkLogSettings({key: 1 for key in NUMERIC_FIELDS}, "原記事", {})

            def defaults(self):
                return WorkLogSettings({key: 0 for key in NUMERIC_FIELDS}, "預設記事", {})

            def case_items(self, schedule_data, settings):
                return []

            def preview(self, values, note, *, vehicle_out_count=0):
                return f"預覽 {values['radio_count']} {note}"

            def save(self, values, note, case_vehicle_counts=None):
                self.saved.append((dict(values), note))
                return WorkLogSettings(dict(values), note, {})

        service = FakeService()
        controller = WorkLogSettingsController(service)
        saved_spy = QSignalSpy(controller.settingsSaved)

        controller.load()
        controller.setValue("radio_count", "35")
        controller.setImportantNote("新記事")
        self.assertTrue(controller.save())

        self.assertEqual(controller.values["radio_count"], 35)
        self.assertEqual(controller.importantNote, "新記事")
        self.assertIn("預覽 35", controller.previewText)
        self.assertEqual(saved_spy.count(), 1)
        self.assertEqual(service.saved[0][0]["radio_count"], 35)

    def test_manual_submission_requires_confirmation_and_uses_confirmation_time(self) -> None:
        from datetime import datetime

        from PySide6.QtTest import QSignalSpy

        from qt_app.controllers.duty_controller import DutyController

        controller = DutyController()
        controller.set_actor_no("10")
        controller.replace_schedule_data(
            {
                "target_date": "1150729",
                "actions": [
                    {
                        "kind": "work_log",
                        "time": "12:00",
                        "actor": "10",
                        "target": "10",
                        "fields": {"工作時間": "12:00", "勤務項目": "巡邏"},
                    }
                ],
            }
        )
        confirmation_spy = QSignalSpy(controller.manualSubmissionConfirmationRequested)
        requested_spy = QSignalSpy(controller.manualSubmissionRequested)
        controller.toggleTaskSelection(0)

        controller.prepareManualSubmission()

        self.assertEqual(confirmation_spy.count(), 1)
        self.assertIn("使用按下確認時的當下時間", controller.manualConfirmationSummary)
        controller.confirmManualSubmission()
        self.assertEqual(requested_spy.count(), 1)
        requests = controller.manual_submission_requests(
            "user10",
            "secret",
            [0],
            submit_at=datetime(2026, 7, 30, 1, 2),
        )
        action = requests[0].schedule_data["actions"][0]
        self.assertEqual(action["time"], "01:02")
        self.assertEqual(action["fields"]["工作時間"], "01:02")
        self.assertEqual(action["submit_target_date"], "1150730")

    def test_qml_logged_out_account_manager_opens_as_legacy_window(self) -> None:
        from PySide6.QtCore import QObject, QPointF, Qt
        from PySide6.QtTest import QTest

        from qt_app.controllers.app_controller import AppController
        from qt_app.main import create_engine

        class MemoryCredentialRepository:
            def __init__(self) -> None:
                self.accounts = [
                    {
                        "actor_no": "10",
                        "user_id": "saved-user",
                        "password": "stored-value",
                        "display_name": "10番 測試員",
                    },
                    *[
                        {
                            "actor_no": str(100 + index),
                            "user_id": f"saved-user-{index}",
                            "password": "stored-value",
                            "display_name": f"{100 + index}番 測試員",
                        }
                        for index in range(15)
                    ],
                ]

            @staticmethod
            def account_identity(account):
                return account.get("user_id") or account.get("actor_no") or ""

            def load(self):
                return SimpleNamespace(
                    accounts=list(self.accounts),
                    last_selected="saved-user",
                    can_persist=True,
                    needs_rewrite=False,
                    invalid_file=False,
                )

            def save(self, accounts, _last_selected=""):
                self.accounts = list(accounts)
                return True

            def enable_persistence(self):
                return None

        with tempfile.TemporaryDirectory() as temp_dir:
            controller = AppController(
                repository=MemoryCredentialRepository(),
                credential_sync_service=SimpleNamespace(enabled=False),
            )
            title_bar_requests = []
            controller.nativeTitleBarRequested.connect(title_bar_requests.append)
            engine = create_engine(controller)
            root = engine.rootObjects()[0]
            account_window = None

            try:
                QTest.qWait(50)
                user_field = root.findChild(QObject, "loginUserIdField")
                password_field = root.findChild(QObject, "loginPasswordField")
                self.assertEqual(user_field.property("text"), "saved-user")
                self.assertEqual(password_field.property("text"), "stored-value")
                account_button = root.findChild(QObject, "savedAccountManagerButton")
                self.assertIsNotNone(account_button)
                controller.sessionController.login("", "", False)
                QTest.qWait(50)
                login_status = root.findChild(QObject, "loginStatusLabel")
                self.assertIsNotNone(login_status)
                self.assertEqual(login_status.property("text"), "請輸入帳號、密碼。")
                self.assertEqual(root.title(), "登入頁面")
                self.assertEqual((root.minimumWidth(), root.maximumWidth()), (550, 550))
                self.assertEqual((root.minimumHeight(), root.maximumHeight()), (352, 352))
                self.assertEqual(root.height(), 352)
                minimize_button = root.findChild(QObject, "titleMinimizeButton")
                close_button = root.findChild(QObject, "titleCloseButton")
                title_window_controls = root.findChild(QObject, "titleWindowControls")
                self.assertIsNotNone(minimize_button)
                self.assertIsNotNone(close_button)
                self.assertIsNotNone(title_window_controls)
                self.assertEqual(title_window_controls.width(), 80)
                self.assertEqual(minimize_button.x(), 0)
                self.assertEqual(close_button.x(), 40)
                self.assertEqual(
                    (minimize_button.width(), minimize_button.height()),
                    (close_button.width(), close_button.height()),
                )
                self.assertEqual((minimize_button.width(), minimize_button.height()), (40, 32))
                self.assertEqual(minimize_button.property("font").pixelSize(), 18)
                self.assertEqual(close_button.property("font").pixelSize(), 18)
                self.assertEqual(minimize_button.property("scale"), 1.0)
                self.assertEqual(close_button.property("scale"), 1.0)
                self.assertIsNone(root.findChild(QObject, "titleMaximizeButton"))
                account_window = next(
                    candidate
                    for candidate in self.app.topLevelWindows()
                    if candidate.objectName() == "accountManagerWindow"
                )
                self.assertFalse(account_window.isVisible())
                account_window.setX(-5000)
                account_window.setY(-5000)

                point = account_button.mapToScene(
                    QPointF(account_button.width() / 2, account_button.height() / 2)
                )
                QTest.mouseClick(root, Qt.LeftButton, Qt.NoModifier, point.toPoint())
                for _ in range(20):
                    if account_window.property("accountCount") == 16:
                        break
                    QTest.qWait(25)

                self.assertTrue(account_window.isVisible())
                self.assertEqual(account_window.title(), "SinpoSmart - 帳號管理")
                self.assertEqual(title_bar_requests, [])
                self.assertEqual(account_window.property("accountCount"), 16)
                account_screen = account_window.screen()
                self.assertIsNotNone(account_screen)
                account_available_area = account_screen.availableGeometry()
                self.assertLessEqual(account_window.width(), account_available_area.width() - 32)
                self.assertLessEqual(account_window.height(), account_available_area.height() - 32)
                self.assertGreaterEqual(account_window.x(), account_available_area.x() + 16)
                self.assertGreaterEqual(account_window.y(), account_available_area.y() + 16)
                self.assertLessEqual(
                    account_window.x() + account_window.width(),
                    account_available_area.x() + account_available_area.width() - 16,
                )
                self.assertLessEqual(
                    account_window.y() + account_window.height(),
                    account_available_area.y() + account_available_area.height() - 16,
                )
                self.assertEqual(root.width(), 550)
                self.assertEqual(root.height(), 352)

                def find_visual_item(parent, object_name):
                    for child in parent.childItems():
                        if child.objectName() == object_name:
                            return child
                        match = find_visual_item(child, object_name)
                        if match is not None:
                            return match
                    return None

                account_content = account_window.contentItem()
                delete_button = find_visual_item(account_content, "savedAccountDeleteButton")
                select_button = find_visual_item(account_content, "savedAccountSelectButton")
                close_button = find_visual_item(account_content, "savedAccountCloseButton")
                title_bar = find_visual_item(account_content, "accountTitleBar")
                title_close_button = find_visual_item(account_content, "accountTitleCloseButton")
                saved_account_grid = find_visual_item(account_content, "savedAccountGrid")
                self.assertIsNotNone(delete_button)
                self.assertIsNotNone(select_button)
                self.assertIsNotNone(close_button)
                self.assertIsNotNone(title_bar)
                self.assertIsNotNone(title_close_button)
                self.assertIsNotNone(saved_account_grid)
                self.assertIsNotNone(find_visual_item(account_content, "savedAccountViewport"))
                self.assertEqual(account_window.property("columnCount"), 2)
                self.assertEqual(saved_account_grid.property("columns"), 2)
                self.assertEqual((title_close_button.width(), title_close_button.height()), (40, 32))
                self.assertEqual(title_close_button.property("tone"), "windowClose")
                self.assertEqual(title_close_button.property("iconKind"), "close")
                self.assertEqual(delete_button.property("tone"), "danger")
                self.assertEqual(delete_button.property("fillColor").name().upper(), "#FFFFFF")
                self.assertEqual(delete_button.property("strokeColor").name().upper(), "#FECACA")
                self.assertEqual(delete_button.property("textColor").name().upper(), "#B91C1C")
                self.assertEqual(select_button.property("tone"), "infoStrong")
                self.assertEqual(select_button.property("fillColor").name().upper(), "#DBEAFE")
                self.assertEqual(select_button.property("strokeColor").name().upper(), "#93C5FD")
                self.assertEqual(select_button.property("textColor").name().upper(), "#1D4ED8")
                self.assertEqual(close_button.property("tone"), "neutralStrong")
                self.assertEqual(close_button.property("fillColor").name().upper(), "#E2E8F0")
                self.assertEqual(close_button.property("hoverColor").name().upper(), "#CBD5E1")

                select_point = select_button.mapToScene(
                    QPointF(select_button.width() / 2, select_button.height() / 2)
                )
                QTest.mouseClick(
                    account_window,
                    Qt.LeftButton,
                    Qt.NoModifier,
                    select_point.toPoint(),
                )
                QTest.qWait(50)
                self.assertFalse(account_window.isVisible())
                user_field = root.findChild(QObject, "loginUserIdField")
                password_field = root.findChild(QObject, "loginPasswordField")
                self.assertEqual(user_field.property("text"), "saved-user")
                self.assertEqual(password_field.property("text"), "stored-value")
                self.assertAlmostEqual(user_field.width(), password_field.width(), delta=1)
            finally:
                if account_window is not None:
                    account_window.close()
                root.close()
                controller.shutdown()

    def test_qml_login_runs_once_and_resolves_actor_from_one_schedule_capture(self) -> None:
        from PySide6.QtCore import QObject, QPointF, Qt
        from PySide6.QtTest import QTest

        from app_core.credential_repository import CredentialRepository
        from app_core.login_verifier import LoginResult
        from app_core.schedule_capture_service import ScheduleCaptureRequest
        from app_core.schedule_repository import ScheduleSnapshot, business_roc_date
        from qt_app.controllers.app_controller import AppController
        from qt_app.main import create_engine

        class FakeVerifier:
            def __init__(self) -> None:
                self.calls = []

            def verify(self, **kwargs):
                self.calls.append(dict(kwargs))
                return LoginResult(
                    actor_no="",
                    user_id=kwargs["user_id"],
                    actor_name="測試員",
                    warning="登入成功，正在查詢勤務資料…",
                )

        class FakeCaptureService:
            def __init__(self) -> None:
                self.requests = []

            def current_request(self, user_id, password, actor_no, actor_name=""):
                return ScheduleCaptureRequest(
                    user_id,
                    password,
                    actor_no,
                    business_roc_date(),
                    actor_name,
                )

            def capture(self, request, *, status_callback=None):
                self.requests.append(request)
                if status_callback is not None:
                    status_callback("假勤務查詢中")
                target = business_roc_date()
                return ScheduleSnapshot(
                    Path(f"schedule_output_{target}.json"),
                    {
                        "target_date": target,
                        "today": {
                            "staff": {"10": {"name": "測試員", "role": "隊員"}},
                            "rows": [
                                {"slot": "8-9", "columns": {"值班": ["10"]}},
                                {"slot": "9-10", "columns": {"值班": ["10"]}},
                            ],
                        },
                        "actions": [],
                    },
                    target,
                    authenticated_actor_no="10",
                    authenticated_actor_name="測試員",
                )

        class FakeOperationalSyncService:
            def enqueue_event(self, record_type, **_fields):
                return {"record_type": record_type}

            def sync_board_async(self, _schedule_data):
                return True

        with tempfile.TemporaryDirectory() as temp_dir:
            verifier = FakeVerifier()
            capture_service = FakeCaptureService()
            controller = AppController(
                repository=CredentialRepository(Path(temp_dir) / "saved_login.json", "SinpoSmart", None),
                verifier=verifier,
                credential_sync_service=SimpleNamespace(enabled=False),
                schedule_capture_service=capture_service,
                operational_sync_service=FakeOperationalSyncService(),
            )
            engine = create_engine(controller)
            root = engine.rootObjects()[0]

            try:
                user_field = root.findChild(QObject, "loginUserIdField")
                password_field = root.findChild(QObject, "loginPasswordField")
                login_button = root.findChild(QObject, "loginSubmitButton")
                self.assertIsNotNone(user_field)
                self.assertIsNotNone(password_field)
                self.assertIsNotNone(login_button)

                user_field.setProperty("text", "fake-user")
                password_field.setProperty("text", "fake-password")
                point = login_button.mapToScene(
                    QPointF(login_button.width() / 2, login_button.height() / 2)
                )
                QTest.mouseClick(root, Qt.LeftButton, Qt.NoModifier, point.toPoint())

                for _ in range(120):
                    if (
                        controller.sessionController.actorNo == "10"
                        and not controller.sessionController._login_workers
                        and not controller.dutyController._schedule_workers
                        and not controller.dutyController._capture_workers
                    ):
                        break
                    QTest.qWait(25)

                self.assertEqual(len(verifier.calls), 1)
                self.assertEqual(verifier.calls[0]["typed_actor_no"], "")
                self.assertEqual(verifier.calls[0]["user_id"], "fake-user")
                self.assertEqual(verifier.calls[0]["password"], "fake-password")
                self.assertEqual(len(capture_service.requests), 1)
                self.assertEqual(capture_service.requests[0].actor_no, "")
                self.assertEqual(capture_service.requests[0].actor_name, "測試員")
                self.assertEqual(controller.sessionController.actorNo, "10")
                self.assertEqual(controller.sessionController.displayName, "10番 測試員")
                self.assertEqual(
                    controller.sessionController.loginStatus,
                    "已登入：隊員 測試員，今日值班時段：08 - 10。",
                )
                self.assertEqual(password_field.property("text"), "")
                self.assertFalse(controller.sessionController.isBusy)
                self.assertFalse(controller.dutyController.isRefreshing)
                self.assertFalse(controller.sessionController._login_workers)
                self.assertFalse(controller.dutyController._capture_workers)
                self.assertEqual((root.width(), root.height()), (550, 872))
            finally:
                root.close()
                controller.shutdown()

    def test_qml_shell_loads_with_app_controller(self) -> None:
        from dataclasses import replace

        from PySide6.QtCore import QMetaObject, QObject, QPointF, Qt
        from PySide6.QtTest import QTest

        from app_core.credential_repository import CredentialRepository
        from app_core.daily_vehicle_service import DailyVehicleDefaults
        from app_core.duty_sheet_service import DutySheetDefaults
        from app_core.rest_monthly_service import RestMonthlyDefaults
        from app_core.rescue_video_service import (
            RescueVideoCheckCard,
            RescueVideoDefaults,
            RescueVideoRunResult,
        )
        from app_core.session import LoginSession
        from app_core.work_log_settings_service import NUMERIC_FIELDS, WorkLogSettings
        from qt_app.controllers.app_controller import AppController
        from qt_app.controllers.tool_controller import ToolController
        from qt_app.main import create_engine

        class FakeOperationalSyncService:
            def __init__(self):
                self.events = []

            def enqueue_event(self, record_type, **_fields):
                self.events.append((record_type, _fields))
                return {"record_type": record_type}

            def sync_board_async(self, _schedule_data):
                return True

        class FakeWorkLogSettingsService:
            case_key = "1150729|09:30|緊急救護"

            def __init__(self):
                self.saved = []
                self.settings = WorkLogSettings(
                    {key: 1 for key in NUMERIC_FIELDS},
                    "驗收記事",
                    {},
                )

            def load(self):
                return self.settings

            def defaults(self):
                return self.settings

            def case_items(self, schedule_data, _settings):
                if not schedule_data.get("target_date"):
                    return []
                count = self.saved[-1][self.case_key] if self.saved else 2
                return [
                    {
                        "key": self.case_key,
                        "date": "1150729",
                        "report_time": "09:30",
                        "category": "緊急救護",
                        "count": count,
                    }
                ]

            def preview(self, _values, note, *, vehicle_out_count=0):
                return f"驗收預覽 {vehicle_out_count} 台 {note}"

            def save(self, values, note, case_vehicle_counts=None):
                self.saved.append(dict(case_vehicle_counts or {}))
                self.settings = WorkLogSettings(dict(values), note, {})
                return self.settings

        class FakeDutySheetService:
            def __init__(self):
                self.requests = []

            def load_defaults(self):
                return DutySheetDefaults(
                    "X:/last-duty.xlsm",
                    "2026/07/30",
                    "91",
                    "11",
                    "92",
                    "93",
                    ("91",),
                    ("11",),
                    ("92", "93"),
                    False,
                )

            def validate(self, request):
                return request

            def confirmation_summary(self, request):
                return f"確認勤務表 {request.target_date}"

            def execute(self, request, *, status_callback=None):
                self.requests.append(request)
                if status_callback is not None:
                    status_callback("勤務表測試執行中")
                return "勤務表完成"

        class FakeRestMonthlyService:
            def __init__(self):
                self.rest_requests = []
                self.monthly_requests = []

            def load_rest_defaults(self):
                return RestMonthlyDefaults(115, ("06", "07", "08"), "07", "")

            def load_monthly_defaults(self):
                return RestMonthlyDefaults(115, ("06", "07", "08"), "07", "")

            def validate_rest(self, request):
                return request

            def validate_monthly(self, request):
                return request

            def confirmation_summary(self, request):
                return f"確認月份 {request.month}"

            def execute_rest(self, request, *, status_callback=None):
                self.rest_requests.append(request)
                if status_callback is not None:
                    status_callback("休息時間測試執行中")
                return "休息時間完成"

            def execute_monthly(self, request, *, status_callback=None):
                self.monthly_requests.append(request)
                if status_callback is not None:
                    status_callback("勤務基準表測試執行中")
                return "勤務基準表完成"

        class FakeDailyVehicleService:
            def __init__(self):
                self.requests = []

            def load_defaults(self):
                return DailyVehicleDefaults("2026/07/29", ("車輛保養檢查", "車輛器材清點"))

            def validate(self, request):
                return request

            def confirmation_summary(self, _request):
                return "將開啟瀏覽器執行車輛保養清點，是否繼續？"

            def execute(self, request, *, status_callback=None):
                self.requests.append(request)
                if status_callback is not None:
                    status_callback("車輛保養測試執行中")
                return "車輛保養完成"

        class FakeRescueVideoService:
            def __init__(self):
                self.requests = []

            def load_defaults(self, *_args, **_kwargs):
                defaults = RescueVideoDefaults(
                    "X:/DCIM",
                    "Z:/救護行車影片",
                    "2026-07-29",
                    ("92", "93"),
                    "92",
                    "6",
                    False,
                    "來源可用\n案件目的地可存取\n工作／返隊紀錄可存取\n自動採用記憶卡偏移：6 分鐘",
                    True,
                    "自動檢查通過",
                )
                return replace(
                    defaults,
                    check_cards=(
                        RescueVideoCheckCard("source", "記憶卡來源", "來源可用", "ok"),
                        RescueVideoCheckCard("destination", "案件目的地", "目的地可用", "ok"),
                        RescueVideoCheckCard("work_log", "工作／返隊紀錄", "工作紀錄可用", "ok"),
                        RescueVideoCheckCard("vehicle_date", "車號與日期", "車號可用", "ok"),
                        RescueVideoCheckCard("report", "報告輸出", "報告可用", "ok"),
                        RescueVideoCheckCard("videos", "影片檢查", "影片可用", "ok"),
                    ),
                )

            def validate(self, request):
                return request, [], {}

            def confirmation_summary(self, request):
                if request.mode == "delete":
                    return "只有複製並完成內容驗證的 .TS 檔案會刪除。\n確定要繼續嗎？"
                return f"確認救護影片 {request.mode}"

            def execute(self, request, *, status_callback=None):
                self.requests.append(request)
                if status_callback is not None:
                    status_callback("救護影片測試執行中")
                return RescueVideoRunResult(
                    summary_text="預覽完成",
                    warning_text="",
                    report_path="report.csv",
                    rows=(
                        {
                            "sourceText": "clip.mp4",
                            "timeText": "09:30",
                            "caseText": "救護案件",
                            "statusText": "已配對",
                            "destinationText": "case/clip.mp4",
                            "noteText": "",
                            "tone": "ok",
                        },
                    ),
                )

        with tempfile.TemporaryDirectory() as temp_dir:
            repository = CredentialRepository(Path(temp_dir) / "saved_login.json", "SinpoSmart", None)
            settings_service = FakeWorkLogSettingsService()
            operational_sync_service = FakeOperationalSyncService()
            duty_sheet_service = FakeDutySheetService()
            rest_monthly_service = FakeRestMonthlyService()
            daily_vehicle_service = FakeDailyVehicleService()
            rescue_video_service = FakeRescueVideoService()
            controller = AppController(
                repository=repository,
                operational_sync_service=operational_sync_service,
                tool_controller=ToolController(Path(temp_dir)),
                work_log_settings_service=settings_service,
                duty_sheet_service=duty_sheet_service,
                rest_monthly_service=rest_monthly_service,
                daily_vehicle_service=daily_vehicle_service,
                rescue_video_service=rescue_video_service,
            )
            controller.workLogSettingsController.settingsSaved.disconnect(
                controller._refresh_after_settings_save
            )
            attempt_id = controller._session_state.begin_login()
            controller._session_state.complete_login(
                attempt_id,
                LoginSession(
                    "10",
                    "acceptance-user",
                    "fake-password",
                    verified=True,
                    actor_name="驗收人員",
                ),
            )
            controller.dutyController.set_actor_no("10")
            schedule_data = {
                "target_date": "1150729",
                "today": {"staff": {"10": {"name": "驗收人員"}}},
                "actions": [
                    {
                        "kind": "work_log",
                        "time": "09:00",
                        "actor": "10",
                        "target": "10",
                        "fields": {"工作時間": "09:00", "勤務項目": "巡邏驗收"},
                    }
                ],
            }
            audit_comparisons = {
                0: {
                    "compare": "人工確認：驗收差異",
                    "group": "review",
                    "matched": [{"勤務項目": "既有巡邏紀錄", "時間": "09:02"}],
                }
            }
            controller.dutyController.replace_schedule_data(schedule_data)
            controller.sessionController._login_status = (
                "已登入：分隊長 驗收人員，今日值班時段：08:00-10:00。"
            )
            controller.sessionController._display_name = "10番 驗收人員"
            controller.workLogSettingsController.set_schedule_data(schedule_data)
            controller.workLogSettingsController.load()
            engine = create_engine(controller)
            root = engine.rootObjects()[0]
            QTest.qWait(100)
            self.assertEqual((root.width(), root.height()), (550, 872))

            def find_visual(name):
                stack = [root.contentItem()]
                while stack:
                    item = stack.pop()
                    if item.objectName() == name:
                        return item
                    stack.extend(item.childItems())
                return None

            def wait_for(name):
                for _ in range(80):
                    item = find_visual(name) or root.findChild(QObject, name)
                    if item is not None:
                        return item
                    QTest.qWait(25)
                self.fail(f"找不到 QML 視覺項目：{name}")

            def wait_for_visible(name):
                for _ in range(80):
                    stack = [root.contentItem()]
                    while stack:
                        item = stack.pop()
                        if item.objectName() == name and item.isVisible():
                            return item
                        stack.extend(item.childItems())
                    QTest.qWait(25)
                self.fail(f"找不到可見的 QML 視覺項目：{name}")

            def click(item, target_window=None):
                point = item.mapToScene(QPointF(item.width() / 2, item.height() / 2))
                QTest.mouseClick(target_window or root, Qt.LeftButton, Qt.NoModifier, point.toPoint())
                QTest.qWait(100)

            def wait_until(predicate, failure_message):
                for _ in range(120):
                    if predicate():
                        return
                    QTest.qWait(25)
                self.fail(failure_message)

            try:
                self.assertTrue(engine.rootObjects())
                settings_button = wait_for("settingsTab")
                self.assertEqual((settings_button.width(), settings_button.height()), (34, 34))
                for button_name, tone in (
                    ("quickDutySheetToolButton", "info"),
                    ("quickDailyVehicleToolButton", "info"),
                    ("quickRescueVideoToolButton", "warning"),
                    ("quickRestTimeToolButton", "monthly"),
                    ("quickMonthlyBaseToolButton", "monthly"),
                ):
                    self.assertEqual(wait_for(button_name).property("tone"), tone)
                controller.sessionController.errorOccurred.emit("驗收錯誤訊息")
                QTest.qWait(50)
                self.assertEqual(
                    wait_for("loggedInStatusLabel").property("text"),
                    "已登入：分隊長 驗收人員，今日值班時段：08:00-10:00。",
                )
                self.assertEqual(root.property("errorMessage"), "驗收錯誤訊息")
                previous_login_status = wait_for("loggedInStatusLabel").property("text")
                previous_login_tone = controller.sessionController.loginStatusTone
                controller.dutyController.errorOccurred.emit("勤務資料讀取失敗")
                QTest.qWait(50)
                self.assertEqual(
                    wait_for("loggedInStatusLabel").property("text"),
                    previous_login_status,
                )
                self.assertEqual(
                    controller.sessionController.loginStatusTone,
                    previous_login_tone,
                )
                controller.sessionController.sessionChanged.disconnect(controller._sync_session_actor)
                try:
                    root.setPosition(321, 123)
                    controller.sessionController.set_logged_in_status(
                        "已登入：分隊長 驗收人員，今日值班時段：08:00-10:00。"
                    )
                    QTest.qWait(50)
                    self.assertEqual((root.x(), root.y()), (321, 123))
                finally:
                    controller.sessionController.sessionChanged.connect(controller._sync_session_actor)

                duty_task_area = wait_for("dutyTaskArea")
                duty_task_header = wait_for("dutyTaskHeader")
                audit_task_header = wait_for("auditTaskHeader")
                mode_menu_button = wait_for("modeMenuButton")
                manual_submit_button = wait_for("manualSubmitButton")
                selected_task_actions = wait_for("selectedTaskActions")
                duty_task_row = wait_for_visible("dutyTaskRow")
                duty_status_pill = wait_for_visible("dutyTaskStatusPill")
                duty_task_time_header = wait_for("dutyTaskTimeHeaderCell")
                duty_task_system_header = wait_for("dutyTaskSystemHeaderCell")
                duty_task_detail_header = wait_for("dutyTaskDetailHeaderCell")
                duty_task_people_header = wait_for("dutyTaskPeopleHeaderCell")
                duty_task_status_header = wait_for("dutyTaskStatusHeaderCell")
                duty_task_time_cell = wait_for_visible("dutyTaskTimeCell")
                duty_task_system_cell = wait_for_visible("dutyTaskSystemCell")
                duty_task_detail_cell = wait_for_visible("dutyTaskDetailCell")
                duty_task_people_cell = wait_for_visible("dutyTaskPeopleCell")
                self.assertTrue(duty_task_header.property("visible"))
                self.assertFalse(audit_task_header.property("visible"))
                for header_cell, task_cell in (
                    (duty_task_time_header, duty_task_time_cell),
                    (duty_task_system_header, duty_task_system_cell),
                    (duty_task_detail_header, duty_task_detail_cell),
                    (duty_task_people_header, duty_task_people_cell),
                    (duty_task_status_header, duty_status_pill),
                ):
                    self.assertEqual(
                        header_cell.mapToScene(QPointF(0, 0)).x(),
                        task_cell.mapToScene(QPointF(0, 0)).x(),
                    )
                    self.assertEqual(header_cell.width(), task_cell.width())
                self.assertEqual(duty_task_time_cell.width(), 82)
                self.assertEqual(
                    duty_task_row.mapToScene(QPointF(duty_task_row.width(), 0)).x()
                    - duty_status_pill.mapToScene(QPointF(duty_status_pill.width(), 0)).x(),
                    4,
                )
                self.assertEqual((mode_menu_button.width(), mode_menu_button.height()), (48, 28))
                system_menu_button = wait_for_visible("systemMenuButton")
                self.assertTrue(system_menu_button.isVisible())
                self.assertEqual((system_menu_button.width(), system_menu_button.height()), (48, 28))
                window_menu_button = wait_for_visible("windowMenuButton")
                self.assertEqual((window_menu_button.width(), window_menu_button.height()), (48, 28))
                click(system_menu_button)
                system_menu = root.findChild(QObject, "systemCommandMenu")
                self.assertIsNotNone(system_menu)
                self.assertTrue(system_menu.property("visible"))
                self.assertTrue(wait_for_visible("checkForUpdatesMenuItem").property("enabled"))
                self.assertTrue(wait_for_visible("exportIssuePackageMenuItem").property("enabled"))
                mode_button_point = mode_menu_button.mapToScene(
                    QPointF(mode_menu_button.width() / 2, mode_menu_button.height() / 2)
                )
                QTest.mouseMove(root, mode_button_point.toPoint())
                QTest.qWait(100)
                mode_menu = root.findChild(QObject, "modeCommandMenu")
                self.assertTrue(mode_menu.property("visible"))
                self.assertFalse(system_menu.property("visible"))
                self.assertTrue(QMetaObject.invokeMethod(mode_menu, "close", Qt.DirectConnection))
                click(window_menu_button)
                window_menu = root.findChild(QObject, "windowCommandMenu")
                self.assertIsNotNone(window_menu)
                self.assertTrue(window_menu.property("visible"))
                self.assertTrue(wait_for_visible("hideToBackgroundMenuItem").property("enabled"))
                self.assertTrue(wait_for_visible("logoutMenuItem").property("enabled"))
                self.assertTrue(wait_for_visible("quitApplicationMenuItem").property("enabled"))
                title_bar = wait_for("appTitleBar")
                title_bar_point = title_bar.mapToScene(
                    QPointF(title_bar.width() * 0.7, title_bar.height() / 2)
                )
                QTest.mouseClick(root, Qt.LeftButton, Qt.NoModifier, title_bar_point.toPoint())
                QTest.qWait(50)
                self.assertFalse(window_menu.property("visible"))
                click(window_menu_button)
                self.assertTrue(window_menu.property("visible"))
                click(window_menu_button)
                self.assertFalse(window_menu.property("visible"))
                click(window_menu_button)
                self.assertTrue(window_menu.property("visible"))
                close_button = wait_for_visible("titleCloseButton")
                close_button_point = close_button.mapToScene(
                    QPointF(close_button.width() / 2, close_button.height() / 2)
                )
                QTest.mouseMove(root, close_button_point.toPoint())
                QTest.qWait(50)
                self.assertTrue(close_button.property("hovered"))
                self.assertEqual(
                    close_button.property("visualFillColor").name().upper(),
                    "#D92D20",
                )
                duty_header_point = duty_task_header.mapToScene(
                    QPointF(duty_task_header.width() / 2, duty_task_header.height() / 2)
                )
                QTest.mouseClick(root, Qt.LeftButton, Qt.NoModifier, duty_header_point.toPoint())
                QTest.qWait(50)
                self.assertFalse(window_menu.property("visible"))
                click(window_menu_button)
                self.assertTrue(window_menu.property("visible"))
                system_button_point = system_menu_button.mapToScene(
                    QPointF(system_menu_button.width() / 2, system_menu_button.height() / 2)
                )
                QTest.mouseMove(root, system_button_point.toPoint())
                QTest.qWait(100)
                self.assertTrue(system_menu.property("visible"))
                self.assertFalse(window_menu.property("visible"))
                self.assertTrue(QMetaObject.invokeMethod(system_menu, "close", Qt.DirectConnection))
                title_bar = wait_for("appTitleBar")
                title_bar_point = title_bar.mapToScene(
                    QPointF(title_bar.width() * 0.7, title_bar.height() / 2)
                )
                window_position = (root.x(), root.y())
                QTest.mouseClick(root, Qt.LeftButton, Qt.NoModifier, title_bar_point.toPoint())
                QTest.qWait(50)
                self.assertEqual((root.x(), root.y()), window_position)
                system_button_point = system_menu_button.mapToScene(
                    QPointF(system_menu_button.width() / 2, system_menu_button.height() / 2)
                )
                QTest.mouseMove(root, system_button_point.toPoint())
                QTest.qWait(50)
                self.assertTrue(system_menu_button.property("hovered"))
                self.assertEqual(
                    system_menu_button.property("visualFillColor").name().upper(),
                    "#D5DFED",
                )
                self.assertLess(
                    mode_menu_button.mapToScene(QPointF(0, 0)).y(),
                    duty_task_header.mapToScene(QPointF(0, 0)).y(),
                )
                for task_button in (manual_submit_button,):
                    self.assertEqual((task_button.width(), task_button.height()), (104, 38))
                self.assertEqual(duty_task_row.height(), 44)
                self.assertEqual(duty_task_row.property("color").name().upper(), "#FFFFFF")
                self.assertEqual((duty_status_pill.width(), duty_status_pill.height()), (96, 28))
                self.assertFalse(selected_task_actions.property("visible"))
                self.assertEqual(manual_submit_button.property("strokeWidth"), 1)
                self.assertEqual(manual_submit_button.property("strokeColor").name().upper(), "#1D4ED8")
                expected_status_color = {
                    "running": "#2563EB",
                    "triggered": "#D1FAE5",
                    "manual": "#FEF3C7",
                }.get(duty_status_pill.property("tone"), "#D1D5DB")
                self.assertEqual(duty_status_pill.property("color").name().upper(), expected_status_color)
                click(duty_task_row)
                self.assertEqual(controller.dutyController.selectedTaskCount, 1)
                self.assertTrue(selected_task_actions.property("visible"))
                self.assertGreaterEqual(
                    manual_submit_button.mapToScene(QPointF(0, 0)).x(),
                    duty_task_area.mapToScene(QPointF(0, 0)).x(),
                )
                self.assertGreaterEqual(
                    manual_submit_button.mapToScene(QPointF(0, 0)).y(),
                    duty_task_area.mapToScene(QPointF(0, 0)).y(),
                )
                self.assertLessEqual(
                    manual_submit_button.mapToScene(
                        QPointF(manual_submit_button.width(), manual_submit_button.height())
                    ).y(),
                    duty_task_area.mapToScene(QPointF(0, duty_task_area.height())).y(),
                )
                click(wait_for("manualSubmitButton"))
                manual_dialog = root.findChild(QObject, "manualSubmissionConfirmation")
                self.assertIsNotNone(manual_dialog)
                self.assertTrue(manual_dialog.property("visible"))
                accept_button = wait_for_visible("appleDialogYesButton")
                reject_button = wait_for_visible("appleDialogNoButton")
                self.assertEqual(accept_button.property("tone"), "primary")
                self.assertEqual(accept_button.property("fillColor").name().upper(), "#2563EB")
                self.assertEqual(accept_button.property("textColor").name().upper(), "#FFFFFF")
                self.assertEqual(reject_button.property("tone"), "neutralStrong")
                self.assertEqual(reject_button.property("fillColor").name().upper(), "#E2E8F0")
                self.assertTrue(QMetaObject.invokeMethod(manual_dialog, "reject", Qt.DirectConnection))
                QTest.qWait(50)
                self.assertEqual(controller.dutyController._pending_manual_indices, [])
                self.assertEqual(controller.dutyController.selectedTaskCount, 1)

                controller.dutyController.replace_schedule_data(
                    schedule_data,
                    comparisons=audit_comparisons,
                )
                QTest.qWait(50)
                click(wait_for("settingsTab"))
                work_log_preview = wait_for("workLogPreviewText")
                self.assertTrue(work_log_preview.isVisible())
                self.assertTrue(str(work_log_preview.property("text")).strip())
                case_field = wait_for("caseVehicleCountField")
                case_field.setProperty("text", "3")
                self.assertTrue(
                    QMetaObject.invokeMethod(case_field, "editingFinished", Qt.DirectConnection)
                )
                click(wait_for("workLogSettingsDiscardButton"))
                work_log_panel = root.findChild(QObject, "workLogSettingsDialog")
                self.assertTrue(work_log_panel.property("visible"))
                QTest.qWait(180)
                self.assertFalse(work_log_panel.property("visible"))
                click(wait_for("settingsTab"))
                case_field = wait_for("caseVehicleCountField")
                self.assertEqual(case_field.property("text"), "2")

                case_field.setProperty("text", "3")
                self.assertTrue(
                    QMetaObject.invokeMethod(case_field, "editingFinished", Qt.DirectConnection)
                )
                click(wait_for("workLogSettingsSaveButton"))
                self.assertEqual(settings_service.saved[-1][settings_service.case_key], 3)

                root.setWidth(root.minimumWidth())
                QTest.qWait(100)
                click(mode_menu_button)
                mode_menu = root.findChild(QObject, "modeCommandMenu")
                self.assertIsNotNone(mode_menu)
                self.assertTrue(mode_menu.property("visible"))
                self.assertFalse(wait_for_visible("dutyModeTab").property("enabled"))
                self.assertTrue(wait_for_visible("auditModeTab").property("enabled"))
                click(wait_for_visible("auditModeTab"))
                self.assertEqual((root.width(), root.height()), (780, 682))
                self.assertEqual(root.title(), "審核模式")
                self.assertFalse(duty_task_header.property("visible"))
                self.assertTrue(audit_task_header.property("visible"))
                maximize_button = wait_for_visible("titleMaximizeButton")
                title_window_controls = wait_for("titleWindowControls")
                self.assertEqual(
                    (maximize_button.width(), maximize_button.height()),
                    (40, 32),
                )
                self.assertEqual(title_window_controls.width(), 120)
                self.assertEqual(wait_for("titleMaximizeButtonLoader").x(), 40)
                self.assertEqual(wait_for("titleCloseButton").x(), 80)
                self.assertEqual(maximize_button.property("iconKind"), "maximize")
                self.assertFalse(maximize_button.property("iconToggled"))
                self.assertEqual(maximize_button.property("text"), "")
                close_button = wait_for_visible("titleCloseButton")
                close_button_point = close_button.mapToScene(
                    QPointF(close_button.width() / 2, close_button.height() / 2)
                )
                QTest.mouseMove(root, close_button_point.toPoint())
                QTest.qWait(50)
                self.assertTrue(close_button.property("hovered"))
                self.assertEqual(close_button.property("visualFillColor").name().upper(), "#D92D20")
                system_menu_button = wait_for_visible("systemMenuButton")
                click(system_menu_button)
                system_menu = root.findChild(QObject, "systemCommandMenu")
                self.assertIsNotNone(system_menu)
                self.assertTrue(system_menu.property("visible"))
                self.assertTrue(wait_for_visible("checkForUpdatesMenuItem").isVisible())
                self.assertTrue(wait_for_visible("exportIssuePackageMenuItem").isVisible())
                self.assertTrue(QMetaObject.invokeMethod(system_menu, "close", Qt.DirectConnection))
                audit_todo_card = wait_for("auditTodoSummaryCard")
                audit_date_card = wait_for("auditDateCard")
                audit_filter_card = wait_for("auditFilterCard")
                self.assertTrue(audit_todo_card.property("visible"))
                self.assertEqual(audit_todo_card.property("tone"), "todo")
                self.assertEqual(wait_for("auditReviewSummaryCard").property("tone"), "review")
                self.assertEqual(wait_for("auditReadySummaryCard").property("tone"), "ready")
                self.assertEqual(wait_for("auditDoneSummaryCard").property("tone"), "done")
                self.assertEqual(audit_date_card.y(), audit_filter_card.y())
                self.assertGreater(
                    audit_todo_card.mapToScene(QPointF(0, 0)).y(),
                    audit_date_card.mapToScene(QPointF(0, 0)).y(),
                )
                self.assertGreater(audit_task_header.width(), 0)
                audit_refresh = wait_for("auditRefreshButton")
                audit_refresh_right = audit_refresh.mapToScene(
                    QPointF(audit_refresh.width(), audit_refresh.height() / 2)
                ).x()
                self.assertLessEqual(audit_refresh_right, root.width())
                self.assertGreaterEqual(audit_refresh.width(), 88)
                audit_date_calendar = root.findChild(QObject, "auditDateCalendarButton")
                self.assertIsNotNone(audit_date_calendar)
                self.assertTrue(
                    QMetaObject.invokeMethod(audit_date_calendar, "openForCurrentDate", Qt.DirectConnection)
                )
                wait_until(
                    lambda: audit_date_calendar.property("popupVisible"),
                    "審核日期月曆未開啟",
                )
                self.assertTrue(QMetaObject.invokeMethod(audit_date_calendar, "closeCalendar", Qt.DirectConnection))
                click(wait_for("auditTaskRow"))
                detail_text = str(wait_for("auditDetailTextArea").property("text"))
                self.assertIn("人工確認：驗收差異", detail_text)
                self.assertIn("既有巡邏紀錄", detail_text)
                self.assertIn("原始預演資料", detail_text)
                audit_dialog = root.findChild(QObject, "auditDetailDialog")
                self.assertIsNotNone(audit_dialog)
                self.assertTrue(QMetaObject.invokeMethod(audit_dialog, "close", Qt.DirectConnection))

                click(wait_for("modeMenuButton"))
                click(wait_for_visible("dutyModeTab"))
                self.assertEqual((root.width(), root.height()), (550, 872))
                self.assertEqual(root.title(), "值班模式")
                self.assertTrue(duty_task_header.property("visible"))
                self.assertFalse(audit_task_header.property("visible"))
                self.assertTrue(wait_for("dutyQuickToolsPanel").property("visible"))
                for usage_history_name in (
                    "duty_sheetUsageHistory",
                    "daily_vehicleUsageHistory",
                    "rest_timeUsageHistory",
                    "monthly_baseUsageHistory",
                ):
                    self.assertIsNotNone(root.findChild(QObject, usage_history_name))
                self.assertIsNone(root.findChild(QObject, "rescue_videoUsageHistory"))
                duty_window_width = root.width()
                main_content_host = wait_for("mainContentHost")
                duty_main_width = main_content_host.width()
                self.assertEqual(duty_window_width, 550)
                self.assertEqual(
                    (main_content_host.x(), main_content_host.y(), duty_main_width),
                    (14, 34, 522),
                )
                for button_name, panel_name in (
                    ("quickDutySheetToolButton", "dutySheetDialog"),
                    ("quickDailyVehicleToolButton", "dailyVehicleDialog"),
                    ("quickRestTimeToolButton", "restTimeDialog"),
                    ("quickMonthlyBaseToolButton", "monthlyBaseDialog"),
                ):
                    click(wait_for(button_name))
                    panel = root.findChild(QObject, panel_name)
                    self.assertIsNotNone(panel)
                    self.assertTrue(panel.property("visible"))
                    self.assertEqual(root.width(), 964)
                    self.assertEqual(
                        (panel.x(), panel.y(), panel.width(), panel.height()),
                    (550, 34, 400, 824),
                    )
                    self.assertEqual(main_content_host.width(), duty_main_width)
                    self.assertTrue(QMetaObject.invokeMethod(panel, "close", Qt.DirectConnection))
                    QTest.qWait(180)
                    self.assertEqual(root.width(), duty_window_width)
                    self.assertEqual(main_content_host.width(), duty_main_width)

                click(wait_for("quickRescueVideoToolButton"))
                rescue_window = root.findChild(QObject, "rescueVideoDialog")
                self.assertIsNotNone(rescue_window)
                self.assertTrue(rescue_window.property("visible"))
                self.assertEqual(root.width(), duty_window_width)
                self.assertEqual(main_content_host.width(), duty_main_width)
                rescue_screen = rescue_window.screen()
                self.assertIsNotNone(rescue_screen)
                rescue_available_area = rescue_screen.availableGeometry()
                self.assertGreaterEqual(rescue_window.x(), rescue_available_area.x() + 16)
                self.assertGreaterEqual(rescue_window.y(), rescue_available_area.y() + 16)
                rescue_date_calendar = rescue_window.findChild(QObject, "rescueVideoDateCalendarButton")
                self.assertIsNotNone(rescue_date_calendar)
                self.assertTrue(
                    QMetaObject.invokeMethod(rescue_date_calendar, "openForCurrentDate", Qt.DirectConnection)
                )
                wait_until(
                    lambda: rescue_date_calendar.property("popupVisible"),
                    "行車紀錄器日期月曆未開啟",
                )
                self.assertTrue(QMetaObject.invokeMethod(rescue_date_calendar, "closeCalendar", Qt.DirectConnection))
                self.assertTrue(QMetaObject.invokeMethod(rescue_window, "close", Qt.DirectConnection))
                QTest.qWait(100)

                click(wait_for("quickDutySheetToolButton"))
                duty_date_field = wait_for("dutyDateField")
                duty_date_calendar = root.findChild(QObject, "dutyDateCalendarButton")
                duty_workbook_browse = wait_for("dutyWorkbookBrowseButton")
                self.assertIsNotNone(duty_date_calendar)
                self.assertEqual(duty_date_field.property("text"), "2026/07/30")
                self.assertEqual(
                    wait_for("dutyWorkbookField").property("text"),
                    "X:/last-duty.xlsm",
                )
                self.assertEqual(wait_for("dutyAttackCombo").property("currentText"), "91")
                self.assertEqual(wait_for("dutyStopCombo").property("currentText"), "11")
                self.assertEqual(wait_for("dutyAmb1Combo").property("currentText"), "92")
                self.assertEqual(wait_for("dutyAmb2Combo").property("currentText"), "93")
                duty_notification_check = wait_for("dutyNotificationCheck")
                self.assertFalse(duty_notification_check.property("checked"))
                click(duty_notification_check)
                self.assertTrue(duty_notification_check.property("checked"))
                self.assertTrue(controller.dutySheetController.notificationEnabled)
                click(duty_notification_check)
                self.assertFalse(duty_notification_check.property("checked"))
                self.assertFalse(controller.dutySheetController.notificationEnabled)
                self.assertTrue(
                    QMetaObject.invokeMethod(duty_date_calendar, "openForCurrentDate", Qt.DirectConnection)
                )
                wait_until(
                    lambda: duty_date_calendar.property("popupVisible"),
                    "勤務表日期月曆未開啟",
                )
                self.assertTrue(QMetaObject.invokeMethod(duty_date_calendar, "closeCalendar", Qt.DirectConnection))
                previous_date_button = wait_for("dutyPreviousDateButton")
                next_date_button = wait_for("dutyNextDateButton")
                for date_button, text in (
                    (previous_date_button, "<"),
                    (next_date_button, ">"),
                ):
                    self.assertEqual(date_button.property("text"), text)
                    self.assertGreater(date_button.property("contentItem").width(), 0)
                click(next_date_button)
                self.assertEqual(duty_date_field.property("text"), "2026/07/31")
                click(previous_date_button)
                self.assertEqual(duty_date_field.property("text"), "2026/07/30")
                self.assertEqual(wait_for("dutySheetStatusBar").property("text"), "準備就緒。")
                self.assertEqual(wait_for("dutySheetRunButton").property("text"), "啟動登打")
                for control_name in (
                    "dutyWorkbookField",
                    "dutyDateField",
                    "dutyAttackCombo",
                    "dutyStopCombo",
                    "dutyAmb1Combo",
                    "dutyAmb2Combo",
                ):
                    self.assertEqual(
                        wait_for(control_name).height(),
                        duty_workbook_browse.height(),
                    )
                self.assertLess(
                    wait_for("dutySheetStatusBar").y(),
                    wait_for("dutySheetRunButton").y(),
                )
                duty_run_button = wait_for("dutySheetRunButton")
                self.assertLessEqual(
                    duty_run_button.mapToScene(QPointF(0, duty_run_button.height())).y(),
                    root.height() - 14,
                )
                click(wait_for("dutyVehicleAddButton"))
                add_dialog = root.findChild(QObject, "dutyVehicleAddDialog")
                self.assertIsNotNone(add_dialog)
                self.assertTrue(add_dialog.property("visible"))
                self.assertGreaterEqual(
                    wait_for("appleDialogYesButton").width(),
                    92,
                )
                self.assertEqual(
                    wait_for("dutyVehicleAddType").height(),
                    duty_workbook_browse.height(),
                )
                wait_for("dutyVehicleAddType").setProperty("currentIndex", 0)
                self.assertTrue(wait_for("dutyVehicleAddFunction").property("visible"))
                self.assertEqual(
                    wait_for("dutyVehicleAddFunction").height(),
                    duty_workbook_browse.height(),
                )
                self.assertTrue(QMetaObject.invokeMethod(add_dialog, "reject", Qt.DirectConnection))
                click(wait_for("dutyVehicleRemoveButton"))
                remove_dialog = root.findChild(QObject, "dutyVehicleRemoveDialog")
                self.assertIsNotNone(remove_dialog)
                self.assertTrue(remove_dialog.property("visible"))
                self.assertEqual(
                    wait_for("dutyVehicleRemoveType").height(),
                    duty_workbook_browse.height(),
                )
                wait_for("dutyVehicleRemoveType").setProperty("currentIndex", 0)
                self.assertTrue(wait_for("dutyVehicleRemoveFunction").property("visible"))
                self.assertEqual(
                    wait_for("dutyVehicleRemoveFunction").height(),
                    duty_workbook_browse.height(),
                )
                self.assertEqual(
                    wait_for("dutyVehicleRemoveValue").height(),
                    duty_workbook_browse.height(),
                )
                self.assertTrue(QMetaObject.invokeMethod(remove_dialog, "reject", Qt.DirectConnection))

                click(wait_for("dutySheetRunButton"))
                duty_confirmation = root.findChild(QObject, "dutySheetConfirmation")
                self.assertIsNotNone(duty_confirmation)
                self.assertTrue(duty_confirmation.property("visible"))
                self.assertTrue(QMetaObject.invokeMethod(duty_confirmation, "accept", Qt.DirectConnection))
                wait_until(
                    lambda: len(duty_sheet_service.requests) == 1
                    and not controller.dutySheetController._workers,
                    "勤務表 QML 執行鍵未完成 worker 呼叫",
                )
                self.assertFalse(duty_sheet_service.requests[0].notification_enabled)
                wait_until(
                    lambda: not duty_confirmation.property("visible")
                    and root.findChild(QObject, "dutySheetDialog").property("visible")
                    and root.width() == 964,
                    "勤務表完成後側邊面板不應收合",
                )

                click(wait_for("quickRestTimeToolButton"))
                rest_run_button = wait_for("restTimeRunButton")
                rest_status_bar = wait_for("restTimeStatusBar")
                rest_month_combo = wait_for("restMonthCombo")
                rest_workbook_title = wait_for("restWorkbookTitle")
                self.assertEqual(rest_run_button.property("text"), "啟動登打")
                self.assertEqual(rest_status_bar.property("text"), "準備就緒。10番 驗收人員")
                self.assertEqual((rest_month_combo.width(), rest_month_combo.height()), (78, 30))
                self.assertEqual(rest_workbook_title.property("font").pixelSize(), 15)
                self.assertGreater(rest_month_combo.property("contentItem").width(), 0)
                self.assertGreater(wait_for("restWorkbookBrowseButton").property("contentItem").width(), 0)
                self.assertLess(rest_status_bar.y(), rest_run_button.y())
                click(rest_run_button)
                rest_confirmation = root.findChild(QObject, "restMonthlyConfirmation")
                self.assertIsNotNone(rest_confirmation)
                self.assertTrue(rest_confirmation.property("visible"))
                self.assertTrue(QMetaObject.invokeMethod(rest_confirmation, "accept", Qt.DirectConnection))
                wait_until(
                    lambda: len(rest_monthly_service.rest_requests) == 1
                    and not controller.restMonthlyController._workers,
                    "休息時間 QML 執行鍵未完成 worker 呼叫",
                )
                wait_until(
                    lambda: not rest_confirmation.property("visible")
                    and root.findChild(QObject, "restTimeDialog").property("visible")
                    and root.width() == 964,
                    "休息時間完成後側邊面板不應收合",
                )

                click(wait_for("quickMonthlyBaseToolButton"))
                monthly_run_button = wait_for("monthlyBaseRunButton")
                monthly_status_bar = wait_for("monthlyBaseStatusBar")
                monthly_month_combo = wait_for("monthlyMonthCombo")
                monthly_source_label = wait_for("monthlySourceLabel")
                monthly_source_title = wait_for("monthlySourceTitle")
                monthly_source_open_button = wait_for("monthlySourceOpenButton")
                self.assertEqual(monthly_run_button.property("text"), "啟動登打")
                self.assertEqual(monthly_status_bar.property("text"), "準備就緒。10番 驗收人員")
                self.assertEqual(
                    monthly_source_label.property("text"),
                    "Google 試算表 / 輪休基準表",
                )
                self.assertTrue(monthly_source_open_button.isEnabled())
                self.assertGreaterEqual(monthly_source_open_button.width(), 112)
                self.assertEqual((monthly_month_combo.width(), monthly_month_combo.height()), (78, 30))
                self.assertEqual(monthly_source_title.property("font").pixelSize(), 15)
                self.assertGreater(monthly_month_combo.property("contentItem").width(), 0)
                self.assertLess(monthly_status_bar.y(), monthly_run_button.y())
                click(monthly_run_button)
                monthly_confirmation = root.findChild(QObject, "restMonthlyConfirmation")
                self.assertIsNotNone(monthly_confirmation)
                self.assertTrue(monthly_confirmation.property("visible"))
                self.assertTrue(QMetaObject.invokeMethod(monthly_confirmation, "accept", Qt.DirectConnection))
                wait_until(
                    lambda: len(rest_monthly_service.monthly_requests) == 1
                    and not controller.restMonthlyController._workers,
                    "勤務基準表 QML 執行鍵未完成 worker 呼叫",
                )
                wait_until(
                    lambda: not monthly_confirmation.property("visible")
                    and root.findChild(QObject, "monthlyBaseDialog").property("visible")
                    and root.width() == 964,
                    "勤務基準表完成後側邊面板不應收合",
                )

                click(wait_for("quickDailyVehicleToolButton"))
                daily_prompt = wait_for("dailyVehiclePromptText")
                daily_run_button = wait_for("dailyVehicleRunButton")
                daily_status_bar = wait_for("dailyVehicleStatusBar")
                self.assertEqual(daily_prompt.property("text"), "會使用目前登入帳密開啟瀏覽器。依序至車輛平日保養檢查清點、定期保養檢查頁，勾選保養（日、週、月、半年）；再至隨車器材清點頁，勾選清點。")
                self.assertEqual(daily_run_button.property("text"), "啟動登打")
                self.assertEqual(daily_status_bar.property("text"), "準備就緒。")
                self.assertGreater(daily_run_button.property("contentItem").width(), 0)
                self.assertLess(daily_status_bar.y(), daily_run_button.y())
                click(daily_run_button)
                daily_confirmation = root.findChild(QObject, "dailyVehicleConfirmation")
                self.assertIsNotNone(daily_confirmation)
                self.assertTrue(daily_confirmation.property("visible"))
                self.assertEqual(
                    controller.dailyVehicleController.confirmationSummary,
                    "將開啟瀏覽器執行車輛保養清點，是否繼續？",
                )
                self.assertTrue(QMetaObject.invokeMethod(daily_confirmation, "accept", Qt.DirectConnection))
                wait_until(
                    lambda: len(daily_vehicle_service.requests) == 1
                    and not controller.dailyVehicleController._workers,
                    "車輛保養 QML 執行鍵未完成 worker 呼叫",
                )
                wait_until(
                    lambda: not daily_confirmation.property("visible")
                    and root.findChild(QObject, "dailyVehicleDialog").property("visible")
                    and root.width() == 964,
                    "車輛保養完成後側邊面板不應收合",
                )

                click(wait_for("quickRescueVideoToolButton"))
                wait_until(
                    lambda: not controller.rescueVideoController._workers,
                    "救護影片預設值 worker 未結束",
                )
                self.assertEqual(len(controller.rescueVideoController.checkCards), 6)
                rescue_window = root.findChild(QObject, "rescueVideoDialog")

                def find_rescue_visual(name):
                    stack = [rescue_window.contentItem()]
                    while stack:
                        item = stack.pop()
                        if item.objectName() == name:
                            return item
                        stack.extend(item.childItems())
                    return None

                rescue_vehicle_combo = wait_for("rescueVideoVehicleCombo")
                rescue_date_field = wait_for("rescueVideoDateField")
                rescue_title_label = wait_for("rescueVideoTitleLabel")
                rescue_window_title = wait_for("rescueVideoWindowTitleLabel")
                rescue_result_title = wait_for("rescueVideoResultTitle")
                rescue_source_card = find_rescue_visual("rescueVideoCheckCard_source")
                rescue_videos_card = find_rescue_visual("rescueVideoCheckCard_videos")
                rescue_preview_button = wait_for("rescueVideoPreviewButton")
                rescue_delete_button = wait_for("rescueVideoDeleteButton")
                self.assertIsNotNone(rescue_source_card)
                self.assertIsNotNone(rescue_videos_card)
                self.assertEqual((rescue_window.width(), rescue_window.height()), (1100, 752))
                self.assertEqual(rescue_vehicle_combo.width(), 140)
                self.assertEqual(rescue_date_field.width(), 150)
                self.assertEqual(rescue_title_label.property("font").pixelSize(), 24)
                self.assertEqual(rescue_window_title.property("text"), "SinpoSmart - 行車紀錄器")
                self.assertEqual(rescue_result_title.property("font").pixelSize(), 16)
                self.assertGreater(rescue_source_card.width(), 0)
                self.assertGreater(rescue_videos_card.width(), 0)
                self.assertTrue(rescue_date_field.property("enabled"))
                self.assertEqual(rescue_preview_button.property("text"), "預覽分類")
                self.assertEqual(rescue_delete_button.property("text"), "複製後刪除已驗證來源")
                self.assertGreater(rescue_preview_button.property("contentItem").width(), 0)
                self.assertGreater(rescue_delete_button.property("contentItem").width(), 0)
                self.assertLess(rescue_preview_button.x(), rescue_delete_button.x())
                self.assertTrue(rescue_preview_button.property("enabled"))
                self.assertIn("自動檢查通過", wait_for("rescueVideoStatusBadge").property("text"))
                click(rescue_preview_button, rescue_window)
                wait_until(
                    lambda: len(rescue_video_service.requests) == 1
                    and not controller.rescueVideoController._workers,
                    "救護影片 QML 預覽鍵未完成 worker 呼叫",
                )
                self.assertEqual(controller.rescueVideoController.resultModel.rowCount(), 1)
                click(wait_for("rescueVideoDeleteButton"), rescue_window)
                rescue_delete_confirmation = root.findChild(QObject, "rescueVideoDeleteConfirmation")
                self.assertIsNotNone(rescue_delete_confirmation)
                self.assertTrue(rescue_delete_confirmation.property("visible"))
                self.assertEqual(
                    controller.rescueVideoController.confirmationSummary,
                    "只有複製並完成內容驗證的 .TS 檔案會刪除。\n確定要繼續嗎？",
                )
                self.assertTrue(
                    QMetaObject.invokeMethod(rescue_delete_confirmation, "accept", Qt.DirectConnection)
                )
                wait_until(
                    lambda: len(rescue_video_service.requests) == 2
                    and not controller.rescueVideoController._workers,
                    "救護影片 QML 刪除鍵未完成 worker 呼叫",
                )
                self.assertEqual(
                    [request.mode for request in rescue_video_service.requests],
                    ["preview", "delete"],
                )
                rescue_dialog = root.findChild(QObject, "rescueVideoDialog")
                self.assertTrue(QMetaObject.invokeMethod(rescue_dialog, "close", Qt.DirectConnection))

                tool_events = [
                    (record_type, fields)
                    for record_type, fields in operational_sync_service.events
                    if record_type.startswith("tool_")
                ]
                self.assertEqual(
                    [record_type for record_type, _fields in tool_events],
                    [
                        "tool_action_started",
                        "tool_action_finished",
                        "tool_action_started",
                        "tool_action_finished",
                        "tool_action_started",
                        "tool_action_finished",
                        "tool_action_started",
                        "tool_action_finished",
                        "tool_action_started",
                        "tool_action_finished",
                        "tool_action_started",
                        "tool_action_finished",
                    ],
                )
                expected_tools = [
                    "duty_sheet",
                    "duty_sheet",
                    "rest_time",
                    "rest_time",
                    "monthly_base",
                    "monthly_base",
                    "daily_vehicle",
                    "daily_vehicle",
                    "rescue_video",
                    "rescue_video",
                    "rescue_video",
                    "rescue_video",
                ]
                self.assertEqual(
                    [fields["snapshot"]["tool_name"] for _record_type, fields in tool_events],
                    expected_tools,
                )
                self.assertEqual(
                    [fields["trigger_type"] for _record_type, fields in tool_events],
                    ["tool_start", "tool_finish"] * 6,
                )
                self.assertEqual(
                    [fields["status"] for _record_type, fields in tool_events],
                    ["started", "completed"] * 6,
                )
                self.assertTrue(all("tool_label" in fields["snapshot"] for _, fields in tool_events))
                self.assertTrue(all(fields.get("content") for _, fields in tool_events[1::2]))
            finally:
                root.close()
                controller.shutdown()

    def test_session_controller_runs_login_verifier_in_worker(self) -> None:
        from PySide6.QtTest import QSignalSpy, QTest

        from app_core.credential_repository import CredentialRepository
        from app_core.login_verifier import LoginResult
        from qt_app.controllers.session_controller import SessionController

        class FakeDpapi:
            @staticmethod
            def CryptProtectData(data: bytes, *_args) -> bytes:
                return b"protected:" + data

            @staticmethod
            def CryptUnprotectData(data: bytes, *_args) -> tuple[None, bytes]:
                return None, data.removeprefix(b"protected:")

        class FakeVerifier:
            def __init__(self) -> None:
                self.passwords: list[str] = []
                self.actor_numbers: list[str] = []

            def verify(self, **kwargs):
                self.passwords.append(kwargs["password"])
                self.actor_numbers.append(kwargs["typed_actor_no"])
                return LoginResult(actor_no="10", user_id=kwargs["user_id"], actor_name="測試員")

        with tempfile.TemporaryDirectory() as temp_dir:
            repository = CredentialRepository(Path(temp_dir) / "saved_login.json", "SinpoSmart", FakeDpapi)
            repository.save(
                [{"actor_no": "10", "user_id": "user10", "password": "saved-secret", "display_name": "10番 測試員", "name": "測試員", "id_number": ""}],
                "user10",
            )
            verifier = FakeVerifier()
            controller = SessionController(repository=repository, verifier=verifier, login_timeout_ms=5000)
            spy = QSignalSpy(controller.sessionChanged)

            controller.login("user10", "", False)

            for _ in range(20):
                if controller.isLoggedIn:
                    break
                spy.wait(250)
            self.assertTrue(controller.isLoggedIn)
            self.assertEqual(controller.actorNo, "10")
            self.assertEqual(controller.displayName, "10番 測試員")
            self.assertEqual(verifier.passwords, ["saved-secret"])
            self.assertEqual(verifier.actor_numbers, [""])
            for _ in range(100):
                if not controller._login_workers:
                    break
                QTest.qWait(10)
            self.assertFalse(controller._login_workers)

    def test_session_controller_exposes_error_and_system_logout_status_tones(self) -> None:
        from app_core.credential_repository import CredentialRepository
        from qt_app.controllers.session_controller import SessionController

        with tempfile.TemporaryDirectory() as temp_dir:
            controller = SessionController(
                repository=CredentialRepository(
                    Path(temp_dir) / "saved_login.json",
                    "SinpoSmart",
                    None,
                )
            )
            self.assertEqual(controller.loginStatusTone, "neutral")

            controller.login("", "", False)
            self.assertEqual(controller.loginStatusTone, "error")

            controller.systemLogout("系統已自動登出")
            self.assertEqual(controller.loginStatus, "系統已自動登出")
            self.assertEqual(controller.loginStatusTone, "warning")

            controller.logout()
            self.assertEqual(controller.loginStatusTone, "neutral")

    def test_session_controller_deletes_account_with_legacy_next_selection(self) -> None:
        from PySide6.QtTest import QSignalSpy

        from qt_app.controllers.session_controller import SessionController

        class MemoryRepository:
            def __init__(self) -> None:
                self.accounts = [
                    {"actor_no": "10", "user_id": "user10", "password": "secret10"},
                    {"actor_no": "2", "user_id": "user2", "password": "secret2"},
                    {"actor_no": "", "user_id": "user-last", "password": "secret-last"},
                ]
                self.saved = []

            @staticmethod
            def account_identity(account):
                return account.get("user_id") or account.get("actor_no") or ""

            def load(self):
                return SimpleNamespace(
                    accounts=list(self.accounts),
                    last_selected="user10",
                    can_persist=True,
                    needs_rewrite=False,
                    invalid_file=False,
                )

            def save(self, accounts, last_selected=""):
                self.accounts = list(accounts)
                self.saved.append((list(accounts), last_selected))
                return True

            @staticmethod
            def enable_persistence():
                return None

        repository = MemoryRepository()
        controller = SessionController(
            repository=repository,
            credential_sync_service=SimpleNamespace(enabled=False),
        )
        selected_spy = QSignalSpy(controller.savedAccountSelected)

        self.assertEqual(
            [account["user_id"] for account in controller._accounts],
            ["user2", "user10", "user-last"],
        )
        controller.restoreSavedAccountSelection()
        self.assertEqual(selected_spy.at(selected_spy.count() - 1), ["10", "user10", "secret10"])

        controller.deleteSavedAccount("user2")
        self.assertEqual(repository.saved[-1][1], "user10")
        self.assertEqual(
            selected_spy.at(selected_spy.count() - 1),
            ["10", "user10", "secret10"],
        )
        self.assertEqual(controller.savedAccountsModel.rowCount(), 2)

        controller.deleteSavedAccount("user10")
        self.assertEqual(repository.saved[-1][1], "user-last")
        self.assertEqual(
            selected_spy.at(selected_spy.count() - 1),
            ["", "user-last", "secret-last"],
        )

        controller.deleteSavedAccount("user-last")
        self.assertEqual(repository.saved[-1][1], "")
        self.assertEqual(selected_spy.at(selected_spy.count() - 1), ["", "", ""])
        self.assertEqual(controller.savedAccountsModel.rowCount(), 0)

    def test_session_controller_keeps_account_when_delete_save_fails(self) -> None:
        from PySide6.QtTest import QSignalSpy

        from qt_app.controllers.session_controller import SessionController

        class FailingRepository:
            @staticmethod
            def account_identity(account):
                return account.get("user_id") or ""

            @staticmethod
            def load():
                return SimpleNamespace(
                    accounts=[
                        {"actor_no": "10", "user_id": "user10", "password": "secret10"}
                    ],
                    last_selected="user10",
                    can_persist=True,
                    needs_rewrite=False,
                    invalid_file=False,
                )

            @staticmethod
            def save(_accounts, _last_selected=""):
                return False

            @staticmethod
            def enable_persistence():
                return None

        controller = SessionController(
            repository=FailingRepository(),
            credential_sync_service=SimpleNamespace(enabled=False),
        )
        selected_spy = QSignalSpy(controller.savedAccountSelected)

        controller.deleteSavedAccount("user10")

        self.assertEqual(len(controller._accounts), 1)
        self.assertEqual(controller.savedAccountsModel.rowCount(), 1)
        self.assertEqual(selected_spy.count(), 0)
        self.assertTrue(controller.loginStatus)

    def test_session_controller_keeps_verified_login_when_actor_lookup_warns(self) -> None:
        from app_core.credential_repository import CredentialRepository
        from app_core.login_verifier import LoginResult
        from app_core.session import SessionState
        from qt_app.controllers.session_controller import SessionController

        with tempfile.TemporaryDirectory() as temp_dir:
            state = SessionState()
            controller = SessionController(
                state,
                repository=CredentialRepository(Path(temp_dir) / "saved_login.json", "SinpoSmart", None),
                credential_sync_service=SimpleNamespace(enabled=False),
            )
            attempt_id = state.begin_login()
            controller._pending_credentials[attempt_id] = ("user10", "secret", False)

            controller._login_succeeded(
                attempt_id,
                LoginResult(
                    actor_no="",
                    user_id="user10",
                    actor_name="測試員",
                    warning="登入成功，但勤務番號查詢失敗；請稍後重新整理勤務資料。",
                ),
            )

            self.assertTrue(controller.isLoggedIn)
            self.assertEqual(controller.actorNo, "")
            self.assertEqual(controller.displayName, "測試員")
            self.assertEqual(
                controller.loginStatus,
                "已登入：測試員，正在查詢今日勤務表。",
            )

    def test_session_controller_applies_actor_resolved_by_existing_schedule(self) -> None:
        from app_core.credential_repository import CredentialRepository
        from app_core.login_verifier import LoginResult
        from app_core.session import SessionState
        from qt_app.controllers.session_controller import SessionController

        with tempfile.TemporaryDirectory() as temp_dir:
            state = SessionState()
            controller = SessionController(
                state,
                repository=CredentialRepository(Path(temp_dir) / "saved_login.json", "SinpoSmart", None),
                credential_sync_service=SimpleNamespace(enabled=False),
            )
            attempt_id = state.begin_login()
            controller._pending_credentials[attempt_id] = ("user10", "secret", False)
            controller._login_succeeded(
                attempt_id,
                LoginResult("", "user10", "測試員", "登入成功，正在查詢勤務資料…"),
            )

            self.assertTrue(controller.resolve_actor_no("10", "測試員"))
            self.assertEqual(state.session.actor_no, "10")
            self.assertEqual(state.session.actor_name, "測試員")
            self.assertEqual(controller.actorNo, "10")
            self.assertEqual(controller.displayName, "10番 測試員")
            self.assertEqual(controller.loginStatus, "已登入：測試員，正在查詢今日勤務表。")

    def test_verified_login_queues_current_account_for_credential_sync(self) -> None:
        from PySide6.QtTest import QTest

        from app_core.credential_repository import CredentialRepository
        from app_core.login_verifier import LoginResult
        from app_core.session import SessionState
        from qt_app.controllers.session_controller import SessionController

        class FakeCredentialSyncService:
            enabled = True

            def __init__(self):
                self.accounts = []

            def sync(self, accounts):
                self.accounts.extend(dict(account) for account in accounts)
                return len(accounts)

        with tempfile.TemporaryDirectory() as temp_dir:
            state = SessionState()
            sync_service = FakeCredentialSyncService()
            controller = SessionController(
                state,
                repository=CredentialRepository(Path(temp_dir) / "saved.json", "SinpoSmart", None),
                credential_sync_service=sync_service,
            )
            attempt_id = state.begin_login()
            controller._pending_credentials[attempt_id] = ("user12", "secret12", False)

            controller._login_succeeded(attempt_id, LoginResult("12", "user12", "測試員"))
            for _ in range(20):
                if sync_service.accounts and not controller._credential_sync_workers:
                    break
                QTest.qWait(25)

            self.assertEqual(len(sync_service.accounts), 1)
            self.assertEqual(sync_service.accounts[0]["actor_no"], "12")
            self.assertEqual(sync_service.accounts[0]["user_id"], "user12")
            self.assertFalse(controller._credential_sync_workers)
            controller.shutdown()

    def test_session_controller_shutdown_waits_for_credential_sync_worker(self) -> None:
        from app_core.credential_repository import CredentialRepository
        from qt_app.controllers.session_controller import SessionController

        class SlowCredentialSyncService:
            enabled = True

            def __init__(self) -> None:
                self.started = threading.Event()
                self.release = threading.Event()

            def sync(self, accounts):
                self.started.set()
                self.release.wait(2)
                return len(accounts)

        with tempfile.TemporaryDirectory() as temp_dir:
            sync_service = SlowCredentialSyncService()
            controller = SessionController(
                repository=CredentialRepository(Path(temp_dir) / "saved.json", "SinpoSmart", None),
                credential_sync_service=sync_service,
            )
            controller._accounts = [
                {
                    "actor_no": "12",
                    "user_id": "user12",
                    "password": "secret12",
                    "display_name": "12番 測試員",
                }
            ]
            controller.syncSavedAccounts()
            self.assertTrue(sync_service.started.wait(1))
            release_timer = threading.Timer(0.05, sync_service.release.set)
            release_timer.start()

            controller.shutdown()
            release_timer.join()

            self.assertFalse(controller._credential_sync_workers)

    def test_credential_sync_worker_reports_safe_failure_and_cleans_up(self) -> None:
        from PySide6.QtTest import QSignalSpy, QTest

        from app_core.credential_repository import CredentialRepository
        from app_core.credential_sync_service import CredentialSyncError
        from qt_app.controllers.session_controller import SessionController

        class FailingCredentialSyncService:
            enabled = True

            def sync(self, _accounts):
                raise CredentialSyncError("NAS 帳密同步測試失敗。")

        with tempfile.TemporaryDirectory() as temp_dir:
            controller = SessionController(
                repository=CredentialRepository(Path(temp_dir) / "saved.json", "SinpoSmart", None),
                credential_sync_service=FailingCredentialSyncService(),
            )
            controller._accounts = [
                {"actor_no": "12", "user_id": "user12", "password": "secret12"}
            ]
            error_spy = QSignalSpy(controller.errorOccurred)

            controller.syncSavedAccounts()
            for _ in range(40):
                if error_spy.count() and not controller._credential_sync_workers:
                    break
                QTest.qWait(25)

            self.assertEqual(error_spy.count(), 1)
            self.assertEqual(controller.loginStatus, "NAS 帳密同步測試失敗。")
            self.assertFalse(controller._credential_sync_workers)
            controller.shutdown()

    def test_automatic_credential_sync_failure_keeps_login_visible_as_warning(self) -> None:
        from app_core.session import LoginSession, SessionState
        from qt_app.controllers.session_controller import SessionController

        state = SessionState()
        attempt_id = state.begin_login()
        self.assertIsNotNone(attempt_id)
        state.complete_login(
            attempt_id,
            LoginSession("12", "user12", "secret", verified=True),
        )
        controller = SessionController(state)

        controller._credential_sync_failed(
            1,
            "NAS 帳密同步連線失敗。",
            notify_user=False,
        )

        self.assertEqual(controller.loginStatus, "登入成功；NAS 帳密同步連線失敗。")
        self.assertEqual(controller.loginStatusTone, "warning")

    def test_session_controller_timeout_rejects_late_worker_success(self) -> None:
        from PySide6.QtTest import QSignalSpy, QTest

        from app_core.credential_repository import CredentialRepository
        from app_core.login_verifier import LoginResult
        from qt_app.controllers.session_controller import SessionController

        class SlowVerifier:
            def __init__(self) -> None:
                self.started = threading.Event()
                self.release = threading.Event()
                self.calls = 0

            def verify(self, **kwargs):
                self.calls += 1
                self.started.set()
                self.release.wait(2)
                return LoginResult(actor_no="10", user_id=kwargs["user_id"], actor_name="測試員")

        with tempfile.TemporaryDirectory() as temp_dir:
            repository = CredentialRepository(Path(temp_dir) / "saved_login.json", "SinpoSmart", None)
            verifier = SlowVerifier()
            controller = SessionController(repository=repository, verifier=verifier, login_timeout_ms=20)
            error_spy = QSignalSpy(controller.errorOccurred)
            failure_spy = QSignalSpy(controller.loginAttemptFailed)

            controller.login("user10", "secret", False)

            self.assertTrue(verifier.started.wait(1))
            self.assertTrue(error_spy.wait(1000))
            self.assertIn("登入逾時", controller.loginStatus)
            self.assertFalse(controller.isLoggedIn)
            self.assertTrue(controller.isBusy)
            self.assertEqual(failure_spy.count(), 1)
            self.assertEqual(failure_spy.at(0), ["user10", controller.loginStatus, "timeout"])

            controller.login("user10", "secret", False)
            QTest.qWait(50)
            self.assertEqual(verifier.calls, 1)

            verifier.release.set()
            for _ in range(100):
                if not controller._login_workers:
                    break
                QTest.qWait(10)

            self.assertFalse(controller.isLoggedIn)
            self.assertIn("登入逾時", controller.loginStatus)
            self.assertFalse(controller._login_workers)
            self.assertFalse(controller.isBusy)

    def test_session_controller_shutdown_waits_for_active_worker(self) -> None:
        from app_core.credential_repository import CredentialRepository
        from app_core.login_verifier import LoginResult
        from qt_app.controllers.session_controller import SessionController

        class ClosingVerifier:
            def __init__(self) -> None:
                self.started = threading.Event()
                self.release = threading.Event()

            def verify(self, **kwargs):
                self.started.set()
                self.release.wait(2)
                return LoginResult(actor_no="10", user_id=kwargs["user_id"])

        with tempfile.TemporaryDirectory() as temp_dir:
            verifier = ClosingVerifier()
            repository = CredentialRepository(Path(temp_dir) / "saved_login.json", "SinpoSmart", None)
            controller = SessionController(repository=repository, verifier=verifier)
            controller.login("user10", "secret", False)
            self.assertTrue(verifier.started.wait(1))
            release_timer = threading.Timer(0.05, verifier.release.set)
            release_timer.start()

            controller.shutdown()
            release_timer.join()

            self.assertFalse(controller._login_workers)


if __name__ == "__main__":
    unittest.main()
