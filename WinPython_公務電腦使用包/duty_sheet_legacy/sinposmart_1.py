from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
import re
import sys
import warnings
import json
import os
import tempfile
import shutil
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta
from urllib.parse import quote
from selenium.common.exceptions import NoAlertPresentException, NoSuchFrameException, TimeoutException, UnexpectedAlertPresentException
import threading
from pathlib import Path
from openpyxl.utils import get_column_letter
from copy import copy
import unicodedata

PACKAGE_DIR = Path(__file__).resolve().parents[1]
if str(PACKAGE_DIR) not in sys.path:
    sys.path.insert(0, str(PACKAGE_DIR))

from duty_rehearsal import build_driver, quit_driver, retry_duty_browser_session_open

# ==========================================
# [區塊一] 模組導入與全域設定 (Imports & Config)
# ==========================================

# 隱藏 Excel 格式警告，保持黑視窗乾淨
warnings.filterwarnings("ignore", category=UserWarning)

# 全域狀態更新函數 (確保執行緒安全)
_runtime_status_callback = None


def log_status(msg):
    try:
        print(msg)
    except Exception:
        pass
    callback = globals().get("_runtime_status_callback")
    if callable(callback):
        try:
            callback(clean_status_message(msg))
        except Exception:
            pass
    # 確保在有視窗的狀態下，將文字更新回 GUI 的狀態列
    if 'root' in globals() and 'status_var' in globals():
        root.after(0, lambda: sync_status_to_gui(msg))

def clean_status_message(msg):
    text = str(msg).strip()
    return re.sub(r'^(?:[➡⏳📂✅⚠❌🧠🖼☁🎉️]\s*)+', '', text).strip()

def truncate_external_duty_name(value, max_units=24):
    """Keep external-duty labels within the website's 12 CJK-character limit."""

    kept = []
    used_units = 0
    for character in str(value or ""):
        character_units = 2 if unicodedata.east_asian_width(character) in {"W", "F"} else 1
        if used_units + character_units > max_units:
            break
        kept.append(character)
        used_units += character_units
    return "".join(kept)


def sync_status_to_gui(msg):
    status_var.set(f"狀態: {clean_status_message(msg)}")
    if 'log_text' in globals():
        log_text.insert(tk.END, f"{msg}\n")
        log_text.see(tk.END)
# ==========================================
# [區塊二] 設定、日期、Excel 截圖與通知工具 (Config, Excel, Notification)
# 放置不直接操作勤務網站的共用工具；網站 Selenium 動作集中在區塊四、五。
# ==========================================
SCRIPT_DIR = Path(__file__).resolve().parent
CONFIG_FILE = SCRIPT_DIR / "config.json"
DEFAULT_CAPTURE_TOP_ROW = 3
DEFAULT_CAPTURE_BOTTOM_ROW = 36
DAILY_SCREENSHOT_DIR = "每日勤務表"
NIGHT_SCREENSHOT_DIR = "夜間勤務"

# 2-1. 設定檔讀寫
def get_default_config():
    return {
        "login": {
            "user_id": "",
            "user_pwd": ""
        },
        "last_selection": {
            "attack": "新坡15/KES-5922",
            "stop": "新坡11/KEC-2608",
            "amb1": "新坡91/BGV-2310",
            "amb2": "新坡93/BSL-9230"
        },
        "car_options": {
            "attack": ["新坡15/KES-5922", "新坡16/981-S5"],
            "stop": ["新坡11/KEC-2608"],
            "amb": ["新坡91/BGV-2310", "新坡92/BXB-7593", "新坡93/BSL-9230", "新坡95/BPE-5951"]
        },
        "hidden_car_options": {
            "attack": [],
            "amb": []
        },
        "notification": {
            "enabled": False,
            "provider": "line",
            "line_channel_access_token": "",
            "line_to_id": "",
            "line_group_id": "",
            "gcs_bucket_name": "sinpo-duty-schedule-images",
            "gcs_service_account_json": "effortless-leaf-353501-63492cc3ece4.json"
        }
    }

def merge_config(defaults, loaded):
    for key, value in defaults.items():
        if key not in loaded:
            loaded[key] = value
        elif isinstance(value, dict) and isinstance(loaded[key], dict):
            loaded[key] = merge_config(value, loaded[key])
    return loaded

def normalize_car_options(config):
    default_config = get_default_config()
    car_options = config.setdefault("car_options", {})
    hidden_options = config.setdefault("hidden_car_options", {})
    if not isinstance(hidden_options, dict):
        hidden_options = {}
        config["hidden_car_options"] = hidden_options

    for group, default_values in default_config.get("car_options", {}).items():
        values = car_options.get(group, [])
        if not isinstance(values, list):
            values = []
        hidden_values = hidden_options.get(group, [])
        if not isinstance(hidden_values, list):
            hidden_values = []

        hidden_set = set(hidden_values)
        merged_values = []
        for item in values + default_values:
            if item and item not in hidden_set and item not in merged_values:
                merged_values.append(item)
        car_options[group] = merged_values
        hidden_options[group] = [item for item in hidden_values if item]

    return config

def resolve_config_path(path_value):
    if not path_value:
        return ""
    path_text = str(path_value).strip()
    path_obj = Path(path_text)
    if path_obj.is_absolute():
        return str(path_obj)
    return str((CONFIG_FILE.parent / path_obj).resolve())

def screenshot_archive_root():
    script_dir = Path(__file__).resolve().parent
    if script_dir.name == "duty_sheet_legacy":
        return script_dir.parent
    return script_dir

def screenshot_date_name(target_date):
    digits = re.sub(r"\D", "", str(target_date or ""))
    if len(digits) >= 4:
        return digits[-4:]
    return datetime.now().strftime("%m%d")

def screenshot_archive_path(folder_name, target_date):
    output_dir = screenshot_archive_root() / folder_name
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir / f"{screenshot_date_name(target_date)}.png"

def load_config():
    """載入設定檔，若不存在則給予預設值"""
    default_config = get_default_config()
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return normalize_car_options(merge_config(default_config, json.load(f)))
    return normalize_car_options(default_config)

def save_config(selection, login_settings=None, notification_settings=None, car_options=None, hidden_car_options=None):
    """儲存本次選擇、登入資訊與通知設定到設定檔"""
    config = load_config()
    if login_settings is None:
        login_settings = config.get("login", get_default_config()["login"])
    config["login"] = login_settings
    config["last_selection"] = selection
    if car_options is not None:
        config["car_options"] = car_options
    if hidden_car_options is not None:
        config["hidden_car_options"] = hidden_car_options
    if notification_settings is None:
        notification_settings = config.get("notification", get_default_config()["notification"])
    config["notification"] = notification_settings
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(config, f, indent=4, ensure_ascii=False)


# 2-2. 日期、文字與 Excel 儲存格工具
def convert_to_minguo(date_obj):
    """西元轉民國格式 (例: 1150426)"""
    return f"{date_obj.year - 1911}{date_obj.strftime('%m%d')}"

def clean_v(v):
    """徹底清理番號：移除中文(全員)、0、標點符號，只留數字與逗號"""
    if v is None: return ""
    v_str = str(v).strip().replace(".0", "")
    v_str = re.sub(r'[，、。．·‧\.\n\r\s]+', ',', v_str)
    v_str = re.sub(r'[^0-9,]', '', v_str) 
    v_str = re.sub(r',+', ',', v_str).strip(',')
    return v_str if v_str not in ["0", "0.0", "nan", ""] else ""

def clean_to_list(v):
    """將清理後的番號字串轉為 List"""
    res = clean_v(v)
    return [x for x in res.split(',') if x] if res else []

def clean_to_list_excluding(v, excluded_numbers):
    """將番號字串轉為 List，並排除輪休基準表標示不需登打的人員。"""
    excluded = {str(no).strip() for no in excluded_numbers if str(no).strip()}
    return [x for x in clean_to_list(v) if x not in excluded]

def clean_v_excluding(v, excluded_numbers):
    """回傳已排除指定番號的逗號字串。"""
    return ",".join(clean_to_list_excluding(v, excluded_numbers))

def normalize_header_text(value):
    return re.sub(r"\s+", "", str(value or "").strip())

def roster_header_columns(sheet):
    """以標題文字定位輪休基準表的人員清單欄位，避免欄位位移。"""
    labels = {"no": "番號", "name": "姓名", "class": "班表欄位"}
    for row in range(1, sheet.max_row + 1):
        columns = {}
        for col in range(1, sheet.max_column + 1):
            text = normalize_header_text(sheet.cell(row=row, column=col).value)
            for key, label in labels.items():
                if text == label:
                    columns[key] = col
        if all(key in columns for key in labels):
            columns["row"] = row
            return columns
    return {}

def trainee_numbers_from_workbook(workbook):
    """從輪休基準表的班表欄位讀取實習生番號。"""
    roster_sheet = next((sheet for sheet in workbook.worksheets if "輪休" in str(sheet.title)), None)
    if roster_sheet is None:
        return set()
    headers = roster_header_columns(roster_sheet)
    if not headers:
        return set()
    trainee_numbers = set()
    for row in range(headers["row"] + 1, roster_sheet.max_row + 1):
        class_text = str(roster_sheet.cell(row=row, column=headers["class"]).value or "")
        if "實習" not in class_text:
            continue
        trainee_numbers.update(clean_to_list(roster_sheet.cell(row=row, column=headers["no"]).value))
    return trainee_numbers

def find_header_column(sheet, row, label):
    """在指定列用標題文字找欄位。"""
    target = normalize_header_text(label)
    for col in range(1, sheet.max_column + 1):
        if normalize_header_text(sheet.cell(row=row, column=col).value) == target:
            return col
    return 0

def duty_status_labels_from_workbook(workbook):
    """讀取班別參數的假別對照，空白假別通常代表上班。"""
    params_sheet = next((sheet for sheet in workbook.worksheets if "班別參數" in str(sheet.title)), None)
    labels = {"": "上班"}
    if params_sheet is None:
        return labels
    for row in range(1, params_sheet.max_row + 1):
        code = normalize_header_text(params_sheet.cell(row=row, column=1).value)
        label = normalize_header_text(params_sheet.cell(row=row, column=2).value)
        if label:
            labels[code] = label
    return labels

def expected_on_duty_numbers_from_roster(workbook, day_int, excluded_numbers):
    """依輪休基準表計算當日應排人員，取代 Excel FILTER/XLOOKUP 結果。"""
    roster_sheet = next((sheet for sheet in workbook.worksheets if "輪休" in str(sheet.title)), None)
    if roster_sheet is None:
        return []

    date_row, date_col = 0, 0
    for row in range(1, roster_sheet.max_row + 1):
        for col in range(1, roster_sheet.max_column + 1):
            if normalize_header_text(roster_sheet.cell(row=row, column=col).value) == "日期":
                date_row, date_col = row, col
                break
        if date_col:
            break
    if not date_col or date_row <= 1:
        return []

    target_row = 0
    for row in range(date_row + 1, roster_sheet.max_row + 1):
        day_values = clean_to_list(roster_sheet.cell(row=row, column=date_col).value)
        if str(day_int) in day_values:
            target_row = row
            break
    if not target_row:
        return []

    status_labels = duty_status_labels_from_workbook(workbook)
    excluded = {str(no).strip() for no in excluded_numbers if str(no).strip()}
    expected = []
    for col in range(date_col + 2, roster_sheet.max_column + 1):
        header = normalize_header_text(roster_sheet.cell(row=date_row, column=col).value)
        if header == "最低人數":
            break
        numbers = clean_to_list(roster_sheet.cell(row=date_row - 1, column=col).value)
        if not header or not numbers:
            continue
        duty_code = normalize_header_text(roster_sheet.cell(row=target_row, column=col).value)
        status = status_labels.get(duty_code, "" if duty_code else status_labels.get("", "上班"))
        if status == "上班" and numbers[0] not in excluded:
            expected.append(numbers[0])
    return expected

def expected_on_duty_numbers_from_daily_sheet(sheet, excluded_numbers):
    """保留巨集原本讀第 22 列「備勤」右側應排名單的備援路徑。"""
    standby_col = find_header_column(sheet, 22, "備勤")
    if not standby_col:
        return []
    return clean_to_list_excluding(get_merged_val(sheet, 22, standby_col + 1), excluded_numbers)

def expected_on_duty_numbers(workbook, sheet, day_int, excluded_numbers):
    expected = expected_on_duty_numbers_from_roster(workbook, day_int, excluded_numbers)
    if expected:
        return expected
    return expected_on_duty_numbers_from_daily_sheet(sheet, excluded_numbers)

def validate_daily_sheet_assignments(workbook, sheet, day_int, excluded_numbers):
    """檢查每日勤務表是否有同時段重複排班或漏排。"""
    issues = []
    start_col = find_header_column(sheet, 5, "值班")
    end_col = find_header_column(sheet, 6, "指揮官")
    if not start_col or not end_col:
        return [f"{sheet.title}：找不到「值班」或「指揮官」欄位，無法執行勤務表檢查。"]

    expected = expected_on_duty_numbers(workbook, sheet, day_int, excluded_numbers)
    if not expected:
        return [f"{sheet.title}：找不到當日應排名單，無法執行漏排檢查。"]

    formula_error_values = {"#NAME?", "#VALUE!", "#REF!", "#DIV/0!", "#N/A"}
    for row in range(10, 34):
        row_title = str(sheet.cell(row=row, column=2).value or "").strip() or f"第 {row} 列"
        seen = {}
        duplicates = []
        formula_errors = []
        for col in range(start_col, end_col + 1):
            value = get_merged_val(sheet, row, col)
            text_value = str(value or "").strip()
            if text_value.upper() in formula_error_values:
                formula_errors.append(sheet.cell(row=row, column=col).coordinate)
            for person in clean_to_list_excluding(value, excluded_numbers):
                seen[person] = seen.get(person, 0) + 1
                if seen[person] == 2:
                    duplicates.append(person)

        missing = [person for person in expected if person not in seen]
        row_errors = []
        if formula_errors:
            row_errors.append(f"【公式錯誤】{','.join(formula_errors)}")
        if duplicates:
            row_errors.append(f"【重複】{','.join(duplicates)}")
        if missing:
            row_errors.append(f"【漏排】{','.join(missing)}")
        if row_errors:
            issues.append(f"{row_title}：" + " ".join(row_errors))
    return issues

def format_daily_sheet_preflight_message(issues):
    shown = issues[:20]
    message = "勤務表檢查未通過，已停止登打。\n\n" + "\n".join(shown)
    if len(issues) > len(shown):
        message += f"\n...另有 {len(issues) - len(shown)} 項未列出"
    return message

def normalize_staff_number(value):
    numbers = clean_to_list(value)
    if not numbers:
        return ""
    try:
        return str(int(numbers[0]))
    except ValueError:
        return numbers[0]

def validate_daily_standby_against_duty_number_leave_types(daily_standby_numbers, website_rows, excluded_numbers):
    """確認每日備勤人員在勤務番號維護的休假別欄位為空白。"""
    excluded = {normalize_staff_number(number) for number in excluded_numbers}
    expected = []
    for number in daily_standby_numbers:
        normalized = normalize_staff_number(number)
        if normalized and normalized not in excluded and normalized not in expected:
            expected.append(normalized)

    website_by_number = {}
    for row in website_rows:
        number = normalize_staff_number(row.get("staff_no", ""))
        if number and number not in website_by_number:
            website_by_number[number] = row

    issues = []
    for number in expected:
        row = website_by_number.get(number)
        if row is None:
            issues.append(f"勤務番號維護找不到每日備勤番號 {number}。")
            continue
        leave_type = str(row.get("leave_type", "") or "").strip()
        if leave_type:
            name = str(row.get("name", "") or "").strip()
            person = f"{number}番 {name}".strip()
            issues.append(f"{person}在勤務番號維護的休假別不是空白（目前：{leave_type}）。")
    return issues

def normalize_mission_cell_value(value):
    return ",".join(clean_to_list(value))

def validate_mission_cell_values(expected_values, actual_values, alert_messages=None):
    """以儲存後欄位值判定編組是否成功；提示文字本身不直接判定失敗。"""
    del alert_messages
    issues = []
    for element_id, expected in expected_values.items():
        if element_id not in actual_values:
            issues.append(f"找不到救災任務編組欄位：{element_id}")
            continue
        actual = actual_values.get(element_id, "")
        if normalize_mission_cell_value(expected) != normalize_mission_cell_value(actual):
            issues.append(
                f"救災任務編組欄位 {element_id} 資料不一致："
                f"預期 {expected!r}，實際 {actual!r}"
            )
    return issues

def get_merged_val(sheet, row, col):
    """處理 Excel 合併儲存格讀取"""
    cell = sheet.cell(row=row, column=col)
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return cell.value

def get_sheet_name_from_target_date(target_date):
    return f"{int(target_date[-2:])}號"

def sanitize_filename(name):
    return re.sub(r'[\\\\/:*?"<>|]+', "_", name)

def normalize_sheet_name(sheet_name, available_sheet_names):
    target_name = str(sheet_name).strip()
    if target_name in available_sheet_names:
        return target_name

    target_day_match = re.search(r"(\d+)", target_name)
    if target_day_match:
        target_day = int(target_day_match.group(1))
        for candidate in available_sheet_names:
            candidate_match = re.match(r"^\s*(\d+)", str(candidate).strip())
            if candidate_match and int(candidate_match.group(1)) == target_day:
                return candidate

    raise KeyError(f"Worksheet {sheet_name} does not exist.")

def resolve_capture_range(sheet):
    """以主表上方那列的大型合併儲存格決定左右邊界。"""
    merged_candidates = []
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if min_row != 3 or max_row != 3:
            continue
        merged_candidates.append((max_col - min_col, min_col, max_col))

    if not merged_candidates:
        return f"B{DEFAULT_CAPTURE_TOP_ROW}:AM{DEFAULT_CAPTURE_BOTTOM_ROW}"

    _, min_col, max_col = max(merged_candidates, key=lambda item: item[0])
    return (
        f"{get_column_letter(min_col)}{DEFAULT_CAPTURE_TOP_ROW}:"
        f"{get_column_letter(max_col)}{DEFAULT_CAPTURE_BOTTOM_ROW}"
    )

def resolve_night_capture_range(sheet):
    """夜間勤務截圖固定列 24-33，右界取第 6 列第一個含指揮官的欄位。"""
    end_col = 31  # AE
    for col in range(2, sheet.max_column + 1):
        cell_value = sheet.cell(row=6, column=col).value
        if cell_value is None:
            continue
        if "指揮官" in str(cell_value).strip():
            end_col = col
            break
    return f"B24:{get_column_letter(end_col)}33"

def fit_summary_cells_for_screenshot(worksheet, sheet_values, min_col, min_row, max_col, max_row):
    """截圖前微調右下角假別統計，避免窄欄把 10 顯示成 1。"""
    summary_labels = {"輪休", "月補", "補休", "請休", "公假", "婚假", "喪假", "身心假", "陪產假"}
    summary_start_row = max(min_row, max_row - 7)
    summary_start_col = max(min_col, max_col - 12)
    for row in range(summary_start_row, max_row + 1):
        for col in range(summary_start_col, max_col + 1):
            value = worksheet.Cells(row, col).Value
            if value is None or str(value).strip() not in summary_labels:
                continue
            end_col = min(max_col, col + 6)
            if end_col <= col:
                continue
            summary_range = worksheet.Range(worksheet.Cells(row, col + 1), worksheet.Cells(row, end_col))
            summary_range.ShrinkToFit = False
            summary_range.WrapText = False
            for value_col in range(col + 1, end_col + 1):
                original_value = sheet_values.cell(row=row, column=value_col).value
                if original_value is None or str(original_value).strip() == "":
                    continue
                value_cell = worksheet.Cells(row, value_col)
                value_cell.NumberFormat = "@"
                value_cell.Value = str(original_value)

def copy_range_picture_with_retry(export_range):
    last_error = None
    for appearance, picture_format in ((1, 2), (2, 2), (1, -4147), (2, -4147)):
        try:
            export_range.CopyPicture(Appearance=appearance, Format=picture_format)
            return
        except Exception as exc:
            last_error = exc
            time.sleep(0.5)
    raise RuntimeError("Excel 範圍複製圖片失敗") from last_error

def export_chart_to_png(chart, output_path):
    output_path = Path(output_path)
    try:
        if chart.Export(str(output_path)):
            return
    except Exception:
        pass

    fallback_path = Path(tempfile.gettempdir()) / f"sinposmart_capture_{datetime.now():%Y%m%d%H%M%S%f}.png"
    try:
        if not chart.Export(str(fallback_path)):
            raise RuntimeError("Excel 工作表匯出圖片失敗")
        shutil.move(str(fallback_path), str(output_path))
    finally:
        if fallback_path.exists():
            try:
                fallback_path.unlink()
            except Exception:
                pass

# 2-3. Excel 截圖輸出
def export_excel_sheet_to_image(excel_path, sheet_name, capture_range=None, output_path=None):
    """使用本機 Excel 將指定工作表輸出為 PNG，並回傳實際擷取範圍。"""
    try:
        import pythoncom
        import win32com.client
    except ImportError as exc:
        raise RuntimeError("Excel 截圖功能需要安裝 pywin32") from exc

    if output_path is None:
        output_dir = Path.cwd() / "screenshots"
        output_dir.mkdir(exist_ok=True)
        output_path = output_dir / f"{sanitize_filename(sheet_name)}_{datetime.now():%Y%m%d_%H%M%S}.png"
    else:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    chart_object = None

    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb_meta = openpyxl.load_workbook(excel_path, read_only=False, keep_vba=False, data_only=False)
        wb_values = openpyxl.load_workbook(excel_path, read_only=False, keep_vba=False, data_only=True)
        try:
            resolved_sheet_name = normalize_sheet_name(sheet_name, wb_meta.sheetnames)
            sheet_meta = wb_meta[resolved_sheet_name]
            sheet_values = wb_values[normalize_sheet_name(sheet_name, wb_values.sheetnames)]
            worksheet_index = wb_meta.sheetnames.index(resolved_sheet_name) + 1
            if not capture_range:
                capture_range = resolve_capture_range(sheet_meta)

            workbook = excel.Workbooks.Open(os.path.abspath(excel_path), ReadOnly=False)
            worksheet = workbook.Worksheets(worksheet_index)

            min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(capture_range)
            value_matrix = []
            for row in range(min_row, max_row + 1):
                row_values = []
                for col in range(min_col, max_col + 1):
                    row_values.append(sheet_values.cell(row=row, column=col).value)
                value_matrix.append(tuple(row_values))
            worksheet.Range(capture_range).Value = tuple(value_matrix)

            on_duty_ids = set()
            for row in range(9, 37):
                duty_id = sheet_values.cell(row=row, column=41).value  # AO
                duty_status = sheet_values.cell(row=row, column=43).value  # AQ
                if duty_id and duty_status in ("上班", "外宿"):
                    try:
                        on_duty_ids.add(int(duty_id))
                    except (TypeError, ValueError):
                        pass

            gray_color = 0xA5A5A5
            for row in range(8, 22):
                for id_col, name_col in ((36, 37), (38, 39)):  # AJ/AK, AL/AM
                    cell_value = sheet_values.cell(row=row, column=id_col).value
                    id_cell = worksheet.Cells(row, id_col)
                    name_cell = worksheet.Cells(row, name_col)
                    id_cell.Interior.Pattern = 0
                    name_cell.Interior.Pattern = 0
                    if cell_value is None:
                        continue
                    try:
                        duty_id = int(cell_value)
                    except (TypeError, ValueError):
                        continue
                    if duty_id in on_duty_ids:
                        id_cell.Interior.Pattern = 1
                        id_cell.Interior.Color = gray_color
            fit_summary_cells_for_screenshot(worksheet, sheet_values, min_col, min_row, max_col, max_row)
        finally:
            wb_meta.close()
            wb_values.close()
        export_range = worksheet.Range(capture_range)
        worksheet.Activate()
        excel.ActiveWindow.DisplayGridlines = False
        excel.ActiveWindow.Zoom = 90
        export_range.Select()
        time.sleep(1)

        copy_range_picture_with_retry(export_range)
        chart_object = worksheet.ChartObjects().Add(
            export_range.Left,
            export_range.Top,
            export_range.Width + 8,
            export_range.Height + 2
        )
        chart = chart_object.Chart
        chart.Paste()
        time.sleep(1)

        export_chart_to_png(chart, output_path)

        return str(output_path), export_range.Address
    finally:
        if chart_object is not None:
            try:
                chart_object.Delete()
            except Exception:
                pass
        if workbook is not None:
            workbook.Close(False)
        if excel is not None:
            excel.Quit()
        pythoncom.CoUninitialize()

# 2-4. GCS 上傳與 LINE 群組通知
def upload_image_to_gcs(image_path, target_date, notification_config):
    try:
        from google.cloud import storage
    except ImportError as exc:
        raise RuntimeError("Google Cloud Storage 上傳功能需要安裝 google-cloud-storage") from exc

    service_account_json = resolve_config_path(notification_config.get("gcs_service_account_json", ""))
    bucket_name = notification_config.get("gcs_bucket_name", "").strip()
    if not service_account_json or not os.path.exists(service_account_json):
        raise ValueError("GCS Service Account JSON 檔案不存在")
    if not bucket_name:
        raise ValueError("GCS Bucket 名稱未設定")

    client = storage.Client.from_service_account_json(service_account_json)
    bucket = client.bucket(bucket_name)
    upload_stamp = datetime.now().strftime("%Y%m%d%H%M%S%f")
    object_name = f"duty-schedules/{target_date}/{upload_stamp}_{Path(image_path).name}"
    blob = bucket.blob(object_name)
    blob.upload_from_filename(image_path, content_type="image/png")

    try:
        blob.make_public()
    except Exception:
        pass

    return f"https://storage.googleapis.com/{bucket_name}/{quote(object_name)}"

def send_line_messages(messages, channel_access_token, to_id):
    try:
        import requests
    except ImportError as exc:
        raise RuntimeError("LINE 發送功能需要安裝 requests") from exc

    if not channel_access_token or not to_id:
        raise ValueError("LINE Channel Access Token 或接收者 ID 未設定")

    response = requests.post(
        "https://api.line.me/v2/bot/message/push",
        headers={
            "Authorization": f"Bearer {channel_access_token}",
            "Content-Type": "application/json"
        },
        json={"to": to_id, "messages": messages},
        timeout=60
    )
    response.raise_for_status()

def send_line_image(image_url, channel_access_token, to_id):
    send_line_messages(
        [{
            "type": "image",
            "originalContentUrl": image_url,
            "previewImageUrl": image_url
        }],
        channel_access_token,
        to_id
    )

def send_line_text(text, channel_access_token, to_id):
    send_line_messages(
        [{
            "type": "text",
            "text": text
        }],
        channel_access_token,
        to_id
    )

def send_group_notification(image_paths, target_date, notification_config):
    provider = (notification_config.get("provider") or "").lower().strip()
    if provider != "line":
        raise ValueError(f"不支援的通知平台: {provider}")

    channel_access_token = notification_config.get("line_channel_access_token", "").strip()
    to_id = (
        notification_config.get("line_to_id")
        or notification_config.get("line_group_id")
        or ""
    ).strip()
    completion_text = f"{target_date}勤務表登打完成"
    if isinstance(image_paths, (str, os.PathLike)):
        image_paths = [str(image_paths)]
    else:
        image_paths = [str(path) for path in image_paths]
    if not image_paths:
        raise ValueError("沒有可發送的截圖")

    image_urls = [
        upload_image_to_gcs(image_path, target_date, notification_config)
        for image_path in image_paths
    ]
    messages = [
        {
            "type": "image",
            "originalContentUrl": image_url,
            "previewImageUrl": image_url
        }
        for image_url in image_urls
    ]
    messages.append({
        "type": "text",
        "text": completion_text
    })
    send_line_messages(messages, channel_access_token, to_id)
    return {
        "provider": "LINE",
        "image_url": image_urls[0],
        "image_urls": image_urls,
        "text": completion_text
    }


def preview_night_excel_capture(excel_path, target_date):
    sheet_name = get_sheet_name_from_target_date(target_date)
    workbook = openpyxl.load_workbook(excel_path, read_only=False, keep_vba=False, data_only=False)
    try:
        resolved_sheet_name = normalize_sheet_name(sheet_name, workbook.sheetnames)
        capture_range = resolve_night_capture_range(workbook[resolved_sheet_name])
    finally:
        workbook.close()
    image_path, exported_range = export_excel_sheet_to_image(
        excel_path,
        sheet_name,
        capture_range=capture_range,
        output_path=screenshot_archive_path(NIGHT_SCREENSHOT_DIR, target_date)
    )
    return {
        "sheet_name": sheet_name,
        "image_path": image_path,
        "capture_range": exported_range
    }
def preview_excel_capture(excel_path, target_date):
    sheet_name = get_sheet_name_from_target_date(target_date)
    image_path, capture_range = export_excel_sheet_to_image(
        excel_path,
        sheet_name,
        output_path=screenshot_archive_path(DAILY_SCREENSHOT_DIR, target_date)
    )
    return {
        "sheet_name": sheet_name,
        "image_path": image_path,
        "capture_range": capture_range
    }


def capture_duty_sheet_images(excel_path, target_date):
    """Capture both duty-sheet images without blocking website submission."""

    return (
        preview_excel_capture(excel_path, target_date),
        preview_night_excel_capture(excel_path, target_date),
    )


# ==========================================
# [區塊三] 演算法大腦 (Core Logic Algorithm)
# 專注於救災任務編組的分配邏輯
# ==========================================

def unique_member_ids(member_ids):
    """依勤務表原順序保留有效且不重複的番號。"""
    return list(dict.fromkeys(str(member).strip() for member in member_ids if str(member).strip()))


def select_ambulance2_members(standby_ids, out_ids, ambulance1_ids):
    """救護車 2 優先取備勤，不足時取同時段外勤，且不得與救護車 1 重複。"""
    ambulance1 = set(unique_member_ids(ambulance1_ids))
    standby = unique_member_ids(standby_ids)
    out_duty = unique_member_ids(out_ids)
    candidates = standby + out_duty
    selected = []
    for member in candidates:
        if member not in ambulance1 and member not in selected:
            selected.append(member)
        if len(selected) == 2:
            break
    return selected


def fire_candidate_pool(med_ids, disaster_ids, out_ids):
    """依備勤人數建立攻擊車與中繼車的五人起始候選池。"""
    standby = unique_member_ids(disaster_ids)
    ambulance1 = unique_member_ids(med_ids)[:2]
    out_duty = unique_member_ids(out_ids)
    standby_count = len(standby)

    if standby_count >= 5:
        return standby[:10]
    if standby_count == 4:
        return unique_member_ids(standby + ambulance1[:1])
    if standby_count == 3:
        return unique_member_ids(standby + ambulance1)
    if standby_count == 2:
        return unique_member_ids(standby + ambulance1 + out_duty[:1])
    if standby_count == 1:
        return unique_member_ids(standby + out_duty[:4])
    return out_duty[:5]


def first_available_member(preferred_ids, candidate_ids, excluded_ids):
    """優先取指定人員；衝突時依候選順序遞補。"""
    excluded = set(excluded_ids)
    for member in unique_member_ids(preferred_ids) + unique_member_ids(candidate_ids):
        if member not in excluded:
            return member
    return ""


def calculate_fire_mission(med_ids, disaster_ids, out_ids, daily_commander):
    """依勤務表規則編組互斥的攻擊車與中繼車，每車最多五人。"""
    standby = unique_member_ids(disaster_ids)
    out_duty = unique_member_ids(out_ids)
    pool = fire_candidate_pool(med_ids, standby, out_duty)
    commander = str(daily_commander).strip()

    if commander and commander not in pool:
        if len(pool) >= 10:
            pool[-1] = commander
        else:
            pool.append(commander)
        pool = unique_member_ids(pool)
    if len(pool) < 2:
        return None

    attack_preferred = standby[1:2] if len(standby) >= 2 else out_duty
    attack_driver = first_available_member(attack_preferred, pool, {commander})
    attack_team = [attack_driver] if attack_driver else []
    if commander and commander not in attack_team:
        attack_team.append(commander)
    if len(attack_team) < 2:
        officer_ids = [member for member in pool if member in {"1", "2", "3", "4", "5"}]
        attack_team.append(first_available_member(officer_ids, pool, set(attack_team)))

    relay_preferred = standby[:1] if standby else out_duty
    relay_driver = first_available_member(relay_preferred, pool, set(attack_team))
    relay_team = [relay_driver] if relay_driver else []

    occupied = set(attack_team + relay_team)
    for member in pool:
        if member in occupied:
            continue
        if len(relay_team) < 5:
            relay_team.append(member)
        elif len(attack_team) < 5:
            attack_team.append(member)
        occupied.add(member)

    return {"relay": ",".join(relay_team), "attack": ",".join(attack_team)}


# ==========================================
# [區塊四] 瀏覽器底層操作 (Browser Core & JS Hooks)
# 放最強的 JS 注入與跨 Frame 搜尋工具
# ==========================================

def super_js_execute(driver, element_id, action="click", value=""):
    """地毯式搜尋全網頁 ID 並執行動作 (支援存在檢查)"""
    js_code = """
    const targetId = arguments[0];
    const action = arguments[1];
    const value = arguments[2];
    function deepScan(win) {
        var el = win.document.getElementById(targetId);
        if (el) {
            if (action == 'click') { el.click(); if (el.onclick) el.onclick(); return true; }
            else if (action == 'set') {
                el.value = value;
                el.dispatchEvent(new Event('input', {bubbles: true}));
                el.dispatchEvent(new Event('change', {bubbles: true}));
                el.dispatchEvent(new Event('blur', {bubbles: true}));
                return true;
            }
            else if (action == 'exists') { return (el.offsetWidth > 0 || el.offsetHeight > 0); }
            return true;
        }
        for (var i = 0; i < win.frames.length; i++) {
            try { if (deepScan(win.frames[i])) return true; } catch(e) {}
        }
        return false;
    }
    return deepScan(window.top);
    """
    return driver.execute_script(js_code, element_id, action, value)

def accept_pending_alerts(driver, timeout=2):
    messages = []
    wait_seconds = max(0.1, float(timeout))
    while True:
        try:
            alert = WebDriverWait(driver, wait_seconds).until(EC.alert_is_present())
        except (NoAlertPresentException, TimeoutException):
            break
        try:
            message = str(alert.text or "").strip()
            if message:
                messages.append(message)
        except Exception:
            pass
        try:
            alert.accept()
        except NoAlertPresentException:
            pass
        wait_seconds = 0.5
        time.sleep(0.2)
    return messages

# ==========================================
# [區塊五] 網頁自動化動作方塊 (Web Automation Steps)
# 拆解成獨立任務，方便主流程呼叫
# ==========================================

def step_login(driver, uid, pwd):
    log_status("➡️ 執行登入勤務管理系統...")
    driver.get("https://dutymgt.tyfd.gov.tw/tyfd119/login119")
    driver.find_element(By.ID, "_txtUsername").send_keys(uid)
    driver.find_element(By.ID, "_txtPassword").send_keys(pwd)
    driver.find_element(By.NAME, "login").click()

    log_status("⏳ 登入資料已送出，等待載入中...")
    time.sleep(5)

def step_navigate_menu(driver, wait):
    log_status("➡️ 正在開啟勤務分配表維護(外勤)...")
    wait.until(EC.frame_to_be_available_and_switch_to_it("ehrFrame"))
    driver.execute_script("""
        function bClick(w){
            var i=w.document.getElementsByName('nodeIcon1');
            for(var j=0;j<i.length;j++){if(i[j].src.indexOf('pnode')>-1)i[j].click();}
            var a=w.document.getElementsByTagName('a');
            for(var k=0;k<a.length;k++){if(a[k].innerText.indexOf('勤務分配表維護(外勤)')>-1){a[k].click();return true;}}
            var f=w.document.getElementsByTagName('frame');
            for(var l=0;l<f.length;l++){if(bClick(f[l].contentWindow))return true;}
            return false;
        }
        bClick(window.top);
    """)
    time.sleep(5)

def step_navigate_to_task_table(driver, wait):
    log_status("➡️ 正在開啟救災任務編組表...")
    driver.switch_to.default_content()
    try: wait.until(EC.frame_to_be_available_and_switch_to_it("ehrFrame"))
    except Exception as e: log_status(f"   ⚠️ 切換 ehrFrame 發生狀況: {e}")

    driver.execute_script("""
        function bClick(w){
            var i=w.document.getElementsByName('nodeIcon1');
            for(var j=0;j<i.length;j++){if(i[j].src.indexOf('pnode')>-1)i[j].click();}
            var a=w.document.getElementsByTagName('a');
            for(var k=0;k<a.length;k++){
                if(a[k].innerText.indexOf('救災任務編組表')>-1){
                    a[k].click();
                    return true;
                }
            }
            var f=w.document.getElementsByTagName('frame');
            for(var l=0;l<f.length;l++){if(bClick(f[l].contentWindow))return true;}
            return false;
        }
        bClick(window.top);
    """)
    time.sleep(5)
    return True

def step_prepare_content(driver, wait):
    log_status("➡️ 等待勤務基準表載入...")
    driver.switch_to.default_content()
    wait.until(EC.frame_to_be_available_and_switch_to_it("ehrFrame"))
    for i in range(5):
        try:
            wait.until(EC.frame_to_be_available_and_switch_to_it("contentFrame"))
            return True
        except TimeoutException:
            time.sleep(2)
    return False


def wait_for_duty_query_button_ready(driver, wait):
    log_status("⏳ 等待勤務基準表查詢按鈕就緒...")
    try:
        wait.until(lambda candidate: super_js_execute(candidate, "_btnQuery", "exists"))
    except TimeoutException as error:
        raise RuntimeError(
            "勤務基準表查詢按鈕逾時，未執行登打。"
        ) from error
    return True


def wait_for_duty_query_completion(driver, wait):
    log_status("⏳ 等待勤務表查詢完成，等待設定按鈕就緒...")

    def query_completed(candidate):
        return candidate.execute_script(r"""
            function findById(win, targetId) {
                let element = null;
                try { element = win.document.getElementById(targetId); } catch (error) {}
                if (element) return element;
                for (let index = 0; index < win.frames.length; index += 1) {
                    try {
                        element = findById(win.frames[index], targetId);
                        if (element) return element;
                    } catch (error) {}
                }
                return null;
            }
            function expectedButtonReady(targetId, expectedValue) {
                const element = findById(window.top, targetId);
                if (!element) return false;
                const style = element.ownerDocument.defaultView.getComputedStyle(element);
                const visible = element.getClientRects().length > 0 &&
                    style.display !== "none" && style.visibility !== "hidden";
                const enabled = !element.disabled &&
                    element.getAttribute("aria-disabled") !== "true";
                const value = String(
                    element.value || element.innerText || element.textContent || ""
                ).trim();
                return visible && enabled && value === expectedValue;
            }
            return expectedButtonReady("_btnOpenWinTaskCode", "設定勤務項目") &&
                expectedButtonReady("_btnOpenWinUserNo", "設定勤務番號");
        """) is True

    try:
        wait.until(query_completed)
    except TimeoutException as error:
        raise RuntimeError(
            "勤務表查詢逾時，尚未進入設定勤務項目、設定勤務番號頁面。"
        ) from error
    return True


def wait_for_duty_result_grid(driver, wait):
    log_status("⏳ 等待勤務設定後的表格載入...")
    try:
        wait.until(lambda candidate: super_js_execute(candidate, "_pln_8_1", "exists"))
    except TimeoutException as error:
        raise RuntimeError(
            "勤務基準表格逾時，尚未開始填寫或儲存。"
        ) from error
    return True


def read_duty_number_leave_rows(driver):
    return driver.execute_script(r"""
        function cellText(cell) {
            return cell ? (cell.innerText || cell.textContent || '').trim() : '';
        }
        return Array.from(document.querySelectorAll('input[name^="_DESIGNATION"]')).map(function(numberInput) {
            var row = numberInput.closest('tr');
            if (!row) return null;
            var nameInput = row.querySelector('input[name^="_hidName"]');
            return {
                staff_no: (numberInput.value || '').trim(),
                leave_type: cellText(row.cells[0]),
                name: nameInput ? (nameInput.value || '').trim() : cellText(row.cells[4])
            };
        }).filter(Boolean);
    """) or []

class DutyNumberPopupOpenError(RuntimeError):
    """The read-only duty-number popup preflight did not reach a usable window."""

    def __init__(self, diagnostic):
        self.diagnostic = dict(diagnostic or {})
        super().__init__()

    def __str__(self):
        diagnostic = self.diagnostic
        yes_no = lambda value: "是" if value is True else "否" if value is False else "未知"
        retry_count = diagnostic.get("attempt", 1)
        retry_total = diagnostic.get("attempts", 2)
        return (
            "勤務番號設定視窗未開啟，已停止登打。"
            f"（按鈕：{yes_no(diagnostic.get('button_found'))}；"
            f"可見：{yes_no(diagnostic.get('button_visible'))}；"
            f"可用：{yes_no(diagnostic.get('button_enabled'))}；"
            f"已點擊：{yes_no(diagnostic.get('click_dispatched'))}；"
            f"視窗：{diagnostic.get('window_count_before', '?')}→{diagnostic.get('window_count_after', '?')}；"
            f"預檢重試：{retry_count}/{retry_total}；未開始勤務表刪除、填寫或儲存。）"
        )


def _click_duty_number_popup_button(driver):
    return driver.execute_script(r"""
        const targetId = arguments[0];
        function findById(win) {
            let element = null;
            try { element = win.document.getElementById(targetId); } catch (error) {}
            if (element) return element;
            for (let index = 0; index < win.frames.length; index += 1) {
                try {
                    const nested = findById(win.frames[index]);
                    if (nested) return nested;
                } catch (error) {}
            }
            return null;
        }
        const element = findById(window.top);
        if (!element) {
            return {button_found: false, button_visible: false, button_enabled: false, click_dispatched: false};
        }
        const style = element.ownerDocument.defaultView.getComputedStyle(element);
        const visible = element.getClientRects().length > 0 && style.display !== 'none' && style.visibility !== 'hidden';
        const enabled = !element.disabled && element.getAttribute('aria-disabled') !== 'true';
        if (!visible || !enabled) {
            return {button_found: true, button_visible: visible, button_enabled: enabled, click_dispatched: false};
        }
        element.click();
        return {button_found: true, button_visible: true, button_enabled: true, click_dispatched: true};
    """, "_btnOpenWinUserNo") or {}


def _duty_number_popup_window_count(driver, fallback=0):
    try:
        return len(driver.window_handles)
    except Exception:
        return fallback


def open_duty_number_popup(driver, wait):
    diagnostic = {}
    try:
        before_handles = list(driver.window_handles)
        diagnostic["window_count_before"] = len(before_handles)
        diagnostic.update(_click_duty_number_popup_button(driver))
    except Exception as error:
        diagnostic["click_error"] = type(error).__name__
        diagnostic.setdefault("window_count_before", 0)
        diagnostic.setdefault("window_count_after", diagnostic["window_count_before"])
        raise DutyNumberPopupOpenError(diagnostic) from error

    if not all(
        diagnostic.get(key) is True
        for key in ("button_found", "button_visible", "button_enabled", "click_dispatched")
    ):
        diagnostic["window_count_after"] = _duty_number_popup_window_count(
            driver,
            diagnostic["window_count_before"],
        )
        raise DutyNumberPopupOpenError(diagnostic)

    try:
        popup_handles = wait.until(
            lambda candidate: [
                handle for handle in candidate.window_handles if handle not in before_handles
            ]
            or False
        )
    except Exception as error:
        diagnostic["window_count_after"] = _duty_number_popup_window_count(
            driver,
            diagnostic["window_count_before"],
        )
        diagnostic["popup_wait_error"] = type(error).__name__
        raise DutyNumberPopupOpenError(diagnostic) from error

    popup_handle = popup_handles[0]
    try:
        driver.switch_to.window(popup_handle)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "_btnSave")))
    except Exception as error:
        diagnostic["window_count_after"] = _duty_number_popup_window_count(
            driver,
            diagnostic["window_count_before"],
        )
        diagnostic["popup_ready_error"] = type(error).__name__
        raise DutyNumberPopupOpenError(diagnostic) from error

    diagnostic["window_count_after"] = _duty_number_popup_window_count(
        driver,
        diagnostic["window_count_before"],
    )
    return popup_handle, diagnostic


def preflight_duty_number_popup(driver, wait, daily_standby_numbers, excluded_numbers):
    main_window = driver.current_window_handle
    _popup_handle, diagnostic = open_duty_number_popup(driver, wait)
    time.sleep(1.5)
    log_status("➡️ 輪休查找：比對每日備勤人員上班日...")
    try:
        leave_rows = read_duty_number_leave_rows(driver)
    except Exception as error:
        diagnostic["window_count_after"] = _duty_number_popup_window_count(
            driver,
            diagnostic["window_count_before"],
        )
        diagnostic["popup_read_error"] = type(error).__name__
        raise DutyNumberPopupOpenError(diagnostic) from error
    leave_issues = validate_daily_standby_against_duty_number_leave_types(
        daily_standby_numbers,
        leave_rows,
        excluded_numbers,
    )
    if leave_issues:
        detail = "\n".join(leave_issues[:20])
        log_status(f"❌ 輪休查找未通過，共 {len(leave_issues)} 項，已停止登打")
        raise RuntimeError(
            "Excel 每日備勤與勤務番號維護休假別不一致，已停止登打。\n" + detail
        )
    log_status("✅ 輪休查找通過：每日備勤人員休假別均為空白")
    return main_window


def retry_duty_number_popup_preflight(open_browser, preflight, cleanup=quit_driver, attempts=2):
    popup_attempts = max(1, min(int(attempts), 2))
    for attempt in range(1, popup_attempts + 1):
        driver = None
        try:
            driver = open_browser()
            preflight(driver)
        except DutyNumberPopupOpenError as error:
            error.diagnostic["attempt"] = attempt
            error.diagnostic["attempts"] = popup_attempts
            if driver is not None:
                try:
                    cleanup(driver)
                except Exception:
                    pass
            if attempt < popup_attempts:
                log_status("⚠️ 勤務番號小視窗預檢未通過，將以全新瀏覽器安全重試一次...")
                continue
            raise
        except Exception:
            if driver is not None:
                try:
                    cleanup(driver)
                except Exception:
                    pass
            raise
        if attempt > 1:
            log_status("✅ 勤務番號小視窗預檢已在全新瀏覽器恢復")
        return driver
    raise RuntimeError("勤務番號小視窗預檢未取得瀏覽器。")


def save_duty_number_popup(driver, wait, daily_commander):
    js_select_v2 = f"""
    (function() {{
        var commanderNo = "{daily_commander}".trim();
        var bossNo = "1";
        var allCbs = document.querySelectorAll('input[type="checkbox"]');
        for (var a = 0; a < allCbs.length; a++) {{ allCbs[a].checked = false; }}
        var rows = document.querySelectorAll('tr');
        for (var i = 0; i < rows.length; i++) {{
            var cells = rows[i].getElementsByTagName('td');
            if (cells.length < 4) continue;
            var inputNo = rows[i].querySelector('input[name^="_DESIGNATION"]');
            if (inputNo) {{
                var currentVal = inputNo.value.trim();
                var cbs = rows[i].querySelectorAll('input[type="checkbox"]');
                if (cbs.length < 2) continue;
                if (currentVal === bossNo || currentVal === "01") cbs[1].checked = true;
                if (commanderNo !== "" && (currentVal === commanderNo || currentVal === ("0"+commanderNo).slice(-2))) cbs[0].checked = true;
            }}
        }}
        return true;
    }})();
    """
    driver.execute_script(js_select_v2)
    time.sleep(0.5)
    driver.find_element(By.ID, "_btnSave").click()

    try:
        WebDriverWait(driver, 3).until(EC.alert_is_present())
        driver.switch_to.alert.accept()
    except TimeoutException:
        pass

    try:
        wait.until(lambda candidate: len(candidate.window_handles) == 1)
    except TimeoutException:
        log_status("   ⚠️ 勤務番號設定視窗未自動關閉")


def step_config_popups(driver, wait, out_duty_names, daily_commander, main_window):
    # --- 1. 設定勤務番號 ---
    log_status("➡️ 正在設定勤務番號...")
    save_duty_number_popup(driver, wait, daily_commander)

    driver.switch_to.window(main_window)
    if not step_prepare_content(driver, wait):
        raise RuntimeError("勤務番號設定後，勤務基準表未重新載入。")

    # --- 2. 設定外勤項目 ---
    log_status(f"➡️ 開始設定外勤項目 (從 Excel 讀取到 {len(out_duty_names)} 項)...")
    if not super_js_execute(driver, "_btnOpenWinTaskCode", "click"):
        raise RuntimeError("外勤設定按鈕未就緒，已停止登打。")

    try:
        wait.until(lambda d: len(d.window_handles) > 1)
    except TimeoutException:
        raise RuntimeError("外勤設定視窗未開啟，已停止登打。")

    for h in driver.window_handles:
        if h != main_window:
            driver.switch_to.window(h)
            try:
                # 給小視窗一點時間載入，避免被系統清空
                time.sleep(1.5)

                for i in range(2, 8):
                    inp = wait.until(EC.presence_of_element_located((By.ID, f"_txtNAME{i}")))
                    inp.clear()
                    if (i-2) < len(out_duty_names):
                        raw_name = out_duty_names[i-2]
                        task_name = truncate_external_duty_name(raw_name)
                        if task_name != str(raw_name or ""):
                            log_status(f"外勤項目超過 12 個中文字限制，已截短：{task_name}")
                        inp.send_keys(task_name)

                time.sleep(0.5)
                driver.find_element(By.ID, "_btnSave").click()

                # 攔截存檔成功的警告窗
                try:
                    WebDriverWait(driver, 3).until(EC.alert_is_present())
                    driver.switch_to.alert.accept()
                except TimeoutException:
                    pass

                # 確保存檔後視窗真的關閉
                try:
                    wait.until(lambda d: len(d.window_handles) == 1)
                except TimeoutException:
                    log_status("   ⚠️ 外勤設定視窗未自動關閉")
            except Exception as e:
                log_status(f"   ❌ 外勤設定發生錯誤: {e}")
                raise RuntimeError("外勤設定儲存失敗，已停止登打。") from e
            break
    else:
        raise RuntimeError("外勤設定視窗未開啟，已停止登打。")

    driver.switch_to.window(main_window)
    if not step_prepare_content(driver, wait):
        raise RuntimeError("外勤設定後，勤務基準表未重新載入。")

def step_select_vehicles_popup(driver, wait, main_window, cars_dict):
    log_status("➡️ 正在鎖定車輛設定視窗...")
    try:
        wait.until(lambda d: len(d.window_handles) > 1)
        for h in driver.window_handles:
            if h != main_window:
                driver.switch_to.window(h)
                break
        time.sleep(1.5)
    except Exception as e:
        log_status(f"❌ 找不到車輛設定視窗: {e}")
        return

    js_select_car_v2 = r"""
    function forceSelect(selectId, targetText) {
        var sel = document.getElementById(selectId);
        if (!sel) return "找不到ID";
        var parts = targetText.split(/[/／]/);
        var mainKey = parts[0].trim();
        for (var i = 0; i < sel.options.length; i++) {
            var optText = sel.options[i].text;
            if (optText.indexOf(mainKey) !== -1 || optText.indexOf(targetText.trim()) !== -1) {
                sel.selectedIndex = i;
                sel.dispatchEvent(new Event('input', { bubbles: true }));
                sel.dispatchEvent(new Event('change', { bubbles: true }));
                if(typeof sel.onchange === 'function') {
                    try { sel.onchange({target: sel, srcElement: sel, type: 'change'}); } catch(e){}
                }
                return "成功選中: " + optText;
            }
        }
        return "找不到匹配項";
    }
    return {
        '攻擊車': forceSelect('_selCALL_TYPEA', arguments[0]),
        '中繼車': forceSelect('_selCALL_TYPEB', arguments[1]),
        '救護1': forceSelect('_selCALL_TYPEF', arguments[2]),
        '救護2': forceSelect('_selCALL_TYPEG', arguments[3])
    };
    """
    try:
        driver.execute_script(js_select_car_v2, cars_dict['attack'], cars_dict['stop'], cars_dict['amb1'], cars_dict['amb2'])
        time.sleep(1)
        driver.execute_script("var btn = document.getElementById('_btnSave'); if (btn) { btn.click(); return true; } return false;")
        
        # 確保存檔後，車輛設定的小視窗真的關閉了才繼續
        wait.until(lambda d: len(d.window_handles) == 1)
        
    except Exception as e: log_status(f"   ❌ 車輛設定執行錯誤: {e}")
    driver.switch_to.window(main_window)

def step_batch_fill_duty(driver, duty_map):
    """批次填寫勤務基準表"""
    if not duty_map: return
    js_fill = """
    const data = arguments[0] || {};
    function deepBatchFill(win, data) {
        var oldAlert = win.alert; var oldConfirm = win.confirm;
        win.alert = function(msg) { console.log("攔截: " + msg); };
        win.confirm = function(msg) { return true; };
        var count = 0;
        for (var id in data) {
            var el = win.document.getElementById(id);
            if (el && data[id] !== "") {
                el.focus(); el.value = data[id];
                el.dispatchEvent(new Event('input', {bubbles: true}));
                el.dispatchEvent(new Event('change', {bubbles: true}));
                el.dispatchEvent(new Event('blur', {bubbles: true}));
                try { if(typeof el.onchange === 'function') el.onchange({target: el, type: 'change'}); } catch(e){}
                count++;
            }
        }
        for(var i=0; i<win.frames.length; i++) {
            try { count += deepBatchFill(win.frames[i], data); } catch(e){}
        }
        win.alert = oldAlert; win.confirm = oldConfirm;
        return count;
    }
    return deepBatchFill(window.top, data);
    """
    try:
        count = driver.execute_script(js_fill, duty_map)
    except Exception as e: log_status(f"   ❌ 批次填寫異常: {e}")

def step_fill_mission_cells(driver, mission_map):
    """批次填寫救災任務編組表"""
    js_fill_and_save = """
    const data = arguments[0] || {};
    function deepProcess(win, data) {
        var foundIds = [];
        for (var id in data) {
            var el = win.document.getElementById(id);
            if (el) {
                el.value = data[id];
                el.dispatchEvent(new Event('input', {bubbles: true}));
                el.dispatchEvent(new Event('change', {bubbles: true}));
                el.dispatchEvent(new Event('blur', {bubbles: true}));
                if(typeof el.onchange === 'function') el.onchange();
                foundIds.push(id);
            }
        }
        var btn = win.document.getElementById('_btnSave');
        if (btn && foundIds.length) {
            btn.click();
            return {status: "saved", found_ids: foundIds, save_found: true};
        }
        for(var i=0; i<win.frames.length; i++) {
            try { var res = deepProcess(win.frames[i], data); if (res) return res; } catch(e){}
        }
        return foundIds.length ? {status: "filled", found_ids: foundIds, save_found: false} : null;
    }
    return deepProcess(window.top, data);
    """
    try:
        result = driver.execute_script(js_fill_and_save, mission_map)
    except UnexpectedAlertPresentException as e:
        alert_text = ""
        try:
            alert_text = str(driver.switch_to.alert.text or "").strip()
        except Exception:
            pass
        log_status(f"   ⚠️ 救災任務編組表提示：{alert_text or e}")
        return {"status": "alert", "found_ids": [], "save_found": False, "alert": alert_text}
    except Exception as e:
        log_status(f"   ❌ 任務填寫報錯: {e}")
        return {"status": "error", "found_ids": [], "save_found": False, "error": str(e)}
    if isinstance(result, dict):
        return result
    return {"status": str(result or "missing"), "found_ids": [], "save_found": False}

def read_mission_cells(driver, mission_map):
    values = driver.execute_script(
        """
        const data = arguments[0] || {};
        function deepRead(win, data, values) {
            for (var id in data) {
                if (Object.prototype.hasOwnProperty.call(values, id)) continue;
                var el = win.document.getElementById(id);
                if (el) values[id] = String(el.value || '');
            }
            for (var i = 0; i < win.frames.length; i++) {
                try { deepRead(win.frames[i], data, values); } catch(e) {}
            }
        }
        var values = {};
        deepRead(window.top, data, values);
        return values;
        """,
        mission_map,
    )
    return values if isinstance(values, dict) else {}

def verify_mission_cells(driver, mission_map, timeout=10):
    deadline = time.monotonic() + max(1, int(timeout))
    issues = []
    while True:
        try:
            actual_values = read_mission_cells(driver, mission_map)
        except UnexpectedAlertPresentException:
            accept_pending_alerts(driver, timeout=0.5)
            actual_values = {}
        issues = validate_mission_cell_values(mission_map, actual_values)
        if not issues:
            return []
        if time.monotonic() >= deadline:
            return issues
        time.sleep(0.5)


# ==========================================
# [區塊六] 主控制流程 (Orchestrator - 流程大總管)
# 負責串聯 Excel 解析、網站填寫、截圖與通知。
# ==========================================

# 6-1. 單次勤務登打流程
def start_automation(
    user_id,
    user_pwd,
    target_date,
    excel_path,
    cars_config,
    status_callback=None,
    success_callback=None,
    error_callback=None,
    show_dialogs=True,
    close_driver=False,
    raise_errors=False,
    stage_callback=None,
):
    def report_stage(stage):
        if callable(stage_callback):
            stage_callback(stage)

    global _runtime_status_callback
    previous_status_callback = _runtime_status_callback
    _runtime_status_callback = status_callback
    # 紀錄流程開始的時間
    start_time = time.time()
    capture_executor = None
    capture_future = None
    driver = None
    # ---------------- 1. 解析 Excel ----------------
    report_stage("source_load")
    day_int = int(target_date[-2:])
    log_status(f"📂 讀取 Excel {day_int}號 分頁...")
    wb = openpyxl.load_workbook(excel_path, data_only=True, keep_vba=True)
    excluded_numbers = trainee_numbers_from_workbook(wb)
    if excluded_numbers:
        ordered_excluded = sorted(excluded_numbers, key=lambda x: (0, int(x)) if x.isdigit() else (1, x))
        log_status(f"ℹ️ 已略過實習生番號：{', '.join(ordered_excluded)}")
    sheet = wb[f"{day_int}號"]
    daily_standby_numbers = expected_on_duty_numbers_from_daily_sheet(sheet, excluded_numbers)
    
    ex_map = {"時間": 2, "值班": 3}
    for r in [5, 6]:
        for c in range(1, 100):
            v = str(sheet.cell(row=r, column=c).value or "").strip()
            if "休息" in v: ex_map["休息_Excel"] = c
            if "備勤緊急救護" in v: ex_map["救護_Excel"] = c
            if "備勤救災" in v: ex_map["備勤_Excel"] = c
            if "指揮官" in v: ex_map["指揮官"] = c

    cmd_all = []
    for r in range(10, 34):
        val = get_merged_val(sheet, r, ex_map["指揮官"])
        cmd_all.extend([int(x) for x in clean_to_list_excluding(val, excluded_numbers) if 1 <= int(x) <= 5])
    daily_commander = min(cmd_all) if cmd_all else ""

    out_names, out_excel_cols = [], []
    for c in range(ex_map["值班"] + 1, ex_map["休息_Excel"]):
        name = str(sheet.cell(row=5, column=c).value or "").strip()
        if name:
            out_names.append(name)
            out_excel_cols.append(c)

    num_out = len(out_names)
    web_idx = {
        "值班": 1, "外勤開始": 2, 
        "救護": 2 + num_out, "備勤": 2 + num_out + 1, "休息": 2 + num_out + 2
    }

    log_status(f"✅ Excel 讀取完成：外勤 {num_out} 項，指揮官為番號 {daily_commander if daily_commander else '無'}")
    report_stage("preflight")
    preflight_issues = validate_daily_sheet_assignments(wb, sheet, day_int, excluded_numbers)
    if not daily_standby_numbers:
        preflight_issues.append(f"{sheet.title}：找不到第 22 列「備勤」欄位人員，無法比對勤務番號維護休假別。")
    if preflight_issues:
        preflight_message = format_daily_sheet_preflight_message(preflight_issues)
        log_status(f"❌ 勤務表檢查未通過，共 {len(preflight_issues)} 項，已停止登打")
        if callable(error_callback):
            error_callback(preflight_message)
        elif show_dialogs and "root" in globals():
            root.after(0, lambda msg=preflight_message: messagebox.showerror("勤務表檢查未通過", msg))
        wb.close()
        _runtime_status_callback = previous_status_callback
        return False
    log_status("✅ 勤務表檢查通過，未發現重複或漏排")
    capture_executor = ThreadPoolExecutor(max_workers=1, thread_name_prefix="sinposmart-duty-capture")
    capture_future = capture_executor.submit(capture_duty_sheet_images, excel_path, target_date)
    log_status("勤務表截圖與網站登打同步進行...")
    
    # ---------------- 2. 瀏覽器自動化 ----------------
    report_stage("browser_start")
    def initialize_browser(candidate: webdriver.Chrome) -> None:
        try:
            candidate.set_page_load_timeout(max(10, int(os.environ.get("SELENIUM_PAGE_LOAD_TIMEOUT_SECONDS", "45"))))
        except Exception:
            pass
        try:
            candidate.set_script_timeout(max(10, int(os.environ.get("SELENIUM_SCRIPT_TIMEOUT_SECONDS", "45"))))
        except Exception:
            pass
        report_stage("login")
        step_login(candidate, user_id, user_pwd)
        report_stage("duty_form_open")
        step_navigate_menu(candidate, WebDriverWait(candidate, 20))

    try:
        popup_main_window = ""

        def open_duty_browser_for_popup_preflight():
            candidate = retry_duty_browser_session_open(
                lambda: build_driver(headless=False),
                initialize_browser,
                cleanup=quit_driver,
            )
            candidate_wait = WebDriverWait(candidate, 20)
            if not step_prepare_content(candidate, candidate_wait):
                raise RuntimeError("勤務表頁面準備失敗，未執行登打。")
            if not super_js_execute(candidate, "_txtTaskDate", "set", target_date):
                raise RuntimeError("勤務基準表日期欄位未就緒，未執行登打。")
            query_wait = WebDriverWait(candidate, 60, poll_frequency=0.5)
            wait_for_duty_query_button_ready(candidate, query_wait)
            if not super_js_execute(candidate, "_btnQuery", "click"):
                raise RuntimeError("勤務基準表查詢按鈕點擊失敗，未執行登打。")
            wait_for_duty_query_completion(
                candidate,
                query_wait,
            )
            return candidate

        def verify_duty_number_popup(candidate):
            nonlocal popup_main_window
            popup_main_window = ""
            report_stage("duty_number_popup_preflight")
            popup_main_window = preflight_duty_number_popup(
                candidate,
                WebDriverWait(candidate, 20),
                daily_standby_numbers,
                excluded_numbers,
            )

        driver = retry_duty_number_popup_preflight(
            open_duty_browser_for_popup_preflight,
            verify_duty_number_popup,
            cleanup=quit_driver,
        )
        wait = WebDriverWait(driver, 20)
        if popup_main_window:
            report_stage("duty_number_popup_config")
            step_config_popups(
                driver,
                wait,
                out_names,
                daily_commander,
                popup_main_window,
            )

            wait_for_duty_result_grid(
                driver,
                WebDriverWait(driver, 60, poll_frequency=0.5),
            )

            if super_js_execute(driver, "_btnDelete", "exists"):
                report_stage("duty_existing_data_delete")
                super_js_execute(driver, "_btnDelete", "click")
                wait.until(EC.alert_is_present())
                driver.switch_to.alert.accept()
                time.sleep(3)
                log_status("✅ 舊資料已刪除")
                wait_for_duty_result_grid(
                    driver,
                    WebDriverWait(driver, 60, poll_frequency=0.5),
                )
            
            log_status("🧠 勤務基準表計算中...")
            
            # --- 收集基準表 24H 購物車 ---
            duty_map = {}
            for r in range(10, 34):
                time_cell = str(sheet.cell(row=r, column=ex_map["時間"]).value or "").strip()
                if "-" not in time_cell: continue
                hour = str(int(time_cell.split("-")[0].strip()))
                
                duty_map[f"_pln_{hour}_{web_idx['值班']}"] = clean_v_excluding(get_merged_val(sheet, r, ex_map["值班"]), excluded_numbers)
                for i, col_idx in enumerate(out_excel_cols):
                    duty_map[f"_pln_{hour}_{web_idx['外勤開始'] + i}"] = clean_v_excluding(get_merged_val(sheet, r, col_idx), excluded_numbers)
                
                med_v = clean_v_excluding(get_merged_val(sheet, r, ex_map["救護_Excel"]), excluded_numbers) + "," + clean_v_excluding(get_merged_val(sheet, r, ex_map["救護_Excel"]+1), excluded_numbers)
                duty_map[f"_pln_{hour}_{web_idx['救護']}"] = med_v.strip(',')
                
                dis_v = ""
                for c in range(ex_map["備勤_Excel"] + 1, ex_map["指揮官"] + 1):
                    val = clean_v_excluding(get_merged_val(sheet, r, c), excluded_numbers)
                    if val: dis_v += str(val) + ","
                duty_map[f"_pln_{hour}_{web_idx['備勤']}"] = dis_v.strip(',')
                duty_map[f"_pln_{hour}_{web_idx['休息']}"] = clean_v_excluding(get_merged_val(sheet, r, ex_map["休息_Excel"]), excluded_numbers)

            team_tra_val = str(sheet["C35"].value or "").strip()
            remark_val = str(sheet["C34"].value or "").strip()
            if team_tra_val: duty_map["_areTeamTra"] = team_tra_val
            if remark_val: duty_map["_arePSREMARK"] = remark_val

            log_status(f"✅ 勤務基準表運算完畢，共填入 {len(duty_map)} 格")

            # --- 填入基準表並儲存 ---
            report_stage("duty_fill")
            step_batch_fill_duty(driver, duty_map)
            time.sleep(1)
            report_stage("duty_save")
            driver.execute_script("""
                function clickSave(win) {
                    var btn = win.document.getElementById('_btnSave');
                    if(btn) { btn.click(); return true; }
                    for(var i=0; i<win.frames.length; i++) {
                        try { if(clickSave(win.frames[i])) return true; } catch(e){}
                    } return false;
                } clickSave(window.top);
            """)

            for _ in range(3): 
                try:
                    wait_alert = WebDriverWait(driver, 3)
                    wait_alert.until(EC.alert_is_present())
                    driver.switch_to.alert.accept()
                    time.sleep(1)
                except (NoAlertPresentException, TimeoutException):
                    break
            time.sleep(2)
            
            # --- 進入救災任務編組表 ---
            report_stage("vehicle_form_open")
            step_navigate_to_task_table(driver, wait) 
            driver.switch_to.default_content()
            wait.until(EC.frame_to_be_available_and_switch_to_it("ehrFrame"))
            wait.until(EC.frame_to_be_available_and_switch_to_it("contentFrame"))
            
            super_js_execute(driver, "_txtTaskDate", "set", target_date)
            super_js_execute(driver, "_btnQuery", "click")
            time.sleep(2)
            
            js_click_car = "function findAndClickBtn(win, id) { var btn = win.document.getElementById(id); if (btn) { btn.click(); return true; } for (var i = 0; i < win.frames.length; i++) { try { if (findAndClickBtn(win.frames[i], id)) return true; } catch(e) {} } return false; } return findAndClickBtn(window.top, '_btnOpenWinTaskCode');"
            if driver.execute_script(js_click_car):
                step_select_vehicles_popup(driver, wait, driver.current_window_handle, cars_config)
            
            log_status("➡️ 等待救災任務編組表載入...")
            driver.switch_to.default_content()
            
            # 🌟 同樣回歸 JS 大法
            for _ in range(15):
                if super_js_execute(driver, "_pln_8_1", "exists"):
                    time.sleep(1)
                    break
                time.sleep(1)
            else:
                log_status("⚠️ 等待表格載入較久，將強制繼續執行")
            
            log_status("🧠 救災任務編組計算中...")
            
            # --- 收集編組 24H 購物車 ---
            mission_map = {}
            for r in range(10, 34):
                time_cell = str(sheet.cell(row=r, column=ex_map["時間"]).value or "").strip()
                if "-" not in time_cell: continue
                hour = str(int(time_cell.split("-")[0].strip()))
                
                amb1_members = []
                for c in range(ex_map["救護_Excel"], ex_map["備勤_Excel"]):
                    amb1_members.extend(clean_to_list_excluding(get_merged_val(sheet, r, c), excluded_numbers))
                amb1_members = unique_member_ids(amb1_members)[:2]
                
                disaster_ids = []
                for c in range(ex_map["備勤_Excel"] + 1, ex_map["指揮官"] + 1):
                    disaster_ids.extend(clean_to_list_excluding(get_merged_val(sheet, r, c), excluded_numbers))
                disaster_ids = unique_member_ids(disaster_ids)
                commander_ids = clean_to_list_excluding(
                    get_merged_val(sheet, r, ex_map["指揮官"]), excluded_numbers
                )
                row_commander = commander_ids[0] if commander_ids else ""
                out_ids = []
                for col_idx in out_excel_cols:
                    out_ids.extend(clean_to_list_excluding(get_merged_val(sheet, r, col_idx), excluded_numbers))
                out_ids = unique_member_ids(out_ids)
                amb2_members = select_ambulance2_members(disaster_ids, out_ids, amb1_members)

                mission = calculate_fire_mission(amb1_members, disaster_ids, out_ids, row_commander)
                if mission:
                    mission_map[f"_pln_{hour}_1"] = mission['attack']
                    mission_map[f"_pln_{hour}_2"] = mission['relay']
                    mission_map[f"_pln_{hour}_6"] = ",".join(amb1_members)
                    mission_map[f"_pln_{hour}_7"] = ",".join(amb2_members)

            if not mission_map:
                raise RuntimeError("救災任務編組表沒有可登打資料，已停止完成流程。")
            log_status(f"✅ 救災任務編組運算完畢，共填入 {len(mission_map)} 格")
            time.sleep(5)
            driver.switch_to.default_content()
            for frame_name in ['main', 'Content', 'contents', 'ehrFrame', 'contentFrame']:
                try:
                    driver.switch_to.frame(frame_name)
                except NoSuchFrameException:
                    continue

            report_stage("vehicle_fill")
            mission_result = step_fill_mission_cells(driver, mission_map)
            alert_messages = accept_pending_alerts(driver, timeout=2)
            for message in alert_messages:
                log_status(f"⚠️ 救災任務編組表提示：{message}")
            if mission_result.get("status") not in ("saved", "alert"):
                raise RuntimeError("救災任務編組表未找到完整欄位或儲存按鈕，已停止完成流程。")
            verification_issues = verify_mission_cells(driver, mission_map)
            if verification_issues:
                detail = "\n".join(verification_issues[:20])
                raise RuntimeError(f"救災任務編組表儲存後資料驗證失敗，已停止完成流程。\n{detail}")
            log_status("✅ 救災任務編組表儲存後資料驗證通過")
            
            report_stage("report")
            notification_status = ""
            notification_config = load_config().get("notification", {})
            try:
                log_status("等待勤務表截圖完成...")
                if capture_future is None:
                    daily_preview, night_preview = capture_duty_sheet_images(excel_path, target_date)
                else:
                    daily_preview, night_preview = capture_future.result()
                image_path = daily_preview["image_path"]
                log_status(f"勤務表截圖完成：{daily_preview['capture_range']}")
                log_status(f"夜間勤務截圖完成：{night_preview['capture_range']}")
                if notification_config.get("enabled"):
                    log_status("開始上傳截圖並發送 LINE 通知...")
                    notification_result = send_group_notification(
                        image_path,
                        target_date,
                        notification_config
                    )
                    notification_status = "\n勤務表截圖已完成，並已發送 LINE 通知。"
                    log_status(
                        f"{notification_result['provider']} 通知已送出，共 {len(notification_result['image_urls'])} 張圖片"
                    )
                else:
                    notification_status = "\n勤務表截圖已完成。"
            except Exception as notify_error:
                notification_status = f"\n勤務表截圖或 LINE 通知失敗：{notify_error}"
                log_status(f"勤務表通知失敗：{notify_error}")

            # 計算總花費秒數
            end_time = time.time()
            elapsed_time = round(end_time - start_time, 1)
            
            log_status(f"🎉 全部完成！耗時 {elapsed_time} 秒")
            
            # 將秒數加入到最後的彈出視窗中
            success_msg = f"已登打並存檔完畢！{notification_status}\n本次自動化總共花費：{elapsed_time} 秒\n請回網頁做最後的複查。"
            if callable(success_callback):
                success_callback(success_msg)
            elif show_dialogs and "root" in globals():
                root.after(0, lambda: messagebox.showinfo("成功", success_msg))
            return True

        error_message = "勤務表頁面準備失敗，未執行登打。"
        log_status(error_message)
        if callable(error_callback):
            error_callback(error_message)
        return False

    except Exception as e:
        log_status(f"❌ 流程中斷：{e}")
        if callable(error_callback):
            error_callback(str(e))
        if raise_errors:
            raise
        return False
    finally:
        wb.close()
        if capture_executor is not None:
            capture_executor.shutdown(
                wait=bool(capture_future is not None and capture_future.done()),
                cancel_futures=True,
            )
        if close_driver:
            try:
                quit_driver(driver)
            except Exception:
                pass
        _runtime_status_callback = previous_status_callback


# ==========================================
# [區塊七] GUI 使用者介面 (Tkinter Setup)
# ==========================================

# 7-1. GUI 事件處理
def browse_file():
    f = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xlsm")])
    entry_file.delete(0, tk.END); entry_file.insert(0, f)

def on_submit():
    uid, pwd, f_path = entry_id.get(), entry_pwd.get(), entry_file.get()
    m_date = convert_to_minguo(cal.get_date())
    
    login_config = {
        "user_id": uid,
        "user_pwd": pwd
    }
    notification_config = load_config().get("notification", get_default_config()["notification"]).copy()
    notification_config["enabled"] = bool(send_group_var.get())
    cars_config = {
        'attack': attack_car_var.get(),
        'stop': stop_car_var.get(),
        'amb1': amb1_car_var.get(),
        'amb2': amb2_car_var.get()
    }
    # 按下啟動時，自動記憶這次選了什麼
    save_config(
        cars_config,
        login_settings=login_config,
        notification_settings=notification_config,
        car_options=opts,
        hidden_car_options=hidden_opts
    )
    
    if not f_path: 
        messagebox.showwarning("提示", "請選擇 Excel 檔案！")
        return
    
    #  防止連點，在執行期間鎖死按鈕
    btn_submit.config(state="disabled", text="⏳ 執行中，請稍候...")

    #  建立一個背景執行緒來跑主流程，避免視窗卡死
    def run_task():
        try:
            start_automation(uid, pwd, m_date, f_path, cars_config)
        finally:
            # 結束後把按鈕恢復原狀
            root.after(0, lambda: btn_submit.config(state="normal", text="⚡ 啟動全自動流程"))

    # 啟動執行緒 (daemon=True 代表關閉視窗時背景也會強制結束)
    threading.Thread(target=run_task, daemon=True).start()

# 7-2. GUI 初始化與畫面配置
if __name__ == "__main__":
    import tkinter as tk
    from tkinter import filedialog, messagebox, scrolledtext, ttk

    from tkcalendar import DateEntry

    # 🌟 1. 先讀取設定檔
    current_config = load_config()
    login = current_config["login"]
    last = current_config["last_selection"]
    opts = current_config["car_options"]
    hidden_opts = current_config["hidden_car_options"]

    root = tk.Tk()
    root.title("🚒 新坡全自動勤務分配表及救災任務編組表V2.0")
    # 稍微加寬拉長，給予元件足夠的呼吸空間
    root.geometry("450x800") 
    # 禁止使用者隨意縮放視窗導致跑版
    root.resizable(False, False) 

    # 設定全局字體，讓中文字體顯示更美觀
    default_font = ("微軟正黑體", 10)
    title_font = ("微軟正黑體", 14, "bold")
    root.option_add("*Font", default_font)

    # 設定按鈕的進階樣式
    style = ttk.Style()
    style.configure("TButton", font=("微軟正黑體", 10), padding=3)
    style.configure("Action.TButton", font=("微軟正黑體", 12, "bold"), padding=10)

    # 主容器：給予邊界留白
    main_frame = ttk.Frame(root, padding="20 15 20 15")
    main_frame.pack(fill="both", expand=True)

    # --- 頂部大標題 ---
    ttk.Label(main_frame, text="🚒 新坡分隊關心您的眼睛", font=title_font, anchor="center").pack(fill="x", pady=(0, 15))

    # ==========================================
    # 區塊 1：系統登入資訊
    # ==========================================
    frame_login = ttk.LabelFrame(main_frame, text="👤 登入資訊", padding="10 10 10 10")
    frame_login.pack(fill="x", pady=5)
    
    # 使用 grid 排版：標籤靠右 (sticky="e")，輸入框靠左 (sticky="w")
    ttk.Label(frame_login, text="系統帳號:").grid(row=0, column=0, sticky="e", padx=5, pady=6)
    entry_id = ttk.Entry(frame_login, width=32)
    entry_id.insert(0, login.get("user_id", ""))
    entry_id.grid(row=0, column=1, sticky="w", padx=5, pady=6)
    
    ttk.Label(frame_login, text="系統密碼:").grid(row=1, column=0, sticky="e", padx=5, pady=6)
    entry_pwd = ttk.Entry(frame_login, width=32, show="*")
    entry_pwd.insert(0, login.get("user_pwd", ""))
    entry_pwd.grid(row=1, column=1, sticky="w", padx=5, pady=6)

    # ==========================================
    # 區塊 2：班表與日期設定
    # ==========================================
    frame_file = ttk.LabelFrame(main_frame, text="📅 班表資料", padding="10 10 10 10")
    frame_file.pack(fill="x", pady=10)

    ttk.Label(frame_file, text="Excel 路徑:").grid(row=0, column=0, sticky="e", padx=5, pady=6)
    
    # 將輸入框與瀏覽按鈕包在一個子框架內，讓它們並排
    file_subframe = ttk.Frame(frame_file)
    file_subframe.grid(row=0, column=1, sticky="w", padx=5, pady=6)
    entry_file = ttk.Entry(file_subframe, width=22)
    entry_file.pack(side="left", padx=(0, 5))
    ttk.Button(file_subframe, text="📁 瀏覽", command=browse_file, width=8).pack(side="left")

    ttk.Label(frame_file, text="班表日期:").grid(row=1, column=0, sticky="e", padx=5, pady=6)
    
    tomorrow = datetime.now() + timedelta(days=1)
    cal = DateEntry(frame_file, width=30, background='darkblue', foreground='white', borderwidth=2, 
                    year=tomorrow.year, month=tomorrow.month, day=tomorrow.day, 
                    date_pattern='yyyy/mm/dd')
    cal.grid(row=1, column=1, sticky="w", padx=5, pady=6)

    send_group_var = tk.BooleanVar(value=current_config.get("notification", {}).get("enabled", True))
    ttk.Checkbutton(
        frame_file,
        text="是否傳送勤務表截圖至值班台",
        variable=send_group_var
    ).grid(row=2, column=1, sticky="w", padx=5, pady=6)

    # ==========================================
    # 區塊 3：主力車設定
    # ==========================================
    frame_car = ttk.LabelFrame(main_frame, text="🚚 主力車設定", padding="10 10 10 10")
    frame_car.pack(fill="x", pady=5)

    # 統一寬度為 32，確保上下對齊
    combo_width = 30 
    
    # 🌟 2. 修改 Combobox 的預設值與選項清單
    tk.Label(frame_car, text="攻擊車:").grid(row=0, column=0, sticky="e", padx=5, pady=6)
    attack_car_var = tk.StringVar(value=last['attack']) # 預設選上次的
    attack_combo = ttk.Combobox(frame_car, textvariable=attack_car_var, values=opts['attack'], width=combo_width)
    attack_combo.grid(row=0, column=1, sticky="w", padx=5, pady=6)
    
    tk.Label(frame_car, text="中繼車:").grid(row=1, column=0, sticky="e", padx=5, pady=6)
    stop_car_var = tk.StringVar(value=last['stop'])
    stop_combo = ttk.Combobox(frame_car, textvariable=stop_car_var, values=opts['stop'], width=combo_width)
    stop_combo.grid(row=1, column=1, sticky="w", padx=5, pady=6)
    
    tk.Label(frame_car, text="救護 1 車:").grid(row=2, column=0, sticky="e", padx=5, pady=6)
    amb1_car_var = tk.StringVar(value=last['amb1'])
    amb1_combo = ttk.Combobox(frame_car, textvariable=amb1_car_var, values=opts['amb'], width=combo_width)
    amb1_combo.grid(row=2, column=1, sticky="w", padx=5, pady=6)
    
    tk.Label(frame_car, text="救護 2 車:").grid(row=3, column=0, sticky="e", padx=5, pady=6)
    amb2_car_var = tk.StringVar(value=last['amb2'])
    amb2_combo = ttk.Combobox(frame_car, textvariable=amb2_car_var, values=opts['amb'], width=combo_width)
    amb2_combo.grid(row=3, column=1, sticky="w", padx=5, pady=6)

    def persist_car_options():
        login_config = {
            "user_id": entry_user.get().strip(),
            "user_pwd": entry_pwd.get()
        }
        notification_config = load_config().get("notification", get_default_config()["notification"]).copy()
        notification_config["enabled"] = bool(send_group_var.get())
        cars_config = {
            'attack': attack_car_var.get(),
            'stop': stop_car_var.get(),
            'amb1': amb1_car_var.get(),
            'amb2': amb2_car_var.get()
        }
        save_config(
            cars_config,
            login_settings=login_config,
            notification_settings=notification_config,
            car_options=opts,
            hidden_car_options=hidden_opts
        )

    vehicle_groups = {
        "消防車": "attack",
        "救護車": "amb"
    }

    def refresh_vehicle_options():
        attack_combo["values"] = opts.get("attack", [])
        amb_values = opts.get("amb", [])
        amb1_combo["values"] = amb_values
        amb2_combo["values"] = amb_values

    def open_add_vehicle_dialog():
        result = {}
        dialog = tk.Toplevel(root)
        dialog.title("新增車輛")
        dialog.transient(root)
        dialog.grab_set()
        dialog.resizable(False, False)

        vehicle_type_var = tk.StringVar(value="救護車")
        code_var = tk.StringVar()
        plate_var = tk.StringVar()

        ttk.Label(dialog, text="車輛類型").grid(row=0, column=0, sticky="e", padx=10, pady=(12, 6))
        type_combo = ttk.Combobox(dialog, textvariable=vehicle_type_var, values=list(vehicle_groups.keys()), state="readonly", width=18)
        type_combo.grid(row=0, column=1, sticky="w", padx=10, pady=(12, 6))

        ttk.Label(dialog, text="車輛代號").grid(row=1, column=0, sticky="e", padx=10, pady=6)
        code_entry = ttk.Entry(dialog, textvariable=code_var, width=22)
        code_entry.grid(row=1, column=1, sticky="w", padx=10, pady=6)

        ttk.Label(dialog, text="車牌號碼").grid(row=2, column=0, sticky="e", padx=10, pady=6)
        plate_entry = ttk.Entry(dialog, textvariable=plate_var, width=22)
        plate_entry.grid(row=2, column=1, sticky="w", padx=10, pady=6)

        button_frame = ttk.Frame(dialog)
        button_frame.grid(row=3, column=0, columnspan=2, sticky="e", padx=10, pady=(8, 12))

        def confirm():
            vehicle_type = vehicle_type_var.get().strip()
            code = code_var.get().strip()
            plate = plate_var.get().strip()
            if not code or not plate:
                messagebox.showwarning("資料不足", "請輸入車輛代號與車牌號碼。", parent=dialog)
                return
            result["type"] = vehicle_type
            result["value"] = f"{code}/{plate}"
            dialog.destroy()

        ttk.Button(button_frame, text="確定", command=confirm).pack(side="left", padx=(0, 6))
        ttk.Button(button_frame, text="取消", command=dialog.destroy).pack(side="left")
        code_entry.focus_set()
        root.wait_window(dialog)
        return result

    def add_vehicle_option():
        result = open_add_vehicle_dialog()
        if not result:
            return
        group = vehicle_groups[result["type"]]
        value = result["value"]
        options = opts.setdefault(group, [])
        hidden_values = hidden_opts.setdefault(group, [])
        if value in hidden_values:
            hidden_values.remove(value)
        if value not in options:
            options.append(value)
        refresh_vehicle_options()
        persist_car_options()
        messagebox.showinfo("已新增", f"已加入{result['type']}選項：{value}", parent=root)

    def open_remove_vehicle_dialog():
        choices = []
        choice_map = {}
        for vehicle_type, group in vehicle_groups.items():
            for value in opts.get(group, []):
                label = f"{vehicle_type} {value}"
                choices.append(label)
                choice_map[label] = (vehicle_type, group, value)
        if not choices:
            messagebox.showwarning("沒有車輛", "目前沒有可移除的車輛。", parent=root)
            return None

        result = {}
        dialog = tk.Toplevel(root)
        dialog.title("移除車輛")
        dialog.transient(root)
        dialog.grab_set()
        dialog.resizable(False, False)

        selected_var = tk.StringVar(value=choices[0])
        ttk.Label(dialog, text="車輛代號/車牌號碼").grid(row=0, column=0, sticky="e", padx=10, pady=(12, 6))
        select_combo = ttk.Combobox(dialog, textvariable=selected_var, values=choices, state="readonly", width=34)
        select_combo.grid(row=0, column=1, sticky="w", padx=10, pady=(12, 6))

        button_frame = ttk.Frame(dialog)
        button_frame.grid(row=1, column=0, columnspan=2, sticky="e", padx=10, pady=(8, 12))

        def confirm():
            selected = selected_var.get()
            if selected in choice_map:
                result["vehicle"] = choice_map[selected]
            dialog.destroy()

        ttk.Button(button_frame, text="確定", command=confirm).pack(side="left", padx=(0, 6))
        ttk.Button(button_frame, text="取消", command=dialog.destroy).pack(side="left")
        select_combo.focus_set()
        root.wait_window(dialog)
        return result.get("vehicle")

    def remove_vehicle_option():
        selected = open_remove_vehicle_dialog()
        if not selected:
            return
        vehicle_type, group, value = selected
        options = opts.setdefault(group, [])
        if value not in options:
            messagebox.showwarning("找不到車輛", f"車輛清單中沒有：{value}", parent=root)
            return
        options.remove(value)
        hidden_values = hidden_opts.setdefault(group, [])
        if value not in hidden_values:
            hidden_values.append(value)
        fallback = options[0] if options else ""
        if group == "attack" and attack_car_var.get().strip() == value:
            attack_car_var.set(fallback)
        if group == "amb":
            if amb1_car_var.get().strip() == value:
                amb1_car_var.set(fallback)
            if amb2_car_var.get().strip() == value:
                amb2_car_var.set(fallback)
        refresh_vehicle_options()
        persist_car_options()
        messagebox.showinfo("已移除", f"已從{vehicle_type}選項移除：{value}", parent=root)

    vehicle_button_frame = ttk.Frame(frame_car)
    vehicle_button_frame.grid(row=4, column=1, sticky="w", padx=5, pady=(0, 6))
    ttk.Button(vehicle_button_frame, text="新增車輛", command=add_vehicle_option).pack(side="left", padx=(0, 6))
    ttk.Button(vehicle_button_frame, text="移除車輛", command=remove_vehicle_option).pack(side="left")

    # ==========================================
    # 區塊 4：執行
    # ==========================================
    action_frame = ttk.Frame(main_frame)
    action_frame.pack(fill="x", pady=(20, 0), padx=15)

    # fill="x" 讓按鈕填滿寬度，更有視覺焦點
    btn_submit = ttk.Button(action_frame, text="⚡ 啟動全自動流程", command=on_submit, style="Action.TButton")
    btn_submit.pack(fill="x")

    log_frame = ttk.LabelFrame(main_frame, text="執行紀錄", padding="10 10 10 10")
    log_frame.pack(fill="both", expand=True, pady=(15, 0))

    log_text = scrolledtext.ScrolledText(log_frame, height=10, wrap="word", state="normal")
    log_text.pack(fill="both", expand=True)
    log_text.insert(tk.END, "準備就緒\n")
    
    # ==========================================
    # 底部狀態列 (Status Bar)
    # ==========================================
    status_var = tk.StringVar(value="狀態: 準備就緒")
    status_bar = ttk.Label(root, textvariable=status_var, relief="sunken", anchor="w", padding=5)
    status_bar.pack(side="bottom", fill="x")

    root.mainloop()
