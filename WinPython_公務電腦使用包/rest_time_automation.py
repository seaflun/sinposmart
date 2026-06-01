# -*- coding: utf-8 -*-
"""Duty-base table helpers for rest-time and monthly leave entry workflows."""

from __future__ import annotations

import json
import csv
import io
import re
import threading
import time
import traceback
from dataclasses import dataclass
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
import tkinter as tk
from typing import Callable
from urllib.parse import urlencode
from urllib.request import urlopen
from urllib.error import HTTPError, URLError

import openpyxl
from selenium.common.exceptions import TimeoutException, UnexpectedAlertPresentException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

from duty_rehearsal import build_driver, js_click, login, open_ap


DUTY_BASE_AP = "wap119.RPS105010"
REST_TIME_CONFIG = Path(__file__).resolve().with_name("rest_time_automation_config.json")
RETAINED_DRIVERS: list[object] = []
MONTHLY_BASE_SHEET_ID = "1m-zy4KNR8_GMO94dYtFotyWPIvuT_tt32J9l7hhGZt0"
MONTHLY_BASE_SHEET_GID = "1587057625"
MONTHLY_BASE_EXPORT_URL = "https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?{query}"
MONTHLY_BASE_SYMBOLS = {
    "": "",
    "休": "○",
    "月": "⊙",
    "請": "△",
    "補": "補",
    "超": "補",
    "公": "公",
    "喪": "喪",
    "心": "❤",
}


@dataclass(frozen=True)
class RestEntry:
    duty_day: int
    start_day: int
    start_hour: int
    end_day: int
    end_hour: int

    @property
    def hours(self) -> int:
        start_abs = self.start_day * 24 + self.start_hour
        end_abs = self.end_day * 24 + self.end_hour
        return end_abs - start_abs

    def summary(self) -> str:
        return f"勤務{self.duty_day:02d}日：{self.start_day:02d}日 {self.start_hour:02d}:00-{self.end_day:02d}日 {self.end_hour:02d}:00，共 {self.hours} 小時"


@dataclass(frozen=True)
class PersonLink:
    name: str
    staff_no: str
    element_index: int


@dataclass(frozen=True)
class MonthlyBasePlan:
    roc_year: int
    month: int
    actor_no: str
    name: str
    day_symbols: dict[int, str]

    @property
    def days(self) -> int:
        return max(self.day_symbols, default=0)

    def filled_days(self) -> int:
        return sum(1 for value in self.day_symbols.values() if value)


def format_automation_error(exc: Exception) -> str:
    text = str(exc).strip()
    if text:
        return text
    details = "".join(traceback.format_exception_only(type(exc), exc)).strip()
    if details:
        return details
    return exc.__class__.__name__


def open_rest_time_dialog(parent: tk.Tk, user_id: str = "", password: str = "", actor_no: str = "", display_name: str = "") -> tk.Toplevel | None:
    existing = getattr(parent, "_rest_time_dialog", None)
    if existing is not None:
        try:
            if existing.winfo_exists():
                existing.deiconify()
                existing.lift()
                existing.focus_force()
                return existing
        except Exception:
            pass
        setattr(parent, "_rest_time_dialog", None)

    dialog = tk.Toplevel(parent)
    setattr(parent, "_rest_time_dialog", dialog)
    dialog.title("SinpoSmart - 休息時間登打")
    dialog.geometry("430x300")
    dialog.minsize(430, 300)
    dialog.configure(bg="#f8fafc")
    dialog.transient(parent)

    def close_dialog() -> None:
        setattr(parent, "_rest_time_dialog", None)
        dialog.destroy()

    dialog.protocol("WM_DELETE_WINDOW", close_dialog)

    root = tk.Frame(dialog, bg="#f8fafc")
    root.pack(fill=tk.BOTH, expand=True)

    header = tk.Frame(root, bg="#eff6ff", highlightbackground="#bfdbfe", highlightthickness=1)
    header.pack(fill=tk.X, padx=10, pady=(10, 0))
    tk.Label(header, text="休息時間登打", bg="#eff6ff", fg="#1e3a8a", font=("Microsoft JhengHei", 11, "bold")).pack(anchor=tk.W, padx=12, pady=(10, 10))

    body = tk.Frame(root, bg="#f8fafc")
    body.pack(fill=tk.BOTH, expand=True, padx=10, pady=(8, 10))

    form = ttk.LabelFrame(body, text="勤務表檔案", padding=8)
    form.pack(fill=tk.X, pady=(10, 8))
    form.columnconfigure(1, weight=1)

    file_var = tk.StringVar(value=str(default_workbook_path()))
    status_var = tk.StringVar(value=f"準備就緒。{display_name or actor_no or user_id}")

    ttk.Label(form, text="Excel").grid(row=0, column=0, sticky=tk.W, padx=(0, 8), pady=4)
    file_entry = ttk.Entry(form, textvariable=file_var)
    file_entry.grid(row=0, column=1, sticky=tk.EW, pady=4)

    def browse_file() -> None:
        current_file = Path(file_var.get().strip())
        initial_dir = current_file.parent if current_file.parent.exists() else Path(__file__).resolve().parent
        path = filedialog.askopenfilename(parent=dialog, filetypes=[("Excel files", "*.xlsx *.xlsm")], initialdir=str(initial_dir))
        if path:
            file_var.set(path)
            save_last_workbook_path(Path(path))
            status_var.set("已選擇勤務表 Excel。")

    def bind_button_hover(button: tk.Button, normal_bg: str, hover_bg: str) -> None:
        button.bind("<Enter>", lambda _event: button.configure(bg=hover_bg))
        button.bind("<Leave>", lambda _event: button.configure(bg=normal_bg))

    browse_button = tk.Button(
        form,
        text="選擇",
        command=browse_file,
        bg="#2563eb",
        fg="#ffffff",
        activebackground="#1d4ed8",
        activeforeground="#ffffff",
        relief=tk.FLAT,
        width=5,
    )
    bind_button_hover(browse_button, "#2563eb", "#1d4ed8")
    browse_button.grid(row=0, column=2, sticky=tk.E, padx=(8, 0), pady=4)

    action_row = tk.Frame(body, bg="#f8fafc")
    action_row.pack(fill=tk.X, pady=(8, 8))
    action_row.columnconfigure(0, weight=1)

    status_bar = ttk.Label(body, textvariable=status_var, relief=tk.SUNKEN, anchor=tk.W, padding=5)
    status_bar.pack(side=tk.BOTTOM, fill=tk.X)

    def set_running(running: bool) -> None:
        start_button.configure(state=tk.DISABLED if running else tk.NORMAL, text="登打中..." if running else "啟動登打")

    def set_status(message: str) -> None:
        dialog.after(0, lambda: status_var.set(message))

    def run_automation() -> None:
        uid = user_id.strip()
        pwd = password
        workbook_path = Path(file_var.get().strip())
        if not uid or not pwd:
            messagebox.showwarning("缺少帳號密碼", "請先在主視窗登入，再啟動休息時間登打。", parent=dialog)
            return
        if not workbook_path.exists():
            messagebox.showwarning("找不到 Excel", "請選擇勤務表 Excel 檔案。", parent=dialog)
            return
        save_last_workbook_path(workbook_path)
        set_running(True)
        set_status("開啟瀏覽器登打休息時間...")

        def worker() -> None:
            try:
                result = submit_rest_entries(uid, pwd, workbook_path, False, set_status, keep_browser_open=True, actor_no=actor_no)
                dialog.after(0, lambda: show_complete_and_close(result))
                set_status(result)
            except Exception as exc:
                error = str(exc)
                dialog.after(0, lambda: messagebox.showerror("休息時間登打失敗", error, parent=dialog))
                set_status(f"失敗：{error}")
            finally:
                dialog.after(0, lambda: set_running(False))

        threading.Thread(target=worker, daemon=True).start()

    def show_complete_and_close(result: str) -> None:
        messagebox.showinfo("完成", result, parent=dialog)
        close_dialog()

    start_button = tk.Button(
        action_row,
        text="啟動登打",
        command=run_automation,
        bg="#16a34a",
        fg="#ffffff",
        activebackground="#15803d",
        activeforeground="#ffffff",
        relief=tk.FLAT,
        font=("Microsoft JhengHei", 11, "bold"),
        height=2,
    )
    bind_button_hover(start_button, "#16a34a", "#15803d")
    start_button.grid(row=0, column=0, sticky=tk.EW, padx=(0, 8))
    close_button = tk.Button(
        action_row,
        text="關閉",
        command=close_dialog,
        bg="#e2e8f0",
        fg="#0f172a",
        activebackground="#cbd5e1",
        activeforeground="#0f172a",
        relief=tk.FLAT,
        font=("Microsoft JhengHei", 11, "bold"),
        width=8,
        height=2,
    )
    bind_button_hover(close_button, "#e2e8f0", "#cbd5e1")
    close_button.grid(row=0, column=1, sticky=tk.E)
    return dialog


def open_monthly_base_dialog(parent: tk.Tk, user_id: str = "", password: str = "", actor_no: str = "", display_name: str = "") -> tk.Toplevel | None:
    existing = getattr(parent, "_monthly_base_dialog", None)
    if existing is not None:
        try:
            if existing.winfo_exists():
                existing.deiconify()
                existing.lift()
                existing.focus_force()
                return existing
        except Exception:
            pass
        setattr(parent, "_monthly_base_dialog", None)

    dialog = tk.Toplevel(parent)
    setattr(parent, "_monthly_base_dialog", dialog)
    dialog.title("SinpoSmart - 勤務基準表登打")
    dialog.geometry("430x280")
    dialog.minsize(430, 280)
    dialog.configure(bg="#f8fafc")
    dialog.transient(parent)

    def close_dialog() -> None:
        setattr(parent, "_monthly_base_dialog", None)
        dialog.destroy()

    dialog.protocol("WM_DELETE_WINDOW", close_dialog)

    root = tk.Frame(dialog, bg="#f8fafc")
    root.pack(fill=tk.BOTH, expand=True)

    header = tk.Frame(root, bg="#eff6ff", highlightbackground="#bfdbfe", highlightthickness=1)
    header.pack(fill=tk.X, padx=10, pady=(10, 0))
    tk.Label(header, text="勤務基準表登打", bg="#eff6ff", fg="#1e3a8a", font=("Microsoft JhengHei", 11, "bold")).pack(anchor=tk.W, padx=12, pady=(10, 10))

    body = tk.Frame(root, bg="#f8fafc")
    body.pack(fill=tk.BOTH, expand=True, padx=10, pady=(8, 10))

    info = ttk.LabelFrame(body, text="固定來源", padding=8)
    info.pack(fill=tk.X, pady=(10, 8))
    info.columnconfigure(0, weight=1)
    ttk.Label(info, text=f"Google 試算表 / 輪休基準表  {display_name or actor_no or user_id}", justify=tk.LEFT).grid(row=0, column=0, sticky=tk.W)

    action_row = tk.Frame(body, bg="#f8fafc")
    action_row.pack(fill=tk.X, pady=(8, 8))
    action_row.columnconfigure(0, weight=1)

    status_var = tk.StringVar(value=f"準備就緒。{display_name or actor_no or user_id}")
    status_bar = ttk.Label(body, textvariable=status_var, relief=tk.SUNKEN, anchor=tk.W, padding=5)
    status_bar.pack(side=tk.BOTTOM, fill=tk.X)

    def bind_button_hover(button: tk.Button, normal_bg: str, hover_bg: str) -> None:
        button.bind("<Enter>", lambda _event: button.configure(bg=hover_bg))
        button.bind("<Leave>", lambda _event: button.configure(bg=normal_bg))

    def set_running(running: bool) -> None:
        start_button.configure(state=tk.DISABLED if running else tk.NORMAL, text="登打中..." if running else "啟動登打")

    def set_status(message: str) -> None:
        dialog.after(0, lambda: status_var.set(message))

    def show_complete_and_close(result: str) -> None:
        messagebox.showinfo("完成", result, parent=dialog)
        close_dialog()

    def run_automation() -> None:
        uid = user_id.strip()
        pwd = password
        actor = actor_no.strip()
        if not uid or not pwd:
            messagebox.showwarning("缺少帳號密碼", "請先在主視窗登入，再啟動每月基準表登打。", parent=dialog)
            return
        if not actor:
            messagebox.showwarning("缺少番號", "請先在主視窗確認番號，再啟動每月基準表登打。", parent=dialog)
            return
        set_running(True)
        set_status("讀取輪休基準表並開啟勤務基準表...")

        def worker() -> None:
            try:
                result = submit_monthly_base_entries(uid, pwd, actor, False, set_status, keep_browser_open=True)
                dialog.after(0, lambda: show_complete_and_close(result))
                set_status(result)
            except Exception as exc:
                error = format_automation_error(exc)
                dialog.after(0, lambda: messagebox.showerror("每月基準表登打失敗", error, parent=dialog))
                set_status(f"失敗：{error}")
            finally:
                dialog.after(0, lambda: set_running(False))

        threading.Thread(target=worker, daemon=True).start()

    start_button = tk.Button(
        action_row,
        text="啟動登打",
        command=run_automation,
        bg="#16a34a",
        fg="#ffffff",
        activebackground="#15803d",
        activeforeground="#ffffff",
        relief=tk.FLAT,
        font=("Microsoft JhengHei", 11, "bold"),
        height=2,
    )
    bind_button_hover(start_button, "#16a34a", "#15803d")
    start_button.grid(row=0, column=0, sticky=tk.EW, padx=(0, 8))
    close_button = tk.Button(
        action_row,
        text="關閉",
        command=close_dialog,
        bg="#e2e8f0",
        fg="#0f172a",
        activebackground="#cbd5e1",
        activeforeground="#0f172a",
        relief=tk.FLAT,
        font=("Microsoft JhengHei", 11, "bold"),
        width=8,
        height=2,
    )
    bind_button_hover(close_button, "#e2e8f0", "#cbd5e1")
    close_button.grid(row=0, column=1, sticky=tk.E)
    return dialog


def default_workbook_path() -> Path:
    saved = load_last_workbook_path()
    if saved and saved.exists():
        return saved
    base_dir = Path(__file__).resolve().parent
    workbooks = sorted(base_dir.glob("*.xlsm"), key=lambda item: item.stat().st_mtime, reverse=True)
    return workbooks[0] if workbooks else Path()


def load_last_workbook_path() -> Path | None:
    if not REST_TIME_CONFIG.exists():
        return None
    try:
        payload = json.loads(REST_TIME_CONFIG.read_text(encoding="utf-8"))
    except Exception:
        return None
    path = Path(str(payload.get("workbook_path", "") or ""))
    return path if str(path) else None


def save_last_workbook_path(path: Path) -> None:
    try:
        REST_TIME_CONFIG.write_text(json.dumps({"workbook_path": str(path)}, ensure_ascii=False, indent=2), encoding="utf-8")
    except OSError:
        pass


def submit_rest_entries(
    user_id: str,
    password: str,
    workbook_path: Path,
    headless: bool,
    status: Callable[[str], None] | None = None,
    keep_browser_open: bool = False,
    actor_no: str = "",
) -> str:
    status = status or (lambda _message: None)
    target_name = workbook_person_name(workbook_path, actor_no)
    driver = build_driver(headless=headless)
    inserted = 0
    skipped = 0
    deleted_duplicates = 0
    success = False
    try:
        login(driver, user_id, password)
        status("登入完成，開啟勤務基準表...")
        open_ap(driver, DUTY_BASE_AP)
        wait_for_main_table(driver)
        person = find_person_link(driver, user_id, target_no=actor_no, target_name=target_name)
        status(f"找到個人連結：{person.name}（系統儲存列 {person.staff_no}）")
        entries = parse_rest_entries(workbook_path, target_name=person.name, target_no=person.staff_no)
        if not entries:
            raise RuntimeError(f"勤務表內找不到 {person.name} 的休息時間。")
        open_person_popup(driver, person)
        deleted_duplicates = delete_duplicate_rest_rows(driver, status)
        for entry in entries:
            status(entry.summary())
            existing = existing_entry_match(driver, entry)
            if existing:
                skipped += 1
                status(f"已存在略過：{entry.summary()}（{existing}）")
                continue
            fill_and_insert_entry(driver, entry)
            inserted += 1
        close_current_popup(driver)
        click_person_save(driver, person.staff_no)
        if keep_browser_open:
            RETAINED_DRIVERS.append(driver)
        success = True
        return f"完成：新增 {inserted} 筆，略過已存在 {skipped} 筆，刪除重複休息 {deleted_duplicates} 筆，已按個人儲存。"
    finally:
        if not keep_browser_open or not success:
            driver.quit()


def submit_monthly_base_entries(
    user_id: str,
    password: str,
    actor_no: str,
    headless: bool,
    status: Callable[[str], None] | None = None,
    keep_browser_open: bool = False,
) -> str:
    status = status or (lambda _message: None)
    actor_no = str(actor_no or "").strip()
    plan = fetch_monthly_base_plan(actor_no)
    driver = build_driver(headless=headless)
    success = False
    try:
        login(driver, user_id, password)
        status("登入完成，開啟勤務基準表...")
        open_ap(driver, DUTY_BASE_AP)
        wait_for_main_table(driver)
        status(f"切換到 {plan.roc_year}年{plan.month:02d}月並查詢...")
        select_base_month(driver, plan.roc_year, plan.month)
        wait_for_person_name_row(driver, plan.name)
        status(f"找到本人列：{plan.name}（{plan.actor_no}番）")
        filled = fill_monthly_base_row(driver, plan.name, plan.day_symbols, plan.days)
        status(f"已填入 {filled} 格，按個人儲存...")
        click_person_row_save(driver, plan.name)
        if keep_browser_open:
            RETAINED_DRIVERS.append(driver)
        success = True
        return f"完成：{plan.roc_year}年{plan.month:02d}月 {plan.name} 已填入 {filled} 格並個人儲存。"
    finally:
        if not keep_browser_open or not success:
            driver.quit()


def delete_duplicate_rest_rows(driver, status: Callable[[str], None]) -> int:
    deleted = 0
    while True:
        duplicate = mark_duplicate_rest_delete(driver)
        if not duplicate:
            break
        if duplicate.startswith("找不到刪除按鈕"):
            raise RuntimeError(duplicate)
        delete_button = driver.find_element(By.CSS_SELECTOR, "[data-rest-duplicate-delete='1']")
        try:
            delete_button.click()
        except UnexpectedAlertPresentException:
            pass
        deleted += 1
        status(f"刪除重複休息：{duplicate}")
        wait_for_alert_if_any(driver)
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "_selOFFTYPE")))
        time.sleep(0.4)
        if deleted > 50:
            raise RuntimeError("重複休息刪除超過 50 筆，已停止避免誤刪。")
    return deleted


def parse_rest_entries(workbook_path: Path, target_name: str | None = None, target_no: str | None = None) -> list[RestEntry]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    _year, _month, days = workbook_date_info(wb)
    person_no = workbook_person_no(wb, target_name)
    if target_name and not person_no:
        raise RuntimeError(f"勤務表輪休表找不到 {target_name} 的番號，已停止避免用勤務系統列編號誤打。")
    person_no = person_no or target_no
    if not person_no:
        raise RuntimeError("無法從勤務表判斷登入者的編號。")
    hours_by_day: dict[int, list[tuple[int, int]]] = {}
    for day in range(1, days + 1):
        sheet_name = f"{day}號"
        if sheet_name not in wb.sheetnames:
            continue
        sheet = wb[sheet_name]
        rest_col = find_rest_column(sheet)
        if not rest_col:
            continue
        day_hours: list[tuple[int, int]] = []
        for row in range(10, 34):
            slot = str(sheet.cell(row=row, column=2).value or "").strip()
            rest_value = sheet.cell(row=row, column=rest_col).value
            if person_no not in split_numbers(rest_value):
                continue
            parsed = parse_slot(slot)
            if parsed:
                day_hours.append(parsed)
        hours_by_day[day] = day_hours
    return group_rest_entries(hours_by_day, days)


def workbook_date_info(wb: openpyxl.Workbook) -> tuple[int, int, int]:
    sheet = wb.worksheets[0]
    year = int(sheet.cell(row=2, column=4).value)
    month = int(sheet.cell(row=2, column=5).value)
    days = int(sheet.cell(row=2, column=7).value or 31)
    return year, month, days


def workbook_person_no(wb: openpyxl.Workbook, target_name: str | None) -> str | None:
    if not target_name:
        return None
    roster = next((sheet for sheet in wb.worksheets if "輪休" in sheet.title), None)
    if roster is None:
        return None
    for col in range(1, roster.max_column + 1):
        name = str(roster.cell(row=5, column=col).value or "").strip()
        if name == target_name.strip():
            number = roster.cell(row=4, column=col).value
            return str(int(number)) if isinstance(number, (int, float)) else str(number).strip()
    return None


def workbook_person_name(workbook_path: Path, target_no: str | None) -> str:
    target_no = str(target_no or "").strip()
    if not target_no:
        return ""
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    roster = next((sheet for sheet in wb.worksheets if "輪休" in sheet.title), None)
    if roster is None:
        return ""
    for col in range(1, roster.max_column + 1):
        number = roster.cell(row=4, column=col).value
        number_text = str(int(number)) if isinstance(number, (int, float)) else str(number or "").strip()
        if number_text == target_no:
            return str(roster.cell(row=5, column=col).value or "").strip()
    return ""


def find_rest_column(sheet: openpyxl.worksheet.worksheet.Worksheet) -> int | None:
    for row in (5, 6, 7):
        for col in range(1, sheet.max_column + 1):
            if "休息" in str(sheet.cell(row=row, column=col).value or ""):
                return col
    return None


def split_numbers(value: object) -> set[str]:
    return {str(int(part)) for part in re.findall(r"\d+", str(value or ""))}


def parse_slot(slot: str) -> tuple[int, int] | None:
    match = re.search(r"(\d{1,2})\s*[-~～]\s*(\d{1,2})", slot)
    if not match:
        return None
    return int(match.group(1)), int(match.group(2))


def group_rest_entries(hours_by_day: dict[int, list[tuple[int, int]]], days: int) -> list[RestEntry]:
    entries: list[RestEntry] = []
    for duty_day in range(1, days + 1):
        slots = sorted(hours_by_day.get(duty_day, []), key=lambda item: duty_relative_hour(item[0]))
        if not slots:
            continue
        start, end = slots[0]
        for next_start, next_end in slots[1:]:
            if duty_relative_hour(next_start) == duty_relative_hour(end):
                end = next_end
            else:
                entries.append(build_entry(duty_day, start, end))
                start, end = next_start, next_end
        entries.append(build_entry(duty_day, start, end))
    return entries


def duty_relative_hour(hour: int) -> int:
    return hour if hour >= 8 else hour + 24


def build_entry(duty_day: int, start_hour: int, end_hour: int) -> RestEntry:
    start_day = duty_day + 1 if start_hour < 8 else duty_day
    end_day = duty_day + 1 if end_hour <= 8 else duty_day
    return RestEntry(duty_day, start_day, start_hour, end_day, end_hour)


def fetch_monthly_base_plan(actor_no: str) -> MonthlyBasePlan:
    csv_text = download_monthly_base_csv()
    rows = list(csv.reader(io.StringIO(csv_text)))
    if len(rows) < 4:
        raise RuntimeError("輪休基準表內容不足，無法判斷月份與人員資料。")
    title = str(rows[0][0] if rows[0] else "").strip()
    month_match = re.search(r"(\d{2,3})年\s*(\d{1,2})月份", title)
    if not month_match:
        raise RuntimeError(f"輪休基準表標題無法判斷年月：{title}")
    roc_year = int(month_match.group(1))
    month = int(month_match.group(2))
    actor_row = rows[1]
    name_row = rows[2]
    actor_no = str(actor_no or "").strip()
    column_index = None
    for index in range(2, max(len(actor_row), len(name_row))):
        current_actor = str(actor_row[index] if index < len(actor_row) else "").strip()
        if current_actor == actor_no:
            column_index = index
            break
    if column_index is None:
        raise RuntimeError(f"輪休基準表找不到 {actor_no} 番。")
    person_name = str(name_row[column_index] if column_index < len(name_row) else "").strip()
    if not person_name:
        raise RuntimeError(f"輪休基準表的 {actor_no} 番沒有姓名。")
    day_symbols: dict[int, str] = {}
    for row in rows[3:]:
        day_text = str(row[0] if row else "").strip()
        if not day_text.isdigit():
            break
        day = int(day_text)
        raw_code = str(row[column_index] if column_index < len(row) else "").strip()
        day_symbols[day] = translate_monthly_base_code(raw_code, day, person_name)
    return MonthlyBasePlan(roc_year=roc_year, month=month, actor_no=actor_no, name=person_name, day_symbols=day_symbols)


def download_monthly_base_csv() -> str:
    query = urlencode({"tqx": "out:csv", "gid": MONTHLY_BASE_SHEET_GID})
    url = MONTHLY_BASE_EXPORT_URL.format(sheet_id=MONTHLY_BASE_SHEET_ID, query=query)
    try:
        with urlopen(url, timeout=20) as response:
            return response.read().decode("utf-8-sig")
    except HTTPError as exc:
        if exc.code == 401:
            raise RuntimeError("固定 Google 試算表目前不允許程式直接讀取（401 Unauthorized）。請先改成可直接檢視/匯出，或改提供本機同步來源。") from exc
        raise RuntimeError(f"讀取固定 Google 試算表失敗：HTTP {exc.code}") from exc
    except URLError as exc:
        raise RuntimeError(f"讀取固定 Google 試算表失敗：{exc.reason}") from exc


def translate_monthly_base_code(code: str, day: int, person_name: str) -> str:
    normalized = str(code or "").strip()
    if normalized not in MONTHLY_BASE_SYMBOLS:
        raise RuntimeError(f"輪休基準表 {person_name} 第 {day} 日出現未支援代碼：{normalized}")
    return MONTHLY_BASE_SYMBOLS[normalized]


def select_base_month(driver, roc_year: int, month: int) -> None:
    result = driver.execute_script(
        """
        function deepFindControls(win) {
          try {
            const year = win.document.querySelector('select[name="_selYEAR"]');
            const month = win.document.querySelector('select[name="_selMONTH"]');
            const query = win.document.querySelector('input[name="_btnQuery"]');
            if (year && month && query) return {win, year, month, query};
          } catch (_error) {}
          try {
            for (let i = 0; i < win.frames.length; i += 1) {
              const found = deepFindControls(win.frames[i]);
              if (found) return found;
            }
          } catch (_error) {}
          return null;
        }
        function setSelectByText(select, wantedText) {
          const wanted = String(wantedText).trim();
          for (const option of select.options) {
            const text = (option.text || '').trim();
            const value = String(option.value || '').trim();
            if (text === wanted || value === wanted) {
              select.value = option.value;
              select.dispatchEvent(new Event('change', {bubbles: true}));
              return true;
            }
          }
          return false;
        }
        const controls = deepFindControls(window.top || window);
        if (!controls) return {ok: false};
        const okYear = setSelectByText(controls.year, arguments[0]);
        const okMonth = setSelectByText(controls.month, arguments[1]);
        controls.query.click();
        return {ok: true, okYear, okMonth};
        """,
        str(roc_year),
        f"{month:02d}",
    )
    if not result or not result.get("ok"):
        raise RuntimeError("勤務基準表找不到年月與查詢欄位。")
    if not result.get("okYear") or not result.get("okMonth"):
        raise RuntimeError(f"勤務基準表無法切換到 {roc_year}年{month:02d}月。")
    expected_pattern = re.compile(rf"目前編輯月份為:\s*{roc_year}年{month:02d}月")
    WebDriverWait(driver, 20).until(lambda d: expected_pattern.search(current_page_text(d)))
    time.sleep(0.8)


def fill_monthly_base_row(driver, person_name: str, day_symbols: dict[int, str], days: int) -> int:
    payload = {str(day): day_symbols.get(day, "") for day in range(1, days + 1)}
    result = driver.execute_script(
        """
        function editableControls(row) {
          return Array.from(row.querySelectorAll('input:not([type=button]):not([type=hidden]), textarea, select'));
        }
        function personRows(win, personName) {
          const rows = Array.from(win.document.querySelectorAll('tr')).filter((row) => {
            const text = (row.innerText || '').trim();
            if (!text.includes(personName)) return false;
            const hasDayCell = Array.from(row.querySelectorAll('input, textarea'))
              .some((el) => /^_pln_\\d+_1$/.test(el.id || el.name || ''));
            const hasSave = Array.from(row.querySelectorAll('input[type=button], button'))
              .some((btn) => ((btn.value || btn.innerText || '').trim() === 'O'));
            return hasDayCell && hasSave;
          });
          rows.sort((a, b) => a.querySelectorAll('input, textarea, select, button').length - b.querySelectorAll('input, textarea, select, button').length);
          return rows;
        }
        function findDayControl(row, day) {
          const candidates = [
            `[name$="_${day}"]`,
            `[name$="_${String(day).padStart(2, '0')}"]`,
            `#_pln_${day}`,
            `#_pln_${String(day).padStart(2, '0')}`,
            `[id$="_${day}"]`,
            `[id$="_${String(day).padStart(2, '0')}"]`,
          ];
          for (const selector of candidates) {
            const el = row.querySelector(selector);
            if (el) return el;
          }
          const controls = editableControls(row);
          if (controls.length >= Number(day)) return controls[Number(day) - 1];
          return null;
        }
        function setRowValues(row, data) {
          let count = 0;
          const missing = [];
          for (const day in data) {
            const el = findDayControl(row, day);
            if (!el) {
              missing.push(day);
              continue;
            }
            try {
              el.focus();
              el.value = data[day];
              el.dispatchEvent(new Event('input', {bubbles: true}));
              el.dispatchEvent(new Event('change', {bubbles: true}));
              el.dispatchEvent(new Event('blur', {bubbles: true}));
              try { if (typeof el.onchange === 'function') el.onchange({target: el, type: 'change'}); } catch (_error) {}
              count += 1;
            } catch (_error) {
              missing.push(day);
            }
          }
          return {count, missing};
        }
        function deepFillByName(win, personName, data) {
          try {
            for (const row of personRows(win, personName)) {
              return setRowValues(row, data);
            }
          } catch (_error) {}
          try {
            for (let i = 0; i < win.frames.length; i += 1) {
              const child = deepFillByName(win.frames[i], personName, data);
              if (child) return child;
            }
          } catch (_error) {}
          return null;
        }
        return deepFillByName(window.top || window, arguments[0], arguments[1]);
        """,
        person_name,
        payload,
    )
    if not result:
        raise RuntimeError(f"勤務基準表找不到 {person_name} 的姓名列。")
    missing = [item for item in (result or {}).get("missing", []) if item]
    if missing:
        preview = ", ".join(missing[:5])
        raise RuntimeError(f"勤務基準表找不到 {person_name} 的日期欄位：{preview}")
    return sum(1 for value in day_symbols.values() if value)


def wait_for_person_name_row(driver, person_name: str) -> None:
    WebDriverWait(driver, 20).until(
        lambda d: d.execute_script(
            """
            function deepHasName(win, personName) {
              try {
                const rows = Array.from(win.document.querySelectorAll('tr')).filter((row) => {
                  const text = (row.innerText || '').trim();
                  if (!text.includes(personName)) return false;
                  const hasDayCell = Array.from(row.querySelectorAll('input, textarea'))
                    .some((el) => /^_pln_\\d+_1$/.test(el.id || el.name || ''));
                  const hasSave = Array.from(row.querySelectorAll('input[type=button], button'))
                    .some((btn) => ((btn.value || btn.innerText || '').trim() === 'O'));
                  return hasDayCell && hasSave;
                });
                if (rows.length) return true;
              } catch (_error) {}
              try {
                for (let i = 0; i < win.frames.length; i += 1) {
                  if (deepHasName(win.frames[i], personName)) return true;
                }
              } catch (_error) {}
              return false;
            }
            return deepHasName(window.top || window, arguments[0]);
            """,
            person_name,
        )
    )
    time.sleep(0.6)


def click_person_row_save(driver, person_name: str) -> None:
    clicked = driver.execute_script(
        """
        function deepClickSave(win, personName) {
          try {
            const rows = Array.from(win.document.querySelectorAll('tr')).filter((row) => {
              const text = (row.innerText || '').trim();
              if (!text.includes(personName)) return false;
              const hasDayCell = Array.from(row.querySelectorAll('input, textarea'))
                .some((el) => /^_pln_\\d+_1$/.test(el.id || el.name || ''));
              return hasDayCell;
            });
            rows.sort((a, b) => a.querySelectorAll('input, textarea, select, button').length - b.querySelectorAll('input, textarea, select, button').length);
            for (const row of rows) {
              const buttons = Array.from(row.querySelectorAll('input[type=button], button'));
              const target = buttons.find((btn) => ((btn.value || btn.innerText || '').trim() === 'O'));
              if (!target) return false;
              target.click();
              return true;
            }
          } catch (_error) {}
          try {
            for (let i = 0; i < win.frames.length; i += 1) {
              if (deepClickSave(win.frames[i], personName)) return true;
            }
          } catch (_error) {}
          return false;
        }
        return deepClickSave(window.top || window, arguments[0]);
        """,
        person_name,
    )
    if not clicked:
        raise RuntimeError(f"找不到 {person_name} 那一列的個人儲存按鈕 O。")
    wait_for_alert_if_any(driver)
    time.sleep(1)


def wait_for_main_table(driver) -> None:
    WebDriverWait(driver, 20).until(lambda d: d.execute_script("return document.body && document.body.innerText.includes('勤務基準表');"))


def current_page_text(driver) -> str:
    try:
        return driver.execute_script(
            """
            const body = document.body ? document.body.innerText : '';
            const frames = Array.from(document.querySelectorAll('frame,iframe'))
              .map(frame => {
                try {
                  return frame.contentWindow && frame.contentWindow.document && frame.contentWindow.document.body
                    ? frame.contentWindow.document.body.innerText
                    : '';
                } catch (_error) {
                  return '';
                }
              })
              .filter(Boolean);
            return [body, ...frames].join('\\n');
            """
        ) or ""
    except Exception:
        return ""


def current_login_name(driver) -> str:
    page_text = current_page_text(driver)
    greeting_match = re.search(r"([^\s,，]+)\s*[,，]\s*您好", page_text)
    if greeting_match:
        return greeting_match.group(1).strip()
    return ""


def find_person_link(driver, user_id: str, target_no: str = "", target_name: str = "") -> PersonLink:
    links = driver.execute_script(
        """
        return Array.from(document.querySelectorAll('a._name')).map((a, index) => ({
          index,
          text: (a.innerText || '').trim(),
          onclick: a.getAttribute('onclick') || '',
          rowText: (a.closest('tr') || document).innerText || '',
          saveId: Array.from((a.closest('tr') || document).querySelectorAll('input[type=button],button'))
            .map(btn => btn.id || '')
            .find(id => id.startsWith('_btnPersonalSave_')) || ''
        }));
        """
    )
    login_name = current_login_name(driver)
    target_no = str(target_no or "").strip()
    target_name = str(target_name or "").strip()

    def link_person(link: dict[str, object]) -> PersonLink | None:
        match = re.search(r"_btnPersonalSave_(\d+)", str(link.get("saveId", "")))
        staff_no = match.group(1) if match else ""
        if not staff_no:
            return None
        return PersonLink(name=str(link.get("text", "") or ""), staff_no=staff_no, element_index=int(link["index"]))

    if target_name:
        matched_links = [link for link in links if str(link.get("text", "")).strip() == target_name]
        if len(matched_links) == 1:
            person = link_person(matched_links[0])
            if person:
                return person
        row_matches = [
            link
            for link in links
            if target_name in str(link.get("rowText", "")) and (not target_no or re.search(rf"(^|\D){re.escape(target_no)}(\D|$)", str(link.get("rowText", ""))))
        ]
        if len(row_matches) == 1:
            person = link_person(row_matches[0])
            if person:
                return person
    for link in links:
        onclick = link.get("onclick", "")
        if user_id and user_id in onclick:
            person = link_person(link)
            if not person:
                raise RuntimeError("找到本人姓名連結，但無法判斷該列的個人儲存按鈕。")
            return person
    if login_name:
        matched_links = [link for link in links if str(link.get("text", "")).strip() == login_name]
        if len(matched_links) == 1:
            person = link_person(matched_links[0])
            if not person:
                raise RuntimeError(f"已找到姓名 {login_name}，但無法判斷該列的個人儲存按鈕。")
            return person
    if len(links) == 1:
        person = link_person(links[0])
        if person:
            return person
    available = "、".join(str(link.get("text", "")).strip() for link in links[:8] if str(link.get("text", "")).strip())
    if target_name or target_no:
        raise RuntimeError(f"勤務基準表找不到番號 {target_no or '-'} / {target_name or '-'} 的個人姓名連結；頁面姓名：{available}")
    if login_name:
        raise RuntimeError(f"勤務基準表找不到帳號 {user_id} 對應的個人姓名連結；登入姓名為 {login_name}；頁面姓名：{available}")
    raise RuntimeError(f"勤務基準表找不到帳號 {user_id} 對應的個人姓名連結；頁面姓名：{available}")


def open_person_popup(driver, person: PersonLink) -> None:
    before = set(driver.window_handles)
    clicked = driver.execute_script(
        """
        const links = Array.from(document.querySelectorAll('a._name'));
        const link = links[arguments[0]];
        if (!link) return false;
        link.click();
        return true;
        """,
        person.element_index,
    )
    if not clicked:
        raise RuntimeError(f"無法點開 {person.name} 的個人連結。")
    WebDriverWait(driver, 20).until(lambda d: len(set(d.window_handles) - before) >= 1)
    popup = list(set(driver.window_handles) - before)[0]
    driver.switch_to.window(popup)
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "_selOFFTYPE")))


def existing_entry_match(driver, entry: RestEntry) -> str | None:
    return driver.execute_script(
        """
        const expected = arguments[0];
        const normalize = (value) => String(value || '').replace(/\\s+/g, ' ').trim();
        const numericEquals = (token, value) => {
          const wanted = String(Number(value));
          return String(token).match(/\\d+/g)?.some(part => String(Number(part)) === wanted) || false;
        };
        const textWithoutFormOptions = (node) => {
          const clone = node.cloneNode(true);
          clone.querySelectorAll('select, option, input, textarea, button, script, style').forEach(el => el.remove());
          return normalize(clone.textContent || '');
        };
        const controlValues = (node) => {
          const values = [];
          node.querySelectorAll('select').forEach((select) => {
            const option = select.options[select.selectedIndex];
            values.push(normalize((option && option.text) || select.value || ''));
          });
          node.querySelectorAll('input, textarea').forEach((input) => {
            const type = (input.type || '').toLowerCase();
            if (['button', 'submit', 'reset', 'hidden'].includes(type)) return;
            values.push(normalize(input.value || ''));
          });
          return values.filter(Boolean);
        };
        const cellTokens = (cell) => [textWithoutFormOptions(cell), ...controlValues(cell)].filter(Boolean);
        const rows = Array.from(document.querySelectorAll('tr'));
        for (const row of rows) {
          if (row.querySelector('#_btnInsert')) continue;
          if (row.querySelector('#_selOFFTYPE, #_selLEAVETYPE, #_selTASKDATE, #_selDATES, #_selTASKHOURS, #_selDATEE, #_selTASKHOURE, #_txtSumHour')) continue;
          const cells = Array.from(row.children);
          const cellTexts = cells.map(cell => cellTokens(cell).join(' ')).filter(Boolean);
          if (cellTexts.length >= 6) {
            for (let index = 0; index <= cellTexts.length - 6; index += 1) {
              const [unit, leaveType, day, startHour, endHour, hours] = cellTexts.slice(index, index + 6);
              const matched =
                unit.includes('小時假') &&
                leaveType === expected.leaveType &&
                numericEquals(day, expected.dutyDay) &&
                numericEquals(startHour, expected.startHour) &&
                numericEquals(endHour, expected.endHour) &&
                numericEquals(hours, expected.hours);
              if (matched) return cellTexts.slice(index, index + 6).join(' | ');
            }
            continue;
          }
          const tokens = (cellTexts.length ? cellTexts : [textWithoutFormOptions(row), ...controlValues(row)]).filter(Boolean);
          const compact = tokens.join(' ');
          if (!compact || compact.length > 500) continue;
          const hasRest = tokens.some(token => token === expected.leaveType || token.includes(expected.leaveType));
          if (!hasRest) continue;
          const matched =
            tokens.some(token => numericEquals(token, expected.dutyDay)) &&
            tokens.some(token => numericEquals(token, expected.startHour)) &&
            tokens.some(token => numericEquals(token, expected.endHour)) &&
            tokens.some(token => numericEquals(token, expected.hours));
          if (matched) return compact.slice(0, 120);
        }
        return null;
        """,
        {
            "leaveType": "休息",
            "dutyDay": f"{entry.duty_day:02d}",
            "startDay": f"{entry.start_day:02d}",
            "startHour": f"{entry.start_hour:02d}",
            "endDay": f"{entry.end_day:02d}",
            "endHour": f"{entry.end_hour:02d}",
            "hours": str(entry.hours),
        },
    )


def mark_duplicate_rest_delete(driver) -> str | None:
    return driver.execute_script(
        """
        const normalize = (value) => String(value || '').replace(/\\s+/g, ' ').trim();
        const textWithoutFormOptions = (node) => {
          const clone = node.cloneNode(true);
          clone.querySelectorAll('select, option, input, textarea, button, script, style').forEach(el => el.remove());
          return normalize(clone.textContent || '');
        };
        const cellText = (cell) => textWithoutFormOptions(cell);
        document.querySelectorAll('[data-rest-duplicate-delete]').forEach(el => el.removeAttribute('data-rest-duplicate-delete'));
        const rows = Array.from(document.querySelectorAll('tr'));
        const seen = new Set();
        for (const row of rows) {
          if (row.querySelector('#_btnInsert')) continue;
          if (row.querySelector('#_selOFFTYPE, #_selLEAVETYPE, #_selTASKDATE, #_selDATES, #_selTASKHOURS, #_selDATEE, #_selTASKHOURE, #_txtSumHour')) continue;
          const cells = Array.from(row.children);
          const cellTexts = cells.map(cellText).filter(Boolean);
          if (cellTexts.length < 6) continue;
          for (let index = 0; index <= cellTexts.length - 6; index += 1) {
            const [unit, leaveType, day, startHour, endHour, hours] = cellTexts.slice(index, index + 6);
            if (!unit.includes('小時假') || leaveType !== '休息') continue;
            const key = [Number(day), Number(startHour), Number(endHour), Number(hours)].join('|');
            if (!seen.has(key)) {
              seen.add(key);
              break;
            }
            const buttons = Array.from(row.querySelectorAll('input[type=button], button'));
            const deleteButton = buttons.find((button) => normalize(button.value || button.innerText || '').includes('刪除'));
            if (!deleteButton) return `找不到刪除按鈕：${cellTexts.slice(index, index + 6).join(' | ')}`;
            deleteButton.setAttribute('data-rest-duplicate-delete', '1');
            return cellTexts.slice(index, index + 6).join(' | ');
          }
        }
        return null;
        """
    )


def fill_and_insert_entry(driver, entry: RestEntry) -> None:
    set_form_values(driver, entry)
    if not js_click(driver, "_btnInsert"):
        raise RuntimeError("找不到新增按鈕。")
    wait_for_alert_if_any(driver)
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.ID, "_selOFFTYPE")))


def set_form_values(driver, entry: RestEntry) -> None:
    result = driver.execute_script(
        """
        function setByTextOrValue(id, value) {
          const el = document.getElementById(id);
          if (!el) return false;
          const want = String(value).padStart(2, '0');
          for (const option of el.options) {
            const text = (option.text || '').trim();
            if (option.value == value || option.value == want || text == value || text == want || text.includes(want)) {
              el.value = option.value;
              el.dispatchEvent(new Event('change', {bubbles: true}));
              return true;
            }
          }
          return false;
        }
        function setByOptionText(id, text) {
          const el = document.getElementById(id);
          if (!el) return false;
          for (const option of el.options) {
            if ((option.text || '').trim() == text) {
              el.value = option.value;
              el.dispatchEvent(new Event('change', {bubbles: true}));
              return true;
            }
          }
          return false;
        }
        const missing = [];
        if (!setByOptionText('_selOFFTYPE', '小時假')) missing.push('假別單位');
        if (!setByOptionText('_selLEAVETYPE', '休息')) missing.push('項目');
        if (!setByTextOrValue('_selTASKDATE', arguments[0])) missing.push('勤務日期');
        return missing;
        """,
        f"{entry.duty_day:02d}",
    )
    if result:
        raise RuntimeError(f"小視窗欄位設定失敗：{', '.join(result)}")

    WebDriverWait(driver, 10).until(
        lambda d: d.execute_script(
            """
            const ids = ['_selDATES', '_selTASKHOURS', '_selDATEE', '_selTASKHOURE'];
            return ids.every((id) => {
              const el = document.getElementById(id);
              return el && !el.disabled && el.options && el.options.length > 0;
            });
            """
        )
    )
    time.sleep(0.4)

    result = driver.execute_script(
        """
        function setByTextOrValue(id, value) {
          const el = document.getElementById(id);
          if (!el) return false;
          const want = String(value).padStart(2, '0');
          for (const option of el.options) {
            const text = (option.text || '').trim();
            if (option.value == value || option.value == want || text == value || text == want || text.includes(want)) {
              el.value = option.value;
              el.dispatchEvent(new Event('change', {bubbles: true}));
              return true;
            }
          }
          return false;
        }
        const missing = [];
        if (!setByTextOrValue('_selDATES', arguments[0])) missing.push('起始日期');
        if (!setByTextOrValue('_selTASKHOURS', arguments[1])) missing.push('起始小時');
        if (!setByTextOrValue('_selDATEE', arguments[2])) missing.push('結束日期');
        if (!setByTextOrValue('_selTASKHOURE', arguments[3])) missing.push('結束小時');
        const sum = document.getElementById('_txtSumHour');
        if (sum) {
          sum.value = arguments[4];
          sum.dispatchEvent(new Event('change', {bubbles: true}));
        } else {
          missing.push('時數');
        }
        return missing;
        """,
        f"{entry.start_day:02d}",
        f"{entry.start_hour:02d}",
        f"{entry.end_day:02d}",
        f"{entry.end_hour:02d}",
        str(entry.hours),
    )
    if result:
        raise RuntimeError(f"小視窗欄位設定失敗：{', '.join(result)}")


def wait_for_alert_if_any(driver) -> None:
    try:
        alert = WebDriverWait(driver, 2).until(EC.alert_is_present())
        alert.accept()
    except TimeoutException:
        pass


def close_current_popup(driver) -> None:
    popup = driver.current_window_handle
    driver.close()
    for handle in driver.window_handles:
        if handle != popup:
            driver.switch_to.window(handle)
            return


def click_person_save(driver, staff_no: str) -> None:
    wait_for_main_table(driver)
    save_id = f"_btnPersonalSave_{staff_no}"
    if not js_click(driver, save_id):
        clicked = driver.execute_script(
            """
            const buttons = Array.from(document.querySelectorAll('input[type=button],button'));
            const target = buttons.find(btn => (btn.id || '').includes(arguments[0]) || (btn.value || btn.innerText || '').trim() === 'O');
            if (!target) return false;
            target.click();
            return true;
            """,
            save_id,
        )
        if not clicked:
            raise RuntimeError("找不到個人儲存按鈕 O。")
    wait_for_alert_if_any(driver)
    time.sleep(1)
